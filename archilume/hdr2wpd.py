"""HDR to Working Plan Data (WPD) Converter.

This module converts High Dynamic Range (HDR) images to Working Plan Data (WPD) files
by extracting illuminance data, filtering points within Areas of Interest (AOI),
and generating compliance reports.
"""

# Archilume imports
from archilume import (
    utils, 
    config
    )

# Standard library imports
import subprocess
import sys
import math
import numpy as np
from matplotlib.path import Path as MplPath
from pathlib import Path
import pandas as pd
from concurrent.futures import ProcessPoolExecutor, as_completed
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Tuple, Optional, Dict
import logging
import os
from skimage.draw import polygon

# Third-party imports
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import ColorScaleRule

logger = logging.getLogger(__name__)

@dataclass
class ViewGroupProcessor:
    """
    Optimized Processor. 
    Uses pre-computed boolean masks and bitwise operations for maximum speed.
    """
    view_name: str
    aoi_files: List[Path]
    hdr_files: List[Path]
    pixel_threshold_value: float
    wpd_output_dir: Path

    # Cache for boolean masks: {Path: np.ndarray(bool)}
    _aoi_masks: Dict[Path, np.ndarray] = field(default_factory=dict)
    # Cache for pixel counts: {Path: int}
    _aoi_pixel_counts: Dict[Path, int] = field(default_factory=dict)

    def __post_init__(self):
        """Pre-rasterize all AOI polygons into boolean masks."""
        if not self.hdr_files or not self.aoi_files:
            return

        # 1. Get resolution from the first HDR file to initialize masks
        # Assumption: All HDRs in a specific view group share the same resolution.
        try:
            self.width, self.height = utils.get_hdr_resolution(self.hdr_files[0])
        except Exception as e:
            print(f"Error getting resolution from {self.hdr_files[0]}: {e}")
            return

        # Silent - only print if there's an error

        # 2. Convert every AOI file into a Boolean Mask immediately
        for aoi_file in self.aoi_files:
            try:
                with open(aoi_file, 'r') as f:
                    lines = f.readlines()

                # Parse polygon coordinates
                # Format expected: ... pixel_x pixel_y ...
                xs = []
                ys = []
                for line in lines[5:]: # Skip header
                    parts = line.strip().split()
                    if len(parts) >= 4:
                        xs.append(int(parts[2]))
                        ys.append(int(parts[3]))

                if xs:
                    # Create a blank boolean mask (False = black)
                    mask = np.zeros((self.height, self.width), dtype=bool)
                    
                    # Efficiently fill the polygon with True
                    # note: polygon takes (row, col) which corresponds to (y, x)
                    rr, cc = polygon(ys, xs, shape=(self.height, self.width))
                    mask[rr, cc] = True

                    # Store the mask and the total pixel count
                    self._aoi_masks[aoi_file] = mask
                    self._aoi_pixel_counts[aoi_file] = np.count_nonzero(mask)
                    
            except Exception as e:
                print(f"Error parsing/masking {aoi_file.name}: {e}")

    def process(self) -> pd.DataFrame:
        """Process HDRs using vectorized bitwise operations."""
        if not self._aoi_masks:
            return pd.DataFrame()

        aoi_results = {aoi: [] for aoi in self.aoi_files if aoi in self._aoi_masks}
        results = []

        total_ops = len(self.hdr_files)
        # Silent - progress will be shown by parent process

        for idx, hdr_file in enumerate(self.hdr_files):
            # 1. READ BINARY DATA (Fastest Method)
            try:
                cmd = ['pvalue', '-h', '-H', '-b', '-df', str(hdr_file)]
                process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                raw_data, _ = process.communicate()

                if process.returncode != 0:
                    continue

                # 2. CONVERT TO NUMPY
                # buffer -> flat array -> 2D array
                data = np.frombuffer(raw_data, dtype=np.float32)
                data_2d = data.reshape((self.height, self.width))

            except (ValueError, IndexError, OSError) as e:
                print(f"  Skipping {hdr_file.name}: {e}")
                continue

            # 3. GLOBAL THRESHOLD (Vectorized)
            # Create a True/False map of the whole image once.
            # "Is the pixel bright enough?"
            bright_pixels_mask = data_2d > self.pixel_threshold_value

            # Clean filename for output
            hdr_name = hdr_file.name
            if 'plan_' in hdr_name:
                hdr_name = 'plan_' + hdr_name.split('plan_')[1]

            # 4. APPLY MASKS (The Optimized Loop)
            # Instead of ray-casting, we just overlay the masks.
            for aoi_file, aoi_mask in self._aoi_masks.items():
                
                # BITWISE AND: Intersection of Room and Bright Pixels
                # This runs at memory bandwidth speed (extremely fast)
                passing_mask = np.logical_and(aoi_mask, bright_pixels_mask)
                
                passing_count = np.count_nonzero(passing_mask)
                total_count = self._aoi_pixel_counts[aoi_file]

                # Store result
                entry = {
                    'hdr_file': hdr_name,
                    'total_pixels': total_count,
                    'passing_pixels': passing_count
                }
                aoi_results[aoi_file].append(entry)
                
                # Add to flat results for DataFrame
                results.append({
                    'aoi_file': aoi_file.name,
                    'hdr_file': hdr_name,
                    'total_pixels': total_count,
                    'passing_pixels': passing_count,
                })

            # Progress updates removed for cleaner output

        # Write physical files
        self._write_wpd_files(aoi_results)
        return pd.DataFrame(results)

    def _write_wpd_files(self, aoi_results: Dict[Path, List[Dict]]) -> None:
        """Write .wpd files (Unchanged logic, just cleaner)."""
        for aoi_file, data_list in aoi_results.items():
            if not data_list:
                continue

            # Sort by filename to ensure temporal order
            data_list.sort(key=lambda x: x['hdr_file'])
            
            # Use cached total pixels
            total_pixels = self._aoi_pixel_counts[aoi_file]
            
            wpd_path = self.wpd_output_dir / (aoi_file.stem + '.wpd')
            
            with open(wpd_path, 'w') as f:
                f.write(f"total_pixels_in_polygon: {total_pixels}\n")
                f.write("hdr_file passing_pixels\n")
                for row in data_list:
                    f.write(f"{row['hdr_file']} {row['passing_pixels']}\n")

@dataclass
class Hdr2Wpd:
    """Converts HDR images to Working Plan Data (WPD) files.

    This class handles:
    - Extracting point data from HDR images using pvalue
    - Filtering non-zero illuminance points
    - Matching points to AOI polygons using ray casting
    - Writing individual .wpd files per timestep
    - Aggregating results into Excel reports
    """

    # Input parameters
    pixel_to_world_map: Path
    pixel_threshold_value: float = 0.0

    def __post_init__(self):
        """Initialize output directories from config."""
        self.aoi_dir = config.AOI_DIR
        self.wpd_dir = config.WPD_DIR
        self.max_workers = config.WORKERS["wpd_processing"]

        self.wpd_dir.mkdir(parents=True, exist_ok=True)

        # Calculate area per pixel and pixel increments from the pixel_to_world_map file
        self.area_per_pixel, self.pixel_increment_x, self.pixel_increment_y = self._calculate_area_per_pixel()

    def sunlight_sequence_wpd_extraction(self) -> None:
        """Execute the complete WPD generation pipeline.

        Phases:
        1. Scan directories for HDR and AOI files
        2. Group files by view
        3. Process view groups in parallel
        4. Generate Excel report from .wpd files
        """
        print("=" * 80 + "\nPARALLEL BATCH PROCESSING (ProcessPoolExecutor)\n" + "=" * 80)

        # Scan directories
        hdr_files_all, aoi_files_all = self._scan_directories()

        if not hdr_files_all or not aoi_files_all:
            print("No files found to process.")
            sys.exit(0)

        # Group files by view
        view_groups = self._group_aoi_by_view(aoi_files_all, hdr_files_all)

        # Process view groups in parallel
        self._process_sunlight_view_groups(view_groups)

        # Generate Excel report
        self._generate_sunlight_excel_report()

        print("\n" + "=" * 80 + "\nSunlight sequence wpd extraction completed successfully.\n" + "=" * 80)

    def daylight_wpd_extraction(self, df_thresholds: List[float] = None) -> None:
        """Extract per-pixel daylight factor (DF%) from single-HDR-per-view renders.

        For each view group (1 HDR + N AOIs):
        1. Read HDR via pvalue → float32 numpy array
        2. Scale by 1.79 (179 luminous efficacy x 100 / 10000 lux) to convert radiometric W/m² to DF%
        3. For each AOI: rasterize polygon, write per-pixel DF% to .wpd
        4. Generate Excel report from .wpd files

        Args:
            df_thresholds: DF% compliance thresholds. For each value, the summary will include
                           the pixel count and area >= that threshold. Defaults to [0.5, 1.0, 2.0].
        """
        if df_thresholds is None:
            df_thresholds = [0.5, 1.0, 2.0]

        print("=" * 80 + "\nDAYLIGHT FACTOR WPD EXTRACTION\n" + "=" * 80)

        hdr_files, aoi_files = self._scan_directories()
        if not hdr_files or not aoi_files:
            print("No files found to process.")
            return

        view_groups = self._group_aoi_by_view(aoi_files, hdr_files)
        self._process_daylight_view_groups(view_groups)
        self._generate_daylight_excel_report(df_thresholds)

        print("\n" + "=" * 80 + "\nDaylight factor wpd extraction completed.\n" + "=" * 80)

    def _process_daylight_view_groups(self, view_groups: Dict) -> None:
        """Process view groups for daylight factor extraction.

        For each view group (expects 1 HDR + N AOIs):
        1. Read HDR via pvalue → float32 numpy array
        2. Convert to illuminance (lux) and DF%
        3. For each AOI: rasterize polygon, write per-pixel DF% to .wpd

        Args:
            view_groups: Dictionary of view groups from _group_aoi_by_view
        """
        for view_name, group in view_groups.items():
            aoi_list, hdr_list = group['aoi_files'], group['hdr_files']
            if not aoi_list or not hdr_list:
                continue

            if len(hdr_list) != 1:
                print(f"  WARNING: View '{view_name}' has {len(hdr_list)} HDRs, expected 1. Using first.")
            hdr_file = hdr_list[0]

            # Read HDR → numpy float32 (height, width)
            width, height = utils.get_hdr_resolution(hdr_file)
            try:
                result = subprocess.run(
                    ['pvalue', '-h', '-H', '-b', '-df', str(hdr_file)],
                    capture_output=True, check=True
                )
                raw_image = np.frombuffer(result.stdout, dtype=np.float32).reshape((height, width))
                illuminance_image = raw_image * 179  # Convert radiometric W/m² to illuminance (lux)
                df_image = illuminance_image * 100 / 10_000  # Convert lux to DF% (against 10K lux sky)
            except Exception as e:
                print(f"  Skipping view '{view_name}': {e}")
                continue

            print(f"\n  View: {view_name} | HDR: {hdr_file.name} | AOIs: {len(aoi_list)}")

            # Per-AOI: rasterize polygon, write per-pixel DF% to .wpd
            for aoi_file in aoi_list:
                with open(aoi_file, 'r') as f:
                    lines = f.readlines()

                xs, ys = [], []
                for line in lines[5:]:
                    parts = line.strip().split()
                    if len(parts) >= 4:
                        xs.append(int(parts[2]))
                        ys.append(int(parts[3]))

                if not xs:
                    print(f"    Skipping {aoi_file.name}: no polygon points")
                    continue

                rr, cc = polygon(ys, xs, shape=(height, width))
                total_pixels = len(rr)

                wpd_path = self.wpd_dir / (aoi_file.stem + '.wpd')
                with open(wpd_path, 'w') as f:
                    f.write(f"total_pixels_in_polygon: {total_pixels}\n")
                    f.write("pixel_x pixel_y illuminance df_percent\n")
                    for i in range(total_pixels):
                        f.write(f"{cc[i]} {rr[i]} {illuminance_image[rr[i], cc[i]]:.4f} {df_image[rr[i], cc[i]]:.4f}\n")

                print(f"    {aoi_file.stem}: {total_pixels} px -> {wpd_path.name}")

    def _generate_daylight_excel_report(self, df_thresholds: List[float] = None) -> None:
        """Generate Excel report from daylight factor .wpd files.

        Reads each .wpd file (format: pixel_x pixel_y illuminance df_percent)
        and produces two sheets:
        - Summary: per-AOI statistics (mean/min/max illuminance and DF%, pixel count, area,
                   and per-threshold compliance columns)
        - Raw Data: all per-pixel values from every .wpd file

        Args:
            df_thresholds: DF% compliance thresholds for summary columns. Defaults to [0.5, 1.0, 2.0].
        """
        if df_thresholds is None:
            df_thresholds = [0.5, 1.0, 2.0]
        print("\n" + "=" * 80 + "\nGENERATING DAYLIGHT EXCEL REPORT\n" + "=" * 80)

        wpd_files = sorted(self.wpd_dir.glob("*.wpd"))
        print(f"\nFound {len(wpd_files)} .wpd files in {self.wpd_dir}")

        if not wpd_files:
            print("No .wpd files found.")
            return

        summary_rows = []
        raw_rows = []

        for wpd_file in wpd_files:
            aoi_name = wpd_file.stem

            with open(wpd_file, 'r') as f:
                lines = f.readlines()

            if len(lines) < 2:
                continue

            # Line 0: "total_pixels_in_polygon: N"
            total_pixels = int(lines[0].split(':')[1].strip())

            # Lines 2+: pixel_x pixel_y illuminance df_percent
            illuminance_vals = []
            df_vals = []
            for line in lines[2:]:
                parts = line.strip().split()
                if len(parts) == 4:
                    px, py = int(parts[0]), int(parts[1])
                    illum, df = float(parts[2]), float(parts[3])
                    illuminance_vals.append(illum)
                    df_vals.append(df)
                    raw_rows.append({
                        'aoi': aoi_name,
                        'pixel_x': px,
                        'pixel_y': py,
                        'illuminance_lux': illum,
                        'df_percent': df,
                    })

            if not illuminance_vals:
                continue

            illum_arr = np.array(illuminance_vals)
            df_arr = np.array(df_vals)
            area_m2 = total_pixels * self.area_per_pixel

            row = {
                'aoi': aoi_name,
                'total_pixels': total_pixels,
                'area_m2': round(area_m2, 4),
                'mean_illuminance_lux': round(float(illum_arr.mean()), 2),
                'min_illuminance_lux': round(float(illum_arr.min()), 2),
                'max_illuminance_lux': round(float(illum_arr.max()), 2),
                'mean_df_percent': round(float(df_arr.mean()), 4),
                'min_df_percent': round(float(df_arr.min()), 4),
                'max_df_percent': round(float(df_arr.max()), 4),
                'median_df_percent': round(float(np.median(df_arr)), 4),
            }
            for t in df_thresholds:
                passing = int((df_arr >= t).sum())
                label = f"{t:g}pct"
                row[f'pixels_df_gte_{label}'] = passing
                row[f'pct_area_df_gte_{label}'] = round(passing / total_pixels * 100, 2) if total_pixels else 0.0
                row[f'area_df_gte_{label}_m2'] = round(passing * self.area_per_pixel, 4)
            summary_rows.append(row)

        if not summary_rows:
            print("No valid data found in .wpd files.")
            return

        summary_df = pd.DataFrame(summary_rows).sort_values('aoi')
        raw_df = pd.DataFrame(raw_rows)

        # Write Excel file
        def write_excel_file(summary_df: pd.DataFrame, raw_df: pd.DataFrame) -> None:
            """Write formatted Excel file with Summary and Raw Data sheets."""

            def format_sheet(ws) -> None:
                """Apply metadata header, hide gridlines, and autofit columns."""
                area_per_pixel_mm2 = self.area_per_pixel * 1_000_000
                pixel_x_mm = round(self.pixel_increment_x * 1_000)
                pixel_y_mm = round(self.pixel_increment_y * 1_000)
                ws['B1'] = (f"{pixel_x_mm} mm × {pixel_y_mm} mm grid | "
                            f"Area per pixel: {self.area_per_pixel} m² ({area_per_pixel_mm2:.2f} mm²) | "
                            f"Source: {self.pixel_to_world_map}")
                ws.sheet_view.showGridLines = False
                for column in ws.columns:
                    max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
                    ws.column_dimensions[column[0].column_letter].width = max_length + 2

            output_excel = self.wpd_dir / "daylight_factor_results.xlsx"
            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=1, startcol=1)
                raw_df.to_excel(writer, sheet_name='Raw Data', index=False, startrow=1, startcol=1)
                format_sheet(writer.sheets['Summary'])
                format_sheet(writer.sheets['Raw Data'])

            print(f"\nResults saved to: {output_excel}")
            print(f"  Sheet 1: 'Summary'  - per-AOI illuminance & DF% statistics")
            print(f"  Sheet 2: 'Raw Data' - all per-pixel values")
            print(f"\n{summary_df.to_string(index=False)}")

        write_excel_file(summary_df, raw_df)

    def _calculate_area_per_pixel(self) -> tuple[float, float, float]:
        """Calculate the area represented by a single pixel from the pixel_to_world_map file.

        Reads the header metadata from the pixel_to_world_map file to extract:
        - Image dimensions in pixels (width, height)
        - World dimensions in meters (width, height)

        Then calculates:
        - pixel_width = world_width / image_width
        - pixel_height = world_height / image_height
        - area_per_pixel = pixel_width × pixel_height

        Returns:
            Tuple of (area_per_pixel, pixel_increment_x, pixel_increment_y) in m² and mm²
        """
        with open(self.pixel_to_world_map, 'r') as f:
            lines = f.readlines()

            # Parse image dimensions from line 2
            # Format: # Image dimensions in pixels: width=2048, height=1778
            image_dims_line = lines[1].strip()
            image_width = int(image_dims_line.split('width=')[1].split(',')[0])
            image_height = int(image_dims_line.split('height=')[1])

            # Parse world dimensions from line 3
            # Format: # World dimensions in meters: width=29.480000, height=25.590000
            world_dims_line = lines[2].strip()
            world_width = float(world_dims_line.split('width=')[1].split(',')[0])
            world_height = float(world_dims_line.split('height=')[1])

        # Calculate pixel spacing in each direction
        pixel_width_meters = world_width / image_width
        pixel_height_meters = world_height / image_height

        # Calculate area per pixel
        area_per_pixel = pixel_width_meters * pixel_height_meters

        # Round to 6 decimal places to preserve sub-mm² precision (e.g. 0.000196 not 0.0002)
        area_per_pixel_rounded = round(area_per_pixel, 6)

        # Convert to mm (1 m = 1,000 mm) and mm² (1 m² = 1,000,000 mm²)
        pixel_width_mm = round(pixel_width_meters * 1_000)
        pixel_height_mm = round(pixel_height_meters * 1_000)
        area_per_pixel_mm2 = area_per_pixel * 1_000_000

        print(f"\nPixel-to-World Mapping:")
        print(f"  Image dimensions: {image_width} x {image_height} pixels")
        print(f"  World dimensions: {world_width} x {world_height} meters")
        print(f"  Pixel increment X: {pixel_width_meters:.6f} m ({pixel_width_mm} mm)")
        print(f"  Pixel increment Y: {pixel_height_meters:.6f} m ({pixel_height_mm} mm)")
        print(f"  Area per pixel (unrounded): {area_per_pixel:.10f} m²")
        print(f"  Area per pixel (rounded): {area_per_pixel_rounded} m² ({area_per_pixel_mm2:.2f} mm²)")
        print(f"  Verification: {image_width} × {image_height} = {image_width * image_height} pixels")
        print(f"  Total world area: {world_width * world_height:.4f} m²")
        print(f"  Calculated area: {area_per_pixel_rounded * image_width * image_height:.4f} m²")
        print(f"  Source file: {self.pixel_to_world_map}\n")

        return area_per_pixel_rounded, pixel_width_meters, pixel_height_meters

    def _scan_directories(self) -> Tuple[List[Path], List[Path]]:
        """Scan directories for HDR and AOI files.

        Returns:
            Tuple of (hdr_files, aoi_files)
        """
        print("Scanning directories...")
        hdr_files_all = sorted(config.IMAGE_DIR.glob("*.hdr"))
        aoi_files_all = sorted(self.aoi_dir.glob("*.aoi"))

        print(f"Found {len(hdr_files_all)} HDR files")
        print(f"Found {len(aoi_files_all)} AOI files\n")

        return hdr_files_all, aoi_files_all

    def _group_aoi_by_view(self, aoi_files: List[Path], hdr_files: List[Path]) -> Dict:
        """Group AOI files and HDR files by their associated view identifier.

        Skips AOI files that already have .wpd files.

        Args:
            aoi_files: List of AOI file paths to group
            hdr_files: List of HDR file paths to match against groups

        Returns:
            Dictionary mapping view names to {'aoi_files': [...], 'hdr_files': [...]}
        """
        def _get_associated_view_file(aoi_file: Path) -> Optional[str]:
            with open(aoi_file, 'r') as f:
                lines = f.readlines()
                if len(lines) >= 2:
                    view_line = lines[1].strip()
                    if 'ASSOCIATED VIEW FILE:' in view_line:
                        view_file = view_line.split('ASSOCIATED VIEW FILE:')[1].strip()
                        return view_file.replace('.vp', '')
            return None

        groups = defaultdict(lambda: {'aoi_files': [], 'hdr_files': []})

        # Group AOI files by view, skipping existing .wpd files
        print("Grouping AOI files by associated view...")
        skipped_count = 0
        warnings = []

        for aoi_file in aoi_files:
            # Check if .wpd file already exists
            wpd_file_path = self.wpd_dir / f"{aoi_file.stem}.wpd"
            if wpd_file_path.exists():
                skipped_count += 1
                continue

            view_name = _get_associated_view_file(aoi_file)
            if view_name:
                groups[view_name]['aoi_files'].append(aoi_file)
            else:
                warnings.append(f"Could not extract view file from {aoi_file.name}")

        # Print AOI grouping summary
        total_grouped = sum(len(group['aoi_files']) for group in groups.values())
        print(f"Grouped {total_grouped} AOI files across {len(groups)} views" +
              (f" (skipped {skipped_count} with existing .wpd files)" if skipped_count > 0 else ""))

        # Match HDR files to groups
        print("Matching HDR files to view groups...")
        unmatched_hdrs = []

        for hdr_file in hdr_files:
            matched = False
            for view_name in groups.keys():
                if view_name in hdr_file.name:
                    groups[view_name]['hdr_files'].append(hdr_file)
                    matched = True
                    break

            if not matched:
                unmatched_hdrs.append(hdr_file.name)

        # Print HDR matching summary
        total_matched = sum(len(group['hdr_files']) for group in groups.values())
        print(f"Matched {total_matched} HDR files to {len(groups)} view groups")

        # Print warnings if any
        if warnings:
            print(f"\nWarnings ({len(warnings)}):")
            for warning in warnings[:3]:  # Show first 3 warnings
                print(f"  - {warning}")
            if len(warnings) > 3:
                print(f"  ... and {len(warnings) - 3} more")

        if unmatched_hdrs:
            print(f"\nUnmatched HDR files ({len(unmatched_hdrs)}):")
            for hdr in unmatched_hdrs[:3]:  # Show first 3 unmatched
                print(f"  - {hdr}")
            if len(unmatched_hdrs) > 3:
                print(f"  ... and {len(unmatched_hdrs) - 3} more")

        # Print summary
        print("\n" + "=" * 80 + "\nGROUPING SUMMARY\n" + "=" * 80)
        for view_name, group in sorted(groups.items()):
            print(f"{view_name}:")
            print(f"  AOI files: {len(group['aoi_files'])}")
            print(f"  HDR files: {len(group['hdr_files'])}")
            print(f"  Total operations: {len(group['aoi_files'])} × {len(group['hdr_files'])} = "
                  f"{len(group['aoi_files']) * len(group['hdr_files'])}")

        return dict(groups)

    def _process_sunlight_view_groups(self, view_groups: Dict) -> None:
        """Process view groups in parallel with dynamic AOI chunking.

        Dynamically splits AOI files into chunks to maximize CPU utilization.
        Each chunk is processed by a separate ViewGroupProcessor instance.

        Args:
            view_groups: Dictionary of view groups from _group_aoi_by_view
        """
        print("\n" + "=" * 80 + f"\nWPD EXTRACTION - Using {self.max_workers} parallel workers\n" + "=" * 80)

        # Calculate optimal chunks per view to match max_workers
        # Example: 20 workers / 5 views = 4 chunks per view
        chunks_per_view = max(1, self.max_workers // len(view_groups)) if view_groups else 1

        # Create chunked processor instances
        processors = []
        total_aois = 0
        for view_name, group in view_groups.items():
            aoi_files = group['aoi_files']
            hdr_files = group['hdr_files']

            if not aoi_files:
                continue

            total_aois += len(aoi_files)

            # Split AOI files into chunks
            chunk_size = max(1, len(aoi_files) // chunks_per_view)
            aoi_chunks = [aoi_files[i:i + chunk_size] for i in range(0, len(aoi_files), chunk_size)]

            # Create processor for each chunk
            for chunk_idx, aoi_chunk in enumerate(aoi_chunks):
                processor = ViewGroupProcessor(
                    view_name=f"{view_name}_chunk_{chunk_idx}",
                    aoi_files=aoi_chunk,
                    hdr_files=hdr_files,  # All HDRs shared across chunks
                    pixel_threshold_value=self.pixel_threshold_value,
                    wpd_output_dir=self.wpd_dir
                )
                processors.append(processor)

        print(f"\nProcessing {total_aois} AOIs across {len(view_groups)} views using {len(processors)} parallel workers...")
        print("=" * 80)

        # Execute all processors in parallel
        all_results = []
        with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            # Submit all processor.process() calls
            futures = {
                executor.submit(processor.process): processor.view_name
                for processor in processors
            }

            # Collect results as they complete
            completed = 0
            for future in as_completed(futures):
                view_name = futures[future]
                try:
                    result_df = future.result()
                    all_results.append(result_df)
                    completed += 1

                    # Show progress every 25% or on first/last completion
                    if completed == 1 or completed == len(futures) or completed % max(1, len(futures) // 4) == 0:
                        print(f"Progress: {completed}/{len(futures)} workers completed")

                except Exception as e:
                    print(f"Error processing {view_name}: {e}")
                    import traceback
                    traceback.print_exc()

        print(f"[OK] All {len(processors)} workers completed successfully\n")

    def _generate_sunlight_excel_report(self) -> None:
        """Generate Excel report from all .wpd files."""
        print("\n" + "=" * 80 + "\nGENERATING EXCEL FILE FROM .WPD FILES\n" + "=" * 80)

        # Read all .wpd files
        wpd_files = sorted(self.wpd_dir.glob("*.wpd"))
        print(f"\nFound {len(wpd_files)} .wpd files in {self.wpd_dir}")

        if not wpd_files:
            print("\nNo .wpd files found to generate Excel file")
            return

        # Parse .wpd files
        all_wpd_data = []
        for wpd_file in wpd_files:
            aoi_name = wpd_file.stem + ".aoi"

            with open(wpd_file, 'r') as f:
                lines = f.readlines()

                # Skip header lines (first 2 lines)
                for line in lines[2:]:
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        hdr_file = parts[0]
                        passing_pixels = int(parts[1])
                        total_pixels = int(lines[0].split(':')[1].strip())

                        all_wpd_data.append({
                            'aoi_file': aoi_name,
                            'hdr_file': hdr_file,
                            'total_pixels': total_pixels,
                            'passing_pixels': passing_pixels
                        })

        # Create DataFrame
        combined_results = pd.DataFrame(all_wpd_data)
        combined_results = combined_results.sort_values(['aoi_file', 'hdr_file'])

        # Add area calculation column (passing_pixels × area_per_pixel)
        combined_results['passing_area_m2'] = combined_results['passing_pixels'] * self.area_per_pixel

        # Display summary
        print(f"\nTotal combinations: {len(combined_results)}")
        print(f"Unique AOI files: {combined_results['aoi_file'].nunique()}")
        print(f"Unique HDR files: {combined_results['hdr_file'].nunique()}")
        print(f"Area per pixel: {self.area_per_pixel} m²")

        # Write Excel file
        def write_excel_file(combined_results: pd.DataFrame) -> None:
            """Write formatted Excel file with raw data and pivot table."""

            def format_raw_data_sheet(worksheet) -> None:
                """Format Raw Data sheet with autofit and no gridlines."""
                worksheet.sheet_view.showGridLines = False
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    worksheet.column_dimensions[column_letter].width = max_length + 2

            def format_pivot_sheet(worksheet, summary_start_col: int, summary_num_cols: int, summary_rows: int, detail_start_row: int, detail_start_col: int) -> None:
                """Format Pivot sheet with rotated headers, fixed width columns, highlighting, and gridlines."""
                worksheet.sheet_view.showGridLines = False

                green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                black_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )

                max_col = worksheet.max_column
                max_row = worksheet.max_row

                summary_index_col = summary_start_col + 1
                summary_first_data_col = summary_index_col + 1
                summary_last_data_col = summary_first_data_col + summary_num_cols - 1
                summary_end_row = summary_rows

                # Rotate summary column headers to vertical
                summary_header_row = 2
                for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
                    cell = worksheet.cell(row=summary_header_row, column=col_idx)
                    if cell.value is not None:
                        cell.alignment = Alignment(textRotation=90)

                # Green highlight summary values >= 2
                for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
                    for row_idx in range(3, summary_end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            if cell.value >= 2.0:
                                cell.fill = green_fill

                # Autofit summary index column (Apartment)
                max_length = 0
                for row_idx in range(2, summary_end_row + 1):
                    cell = worksheet.cell(row=row_idx, column=summary_index_col)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                if max_length > 0:
                    column_letter = worksheet.cell(row=1, column=summary_index_col).column_letter
                    worksheet.column_dimensions[column_letter].width = max_length + 2

                # Autofit summary data columns
                for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
                    max_length = 0
                    for row_idx in range(2, summary_end_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    if max_length > 0:
                        column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                        worksheet.column_dimensions[column_letter].width = max(max_length + 2, 8)

                # Rotate detail section timestep column headers
                detail_header_row = detail_start_row + 1
                first_timestep_col = detail_start_col + 4 + 1
                for idx, cell in enumerate(worksheet[detail_header_row]):
                    if cell.value is not None and idx + 1 >= first_timestep_col:
                        cell.alignment = Alignment(textRotation=90)

                detail_aoi_col = detail_start_col + 1 + 1
                detail_consecutive_col = detail_start_col + 2 + 1
                detail_hours_col = detail_start_col + 3 + 1

                for col_idx in range(1, max_col + 1):
                    column_letter = worksheet.cell(row=1, column=col_idx).column_letter

                    if col_idx == 1:
                        worksheet.column_dimensions[column_letter].width = 2

                    elif col_idx in (detail_aoi_col, detail_consecutive_col, detail_hours_col):
                        max_length = 0
                        for row_idx in range(2, max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            if row_idx >= detail_start_row + 2:
                                cell.border = black_border
                        worksheet.column_dimensions[column_letter].width = max_length + 2

                    elif col_idx >= first_timestep_col:
                        worksheet.column_dimensions[column_letter].width = 7.29
                        for row_idx in range(3, max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                if cell.value >= 1.0:
                                    cell.fill = green_fill

                # Color scale on Hours of Direct Sun column
                detail_hours_start = detail_start_row + 2
                hours_col_letter = worksheet.cell(row=1, column=detail_hours_col).column_letter
                hours_range = f"{hours_col_letter}{detail_hours_start}:{hours_col_letter}{max_row}"
                color_scale_rule = ColorScaleRule(
                    start_type='num', start_value=2, start_color='FFFFFF',
                    mid_type='num', mid_value=3, mid_color='FFFF99',
                    end_type='max', end_color='FFB6C1'
                )
                worksheet.conditional_formatting.add(hours_range, color_scale_rule)

            output_excel = self.wpd_dir / "sunlight_analysis_results.xlsx"

            with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                combined_results.to_excel(writer, sheet_name='Raw Data', index=False, startrow=3, startcol=1)

                ws_raw = writer.sheets['Raw Data']
                area_per_pixel_mm2 = self.area_per_pixel * 1_000_000
                pixel_increment_x_mm = round(self.pixel_increment_x * 1_000)
                pixel_increment_y_mm = round(self.pixel_increment_y * 1_000)
                ws_raw['B1'] = (f"{pixel_increment_x_mm} mm × {pixel_increment_y_mm} mm grid with an area per pixel of {self.area_per_pixel} m² ({area_per_pixel_mm2:.2f} mm²) | "
                               f"Source: {self.pixel_to_world_map}")

                combined_results['hdr_file'] = combined_results['hdr_file'].str.split('_SS_').str[1]

                pivot_data = combined_results.pivot_table(
                    values='passing_area_m2',
                    index='aoi_file',
                    columns='hdr_file',
                    aggfunc='sum',
                    fill_value=0
                )

                if len(pivot_data.columns) >= 2:
                    col1 = pivot_data.columns[0]
                    col2 = pivot_data.columns[1]
                    time1_str = col1.split('_')[-1].replace('.hdr', '')
                    time2_str = col2.split('_')[-1].replace('.hdr', '')
                    hour1 = int(time1_str[:2]) + int(time1_str[2:]) / 60.0
                    hour2 = int(time2_str[:2]) + int(time2_str[2:]) / 60.0
                    timestep_hours = abs(hour2 - hour1)
                    print(f"\nDetected timestep interval: {timestep_hours} hours")
                else:
                    timestep_hours = 1.0
                    print(f"\nUsing default timestep interval: {timestep_hours} hours")

                consecutive_counts = []
                consecutive_hours = []
                for aoi_name in pivot_data.index:
                    row_values = pivot_data.loc[aoi_name].values
                    max_consecutive = 0
                    current_consecutive = 0
                    for value in row_values:
                        if value >= 1.0:
                            current_consecutive += 1
                            max_consecutive = max(max_consecutive, current_consecutive)
                        else:
                            current_consecutive = 0
                    consecutive_counts.append(max_consecutive)
                    hours_rounded = math.floor(max_consecutive * timestep_hours * 10) / 10
                    consecutive_hours.append(hours_rounded)

                pivot_data_reset = pivot_data.reset_index()
                pivot_data_reset.insert(1, 'Consecutive Timesteps ≥1m²', consecutive_counts)
                pivot_data_reset.insert(2, 'Hours of Direct Sun', consecutive_hours)
                pivot_data_reordered = pivot_data_reset.set_index('aoi_file')

                summary_data = pivot_data_reset.copy()
                summary_data['Apartment'] = summary_data['aoi_file'].str.split('_', n=1).str[0]
                summary_data['Sub-Space'] = summary_data['aoi_file'].str.split('_', n=1).str[1]
                summary_data['Sub-Space'] = summary_data['Sub-Space'].str.replace('.aoi', '', regex=False)

                summary_pivot = summary_data.pivot_table(
                    values='Hours of Direct Sun',
                    index='Apartment',
                    columns='Sub-Space',
                    aggfunc='first',
                    fill_value=0
                )

                summary_start_col = 1
                summary_pivot.to_excel(writer, sheet_name='Pivot - Passing Area (m²)', startrow=1, startcol=summary_start_col)

                summary_rows = len(summary_pivot) + 2
                detail_start_row = 1
                detail_start_col = summary_start_col + 1 + summary_pivot.shape[1] + 1
                pivot_data_reordered.to_excel(writer, sheet_name='Pivot - Passing Area (m²)', startrow=detail_start_row, startcol=detail_start_col)

                format_raw_data_sheet(writer.sheets['Raw Data'])
                format_pivot_sheet(writer.sheets['Pivot - Passing Area (m²)'], summary_start_col, summary_pivot.shape[1], summary_rows, detail_start_row, detail_start_col)

            print(f"\nResults saved to: {output_excel}")
            print(f"  Sheet 1: 'Raw Data' - All combinations with passing_area_m2")
            print(f"  Sheet 2: 'Pivot - Passing Area (m²)' - AOI x HDR matrix")
            print(f"  Metadata: Area per pixel = {self.area_per_pixel} m²")
            print("\n" + "=" * 80 + "\nSAMPLE RESULTS (first 10 rows)\n" + "=" * 80)
            print(combined_results.head(10).to_string(index=False))

        write_excel_file(combined_results)
