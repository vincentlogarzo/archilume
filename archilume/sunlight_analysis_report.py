#!/usr/bin/env python3
"""
Generate sunlight analysis Excel report from .wpd files.

This script reads .wpd files containing sunlight analysis results and produces
a formatted Excel workbook with:
- Raw Data sheet: All AOI/HDR combinations with passing area calculations
- Pivot sheet: Summary tables, compliance checks, and detailed pivot data
"""

import math
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter


class SunlightAnalysisReport:
    """Generate Excel report from sunlight analysis .wpd files."""

    def __init__(
        self,
        wpd_dir: Path,
        area_per_pixel: float = 0.0002,
        pixel_increment_x: float = 0.014,
        pixel_increment_y: float = 0.014,
        pixel_to_world_map: str = "",
    ):
        self.wpd_dir = Path(wpd_dir)
        self.area_per_pixel = area_per_pixel
        self.pixel_increment_x = pixel_increment_x
        self.pixel_increment_y = pixel_increment_y
        self.pixel_to_world_map = pixel_to_world_map

        # Styles
        self.green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        self.black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
        )
        self.bold_font = Font(bold=True)
        self.vertical_align = Alignment(textRotation=90)

    def generate_report(self) -> Path:
        """Generate the full Excel report."""
        # Parse .wpd files
        combined_results = self._parse_wpd_files()
        if combined_results.empty:
            raise ValueError("No .wpd files found")

        # Create workbook
        wb = Workbook()
        ws_intro = wb.active
        ws_intro.title = "Methodology"

        # Build Methodology sheet
        self._build_methodology_sheet(ws_intro)

        # Build Raw Data sheet
        ws_raw = wb.create_sheet("Raw Data")
        self._build_raw_data_sheet(ws_raw, combined_results)

        # Build Pivot sheet
        ws_pivot = wb.create_sheet("Pivot - Passing Area (m²)")
        self._build_pivot_sheet(ws_pivot, combined_results)

        # Set Pivot sheet as active
        wb.active = wb.sheetnames.index("Pivot - Passing Area (m²)")

        # Save
        output_path = self.wpd_dir / "sunlight_analysis_results.xlsx"
        wb.save(output_path)
        print(f"Report saved to: {output_path}")
        return output_path

    def _parse_wpd_files(self) -> pd.DataFrame:
        """Parse all .wpd files and return combined DataFrame."""
        wpd_files = sorted(self.wpd_dir.glob("*.wpd"))
        print(f"Found {len(wpd_files)} .wpd files")

        all_data = []
        for wpd_file in wpd_files:
            aoi_name = wpd_file.stem + ".aoi"
            with open(wpd_file, 'r') as f:
                lines = f.readlines()
                total_pixels_in_aoi = int(lines[0].split(':')[1].strip())
                for line in lines[2:]:
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        all_data.append({
                            'aoi_file': aoi_name,
                            'hdr_file': parts[0],
                            'total_pixels_in_aoi': total_pixels_in_aoi,
                            'sunlit_pixels': int(parts[1]),
                        })

        df = pd.DataFrame(all_data)
        df = df.sort_values(['aoi_file', 'hdr_file'])
        df['passing_area_m2'] = df['sunlit_pixels'] * self.area_per_pixel
        return df

    def _build_methodology_sheet(self, ws):
        """Build the Methodology and Metadata sheet."""
        ws.sheet_view.showGridLines = False

        # Title
        ws['B2'] = 'Sunlight Analysis Methodology & Metadata'
        ws['B2'].font = Font(size=16, bold=True)

        # Analysis Parameters section
        ws['B4'] = 'Analysis Parameters'
        ws['B4'].font = Font(size=14, bold=True, underline='single')

        area_mm2 = self.area_per_pixel * 1_000_000
        px_x_mm = round(self.pixel_increment_x * 1000)
        px_y_mm = round(self.pixel_increment_y * 1000)

        params = [
            ['Grid Resolution (X)', f'{px_x_mm} mm', f'{self.pixel_increment_x} m'],
            ['Grid Resolution (Y)', f'{px_y_mm} mm', f'{self.pixel_increment_y} m'],
            ['Area per Pixel', f'{area_mm2:.2f} mm²', f'{self.area_per_pixel} m²'],
            ['Source Map', str(self.pixel_to_world_map), ''],
        ]

        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        row = 6
        ws.cell(row=row, column=2, value='Parameter').font = Font(bold=True)
        ws.cell(row=row, column=2).fill = grey_fill
        ws.cell(row=row, column=3, value='Value (mm)').font = Font(bold=True)
        ws.cell(row=row, column=3).fill = grey_fill
        ws.cell(row=row, column=4, value='Value (m)').font = Font(bold=True)
        ws.cell(row=row, column=4).fill = grey_fill

        for param in params:
            row += 1
            ws.cell(row=row, column=2, value=param[0])
            ws.cell(row=row, column=3, value=param[1])
            ws.cell(row=row, column=4, value=param[2])

        # Methodology section
        row += 3
        ws.cell(row=row, column=2, value='Calculation Methodology').font = Font(size=14, bold=True, underline='single')

        methodology_steps = [
            ('Step 1: Data Collection',
             'Radiance simulation results are parsed from .wpd files containing pixel-by-pixel sunlight analysis data for each area of interest (AOI) across multiple timesteps.'),

            ('Step 2: Area Calculation',
             f'Each pixel represents {area_mm2:.2f} mm² ({self.area_per_pixel} m²). The total sunlit area is calculated by multiplying the number of sunlit pixels by the area per pixel.'),

            ('Step 3: Timestep Analysis',
             'For each AOI and timestep combination, the analysis identifies pixels meeting the minimum illuminance threshold, calculating the total passing area in square meters.'),

            ('Step 4: Consecutive Duration Analysis',
             'The system identifies the longest consecutive sequence of timesteps where the sunlit area ≥1m². This represents the maximum continuous duration of direct sunlight exposure.'),

            ('Step 5: Moving Window Analysis (15-minute blocks)',
             'A 3-timestep rolling window checks for continuous 15-minute periods where each timestep maintains ≥1m² of sunlight. This validates sustained sun exposure rather than intermittent illumination.'),

            ('Step 6: 2-Hour Block Assessment',
             'The analysis evaluates whether spaces achieve 2 consecutive 1-hour blocks (each with continuous 15-minute validation). This determines compliance with minimum sunlight duration requirements.'),

            ('Step 7: Compliance Reporting',
             'Results are aggregated by apartment and space type to determine:\n  • Percentage of apartments achieving ≥2 hours direct sun\n  • Percentage of apartments achieving ≥3 hours direct sun\n  • Percentage of apartments with zero sun exposure'),
        ]

        row += 2
        for title, description in methodology_steps:
            ws.cell(row=row, column=2, value=title).font = Font(bold=True)
            row += 1
            ws.cell(row=row, column=2, value=description)
            row += 2

        # Compliance Criteria section
        row += 1
        ws.cell(row=row, column=2, value='Compliance Criteria').font = Font(size=14, bold=True, underline='single')

        criteria = [
            ['Minimum Area Threshold', '≥1 m²', 'Minimum sunlit area required per timestep'],
            ['Continuous Duration', '15 minutes', 'Minimum consecutive timesteps (3 × 5-minute intervals)'],
            ['Minimum Daily Exposure', '≥2 hours', 'Required for 70% of apartments'],
            ['Extended Exposure Target', '≥3 hours', 'Target for 70% of apartments'],
            ['Maximum Zero Sun', '≤15%', 'Maximum percentage of apartments with no direct sun'],
        ]

        row += 2
        ws.cell(row=row, column=2, value='Criterion').font = Font(bold=True)
        ws.cell(row=row, column=2).fill = grey_fill
        ws.cell(row=row, column=3, value='Requirement').font = Font(bold=True)
        ws.cell(row=row, column=3).fill = grey_fill
        ws.cell(row=row, column=4, value='Description').font = Font(bold=True)
        ws.cell(row=row, column=4).fill = grey_fill

        for criterion in criteria:
            row += 1
            ws.cell(row=row, column=2, value=criterion[0])
            ws.cell(row=row, column=3, value=criterion[1])
            ws.cell(row=row, column=4, value=criterion[2])

        # Sheet Guide section
        row += 3
        ws.cell(row=row, column=2, value='Workbook Structure').font = Font(size=14, bold=True, underline='single')

        sheets_info = [
            ['Methodology', 'This sheet - explains analysis parameters and calculation methods'],
            ['Raw Data', 'Parsed .wpd file data with pixel counts and calculated areas'],
            ['Pivot - Passing Area', 'Summary tables with compliance assessment and detailed timestep results:\n  • Compliance Summary\n  • Table 1: 2-hour block assessment\n  • Table 2: Continuous hours analysis\n  • Table 3: Timestep-by-timestep area calculations\n  • Table 4: Moving average validation'],
        ]

        row += 2
        ws.cell(row=row, column=2, value='Sheet Name').font = Font(bold=True)
        ws.cell(row=row, column=2).fill = grey_fill
        ws.cell(row=row, column=3, value='Contents').font = Font(bold=True)
        ws.cell(row=row, column=3).fill = grey_fill

        for sheet_info in sheets_info:
            row += 1
            ws.cell(row=row, column=2, value=sheet_info[0])
            ws.cell(row=row, column=3, value=sheet_info[1])

        # Column widths
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50

    def _build_raw_data_sheet(self, ws, df: pd.DataFrame):
        """Build the Raw Data sheet."""
        ws.sheet_view.showGridLines = False

        # Metadata in B1
        area_mm2 = self.area_per_pixel * 1_000_000
        px_x_mm = round(self.pixel_increment_x * 1000)
        px_y_mm = round(self.pixel_increment_y * 1000)
        ws['B1'] = (
            f"{px_x_mm} mm × {px_y_mm} mm grid with an area per pixel of "
            f"{self.area_per_pixel} m² ({area_mm2:.2f} mm²) | Source: {self.pixel_to_world_map}"
        )

        # Headers in row 4
        headers = ['aoi_file', 'hdr_file', 'total_pixels_in_aoi', 'sunlit_pixels', 'passing_area_m2']
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col_idx, header in enumerate(headers, start=2):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.alignment = Alignment(wrap_text=True)
            cell.fill = grey_fill

        # Data from row 5
        for row_idx, row in enumerate(df.itertuples(index=False), start=5):
            ws.cell(row=row_idx, column=2, value=row.aoi_file)
            ws.cell(row=row_idx, column=3, value=row.hdr_file)
            ws.cell(row=row_idx, column=4, value=row.total_pixels_in_aoi)
            ws.cell(row=row_idx, column=5, value=row.sunlit_pixels)
            ws.cell(row=row_idx, column=6, value=row.passing_area_m2)

        # Set fixed column widths (126 pixels ≈ 18 Excel width units)
        for col_idx in range(2, 2 + len(headers)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

    def _build_pivot_sheet(self, ws, df: pd.DataFrame):
        """Build the Pivot sheet with all sections."""
        ws.sheet_view.showGridLines = False

        # Extract apartment and space from aoi_file
        df = df.copy()
        df['apartment'] = df['aoi_file'].str.split('_', n=1).str[0]
        df['space'] = df['aoi_file'].str.replace('.aoi', '', regex=False).str.split('_', n=1).str[1]

        # Get unique values
        apartments = sorted(df['apartment'].unique())
        spaces = sorted(df['space'].unique())
        
        # Transform HDR filenames for pivot (remove prefix before _SS_)
        df['hdr_short'] = df['hdr_file'].str.split('_SS_').str[1]
        df = df.dropna(subset=['hdr_short'])  # Remove rows with no valid hdr_short
        hdr_files = sorted(df['hdr_short'].unique().tolist())

        # Calculate timestep hours
        timestep_hours = self._calculate_timestep_hours(hdr_files)

        # Create pivot table
        pivot_df = df.pivot_table(
            values='passing_area_m2',
            index='aoi_file',
            columns='hdr_short',
            aggfunc='sum',
            fill_value=0
        )

        # Calculate consecutive timesteps and hours for each AOI
        consecutive_data = self._calculate_consecutive_data(pivot_df, timestep_hours)

        # Build sheet sections
        num_aois = len(pivot_df)
        self._build_compliance_section(ws, apartments, spaces)
        self._build_table1_section(ws, apartments, spaces, num_aois)
        self._build_table2_section(ws, apartments, spaces, num_aois)
        self._build_detailed_pivot_section(ws, pivot_df, consecutive_data, hdr_files)
        self._build_moving_average_section(ws, pivot_df, hdr_files)

        # Apply formatting
        self._format_pivot_sheet(ws, apartments, spaces, hdr_files, len(pivot_df))

    def _calculate_timestep_hours(self, hdr_files: list) -> float:
        """Calculate timestep duration in hours from HDR filenames."""
        if len(hdr_files) < 2:
            return 1.0
        t1 = hdr_files[0].split('_')[-1].replace('.hdr', '')
        t2 = hdr_files[1].split('_')[-1].replace('.hdr', '')
        h1 = int(t1[:2]) + int(t1[2:]) / 60.0
        h2 = int(t2[:2]) + int(t2[2:]) / 60.0
        return abs(h2 - h1)

    def _calculate_consecutive_data(self, pivot_df: pd.DataFrame, timestep_hours: float) -> dict:
        """Calculate consecutive timesteps and hours for each AOI."""
        result = {}
        for aoi in pivot_df.index:
            values = pivot_df.loc[aoi].values
            max_consec = current = 0
            for v in values:
                if v >= 1.0:
                    current += 1
                    max_consec = max(max_consec, current)
                else:
                    current = 0
            hours = math.floor(max_consec * timestep_hours * 10) / 10
            result[aoi] = {'consecutive': max_consec, 'hours': hours}
        return result

    def _build_compliance_section(self, ws, apartments: list, spaces: list):
        """Build compliance summary section (rows 2-5)."""
        # Headers
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws['G2'] = ''

        ws['H2'] = 'Target'
        ws['H2'].font = self.bold_font
        ws['H2'].fill = grey_fill
        ws['H2'].alignment = Alignment(textRotation=90, wrap_text=True)

        ws['I2'] = 'Actual'
        ws['I2'].font = self.bold_font
        ws['I2'].fill = grey_fill
        ws['I2'].alignment = Alignment(textRotation=90, wrap_text=True)

        ws['J2'] = 'Compliant'
        ws['J2'].font = self.bold_font
        ws['J2'].fill = grey_fill
        ws['J2'].alignment = Alignment(textRotation=90, wrap_text=True)

        # Row 3: % of Apts ≥2 hrs
        ws['G3'] = '% of Apts ≥2 hrs Direct Sun'
        ws['G3'].fill = grey_fill
        ws['G3'].alignment = Alignment(wrap_text=True)
        ws['H3'] = '≥70%'
        ws['I3'] = '=COUNTIF($F$10:$F$24,"PASS")/COUNTA($E$10:$E$24)'
        ws['I3'].number_format = '0%'
        ws['J3'] = '=IF(I3>=0.7, "PASS","FAIL")'
        ws['J3'].font = self.bold_font

        # Row 4: % of Apts ≥3 hrs
        ws['G4'] = '% of Apts ≥3 hrs Direct Sun'
        ws['G4'].fill = grey_fill
        ws['G4'].alignment = Alignment(wrap_text=True)
        ws['H4'] = '≥70%'

        # Row 5: % of Zero Sun
        ws['G5'] = '% of Zero Sun Apartments'
        ws['G5'].fill = grey_fill
        ws['G5'].alignment = Alignment(wrap_text=True)
        ws['H5'] = '≤ 15%'
        ws['I5'] = '=COUNTIF($F$29:$F$43,"Zero Sun")/COUNTA($E$29:$E$43)'
        ws['I5'].number_format = '0%'
        ws['J5'] = '=IF(I5<=0.15, "PASS","FAIL")'
        ws['J5'].font = self.bold_font

    def _build_table1_section(self, ws, apartments: list, spaces: list, num_aois: int):
        """Build Table 1: No. consecutive 2hr blocks section (rows 7-24)."""
        # Title
        ws['E7'] = 'Table 1: No. consecutive 2 hr blocks with ≥ 1m2 of Sun for a continuous 15 mins '

        # Headers row 9
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws['E9'] = 'Apt'
        ws['E9'].font = self.bold_font
        ws['E9'].fill = grey_fill
        ws['E9'].alignment = Alignment(wrap_text=True)

        ws['F9'] = 'Living + Private Outdoor Area with ≥2 hours direct sun'
        ws['F9'].font = self.bold_font
        ws['F9'].fill = grey_fill
        ws['F9'].alignment = Alignment(wrap_text=True)

        ws['G9'] = 'Living + Private Outdoor Area with ≥3 hours direct sun'
        ws['G9'].font = self.bold_font
        ws['G9'].fill = grey_fill
        ws['G9'].alignment = Alignment(wrap_text=True)

        # Space headers with rotation
        for col_idx, space in enumerate(spaces, start=8):
            cell = ws.cell(row=9, column=col_idx, value=space)
            cell.font = self.bold_font
            cell.fill = grey_fill
            cell.alignment = Alignment(textRotation=90, wrap_text=True)

        # Data rows (10-24 for apartments)
        ma_start = 87  # Row where moving average section starts
        ma_end = ma_start + num_aois - 1
        for row_offset, apt in enumerate(apartments):
            row = 10 + row_offset
            ws.cell(row=row, column=5, value=apt)

            # Living + Private Outdoor columns with PASS/FAIL formulas
            col_range_living = f"H{row}:M{row}"  # Beds
            col_range_outdoor = f"N{row}:P{row}"  # T, T1, T2
            ws.cell(row=row, column=6, value=f'=IF(AND(COUNTIF({col_range_living},"PASS")>0,COUNTIF({col_range_outdoor},"PASS")>0),"PASS","FAIL")')

            # Space formulas - lookup from moving average section (G column = consecutive hours result)
            for col_idx, space in enumerate(spaces, start=8):
                col_letter = get_column_letter(col_idx)
                ws.cell(row=row, column=col_idx,
                        value=f'=IFERROR(INDEX($G${ma_start}:$G${ma_end},'
                              f'MATCH($E{row}&"_"&{col_letter}$9&".aoi",$E${ma_start}:$E${ma_end},0)),"")')

    def _build_table2_section(self, ws, apartments: list, spaces: list, num_aois: int):
        """Build Table 2: Continuous Hours section (rows 26-43)."""
        # Title
        ws['E26'] = 'Table 2: Continuous Hours of >1m2 of Direct Sun'

        # Headers row 28
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        ws['E28'] = 'Apt'
        ws['E28'].font = self.bold_font
        ws['E28'].fill = grey_fill
        ws['E28'].alignment = Alignment(wrap_text=True)

        ws['F28'] = 'Living + Private Outdoor Area with ≥2 hours direct sun'
        ws['F28'].font = self.bold_font
        ws['F28'].fill = grey_fill
        ws['F28'].alignment = Alignment(wrap_text=True)

        ws['G28'] = 'Living + Private Outdoor Area with ≥3 hours direct sun'
        ws['G28'].font = self.bold_font
        ws['G28'].fill = grey_fill
        ws['G28'].alignment = Alignment(wrap_text=True)

        for col_idx, space in enumerate(spaces, start=8):
            cell = ws.cell(row=28, column=col_idx, value=space)
            cell.font = self.bold_font
            cell.fill = grey_fill
            cell.alignment = Alignment(textRotation=90, wrap_text=True)

        # Data rows (29-43)
        detail_start = 47  # Row where first detailed pivot starts
        detail_end = detail_start + num_aois - 1
        for row_offset, apt in enumerate(apartments):
            row = 29 + row_offset
            ws.cell(row=row, column=5, value=apt)

            # Living + Private Outdoor columns
            col_range_living = f"H{row}:M{row}"
            col_range_outdoor = f"N{row}:P{row}"
            ws.cell(row=row, column=6, 
                    value=f'=IF(AND(COUNTIF({col_range_living},">=2")>0,COUNTIF({col_range_outdoor},">=2")>0),"PASS",'
                          f'IF(SUM(H{row}:P{row})=0,"Zero Sun","FAIL"))')
            ws.cell(row=row, column=7,
                    value=f'=IF(AND(COUNTIF({col_range_living},">=3")>0,COUNTIF({col_range_outdoor},">=3")>0),"PASS",'
                          f'IF(SUM(H{row}:P{row})=0,"Zero Sun","FAIL"))')

            # Space values - lookup hours from first detail section
            for col_idx, space in enumerate(spaces, start=8):
                col_letter = get_column_letter(col_idx)
                ws.cell(row=row, column=col_idx,
                        value=f'=IFERROR(INDEX($G${detail_start}:$G${detail_end},'
                              f'MATCH($E{row}&"_"&{col_letter}$28&".aoi",$E${detail_start}:$E${detail_end},0)),0)')

    def _build_detailed_pivot_section(self, ws, pivot_df: pd.DataFrame, consecutive_data: dict, hdr_files: list):
        """Build detailed pivot section with raw timestep data (rows 47-80)."""
        # Section title
        ws.cell(row=45, column=5, value='Table 3: Area (m²) of Direct Sun at specified height above FFL per Timestep')

        
        header_row = 47
        data_start_row = 48
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Headers
        cell_c = ws.cell(row=header_row, column=3, value='Apartment')
        cell_c.font = self.bold_font
        cell_c.fill = grey_fill
        cell_c.alignment = Alignment(wrap_text=True)

        cell_d = ws.cell(row=header_row, column=4, value='Space')
        cell_d.font = self.bold_font
        cell_d.fill = grey_fill
        cell_d.alignment = Alignment(wrap_text=True)

        cell_e = ws.cell(row=header_row, column=5, value='aoi_file')
        cell_e.font = self.bold_font
        cell_e.fill = grey_fill
        cell_e.alignment = Alignment(wrap_text=True)

        cell_f = ws.cell(row=header_row, column=6, value='Consecutive Timesteps ≥1m²')
        cell_f.font = self.bold_font
        cell_f.fill = grey_fill
        cell_f.alignment = Alignment(wrap_text=True)

        cell_g = ws.cell(row=header_row, column=7, value='Continuous Hours of Direct Sun')
        cell_g.font = self.bold_font
        cell_g.fill = grey_fill
        cell_g.alignment = Alignment(wrap_text=True)

        # HDR file headers with rotation
        for col_idx, hdr in enumerate(hdr_files, start=8):
            cell = ws.cell(row=header_row, column=col_idx, value=hdr)
            cell.font = self.bold_font
            cell.fill = grey_fill
            cell.alignment = Alignment(textRotation=90, wrap_text=True)

        # Data rows
        for row_offset, aoi in enumerate(pivot_df.index):
            row = data_start_row + row_offset
            ws.cell(row=row, column=5, value=aoi)
            # Apartment formula
            ws.cell(row=row, column=3, value=f'=LEFT(E{row},SEARCH("_",E{row})-1)')
            # Space formula
            ws.cell(row=row, column=4, value=f'=LEFT(RIGHT(E{row},LEN(E{row})-SEARCH("_",E{row})),LEN(RIGHT(E{row},LEN(E{row})-SEARCH("_",E{row})))-4)')
            # Consecutive timesteps
            ws.cell(row=row, column=6, value=consecutive_data[aoi]['consecutive'])
            # Hours
            ws.cell(row=row, column=7, value=consecutive_data[aoi]['hours'])

            # Timestep values
            for col_idx, hdr in enumerate(hdr_files, start=8):
                val = pivot_df.loc[aoi, hdr] if hdr in pivot_df.columns else 0
                ws.cell(row=row, column=col_idx, value=val)

    def _build_moving_average_section(self, ws, pivot_df: pd.DataFrame, hdr_files: list):
        """Build moving average section for consecutive hour calculations (rows 83-120)."""
        header_row = 87
        data_start_row = 88
        detail_row_ref = 48  # Reference to first detail section

        # Section title
        ws.cell(row=84, column=5, value='Table 4: Moving Average table to calculate No. consecutive hours with ≥ 1m2 of Sun for a continuous 15 mins ')

        # Helper rows for HH and HHMM extraction
        ws.cell(row=85, column=7, value='HH')
        ws.cell(row=86, column=7, value='HHMM')

        # Headers
        grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        cell_c = ws.cell(row=header_row, column=3, value='Apartment')
        cell_c.font = self.bold_font
        cell_c.fill = grey_fill
        cell_c.alignment = Alignment(wrap_text=True)

        cell_d = ws.cell(row=header_row, column=4, value='Space')
        cell_d.font = self.bold_font
        cell_d.fill = grey_fill
        cell_d.alignment = Alignment(wrap_text=True)

        cell_e = ws.cell(row=header_row, column=5, value='aoi_file')
        cell_e.font = self.bold_font
        cell_e.fill = grey_fill
        cell_e.alignment = Alignment(wrap_text=True)

        cell_g = ws.cell(row=header_row, column=7, value='Consecutive hours')
        cell_g.font = self.bold_font
        cell_g.fill = grey_fill
        cell_g.alignment = Alignment(wrap_text=True)

        # HDR file headers
        for col_idx, hdr in enumerate(hdr_files, start=8):
            cell = ws.cell(row=header_row, column=col_idx, value=hdr)
            cell.font = self.bold_font
            cell.fill = grey_fill
            cell.alignment = Alignment(textRotation=90, wrap_text=True)
            # HH extraction formula
            ws.cell(row=85, column=col_idx, value=f'=LEFT({get_column_letter(col_idx)}86,2)')
            # HHMM extraction formula
            ws.cell(row=86, column=col_idx,
                    value=f'=LEFT(RIGHT({get_column_letter(col_idx)}87,LEN({get_column_letter(col_idx)}87)-SEARCH("_",{get_column_letter(col_idx)}87)),LEN(RIGHT({get_column_letter(col_idx)}87,LEN({get_column_letter(col_idx)}87)-SEARCH("_",{get_column_letter(col_idx)}87)))-4)')

        # Data rows with formulas
        num_timesteps = len(hdr_files)
        for row_offset, aoi in enumerate(pivot_df.index):
            row = data_start_row + row_offset
            detail_row = detail_row_ref + row_offset

            ws.cell(row=row, column=5, value=aoi)
            ws.cell(row=row, column=3, value=f'=LEFT(E{row},SEARCH("_",E{row})-1)')
            ws.cell(row=row, column=4, value=f'=LEFT(RIGHT(E{row},LEN(E{row})-SEARCH("_",E{row})),LEN(RIGHT(E{row},LEN(E{row})-SEARCH("_",E{row})))-4)')

            # Consecutive hours formula - check for 2-hour blocks (24 timesteps at 5-min intervals)
            # Each 2-hour block = 24 timesteps, we check for overlapping windows
            block_size = 12  # 12 timesteps = 1 hour at 5-min intervals
            start_col = 8
            end_col = start_col + num_timesteps - 1

            # Build OR conditions for consecutive 2-hour blocks
            or_conditions = []
            for block_start in range(0, num_timesteps - 24 + 1, block_size):
                first_half_start = start_col + block_start
                first_half_end = first_half_start + block_size - 1
                second_half_start = first_half_end + 1
                second_half_end = second_half_start + block_size - 1

                if second_half_end <= end_col:
                    range1 = f"{get_column_letter(first_half_start)}{row}:{get_column_letter(first_half_end)}{row}"
                    range2 = f"{get_column_letter(second_half_start)}{row}:{get_column_letter(second_half_end)}{row}"
                    or_conditions.append(f'AND(COUNTIF({range1},"PASS")>0,COUNTIF({range2},"PASS")>0)')

            if or_conditions:
                formula = f'=IF(OR({",".join(or_conditions)}),"PASS","FAIL")'
                ws.cell(row=row, column=7, value=formula)
            else:
                ws.cell(row=row, column=7, value='"FAIL"')

            # 15-minute rolling window PASS formulas (3 consecutive timesteps ≥1)
            for col_idx in range(8, 8 + num_timesteps):
                col_letter = get_column_letter(col_idx)
                if col_idx == 8:  # First timestep - just check if ≥1
                    ws.cell(row=row, column=col_idx,
                            value=f'=IF({col_letter}{detail_row}>=1,"PASS","")')
                elif col_idx == 9:  # Second timestep - check current and previous
                    prev1_col = get_column_letter(col_idx - 1)
                    ws.cell(row=row, column=col_idx,
                            value=f'=IF(AND({prev1_col}{detail_row}>=1,{col_letter}{detail_row}>=1),"PASS","")')
                else:  # Third timestep onwards - check 3 consecutive
                    prev2_col = get_column_letter(col_idx - 2)
                    prev1_col = get_column_letter(col_idx - 1)
                    ws.cell(row=row, column=col_idx,
                            value=f'=IF(AND({prev2_col}{detail_row}>=1,{prev1_col}{detail_row}>=1,{col_letter}{detail_row}>=1),"PASS","")')

    def _format_pivot_sheet(self, ws, apartments: list, spaces: list, hdr_files: list, num_aois: int):
        """Apply formatting to the pivot sheet."""
        # Column widths (177 pixels ≈ 25 Excel width units for C-G, 40 pixels ≈ 5.7 for H+)
        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 2
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['C'].hidden = True
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['D'].hidden = True
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25

        # Columns H onwards (40 pixels ≈ 5.7 Excel width units)
        for col_idx in range(8, 8 + len(hdr_files)):
            ws.column_dimensions[get_column_letter(col_idx)].width = 5.7

        # Row heights for table headers (100 pixels ≈ 75 Excel height units)
        ws.row_dimensions[2].height = 75  # Compliance section header
        ws.row_dimensions[9].height = 75  # Table 1 header
        ws.row_dimensions[28].height = 75  # Table 2 header
        ws.row_dimensions[47].height = 75  # Detailed pivot header
        ws.row_dimensions[87].height = 75  # Moving average header

        # Conditional formatting for compliance cells
        green_text = Font(color='006400')
        red_text = Font(color='8B0000')

        # Color scale for Hours column (G47:G80 and G87:G120)
        detail_end = 47 + num_aois - 1
        ma_end = 87 + num_aois - 1

        color_scale = ColorScaleRule(
            start_type='num', start_value=2, start_color='FFFFFF',
            mid_type='num', mid_value=3, mid_color='FFFF99',
            end_type='max', end_color='FFB6C1'
        )
        ws.conditional_formatting.add(f'G47:G{detail_end}', color_scale)
        ws.conditional_formatting.add(f'G87:G{ma_end}', color_scale)

        # PASS/FAIL formatting for compliance summary
        pass_rule = FormulaRule(formula=['$J3="PASS"'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'))
        fail_rule = FormulaRule(formula=['$J3="FAIL"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'))
        ws.conditional_formatting.add('J3:J5', pass_rule)
        ws.conditional_formatting.add('J3:J5', fail_rule)

        # PASS highlighting for Table 1 and Table 2
        table_pass = FormulaRule(formula=['H10="PASS"'], fill=self.green_fill)
        ws.conditional_formatting.add(f'H10:P{10 + len(apartments) - 1}', table_pass)

        # Green fill for timestep values ≥1
        for col_idx in range(8, 8 + len(hdr_files)):
            col_letter = get_column_letter(col_idx)
            # First detail section
            ws.conditional_formatting.add(
                f'{col_letter}47:{col_letter}{detail_end}',
                FormulaRule(formula=[f'{col_letter}48>=1'], fill=self.green_fill)
            )
            # Moving average section
            ws.conditional_formatting.add(
                f'{col_letter}87:{col_letter}{ma_end}',
                FormulaRule(formula=[f'{col_letter}87="PASS"'], fill=self.green_fill)
            )


def main():
    # Configuration variables
    wpd_dir = Path(__file__).parent.parent / "outputs" / "wpd"
    area_per_pixel = 0.0002  # Area per pixel in m²
    pixel_increment_x = 0.014  # Pixel increment X in meters
    pixel_increment_y = 0.014  # Pixel increment Y in meters
    pixel_to_world_map = Path(__file__).parent.parent / "outputs" / "aoi" / "pixel_to_world_coordinate_map.txt"

    report = SunlightAnalysisReport(
        wpd_dir=wpd_dir,
        area_per_pixel=area_per_pixel,
        pixel_increment_x=pixel_increment_x,
        pixel_increment_y=pixel_increment_y,
        pixel_to_world_map=pixel_to_world_map,
    )
    output_path = report.generate_report()
    print(f"\nGenerated: {output_path}")

    #TODO: consecutive hours column is not dynamic to the data,
    #TODO: cell look up are not dynamic to the data. 
    #TODO: pass calcualtion in table 4 is not dynamic to the timeste that a user may put in . 
if __name__ == '__main__':
    main()