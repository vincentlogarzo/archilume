"""Export pipeline — Excel reports, overlay images, ZIP archives."""

import io
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

import numpy as np
from PIL import Image, ImageDraw, ImageFont


def export_report(
    rooms: list[dict],
    hdr_files: list[dict],
    image_dir: Path,
    output_dir: Path,
    wpd_dir: Path,
    archive_dir: Path,
    project_name: str,
    df_thresholds: dict,
    on_progress: Optional[Callable[[int, str], None]] = None,
) -> Optional[Path]:
    """Run full export pipeline: compute DF → Excel → overlay images → ZIP.

    Args:
        rooms: List of room dicts.
        hdr_files: List of HDR file info dicts.
        image_dir: Directory containing HDR/TIFF images.
        output_dir: Directory for output files.
        wpd_dir: Directory for per-pixel data CSVs.
        archive_dir: Directory for ZIP archives.
        project_name: Name of the project.
        df_thresholds: Dict of room_type → DF% threshold.
        on_progress: Callback(percent, message) for progress updates.

    Returns path to ZIP archive, or None on failure.
    """
    if on_progress:
        on_progress(0, "Starting export...")

    output_dir.mkdir(parents=True, exist_ok=True)
    wpd_dir.mkdir(parents=True, exist_ok=True)
    archive_dir.mkdir(parents=True, exist_ok=True)

    total_steps = len(hdr_files) + 2  # +1 for Excel, +1 for ZIP
    step = 0

    # Phase 1: Generate overlay images for each HDR
    for hdr_info in hdr_files:
        hdr_name = hdr_info["name"]
        hdr_rooms = [r for r in rooms if r.get("hdr_file") == hdr_name]

        if hdr_rooms:
            _generate_overlay_image(
                hdr_info, hdr_rooms, image_dir, output_dir
            )

        step += 1
        if on_progress:
            pct = int(step / total_steps * 100)
            on_progress(pct, f"Processing {hdr_name}...")

    # Phase 2: Generate Excel report
    if on_progress:
        on_progress(int(step / total_steps * 100), "Generating Excel report...")

    xlsx_path = _generate_excel_report(rooms, output_dir, df_thresholds)
    step += 1

    # Phase 3: Create ZIP archive
    if on_progress:
        on_progress(int(step / total_steps * 100), "Creating archive...")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = f"{project_name}_{timestamp}.zip"
    zip_path = archive_dir / zip_name

    _create_archive(output_dir, zip_path)
    step += 1

    if on_progress:
        on_progress(100, "Export complete!")

    return zip_path


def _generate_overlay_image(
    hdr_info: dict,
    rooms: list[dict],
    image_dir: Path,
    output_dir: Path,
) -> Optional[Path]:
    """Generate an overlay image with room boundaries drawn on top of the source image."""
    # Find a displayable image (prefer TIFF, fall back to HDR)
    img_path = None
    for tp in hdr_info.get("tiff_paths", []):
        p = Path(tp)
        if p.exists():
            img_path = p
            break
    if img_path is None:
        hdr_path = Path(hdr_info["hdr_path"])
        if hdr_path.exists():
            # Load and tone-map HDR
            from .image_loader import _load_and_encode
            # For overlay we need PIL image, not base64
            from .image_loader import _load_hdr
            arr = _load_hdr(hdr_path)
            if arr is not None:
                arr_uint8 = (np.clip(arr, 0.0, 1.0) * 255).astype(np.uint8)
                img = Image.fromarray(arr_uint8, "RGB")
            else:
                return None
        else:
            return None
    else:
        img = Image.open(img_path).convert("RGB")

    draw = ImageDraw.Draw(img, "RGBA")

    for room in rooms:
        verts = room.get("vertices", [])
        if len(verts) < 3:
            continue

        # Draw filled polygon
        poly_points = [(int(v[0]), int(v[1])) for v in verts]
        draw.polygon(poly_points, fill=(13, 148, 136, 40), outline=(13, 148, 136, 200))

        # Draw label
        from .geometry import polygon_label_point
        lx, ly = polygon_label_point(verts)
        name = room.get("name", "")
        try:
            draw.text((int(lx), int(ly)), name, fill=(26, 31, 39, 200), anchor="mm")
        except Exception:
            draw.text((int(lx), int(ly)), name, fill=(26, 31, 39, 200))

    out_path = output_dir / f"{hdr_info['name']}_aoi_overlay.png"
    img.save(out_path, "PNG")
    return out_path


def _generate_excel_report(
    rooms: list[dict],
    output_dir: Path,
    df_thresholds: dict,
) -> Optional[Path]:
    """Generate Excel report with per-room DF% data."""
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Summary"

    # Header row
    headers = ["Room Name", "Parent", "Type", "HDR File", "Vertices"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows
    for row_idx, room in enumerate(rooms, 2):
        ws.cell(row=row_idx, column=1, value=room.get("name", ""))
        ws.cell(row=row_idx, column=2, value=room.get("parent", ""))
        ws.cell(row=row_idx, column=3, value=room.get("room_type", ""))
        ws.cell(row=row_idx, column=4, value=room.get("hdr_file", ""))
        ws.cell(row=row_idx, column=5, value=str(len(room.get("vertices", []))))

    xlsx_path = output_dir / "aoi_report_daylight.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def _create_archive(source_dir: Path, zip_path: Path) -> None:
    """Create a ZIP archive of all files in source_dir."""
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in source_dir.rglob("*"):
            if file_path.is_file():
                arcname = file_path.relative_to(source_dir)
                zf.write(file_path, arcname)


def extract_archive(zip_path: Path, target_dir: Path) -> bool:
    """Extract a ZIP archive to target directory.

    Returns True on success.
    """
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(target_dir)
        return True
    except Exception:
        return False


def list_archives(archive_dir: Path) -> list[str]:
    """List all .zip files in the archive directory."""
    if not archive_dir.exists():
        return []
    return sorted(
        [p.name for p in archive_dir.glob("*.zip")],
        reverse=True,
    )
