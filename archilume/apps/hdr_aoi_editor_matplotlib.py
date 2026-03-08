"""
Archilume: Interactive Room Boundary Editor for HDR/TIFF Floor Plan Images

The workflow generates the HDR/TIFF images and .aoi boundaries this
editor depends on.

Draw apartment and sub-room boundaries on top of HDR or associated TIFF
rendered floor plan images. JSON and CSV are saved automatically to the
image_dir on every save/delete.

Naming convention:
    U101          apartment boundary
    U101_BED1     sub-room (auto-prefixed when parent is selected)
    U101_DIV1     division sub-room (created by the room divider tool)

Controls:
    Up/Down       Navigate HDR files
    t             Toggle image variant (HDR / TIFFs)
    Left-click    Select room (default) / place vertex (draw mode) / drag vertex (edit mode)
    Shift+click   Drag entire edge in edit mode (moves both endpoints together)
    Right-click   Delete hovered vertex (edit mode) / undo divider point (divider mode)
    Scroll        Zoom centred on cursor
    s             Save room / confirm edit
    e             Toggle Edit Mode
    d             Toggle Draw Mode (default) / room divider — multi-segment ortho split (edit mode)
    Esc           Exit draw mode / exit divider mode / deselect room
    Ctrl+Z        Undo — edit: vertex/edge/division; draw: delete/type/rename/create
    o             Toggle orthogonal lines (H/V snap)
    Ctrl+click    Multi-select rooms in list (bulk type tagging)
    Ctrl+A        Select all rooms on current HDR
    f             Fit zoom to selected room
    r             Reset zoom
    q             Quit

Workflow:
    1. Navigate to the desired HDR file with Up/Down
    2. Draw apartment boundary, name "U101", press s to save
    3. Select "U101" as parent
    4. Draw sub-rooms (e.g. "BED1" auto-saved as "U101_BED1")
    5. Repeat for each HDR file / floor
    6. Export and archive results as CSV for analysis and reporting

Functionality:

    Room Divider:
        Splits a room into two sub-rooms along a multi-segment ortho
        polyline. Enter edit mode (e), select a room, press d, then click
        to place points (each segment H or V). Press s to finalize. Lines
        auto-extend to the boundary when points are placed outside the
        room. Right-click undoes the last point. Ctrl+Z undoes the split.

    Edit Mode:
        Toggle with e. All rooms on the current HDR become editable
        simultaneously. Available actions while in edit mode:
        - Drag a vertex to reposition it (ortho-constrained when o is on)
        - Shift+drag an edge to move both its endpoints together
        - Click on an edge to insert two new vertices straddling the
          click point (edit mode only)
        - Right-click a vertex to delete it
        - Press d to enter room divider mode (edit mode only)
        - Press f to fit-zoom to the selected room boundary
        - Ctrl+Z undoes vertex edits (up to 50 levels)
        Press s or toggle e off to save and exit edit mode.

    Ortho Mode:
        Toggle with o. Constrains drawn lines to horizontal or vertical
        only. The live preview snaps to the nearest axis in real time.
        Useful for architectural floor plans with axis-aligned walls.

    Vertex & Edge Snapping:
        Automatically snaps new vertices to nearby existing vertices or
        edges within 10 pixels. Always active during drawing. Helps
        align room boundaries precisely without manual pixel placement.

    Parent Auto-Detection:
        When the first vertex of a new polygon falls inside an existing
        parent room, that parent is auto-selected. Sub-rooms are then
        auto-prefixed (e.g. drawing "BED1" inside U101 saves as
        U101_BED1). Boundary containment is checked on save.

    Room Type Tagging:
        Assign BED, LIVING, or CIRCULATION types via buttons. Sub-rooms
        default to BED; parents default to LIVING when children exist.
        Ctrl+click rooms for multi-select, then tag in bulk. Types
        drive daylight factor threshold evaluation.

    Export & Archive:
        Exports an Excel summary, per-room pixel CSVs, and overlay TIFFs
        with room boundaries rendered. Everything is zipped into a
        timestamped archive. Use "Extract Archive" to restore a previous
        export. Runs in a background thread with progress bar.

    Image Variant Toggle:
        Press t to cycle between the HDR and associated TIFF images for
        the current floor. Useful for comparing raw HDR data against
        processed TIFF renderings side by side.

    Multi-Select:
        Ctrl+click rooms in the list to toggle multi-selection. Ctrl+A
        selects all rooms on the current HDR. Enables bulk room type
        tagging across multiple rooms at once.

    Undo:
        Ctrl+Z in edit mode undoes vertex/edge edits and room divisions
        (50 levels). In draw mode, Ctrl+Z undoes room deletion, room
        type changes, room renames, and new room creation (including
        any auto-assigned parent type). Both stacks hold up to 50
        entries.
"""

# fmt: off
# autopep8: off

# Standard library imports
import csv
import hashlib
import json
import subprocess
import ctypes
import ctypes.wintypes
import os
import re
import shutil
import sys
import threading
import time
import traceback
import types
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
from typing import List, Optional, Tuple, Union

# Third-party imports
import imageio.v2 as imageio
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import matplotlib.pyplot as plt
from matplotlib.widgets import PolygonSelector, TextBox, Button, RadioButtons
from matplotlib.patches import Polygon, FancyBboxPatch
from matplotlib.path import Path as MplPath
import matplotlib.patheffects as patheffects
import numpy as np

# Archilume imports
from archilume import config, utils, Hdr2Wpd
from archilume.config import OUTPUTS_DIR, ARCHIVE_DIR
from archilume.utils import rasterize_pdf_page, make_lines_only


def _load_pil_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    """Load a Unicode-capable TrueType font at *size* points.

    Resolution order:
      1. Matplotlib's bundled DejaVu Sans (always present, cross-platform).
      2. System DejaVu / Arial as fallback in case font_manager returns a path
         that PIL cannot open (very rare).
      3. PIL bitmap default as last resort.
    """
    from matplotlib import font_manager as _fm
    style = 'bold' if bold else 'normal'
    candidates = []
    try:
        candidates.append(_fm.findfont(_fm.FontProperties(family='DejaVu Sans', weight=style)))
    except Exception:
        pass
    # Common system paths as insurance
    if bold:
        candidates += [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
            "C:/Windows/Fonts/arial.ttf",
        ]
    else:
        candidates += [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "C:/Windows/Fonts/arial.ttf",
        ]
    for path in candidates:
        try:
            return ImageFont.truetype(path, size)
        except (OSError, IOError, TypeError):
            pass
    return ImageFont.load_default()


class HdrAoiEditor:
    """Interactive room boundary drawing tool for HDR/TIFF floor plan images.

    Displays an HDR or associated TIFF image and provides a PolygonSelector
    for drawing room boundary polygons. Supports hierarchical parent/child
    relationships (e.g. U101 → U101_BED1). Auto-saves JSON and CSV on every
    save or delete action.

    Args:
        image_dir: Directory containing .hdr files and associated .tiff files.
        aoi_dir: Directory containing .aoi files with pixel-mapped room boundaries.
        session_path: Optional path for saving/loading editor sessions (JSON).
    """

    def __init__(
        self,
        image_dir:              Union[Path, str]            = config.IMAGE_DIR,
        aoi_dir:                Union[Path, str]            = config.AOI_DIR,
        session_path:           Optional[Path]              = None,
        pdf_path:               Optional[Union[Path, str]]  = None,
        project:                Optional[str]               = None,
        iesve_room_data:        Optional[Union[Path, str]]  = None,
    ):
        # Base project directory within inputs/
        self.project = project
        self.project_input_dir = config.INPUTS_DIR / project if project else config.INPUTS_DIR

        # Protected directory for editor-generated files (session, CSV, cache)
        # This prevents simulation reruns in outputs/ from overwriting manual editor work.
        self.project_aoi_dir = self.project_input_dir / "aoi"
        if project:
            self.project_aoi_dir.mkdir(parents=True, exist_ok=True)

        if pdf_path is not None:
            pdf_path = Path(pdf_path)
            pdf_path = self.project_input_dir / pdf_path if not pdf_path.is_absolute() else pdf_path

        if iesve_room_data is not None:
            iesve_room_data = Path(iesve_room_data)
            iesve_room_data = self.project_input_dir / iesve_room_data if not iesve_room_data.is_absolute() else iesve_room_data
        self._iesve_room_data_path: Optional[Path]          = iesve_room_data

        image_dir = Path(image_dir)
        image_dir = self.project_input_dir / image_dir if not image_dir.is_absolute() else image_dir
        self.image_dir                                  = image_dir

        # Validate required paths before building any UI
        if pdf_path is not None and not pdf_path.exists():
            raise FileNotFoundError(f"pdf_path not found: {pdf_path}")
        if iesve_room_data is not None and not iesve_room_data.exists():
            raise FileNotFoundError(f"iesve_room_data not found: {iesve_room_data}")
        if not self.image_dir.exists():
            raise FileNotFoundError(f"image_dir not found: {self.image_dir}")
        # Use the protected project AOI dir for both session and .aoi exports if project is set
        self.aoi_dir                                    = self.project_aoi_dir if project else Path(aoi_dir)
        
        # Session and CSV paths now default to the protected project input directory
        self.session_path                               = session_path or (self.project_aoi_dir / "aoi_session.json")
        self.csv_path                                   = self.project_aoi_dir / "aoi_boundaries.csv"

        # Scan HDR files in image_dir
        self.hdr_files:             List[dict]          = self._scan_hdr_files()
        self.current_hdr_idx:       int                 = 0

        # Image variant toggle state (HDR + associated TIFFs for current HDR)
        self.image_variants:        List[Path]          = []
        self.current_variant_idx:   int                 = 0
        self._rebuild_image_variants()

        # Room storage
        self.rooms:                 List[dict]          = []
        self.current_polygon_vertices                   = []
        self.selected_room_idx:     Optional[int]       = None

        # Zoom / pan state
        self.original_xlim                              = None
        self.original_ylim                              = None
        self._pan_active:           bool                = False
        self._pan_start:            Optional[tuple]     = None  # (x_event, y_event) in pixels

        # Image dimensions (set on first render)
        self._image_width:          int                 = 1
        self._image_height:         int                 = 1
        self._reference_view_w:     float               = 1.0  # view width at default zoom

        # Snap to existing polygon vertices (always on, no UI control)
        self._snap_distance_px:     float               = 10.0
        self.current_vertices:      np.ndarray          = np.array([])
        self.ortho_mode:            bool                = True
        self._pending_snap:         Optional[tuple]     = None

        # Draw mode (polygon drawing; toggled with 'd' when not in edit mode)
        self.draw_mode:             bool                = False

        # Vertex editing mode
        self.edit_mode:             bool                = False
        self.edit_room_idx:         Optional[int]       = None
        self.edit_vertex_idx:       Optional[int]       = None
        self.hover_room_idx:        Optional[int]       = None
        self.hover_vertex_idx:      Optional[int]       = None
        self.hover_edge_room_idx:   Optional[int]       = None
        self.hover_edge_idx:        Optional[int]       = None
        self.hover_edge_point:      Optional[tuple]     = None
        self.edit_edge_room_idx:    Optional[int]       = None
        self.edit_edge_idx:         Optional[int]       = None
        self.edit_edge_start:       Optional[tuple]     = None
        self._edit_drag_origin:     Optional[tuple]     = None  # vertex pos at drag start (for ortho)

        # Parent apartment selection
        self.selected_parent:       Optional[str]       = None
        self.parent_options:        List[str]           = []

        # Room list scroll state
        self.room_list_scroll_offset: int               = 0
        self._room_list_hit_boxes:  List[Tuple]         = []

        # Undo stack for edit-mode vertex operations
        # Each entry: (room_idx, vertices_snapshot)
        self._edit_undo_stack:      List[Tuple]         = []
        self._edit_undo_max:        int                 = 50

        # General-purpose undo stack for draw-mode operations
        # Each entry: (tag, ...data) where tag is one of:
        #   ('delete',    insertion_index, room_dict)
        #   ('type',      [(idx, old_type), ...])
        #   ('rename',    room_idx, old_name)
        #   ('create',    room_idx)
        self._draw_undo_stack:      List[Tuple]         = []
        self._draw_undo_max:        int                 = 50

        # Image cache: path → numpy array (avoids reloading from disk)
        self._image_cache:          dict                = {}
        self._image_cache_lock:     threading.Lock      = threading.Lock()
        self._image_cache_limit:    int                 = 15

        # Cached matplotlib artists for incremental rendering
        self._room_patch_cache                          = {}
        self._room_label_cache                          = {}
        self._df_text_cache:        list                = []
        self._edit_vertex_scatter                       = None
        self._last_hover_check                          = 0.0
        self._last_drag_draw:       float               = 0.0
        # Window persistence settings
        self.window_settings:       dict                = {}
        self._closing:              bool                = False
        # Pre-built hover arrays (rebuilt on full render)
        self._hover_all_verts:      Optional[np.ndarray]= None
        self._hover_vert_room_idx:  Optional[np.ndarray]= None
        self._hover_vert_local_idx: Optional[np.ndarray]= None
        self._hover_edge_starts:    Optional[np.ndarray]= None
        self._hover_edge_ends:      Optional[np.ndarray]= None
        self._hover_edge_room_idx_arr: Optional[np.ndarray]= None
        self._hover_edge_local_idx: Optional[np.ndarray]= None
        # Blitting state for drag operations
        self._blit_background                           = None
        self._blit_active:          bool                = False
        
        # Prefetching state
        self._prefetching_pages:    set                 = set()
        self._prefetching_hdrs:     set                 = set()
        self._blit_save_timer                           = None
        # Blitting state for DF cursor readout
        self._df_cursor_bg                              = None
        self._image_handle                              = None
        self.ax_legend                                  = None
        # Stamped DF readings: hdr_name → list of (x, y, df_val, px, py)
        self._df_stamps:            dict                = {}
        # Placement mode: when True, left-click stamps DF%; when False, click only selects rooms
        self.placement_mode:        bool                = False

        # Room type tagging (BED / LIVING — LIVING requires sub-rooms)
        self.room_type:             Optional[str]       = None
        self.multi_selected_room_idxs: set              = set()   # Ctrl+click multi-select
        self._last_list_click_idx:  Optional[int]       = None   # for Shift+click range select
        self._room_list_flat_items: List[Tuple]         = []     # ordered items for range select

        # IESVE AOI level assignment state
        self._aoi_level_idx:        int                 = 0       # current FFL group index

        # PDF floor plan overlay state
        self._overlay_pdf_path: Optional[Path]          = Path(pdf_path) if pdf_path else None
        self._overlay_pdf_info: Optional[dict]          = None
        self._overlay_page_idx: int                     = 0
        self._overlay_rgba:     Optional[np.ndarray]    = None   # cached rasterized page (H,W,4)
        self._overlay_visible:  bool                    = False
        self._overlay_alpha:    float                   = 0.6
        self._overlay_raster_dpi: int                  = 150    # global PDF rasterization DPI
        self._overlay_cache_pdf:  Optional[str]        = None   # PDF path used for cached raster
        self._overlay_cache_dpi:  Optional[int]        = None   # DPI used for cached raster
        self._overlay_handle                            = None   # matplotlib AxesImage artist
        self._dpi_dropdown_visible: bool                = False
        self._dpi_radio                                 = None
        # Per-HDR alignment: {hdr_name: {offset_x, offset_y, scale_x, scale_y, rotation_90, page_idx}}
        self._overlay_transforms: dict                  = {}
        # Two-point alignment mode
        self._align_mode:       bool                    = False
        self._align_points_overlay: list                = []
        self._align_points_hdr: list                    = []
        self._align_markers:    list                    = []
        # Arrow-key acceleration for overlay alignment
        self._align_key_press_start: Optional[float]    = None
        self._align_key_last:   Optional[str]           = None

        # Daylight factor analysis — thresholds are fixed per room type
        self.DF_THRESHOLDS          = {'BED': 0.5, 'LIVING': 1.0, 'NON-RESI': 2.0}
        self._df_image:             Optional[np.ndarray]= None   # (H, W) DF% for current HDR
        self._df_image_cache:       dict                = {}     # hdr_path_str → np.ndarray
        self._room_df_results:      dict                = {}     # room_idx → list of result strings
        self._hdr2wpd:              Optional[Hdr2Wpd]   = None
        # Persistent DF cache: keyed by (hdr_file, vertices_tuple, thresholds_tuple)
        # Stored on each room dict as 'df_cache' = {'thresholds': [...], 'results': [...], 'vertices_hash': str}
        # This avoids recomputation when rooms haven't changed.

        # Room divider mode (sub-state of edit mode)
        self.divider_mode:          bool                = False
        self._divider_room_idx:     Optional[int]       = None   # room being divided
        self._divider_points:       list                = []     # placed (x,y) tuples
        self._divider_markers:      list                = []     # scatter artists per point
        self._divider_segments:     list                = []     # solid Line2D per segment
        self._divider_preview_line                      = None   # dashed Line2D to cursor
        self._divider_snap_pt:      Optional[tuple]     = None   # snapped vertex (x, y) or None
        self._divider_snap_marker                       = None   # highlight ring artist

    # === LAYOUT HELPERS ========================================================

    def _axes(self, x, y, w, h):
        """Create figure axes at (x, y) measured from top-left corner.

        x increases right, y increases down — unlike matplotlib's native
        bottom-left origin. Converts to fig.add_axes([left, bottom, w, h]).
        """
        return self.fig.add_axes((x, 1.0 - y - h, w, h))

    def _make_button(self, x, y, w, h, label, callback,
                     fontsize=7, color=None, hovercolor=None):
        """Create a styled Button with rounded corners, shadow, and hover effect."""
        ax = self._axes(x, y, w, h)
        base_color  = color or self._btn_color
        hover_color = hovercolor or self._btn_hover

        # Hide default axes chrome
        ax.set_facecolor('none')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.set_navigate(False)
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)

        # Drop shadow
        shadow = FancyBboxPatch(
            (0.03, 0.03), 0.94, 0.86,
            boxstyle='round,pad=0.04', facecolor='#00000015', edgecolor='none',
            transform=ax.transAxes, clip_on=False, zorder=1)
        ax.add_patch(shadow)

        # Main button body
        body = FancyBboxPatch(
            (0.02, 0.08), 0.96, 0.86,
            boxstyle='round,pad=0.04', facecolor=base_color,
            edgecolor='#B0B0A8', linewidth=0.8,
            transform=ax.transAxes, clip_on=False, zorder=2)
        ax.add_patch(body)

        # Highlight strip (subtle top gradient effect)
        highlight = FancyBboxPatch(
            (0.05, 0.55), 0.90, 0.34,
            boxstyle='round,pad=0.03', facecolor='#FFFFFF30', edgecolor='none',
            transform=ax.transAxes, clip_on=False, zorder=3)
        ax.add_patch(highlight)

        # Label
        txt = ax.text(0.5, 0.48, label, fontsize=fontsize,
                      fontweight='medium', color='#333333',
                      ha='center', va='center',
                      transform=ax.transAxes, zorder=4)

        # Build a button-like object to stay compatible with existing code
        btn = Button(ax, '', color='none', hovercolor='none')
        btn.label = txt
        btn.color = base_color
        btn.hovercolor = hover_color
        btn._body = body
        btn._shadow = shadow
        btn._highlight = highlight

        # Hover enter/leave callbacks
        def _on_enter(event):
            if event.inaxes == ax:
                body.set_facecolor(hover_color)
                body.set_edgecolor('#909088')
                shadow.set_facecolor('#00000025')
                ax.figure.canvas.draw_idle()

        def _on_leave(event):
            body.set_facecolor(btn.color)
            body.set_edgecolor('#B0B0A8')
            shadow.set_facecolor('#00000018')
            ax.figure.canvas.draw_idle()

        ax.figure.canvas.mpl_connect('axes_enter_event', _on_enter)
        ax.figure.canvas.mpl_connect('axes_leave_event', _on_leave)

        btn.on_clicked(callback)
        return btn

    def _make_label(self, x, y, w, h, text, fontsize=9, **kwargs):
        """Create a text label on an invisible axes."""
        ax = self._axes(x, y, w, h)
        ax.axis('off')
        txt = ax.text(0, 0.5, text, fontsize=fontsize, fontweight='bold',
                      color='#404040', **kwargs)
        return ax, txt

    def _is_room_on_current_hdr(self, room: dict) -> bool:
        """Check whether a room belongs to the currently displayed HDR file."""
        return room.get('hdr_file') == self.current_hdr_name

    def _reset_hover_state(self):
        """Clear all hover and drag tracking variables."""
        self.hover_room_idx      = None
        self.hover_vertex_idx    = None
        self.hover_edge_room_idx = None
        self.hover_edge_idx      = None
        self.hover_edge_point    = None
        self.edit_room_idx       = None
        self.edit_vertex_idx     = None
        self.edit_edge_room_idx  = None
        self.edit_edge_idx       = None
        self.edit_edge_start     = None
        self._edit_drag_origin   = None

    def _start_blit_drag(self, room_idx: int):
        """Capture background for blitting before starting a drag operation."""
        canvas = self.fig.canvas
        if not getattr(canvas, 'supports_blit', False):
            return
        patch = self._room_patch_cache.get(room_idx)
        if patch is None:
            return
        # Hide the patch, draw the static background, then restore
        patch.set_visible(False)
        label = self._room_label_cache.get(room_idx)
        if label is not None:
            label.set_visible(False)
        canvas.draw()
        self._blit_background = canvas.copy_from_bbox(self.ax.bbox)
        patch.set_visible(True)
        if label is not None:
            label.set_visible(True)
        self._blit_active = True

    def _end_blit_drag(self):
        """Clear blitting state after a drag ends."""
        self._blit_active = False
        self._blit_background = None

    def _update_dragged_patch(self, room_idx: int):
        """Update a cached patch/label after vertex or edge drag. Returns True if handled.
        Uses blitting when available for fast partial redraws.
        Label position is deferred to drag-end (force_full render)."""
        patch = self._room_patch_cache.get(room_idx)
        if patch is not None:
            patch.set_xy(self.rooms[room_idx]['vertices'])
            # Centroid/label update deferred to button release (_render_section force_full)
            now = time.monotonic()
            if now - self._last_drag_draw >= 0.033:  # ~30 fps throttle
                self._last_drag_draw = now
                canvas = self.fig.canvas
                if self._blit_active and self._blit_background is not None:
                    canvas.restore_region(self._blit_background)
                    self.ax.draw_artist(patch)
                    canvas.blit(self.ax.bbox)
                else:
                    canvas.draw_idle()
            return True
        self._render_section(force_full=True)
        return True

    # === COORDINATE MAPPING ====================================================

    def _load_from_aoi_files(self, aoi_dir: Path):
        """Load room boundaries from .aoi files (pixel coordinates already included).

        AOI format:
            AOI Points File: <name> <level>
            ASSOCIATED VIEW FILE: plan_ffl_<z_mm>.vp
            FFL z height(m): <z>
            CENTRAL x,y: <cx> <cy>
            NO. PERIMETER POINTS <n>: x,y pixel_x pixel_y positions
            <world_x> <world_y> <pixel_x> <pixel_y>
            ...
        """
        aoi_files = sorted(aoi_dir.glob('*.aoi'))
        for aoi_path in aoi_files:
            with open(aoi_path, 'r') as f:
                lines = [l.strip() for l in f.readlines()]
            if len(lines) < 6:
                continue

            # Header: name from line 0
            name_match = re.match(r'AOI Points File:\s*(.+)', lines[0])
            name       = name_match.group(1).strip() if name_match else aoi_path.stem

            # HDR file from view file reference (plan_ffl_28700.vp → model_plan_ffl_28700)
            vp_match = re.search(r'plan_ffl_(\d+)', lines[1])
            hdr_file = self.current_hdr_name
            if vp_match:
                ffl_val = vp_match.group(1)
                for entry in self.hdr_files:
                    if ffl_val in entry['name']:
                        hdr_file = entry['name']
                        break

            # Vertex lines: world_x world_y pixel_x pixel_y
            vertices = []
            for line in lines[5:]:
                parts = line.split()
                if len(parts) >= 4:
                    px, py = float(parts[2]), float(parts[3])
                    vertices.append([px, py])

            if len(vertices) >= 3:
                self.rooms.append({
                    'name':     name,
                    'parent':   None,
                    'vertices': vertices,
                    'hdr_file': hdr_file,
                })

        print(f"Loaded {len(self.rooms)} rooms from {len(aoi_files)} .aoi files in {aoi_dir}")

    def _load_from_iesve_aoi(self) -> int:
        """Load room boundaries from IESVE .aoi files (world X/Y only).

        Reads the IESVE .aoi format (no pixel coords, no z-height) from
        self.aoi_dir and uses iesve_room_data CSV to look up the FFL (finished
        floor level) for each room by Space ID. World coordinates are projected
        to pixel coordinates using the VIEW header of the first .pic/.hdr file.

        Each room dict gains two extra keys beyond the standard set:
            'ffl'            - finished floor level in metres (float)
            'world_vertices' - original world X/Y pairs, preserved for
                               re-projection when the user cycles AOI levels

        All rooms are initially assigned to the first HDR entry. The user then
        uses the AOI Level cycle button to reassign each FFL group to the
        correct .pic level, at which point vertices are re-projected using that
        level's VIEW header and hdr_file is updated.

        Returns the number of rooms loaded.
        """
        csv_path = self._iesve_room_data_path
        if not self.hdr_files:
            print("Warning: no image files found, cannot project IESVE AOI vertices.")
            return 0

        # Build Space ID → FFL lookup from CSV if provided
        ffl_lookup: dict = {}
        if csv_path is not None and csv_path.exists():
            try:
                df_csv = pd.read_csv(csv_path, encoding='utf-8')
                id_col  = 'Space ID'
                ffl_col = 'Min. Height (m) (Real)'
                if id_col in df_csv.columns and ffl_col in df_csv.columns:
                    ffl_lookup = dict(zip(df_csv[id_col].astype(str), df_csv[ffl_col]))
                else:
                    print(f"Warning: iesve_room_data missing '{id_col}' or '{ffl_col}' columns.")
            except Exception as exc:
                print(f"Warning: could not read iesve_room_data CSV: {exc}")

        # Get VIEW params from the first image file for initial projection
        first_entry = self.hdr_files[0]
        view_params = self._read_view_params(first_entry['hdr_path'])
        if view_params is None:
            print(f"Warning: no VIEW params in {first_entry['hdr_path'].name}, cannot project.")
            return 0
        vp_x, vp_y, vh_val, vv_val, img_w, img_h = view_params

        aoi_files = sorted(self.aoi_dir.glob('*.aoi'))
        count = 0
        for aoi_path in aoi_files:
            with open(aoi_path, 'r') as f:
                lines = [l.strip() for l in f.readlines()]

            # IESVE format: line 0 = header, line 1 = ZONE id name, line 2 = POINTS n
            if len(lines) < 4:
                continue
            zone_match = re.match(r'ZONE\s+(\S+)\s+(.*)', lines[1])
            if not zone_match:
                continue
            space_id  = zone_match.group(1)
            room_name = zone_match.group(2).strip()
            ffl       = ffl_lookup.get(space_id, 0.0)

            world_verts = []
            for line in lines[3:]:
                parts = line.split()
                if len(parts) >= 2:
                    try:
                        world_verts.append([float(parts[0]), float(parts[1])])
                    except ValueError:
                        continue

            if len(world_verts) < 3:
                continue

            pixels = self._world_to_pixels(world_verts, vp_x, vp_y, vh_val, vv_val, img_w, img_h)
            self.rooms.append({
                'name':           f"{space_id} {room_name}",
                'parent':         None,
                'vertices':       pixels,
                'world_vertices': world_verts,
                'ffl':            ffl,
                'hdr_file':       first_entry['name'],
            })
            count += 1

        # Auto-distribute FFL groups across HDR levels (sorted by height → sorted
        # HDR entries).  Each group is re-projected using the target level's VIEW
        # params so rooms appear on the correct image from the start.
        ffl_groups = sorted({r['ffl'] for r in self.rooms if 'ffl' in r})
        if len(ffl_groups) > 1 and len(self.hdr_files) >= len(ffl_groups):
            ffl_to_entry = dict(zip(ffl_groups, self.hdr_files))
            view_cache: dict = {}
            for ffl_val, entry in ffl_to_entry.items():
                vp = self._read_view_params(entry['hdr_path'])
                if vp is not None:
                    view_cache[ffl_val] = (entry, vp)
            for room in self.rooms:
                cached = view_cache.get(room.get('ffl'))
                if cached and 'world_vertices' in room:
                    entry, (vx, vy, vh_v, vv_v, w, h) = cached
                    room['vertices'] = self._world_to_pixels(
                        room['world_vertices'], vx, vy, vh_v, vv_v, w, h)
                    room['hdr_file'] = entry['name']
            assigned = [f"{ffl} m -> {ffl_to_entry[ffl]['name']}" for ffl in ffl_groups if ffl in view_cache]
            print(f"Auto-assigned {len(assigned)} FFL groups: {', '.join(assigned)}")

        print(f"Loaded {count} IESVE AOI rooms from {self.aoi_dir}")
        return count

    @staticmethod
    def _read_view_params(pic_path: Path):
        """Extract (vp_x, vp_y, vh, vv, img_w, img_h) from a .pic/.hdr VIEW header.

        Returns None if the VIEW line is missing or incomplete.
        """
        try:
            with open(pic_path, 'r', encoding='utf-8', errors='ignore') as f:
                view_line = next((l.strip() for l in f if l.startswith('VIEW=')), None)
            if not view_line:
                return None
            vp = re.search(r'-vp\s+([\d.-]+)\s+([\d.-]+)', view_line)
            vh = re.search(r'-vh\s+([\d.-]+)', view_line)
            vv = re.search(r'-vv\s+([\d.-]+)', view_line)
            if not (vp and vh and vv):
                return None
            vp_x, vp_y = float(vp.group(1)), float(vp.group(2))
            vh_val = float(vh.group(1))
            vv_val = float(vv.group(1))
            img = imageio.imread(str(pic_path))
            img_h, img_w = img.shape[:2]
            return vp_x, vp_y, vh_val, vv_val, img_w, img_h
        except Exception as exc:
            print(f"Warning: could not read VIEW params from {pic_path}: {exc}")
            return None

    @staticmethod
    def _world_to_pixels(world_verts, vp_x, vp_y, vh_val, vv_val, img_w, img_h):
        """Convert a list of [world_x, world_y] to pixel [px, py] coordinates.

        Uses the inverse of the Radiance orthographic (-vtl) projection:
            px = (world_x - vp_x) / (vh / img_w) + img_w / 2
            py = img_h / 2 - (world_y - vp_y) / (vv / img_h)
        """
        pixels = []
        for wx, wy in world_verts:
            px = (wx - vp_x) / (vh_val / img_w) + img_w / 2
            py = img_h / 2 - (wy - vp_y) / (vv_val / img_h)
            pixels.append([px, py])
        return pixels

    # === IMAGE SCANNING & LOADING ==============================================

    def _scan_hdr_files(self) -> List[dict]:
        """Scan image_dir for HDR files and associated TIFFs.

        Returns:
            Sorted list of dicts with keys: hdr_path, tiff_paths, name (stem).
        """
        hdr_paths = sorted([*self.image_dir.glob("*.hdr"), *self.image_dir.glob("*.pic")])
        result = []
        for hdr_path in hdr_paths:
            stem = hdr_path.stem
            # Associated PNGs: any .png in same dir whose stem starts with stem + '_'
            # Exclude previously-exported aoi_overlay files to avoid re-processing them
            tiff_paths = sorted(
                p for p in self.image_dir.glob("*.png")
                if p.stem.startswith(stem + "_") and not p.stem.endswith("_aoi_overlay")
            )
            result.append({
                'hdr_path': hdr_path,
                'tiff_paths': tiff_paths,
                'name': stem,
            })

        # Build legend map: key → Path for files matching '*_legend.png'
        # e.g. 'df_cntr' → Path('df_cntr_legend.png')
        self.legend_map: dict = {}
        for legend_path in sorted(self.image_dir.glob("*_legend.png")):
            key = legend_path.stem[: -len("_legend")]  # strip trailing '_legend'
            self.legend_map[key] = legend_path
        if self.legend_map:
            print(f"Found legend(s): {list(self.legend_map.keys())}")

        print(f"Found {len(result)} HDR file(s) in {self.image_dir}")
        return result

    def _rebuild_image_variants(self):
        """Rebuild the image_variants list for the current HDR index.

        Preserves the active layer type across HDR navigation by matching the
        suffix of the current variant (e.g. '_df_false') against the new HDR's
        variants. Falls back to index 0 (the HDR itself) if no match is found.
        """
        if not self.hdr_files:
            self.image_variants = []
            self.current_variant_idx = 0
            return

        # Remember which layer suffix is currently active before rebuilding
        active_suffix = None
        if self.image_variants and 0 < self.current_variant_idx < len(self.image_variants):
            old_path = self.image_variants[self.current_variant_idx]
            if old_path.suffix.lower() != '.hdr':
                # Extract the part after the old HDR stem, e.g. '_df_false'
                old_hdr_stem = self.hdr_files[self.current_hdr_idx]['name']
                active_suffix = old_path.stem[len(old_hdr_stem):]

        entry = self.hdr_files[self.current_hdr_idx]
        self.image_variants = [entry['hdr_path']] + list(entry['tiff_paths'])

        # Try to restore the same layer type in the new HDR's variants
        if active_suffix:
            new_hdr_stem = entry['name']
            for i, path in enumerate(self.image_variants):
                if path.suffix.lower() != '.hdr' and path.stem[len(new_hdr_stem):] == active_suffix:
                    self.current_variant_idx = i
                    return

        self.current_variant_idx = 0

    @property
    def current_hdr_name(self) -> str:
        """Stem of the currently active HDR file."""
        if not self.hdr_files:
            return ""
        return self.hdr_files[self.current_hdr_idx]['name']

    @property
    def current_variant_path(self) -> Optional[Path]:
        """Path of the currently displayed image variant."""
        if not self.image_variants:
            return None
        idx = self.current_variant_idx % len(self.image_variants)
        return self.image_variants[idx]

    def _load_image(self, path: Path) -> Optional[np.ndarray]:
        """Load an image file as a normalised float32 numpy array.

        Results are cached in memory so repeated renders don't hit disk.
        Thread-safe: uses a lock and maintains an LRU-style cache size limit.

        Args:
            path: Path to .hdr or .tiff file.

        Returns:
            Array of shape (H, W, 3) with values in [0, 1], or None on failure.
        """
        key = str(path)
        with self._image_cache_lock:
            if key in self._image_cache:
                # Move to end to maintain LRU order (if using OrderedDict, but here
                # we just use a simple dict for compatibility)
                return self._image_cache[key]

        try:
            if path.suffix.lower() in ('.hdr', '.pic'):
                img = imageio.imread(str(path)).astype(np.float32)
                if img.ndim == 2:
                    img = np.stack([img, img, img], axis=-1)
                p99 = np.percentile(img, 99)
                if p99 > 0:
                    img = img / p99
                img = np.clip(img ** (1.0 / 2.2), 0.0, 1.0)
            else:
                pil_img = Image.open(path).convert('RGB')
                img = np.array(pil_img, dtype=np.float32) / 255.0

            with self._image_cache_lock:
                # Enforce cache size limit (LRU-ish)
                if len(self._image_cache) >= self._image_cache_limit:
                    # Drop a random item (or first key if Python 3.7+)
                    first_key = next(iter(self._image_cache))
                    self._image_cache.pop(first_key)
                
                self._image_cache[key] = img
            return img
        except Exception as exc:
            print(f"Warning: could not load image {path}: {exc}")
            return None

    def _start_hdr_prefetch(self):
        """Trigger background loading for adjacent HDR and TIFF images.

        Identifies the next and previous floor levels and spawns a thread
        to pre-load their assets into the memory cache.
        """
        if not self.hdr_files:
            return

        targets = []
        curr = self.current_hdr_idx
        # Check neighbors: next and previous
        for idx in [curr + 1, curr - 1]:
            if 0 <= idx < len(self.hdr_files):
                entry = self.hdr_files[idx]
                # Priority 1: the HDR itself
                targets.append((idx, entry['hdr_path']))
                # Priority 2: associated TIFFs (e.g. falsecolor)
                for tiff_path in entry.get('tiff_paths', []):
                    targets.append((idx, tiff_path))

        # Filter out what's already cached or already being fetched
        with self._image_cache_lock:
            to_fetch = []
            for idx, path in targets:
                path_str = str(path)
                if path_str not in self._image_cache and path_str not in self._prefetching_hdrs:
                    to_fetch.append((idx, path))
                    self._prefetching_hdrs.add(path_str)

        if not to_fetch:
            return

        def _prefetch_worker():
            try:
                for idx, path in to_fetch:
                    # RELEVANCE VALIDATION:
                    # If the user has spammed buttons and moved far away from this floor,
                    # abort the expensive loading process immediately.
                    if abs(idx - self.current_hdr_idx) > 2:
                        continue
                    
                    self._load_image(path)
                    
                    with self._image_cache_lock:
                        self._prefetching_hdrs.discard(str(path))
            except Exception as e:
                print(f"Prefetch error: {e}")

        thread = threading.Thread(target=_prefetch_worker, daemon=True)
        thread.start()

    def _get_legend_for_variant(self, path: Optional[Path]) -> Optional[Path]:
        """Return the legend TIFF matching the given image variant, or None.

        Matching rule: for a TIFF whose stem is '<hdr_stem>_<suffix>', find
        the first legend key that is a substring of '<suffix>'.
        HDR variants never have an associated legend.
        """
        if path is None or path.suffix.lower() == '.hdr':
            return None
        hdr_stem = self.current_hdr_name
        suffix = path.stem[len(hdr_stem) + 1:] if path.stem.startswith(hdr_stem + "_") else path.stem
        for key, legend_path in self.legend_map.items():
            if key in suffix:
                return legend_path
        return None

    # === LAUNCH ================================================================

    _WINDOW_TITLE = "Archilume - HDR AOI Editor"

    @staticmethod
    def _get_monitor_rects() -> list:
        """Return a list of (left, top, right, bottom) rects for all active monitors.

        Uses EnumDisplayMonitors via ctypes on Windows. Falls back to an empty
        list on other platforms (caller should handle the fallback).
        """
        rects = []
        if sys.platform != "win32":
            return rects
        try:
            MonitorEnumProc = ctypes.WINFUNCTYPE(
                ctypes.c_bool,
                ctypes.c_ulong, ctypes.c_ulong,
                ctypes.POINTER(ctypes.wintypes.RECT),
                ctypes.c_double,
            )
            def _callback(hMonitor, hdcMonitor, lprcMonitor, dwData):
                r = lprcMonitor.contents
                rects.append((r.left, r.top, r.right, r.bottom))
                return True
            ctypes.windll.user32.EnumDisplayMonitors(None, None, MonitorEnumProc(_callback), 0)
        except Exception:
            pass
        return rects

    def launch(self):
        """Open the interactive editor window."""
        if not self.hdr_files:
            raise FileNotFoundError(f"No .hdr files found in {self.image_dir}")

        # Close any previous editor window that was left open
        for fig_num in plt.get_fignums():
            fig = plt.figure(fig_num)
            if getattr(fig, '_archilume_editor', False):
                plt.close(fig)

        # Initialise daylight factor analysis (Phase 3 from daylight_workflow_iesve)
        try:
            coordinate_map_path = utils.create_pixel_to_world_coord_map(self.image_dir)
            self._hdr2wpd = Hdr2Wpd(pixel_to_world_map=coordinate_map_path)
            print(f"DF analysis ready (area_per_pixel={self._hdr2wpd.area_per_pixel} m2)")
        except Exception as exc:
            print(f"Warning: DF analysis unavailable ({exc}). Results will not be shown.")
            self._hdr2wpd = None

        # Capture the pdf_path passed to __init__ before _load_session() can overwrite it.
        # An explicit constructor argument always takes priority over the session-stored path.
        _init_pdf_path = self._overlay_pdf_path

        # Load existing session to get window settings before creating the figure
        self._load_session()

        # Re-apply constructor pdf_path if one was explicitly provided
        if _init_pdf_path is not None:
            self._overlay_pdf_path = _init_pdf_path
            self._overlay_needs_rasterize = True

        # Setup matplotlib figure — wide to match ~2.6:1 floor plan aspect ratio
        plt.rcParams['savefig.directory'] = str(self.image_dir)
        plt.rcParams['keymap.save']       = []  # disable 's' / 'ctrl+s' save-figure hotkey
        plt.rcParams['keymap.fullscreen'] = []  # disable 'f' fullscreen hotkey
        # Small initial size (matplotlib figsize API requires inches); window is
        # immediately maximised so this value is only used for the brief render before
        # _force_resize_update corrects it to the actual screen dimensions.
        self.fig = plt.figure(figsize=(8, 5), facecolor='#F5F5F0')
        self.fig._archilume_editor = True
        self.fig.canvas.manager.set_window_title(self._WINDOW_TITLE)
        self.fig.subplots_adjust(left=0.02, right=0.98, top=0.95, bottom=0.02)

        # Restore window position and maximization state with safety check
        try:
            manager = self.fig.canvas.manager
            if manager and hasattr(manager, 'window'):
                window = manager.window
                
                # Bind configure event to track moves/resizes while window is healthy
                window.bind('<Configure>', self._on_window_configure)

                # Apply saved geometry before maximization
                ws = self.window_settings
                if ws and all(k in ws for k in ('x', 'y', 'width', 'height')):
                    x, y, w, h = ws['x'], ws['y'], ws['width'], ws['height']
                    
                    # Ignore tiny/square "dying" states (e.g. 200x200) captured during close
                    if w > 400 and h > 400:
                        # Safety check: ensure window top-left falls within an active monitor.
                        # Uses EnumDisplayMonitors on Windows for accurate multi-monitor rects;
                        # falls back to winfo_screenwidth/height on other platforms.
                        try:
                            monitor_rects = self._get_monitor_rects()
                            if monitor_rects:
                                on_screen = any(
                                    l <= x < r and t <= y < b
                                    for l, t, r, b in monitor_rects
                                )
                                if not on_screen:
                                    print(f"Window position ({x}, {y}) is off-screen (no active monitor). Resetting to primary.")
                                    l, t, r, b = monitor_rects[0]
                                    sw, sh = r - l, b - t
                                    w = min(w, sw)
                                    h = min(h, sh)
                                    x = l + (sw - w) // 2
                                    y = t + (sh - h) // 2
                            else:
                                # Non-Windows fallback
                                sw = window.winfo_screenwidth()
                                sh = window.winfo_screenheight()
                                if not (0 <= x < sw and 0 <= y < sh):
                                    print(f"Window position ({x}, {y}) is off-screen. Resetting to primary.")
                                    w = min(w, sw)
                                    h = min(h, sh)
                                    x = (sw - w) // 2
                                    y = (sh - h) // 2
                        except Exception:
                            pass
                        
                        # Apply geometry using absolute coordinates (works across multiple monitors)
                        window.geometry(f"{w}x{h}+{x:d}+{y:d}")
                        window.update_idletasks()
                
                # Apply maximization state
                is_maximized = ws.get('maximized', True) # default to maximized
                if sys.platform == "win32":
                    if is_maximized:
                        window.state('zoomed')
                else:
                    if is_maximized:
                        window.attributes('-zoomed', True)
                    else:
                        window.attributes('-zoomed', False)
                
                window.update_idletasks()

                self.fig.canvas.mpl_connect('resize_event', self._on_resize)
                # Fire once early (handles non-maximised restores), then again after
                # the OS has finished processing the zoomed state (Windows is async).
                window.after(100, self._force_resize_update)
                window.after(400, self._force_resize_update)
                # Intercept the OS window-close (X button)
                window.protocol('WM_DELETE_WINDOW', lambda: (self._on_close(None), plt.close(self.fig)))
        except (AttributeError, Exception) as exc:
            print(f"Warning: could not restore window state: {exc}")

        # Main plot area — maximised to fill available space
        self.ax = self._axes(0.02, 0.21, 0.96, 0.66)
        self.ax.set_aspect('equal', adjustable='box')
        self.ax.set_facecolor('#FAFAF8')

        # DF% cursor readout — bottom-left corner of the main axes
        self._df_cursor_text = self.ax.text(
            0, 0, '', transform=self.ax.transData,
            fontsize=8, color='white', va='bottom', ha='left',
            bbox=dict(boxstyle='round,pad=0.2', facecolor='black', alpha=0.55, edgecolor='none'),
            zorder=300, visible=False, animated=True)

        # DF% legend axes: top-right, just above the main image (rotated 90°)
        # DF legend — positioned in _setup_bottom_toolbar after buttons are laid out
        self.ax_legend = None

        # Setup side panel
        self._setup_side_panel()

        # Apply initial bevel styling to toggle buttons
        self._style_toggle_button(self.btn_edit_mode, self.edit_mode)
        self._style_toggle_button(self.btn_ortho, self.ortho_mode)
        self._style_toggle_button(self.btn_placement, self.placement_mode)
        self._update_room_type_buttons()

        # Initial render - force_full=True is critical to load the image
        self._render_section(force_full=True)

        # Store original limits
        self.original_xlim = self.ax.get_xlim()
        self.original_ylim = self.ax.get_ylim()

        # Polygon selector for drawing — created inactive; 'd' activates draw mode
        self._create_polygon_selector()

        # Event handlers
        self.fig.canvas.mpl_connect('button_press_event', self._on_click_with_snap)
        self.fig.canvas.mpl_connect('button_press_event', self._on_list_click)
        self.fig.canvas.mpl_connect('button_release_event', self._on_button_release)
        self.fig.canvas.mpl_connect('motion_notify_event', self._on_mouse_motion)
        self.fig.canvas.mpl_connect('key_press_event', self._on_key_press)
        self.fig.canvas.mpl_connect('key_release_event', self._on_key_release)
        self.fig.canvas.mpl_connect('scroll_event', self._on_scroll)
        self.fig.canvas.mpl_connect('close_event', self._on_close)

        # Apply loaded PDF resolution to button label
        self.btn_overlay_dpi.label.set_text(f'PDF Res: {self._overlay_raster_dpi} DPI')

        # Sync overlay button label and style with loaded state
        if self._overlay_visible:
            self.btn_overlay_toggle.label.set_text('Floor Plan: ON')
            self._style_toggle_button(self.btn_overlay_toggle, True)
        else:
            self.btn_overlay_toggle.label.set_text('Floor Plan: OFF')
            self._style_toggle_button(self.btn_overlay_toggle, False)

        # Deferred PDF overlay rasterization (from session restore or __init__ pdf_path)
        if getattr(self, '_overlay_needs_rasterize', False) or (
                self._overlay_pdf_path is not None and self._overlay_rgba is None):
            try:
                from archilume.utils import get_pdf_info
                self._overlay_pdf_info = get_pdf_info(self._overlay_pdf_path)
                self._rasterize_overlay_page()
                self._update_overlay_page_label()
                self._render_section(force_full=True)
            except Exception as e:
                traceback.print_exc()
                print(f"Warning: could not load PDF overlay: {e}")
            self._overlay_needs_rasterize = False

        self._update_room_list()
        self._update_hdr_list()

        print("\n=== HDR Boundary Editor ===")
        print(f"Loaded {len(self.hdr_files)} HDR file(s) from {self.image_dir}")
        print("Use Up/Down to navigate HDR files, 't' to toggle image variant.")
        print("Scroll: zoom | Right-click: select room | s: save | d: delete | q: quit")
        print("===========================\n")
        
        # Bind configure event to track moves/resizes while window is healthy
        try:
            self.fig.canvas.manager.window.bind('<Configure>', self._on_window_configure)
        except Exception:
            pass

        try:
            plt.show()
        except KeyboardInterrupt:
            # Clean exit on Ctrl+C or window force-close
            pass

    def _on_window_configure(self, event):
        """Track window geometry and maximization state while window is healthy."""
        if self._closing:
            return
        
        # Debounce: only capture at most once every 200ms to avoid GIL churn
        now = time.monotonic()
        if now - getattr(self, '_last_win_capture', 0.0) < 0.2:
            return
        self._last_win_capture = now

        try:
            window = self.fig.canvas.manager.window
            
            # Maximization state
            is_maximized = False
            if sys.platform == "win32":
                is_maximized = window.state() == 'zoomed'
            else:
                is_maximized = window.attributes('-zoomed')

            # Only capture normal/maximized geometry, not minimized or intermediate closing states
            geom_str = window.geometry()
            match = re.match(r'(\d+)x(\d+)([+-]-?\d+)([+-]-?\d+)', geom_str)
            if match:
                w, h, x, y = match.groups()
                # 300x300 is a common dying state or tiny default; ignore it
                if int(w) > 300 and int(h) > 300:
                    self.window_settings = {
                        'x': int(x.lstrip('+').replace('-', '-')), # handle negative signs correctly
                        'y': int(y.lstrip('+').replace('-', '-')),
                        'width': int(w),
                        'height': int(h),
                        'maximized': bool(is_maximized)
                    }
        except Exception:
            pass

    # === UI SETUP ==============================================================

    def _setup_side_panel(self):
        """Create the side panel with inputs, buttons, and room list."""
        self._btn_color    = '#E8E8E0'
        self._btn_hover    = '#D8D8D0'
        self._btn_on_color = '#8A8A84'
        self._btn_on_hover = '#7A7A74'

        self._setup_instructions_panel()
        status_y = self._setup_input_panels()
        action_bot_y = self._setup_action_buttons(status_y)
        self._setup_room_list_panel(action_bot_y)
        self._setup_bottom_toolbar()
        self._setup_colour_key_legend()

    # Layout constants shared across sub-methods
    _PL     = 0.02    # panel left
    _PW     = 0.28    # panel width
    _LBL_H  = 0.016
    _INP_H  = 0.032
    _GAP    = 0.008
    _SUB_H  = 0.014
    _ROW_Y  = 0.02    # top row y
    _INSTR_W = 0.12
    _COL2_X = _PL + _INSTR_W + 0.01
    _PRNT_X   = _COL2_X + 0.18
    _FIELD_W  = 0.150 * 0.75            # single field width (Parent / Name)
    _PRNT_W   = _FIELD_W * 2 + _GAP     # full row width (spans both fields)

    def _setup_instructions_panel(self):
        """Create the keyboard shortcut reference panel (top-left)."""
        ax = self._axes(self._PL, 0.02, self._INSTR_W, 0.15)
        ax.axis('off')
        ax.patch.set_visible(False)
        ax.text(0, 0.95, "HDR BOUNDARY EDITOR", fontsize=9, fontweight='bold',
                color='#404040', transform=ax.transAxes)
        controls = [
            ("\u2191/\u2193", "Navigate HDR files"),   ("t",           "Toggle image (HDR / TIFFs)"),
            ("Left-click",    "Place vertex / drag"),   ("Shift+click", "Drag edge (edit mode)"),
            ("Right-click",   "Select existing room"),  ("Scroll",      "Zoom centred on cursor"),
            ("s",             "Save room / confirm edit"),
            ("e",             "Toggle Edit Mode"),      ("d",           "Room divider (edit mode)"),
            ("ctrl+z",        "Undo (edit/type/name/create/del)"),
            ("o",             "Toggle orthogonal lines"), ("f",         "Fit zoom to selected room"),
            ("r",             "Reset zoom"),              ("p",         "Toggle DF% stamp on/off"),
            ("Ctrl+click",    "Multi-select rooms"),     ("ctrl+a",    "Select all rooms"),
            ("q",             "Quit"),
        ]
        for i, (key, desc) in enumerate(controls):
            y = 0.87 - i * 0.08
            ax.text(0.00, y, key,  fontsize=7.5, color='#404040', fontweight='bold', transform=ax.transAxes)
            ax.text(0.38, y, desc, fontsize=7.5, color='#505050', transform=ax.transAxes)

    def _setup_input_panels(self) -> float:
        """Create HDR nav, parent, name, room-type inputs. Returns status_y."""
        col2_x, row_y = self._COL2_X, self._ROW_Y
        lbl_h, inp_h, gap = self._LBL_H, self._INP_H, self._GAP
        prnt_x, prnt_w = self._PRNT_X, self._PRNT_W
        arrow_w = 0.025
        arrow_h = inp_h * 2 + gap

        # HDR FILES label + nav arrows + list
        self._make_label(col2_x, row_y, arrow_w * 2 + 0.004, lbl_h, "HDR FILES:")
        arrows_y = row_y + lbl_h + gap
        self.btn_next_hdr = self._make_button(col2_x, arrows_y, arrow_w, arrow_h,
                                              '\u25b2', self._on_next_hdr_click)
        self.btn_prev_hdr = self._make_button(col2_x + arrow_w + 0.004, arrows_y, arrow_w, arrow_h,
                                              '\u25bc', self._on_prev_hdr_click)
        hdr_list_x = col2_x + arrow_w * 2 + 0.008
        self.ax_hdr_list = self._axes(hdr_list_x, row_y, 0.060, lbl_h + gap + arrow_h)
        self.ax_hdr_list.axis('off')

        # Parent Apartment + Apartment Name (side by side)
        field_w = self._FIELD_W
        name_x  = prnt_x + field_w + gap

        self._make_label(prnt_x, row_y, field_w, lbl_h, "Parent Apartment:")
        self.btn_parent = self._make_button(prnt_x, row_y + lbl_h + gap, field_w, inp_h,
                                            '(None)', self._on_parent_cycle, fontsize=8)

        _, self.name_label_text = self._make_label(name_x, row_y, field_w, lbl_h, "Apartment Name:")
        ax_name = self._axes(name_x, row_y + lbl_h + gap, field_w, inp_h)
        self.name_textbox = TextBox(ax_name, '', initial='')
        self.name_textbox.on_text_change(self._on_name_changed)

        # Room Type (BED / LIVING / CIRC)
        rtype_y = row_y + lbl_h + gap + inp_h + gap
        self._make_label(prnt_x, rtype_y, prnt_w, lbl_h, "Room Type:")
        rtype_btn_y = rtype_y + lbl_h + gap
        n_type_btns = 4
        rtype_btn_w = (prnt_w - gap * (n_type_btns - 1)) / n_type_btns
        self.btn_room_type_bed = self._make_button(
            prnt_x, rtype_btn_y, rtype_btn_w, inp_h, 'BED',
            lambda e: self._on_room_type_toggle('BED'))
        self.btn_room_type_living = self._make_button(
            prnt_x + (rtype_btn_w + gap), rtype_btn_y, rtype_btn_w, inp_h, 'LIVING',
            lambda e: self._on_room_type_toggle('LIVING', requires_children=True))
        self.btn_room_type_nonresi = self._make_button(
            prnt_x + 2 * (rtype_btn_w + gap), rtype_btn_y, rtype_btn_w, inp_h, 'NON-RESI',
            lambda e: self._on_room_type_toggle('NON-RESI'))
        self.btn_room_type_circ = self._make_button(
            prnt_x + 3 * (rtype_btn_w + gap), rtype_btn_y, rtype_btn_w, inp_h, 'CIRC',
            lambda e: self._on_room_type_toggle('CIRC'))

        # Status / preview
        status_y = rtype_y + lbl_h + gap + inp_h + gap
        ax_preview = self._axes(prnt_x, status_y, prnt_w, self._SUB_H)
        ax_preview.axis('off')
        self.name_preview_text = ax_preview.text(0, 0.5, "", fontsize=8, color='#666666', style='italic')
        ax_status = self._axes(prnt_x, status_y, prnt_w, self._SUB_H)
        ax_status.axis('off')
        self.status_text = ax_status.text(
            0, 0.5, "Status: Ready to draw", fontsize=8, color='blue', style='italic')
        return status_y

    def _setup_action_buttons(self, status_y: float) -> float:
        """Create Save / Clear / Delete button row below status. Returns bottom y."""
        gap    = self._GAP
        prnt_x = self._PRNT_X
        prnt_w = self._PRNT_W
        btn_y  = status_y + self._SUB_H + gap
        btn_h  = self._INP_H
        n_btns = 3
        btn_w  = (prnt_w - gap * (n_btns - 1)) / n_btns

        for i, (attr, label, cb) in enumerate([
            ('btn_save',   'Save',   self._on_save_click),
            ('btn_clear',  'Clear',  self._on_clear_click),
            ('btn_delete', 'Delete', self._on_delete_click),
        ]):
            setattr(self, attr, self._make_button(
                prnt_x + i * (btn_w + gap), btn_y, btn_w, btn_h, label, cb))
        return btn_y + btn_h

    def _setup_room_list_panel(self, action_bot_y: float):
        """Create the scrollable saved-rooms list (right of input column)."""
        gap     = self._GAP
        list_x  = self._PRNT_X + self._PRNT_W + gap
        list_w  = 0.270
        list_y  = self._ROW_Y
        label_h = 0.020
        self._make_label(list_x, list_y, list_w, label_h, "SAVED ROOMS:")
        list_top = list_y + label_h + gap * 0.5
        list_h   = action_bot_y - list_top
        self.ax_list = self._axes(list_x, list_top, list_w, list_h)
        self.ax_list.set_facecolor('#FAFAF8')
        self.ax_list.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
        for spine in self.ax_list.spines.values():
            spine.set_edgecolor('#CCCCCC')
            spine.set_linewidth(0.5)

    def _setup_bottom_toolbar(self):
        """Create the three-row bottom toolbar with DF legend filling the right."""
        gap   = self._GAP
        tr_x  = 0.02                     # push buttons to far left
        tr_h  = 0.025                     # single row height
        row1_y = 0.88                     # top row
        row2_y = row1_y + tr_h + gap * 2
        row3_y = row2_y + tr_h + gap * 2

        # Button width: 5 columns across to fit the Reset button
        n_cols   = 5
        btn_w    = 0.115
        btn_step = btn_w + gap

        # Row 1: Toggle, Edit, Draw, Ortho
        self.btn_image_toggle = self._make_button(
            tr_x, row1_y, btn_w, tr_h,
            'Toggle Image Layers: HDR (Press T)', self._on_image_toggle_click)
        self.btn_edit_mode = self._make_button(
            tr_x + btn_step, row1_y, btn_w, tr_h,
            'Boundary Edit Mode: OFF (Press E)', self._on_edit_mode_toggle)
        self.btn_draw_mode = self._make_button(
            tr_x + 2 * btn_step, row1_y, btn_w, tr_h,
            'Draw Mode: OFF (Press D)', self._on_draw_mode_toggle)

        ortho_label = 'Ortho Lines: ON (Press O)' if self.ortho_mode else 'Ortho Lines: OFF (Press O)'
        ortho_color = self._btn_on_color if self.ortho_mode else self._btn_color
        ortho_hover = self._btn_on_hover if self.ortho_mode else self._btn_hover
        self.btn_ortho = self._make_button(
            tr_x + 3 * btn_step, row1_y, btn_w, tr_h,
            ortho_label, self._on_ortho_toggle, color=ortho_color, hovercolor=ortho_hover)

        # Row 2: Export & Archive, Extract Archive, Reset Zoom, Place DF%
        self.btn_export = self._make_button(
            tr_x, row2_y, btn_w, tr_h,
            'Export & Archive', self._on_export_report, color='#C8E6C9', hovercolor='#A5D6A7')
        self.btn_extract = self._make_button(
            tr_x + btn_step, row2_y, btn_w, tr_h,
            'Extract Archive', self._on_extract_click)
        self.btn_reset_zoom = self._make_button(
            tr_x + 2 * btn_step, row2_y, btn_w, tr_h,
            'Reset Zoom (Press R)', self._on_reset_zoom_click)

        place_label = 'Place DF% Point: ON (Press P)' if self.placement_mode else 'Place DF% Point: OFF (Press P)'
        place_color = self._btn_on_color if self.placement_mode else self._btn_color
        place_hover = self._btn_on_hover if self.placement_mode else self._btn_hover
        self.btn_placement = self._make_button(
            tr_x + 3 * btn_step, row2_y, btn_w, tr_h,
            place_label, self._on_placement_toggle, color=place_color, hovercolor=place_hover)

        # Row 3: PDF Overlay controls
        self.btn_overlay_toggle = self._make_button(
            tr_x, row3_y, btn_w, tr_h,
            'Floor Plan: OFF', self._on_overlay_toggle)
        self.btn_overlay_page = self._make_button(
            tr_x + btn_step, row3_y, btn_w, tr_h,
            'Cycle Floor Plans: -/-', self._on_overlay_page_cycle)
        self.btn_overlay_align = self._make_button(
            tr_x + 2 * btn_step, row3_y, btn_w, tr_h,
            'Plan Alignment Edit Mode: OFF', self._on_overlay_align_toggle)
        self.btn_overlay_dpi = self._make_button(
            tr_x + 3 * btn_step, row3_y, btn_w, tr_h,
            f'PDF Res: {self._overlay_raster_dpi} DPI', self._on_overlay_dpi_click)

        # 5th Column on Row 3: Reset Level Alignment
        self.btn_overlay_reset = self._make_button(
            tr_x + 4 * btn_step, row3_y, btn_w, tr_h,
            'Reset Level Alignment', self._on_overlay_reset_click,
            color='#FFEBEE', hovercolor='#FFCDD2') # Subtle red tint for "reset" action
        self.btn_overlay_reset.ax.set_visible(False) # Only show when Align mode is ON

        # 6th Column on Row 3: IESVE AOI level cycle — only shown when iesve_room_data provided
        self.btn_aoi_level = self._make_button(
            tr_x + 5 * btn_step, row3_y, btn_w, tr_h,
            'AOI Level: -/-', self._on_aoi_level_cycle)
        self.btn_aoi_level.ax.set_visible(self._iesve_room_data_path is not None)
        self._update_aoi_level_label()

        # Create DPI RadioButtons (initially hidden)
        # Position it vertically above the main DPI button
        radio_h = 0.20  # increased from 0.12 for more space
        radio_y = row3_y - radio_h - gap
        self.ax_dpi_radio = self._axes(tr_x + 3 * btn_step, radio_y, btn_w, radio_h)
        self.ax_dpi_radio.set_facecolor('#FFFFFF') # White background for better contrast
        self.ax_dpi_radio.set_zorder(1000) # Ensure it's on top
        self.ax_dpi_radio.set_navigate(False)
        # Draw a subtle border
        for spine in self.ax_dpi_radio.spines.values():
            spine.set_visible(True)
            spine.set_color('#B0B0B0')
            spine.set_linewidth(1.0)

        labels = [f"{d} DPI" for d in self._DPI_PRESETS]
        active_idx = self._DPI_PRESETS.index(self._overlay_raster_dpi) if self._overlay_raster_dpi in self._DPI_PRESETS else 2
        self._dpi_radio = RadioButtons(self.ax_dpi_radio, labels, active=active_idx,
                                       activecolor='#4CAF50', radio_props={'s': 100}) # s=100 for very large buttons
        # Style the labels
        for txt in self._dpi_radio.labels:
            txt.set_fontsize(9.5) # increased from 8.5
            txt.set_color('#111111')

        self._dpi_radio.on_clicked(self._on_dpi_radio_select)
        self._toggle_dpi_dropdown(False)

        # DF% legend — right of buttons, constrained within toolbar
        legend_x = tr_x + n_cols * btn_step
        legend_w = min(0.98 - legend_x, 0.18)
        legend_h = 0.99 - row1_y
        self.ax_legend = self._axes(legend_x, row1_y, legend_w, legend_h)
        self.ax_legend.axis('off')
        self.ax_legend.set_visible(False)

        # Progress bar (hidden until export)
        prog_y = row3_y + tr_h + gap * 0.5
        self.ax_progress = self._axes(tr_x, prog_y, 0.96, 0.010)
        self.ax_progress.set_xlim(0, 1)
        self.ax_progress.set_ylim(0, 1)
        self.ax_progress.axis('off')
        self.ax_progress.set_visible(False)
        self._progress_bar_patch = None
        self._progress_text = None

    def _setup_colour_key_legend(self):
        """Create the colour-key legend (far right, vertical stack)."""
        ax_legend = self._axes(0.88, 0.02, 0.11, 0.14)
        ax_legend.axis('off')
        ax_legend.set_facecolor('#F0F0EC')
        ax_legend.text(0.05, 0.95, "LEGEND", fontsize=7, fontweight='bold', color='#404040',
                       va='top', transform=ax_legend.transAxes)
        items = [
            ('red',     0.6, 'Room boundary'),
            ('yellow',  0.5, 'Selected'),
            ('cyan',    0.5, 'Being edited'),
            ('magenta', 0.5, 'Hover (edit mode)'),
        ]
        n = len(items)
        for i, (color, alpha, label) in enumerate(items):
            y0 = 0.78 - i * (0.75 / max(n - 1, 1))
            rect = FancyBboxPatch((0.05, y0 - 0.04), 0.08, 0.08,
                                  boxstyle='round,pad=0.01',
                                  facecolor=color, edgecolor=color, alpha=alpha,
                                  transform=ax_legend.transAxes, clip_on=False)
            ax_legend.add_patch(rect)
            ax_legend.text(0.18, y0, label, fontsize=6, color='#404040',
                           va='center', transform=ax_legend.transAxes)

    # === HDR NAVIGATION ========================================================

    def _on_prev_hdr_click(self, event):
        """Navigate to the previous HDR file (lower index)."""
        if not self.hdr_files:
            self._update_status("No HDR files loaded", 'red')
            return
        if self.current_hdr_idx > 0:
            self._jump_to_hdr(self.current_hdr_idx - 1)
        else:
            self._update_status("Already at first HDR file", 'orange')

    def _on_next_hdr_click(self, event):
        """Navigate to the next HDR file (higher index)."""
        if not self.hdr_files:
            self._update_status("No HDR files loaded", 'red')
            return
        if self.current_hdr_idx < len(self.hdr_files) - 1:
            self._jump_to_hdr(self.current_hdr_idx + 1)
        else:
            self._update_status("Already at last HDR file", 'orange')

    def _jump_to_hdr(self, hdr_idx: int):
        """Switch to the given HDR file index."""
        if self.divider_mode:
            self._exit_divider_mode(cancelled=True)
        if not (0 <= hdr_idx < len(self.hdr_files)):
            return
        self.current_hdr_idx = hdr_idx
        self._rebuild_image_variants()
        self._update_image_toggle_label()
        self.selected_parent = None
        self.btn_parent.label.set_text('(None)')
        self.name_label_text.set_text("Apartment Name:")
        self._update_status(f"HDR: {self.current_hdr_name}", 'green')
        # Restore per-HDR overlay page assignment
        if self._overlay_pdf_path is not None and self._overlay_pdf_info is not None:
            hdr = self.current_hdr_name
            tf = self._overlay_transforms.get(hdr, {})
            page_idx = tf.get('page_idx', self._overlay_page_idx)
            if page_idx != self._overlay_page_idx:
                self._overlay_page_idx = page_idx
                self._rasterize_overlay_page()
            self._update_overlay_page_label()
        self._update_hdr_list()
        self._update_room_list()
        self._render_section(reset_view=True, force_full=True)
        self._create_polygon_selector()
        self._start_hdr_prefetch()

    def _update_hdr_list(self):
        """Refresh the HDR file list in the side panel."""
        self.ax_hdr_list.clear()
        self.ax_hdr_list.axis('off')

        if not self.hdr_files:
            self.ax_hdr_list.text(0, 0.95, "(no HDR files)", fontsize=8, style='italic', color='gray')
        else:
            max_display = 8
            total = len(self.hdr_files)
            for display_idx in range(min(max_display, total)):
                i    = total - 1 - display_idx  # descending order (latest first)
                name = self.hdr_files[i]['name']
                y_pos = 0.95 - (display_idx * 0.12)
                is_current = (self.current_hdr_idx == i)
                room_count = sum(1 for r in self.rooms if r.get('hdr_file') == name)
                indicator  = "*" if is_current else "o"
                text       = f"{indicator} {name}"
                if room_count > 0:
                    text += f" ({room_count})"
                self.ax_hdr_list.text(
                    0, y_pos, text, fontsize=7,
                    fontweight='bold' if is_current else 'normal',
                    color='green' if is_current else 'darkgray',
                )
            if total > max_display:
                self.ax_hdr_list.text(0, 0.02, f"... and {total - max_display} more",
                                      fontsize=7, style='italic')

        self.fig.canvas.draw_idle()

    # === IMAGE TOGGLE ==========================================================

    def _on_image_toggle_click(self, event):
        """Cycle through image variants (HDR then associated TIFFs)."""
        if not self.image_variants:
            return
        self.current_variant_idx = (self.current_variant_idx + 1) % len(self.image_variants)
        self._update_image_toggle_label()
        self._render_section(force_full=True)

    def _update_image_toggle_label(self):
        """Update the toggle button label to show the active variant."""
        if not self.image_variants:
            self.btn_image_toggle.label.set_text('Toggle Image Layers: (none)')
            return
        idx  = self.current_variant_idx % len(self.image_variants)
        path = self.image_variants[idx]
        # Show a short label: HDR stem or TIFF suffix portion
        if path.suffix.lower() == '.hdr':
            label = 'Toggle Image Layers: HDR (Press T)'
        else:
            # e.g. model_plan_ffl_25300_df_false → show "df_false"
            hdr_stem = self.hdr_files[self.current_hdr_idx]['name']
            suffix   = path.stem[len(hdr_stem):]  # e.g. "_df_false"
            suffix   = suffix.lstrip('_')
            label    = f'Toggle Image Layers: {suffix} (Press T)'
        self.btn_image_toggle.label.set_text(label)
        self.fig.canvas.draw_idle()

    # === ROOM DATA MANAGEMENT ==================================================

    def _get_apartments_for_hdr(self, hdr_name: str) -> List[str]:
        """Return names of apartment-level rooms (no parent) for the given HDR file."""
        return [
            room['name'] for room in self.rooms
            if room.get('parent') is None and room.get('hdr_file') == hdr_name
        ]

    def _get_children(self, parent_name: str) -> List[dict]:
        """Return all rooms that have the given parent."""
        return [r for r in self.rooms if r.get('parent') == parent_name]

    def _get_parent_room(self, parent_name: str) -> Optional[dict]:
        """Return the room dict for the given parent name, or None."""
        for r in self.rooms:
            if r.get('name') == parent_name and r.get('parent') is None:
                return r
        return None

    # --- Name helpers ---

    def _make_unique_name(self, base_name: str, exclude_idx: Optional[int] = None) -> str:
        """Ensure room name is unique by appending numeric suffix if needed."""
        existing = {r['name'] for i, r in enumerate(self.rooms) if i != exclude_idx}
        if base_name not in existing:
            return base_name

        match = re.match(r'^(.*?)(\d+)$', base_name)
        root  = match.group(1) if match else base_name
        counter = 1
        while f"{root}{counter}" in existing:
            counter += 1
        return f"{root}{counter}"

    def _enforce_unique_names(self) -> int:
        """Ensure all room names are unique; returns count of rooms renamed."""
        seen       = set()
        renamed    = 0
        for room in self.rooms:
            name = room['name']
            if name not in seen:
                seen.add(name)
                continue
            match = re.match(r'^(.*?)(\d+)$', name)
            root  = match.group(1) if match else name
            counter = 1
            while f"{root}{counter}" in seen:
                counter += 1
            new_name = f"{root}{counter}"
            print(f"Renamed duplicate '{name}' -> '{new_name}'")
            room['name'] = new_name
            seen.add(new_name)
            renamed += 1
        return renamed

    def _check_boundary_containment(self, child_verts, parent_verts) -> bool:
        """Return True if all child vertices lie inside the parent polygon."""
        if not parent_verts or len(parent_verts) < 3:
            return True
        if not child_verts or len(child_verts) < 3:
            return True
        arr = np.array(parent_verts)
        if not np.allclose(arr[0], arr[-1]):
            arr = np.vstack([arr, arr[0]])
        path = MplPath(arr)
        return all(path.contains_point(v) for v in child_verts)

    def _update_parent_options(self):
        """Update the list of available parent apartments for the current HDR file."""
        self.parent_options = self._get_apartments_for_hdr(self.current_hdr_name)

    def _on_parent_cycle(self, event):
        """Cycle through parent apartment options."""
        self._update_parent_options()
        if not self.parent_options:
            self.selected_parent = None
            self.btn_parent.label.set_text('(None - New Apartment)')
            self.name_label_text.set_text("Apartment Name (Space ID):")
        else:
            options = [None] + self.parent_options
            try:
                current_idx = options.index(self.selected_parent)
                next_idx    = (current_idx + 1) % len(options)
            except ValueError:
                next_idx = 0
            self.selected_parent = options[next_idx]
            if self.selected_parent is None:
                self.btn_parent.label.set_text('(None - New Apartment)')
                self.name_label_text.set_text("Apartment Name (Space ID):")
            else:
                self.btn_parent.label.set_text(self.selected_parent)
                self.name_label_text.set_text("Room Name:")
        self._update_name_preview()
        self.fig.canvas.draw_idle()

    def _on_name_changed(self, _text):
        """Update name preview when name textbox changes."""
        self._update_name_preview()

    def _update_name_preview(self):
        """Update the name preview text, hiding status while preview is shown."""
        name = self.name_textbox.text.strip().upper()
        if not name:
            self.name_preview_text.set_text("")
            self.status_text.set_visible(True)
        elif self.selected_parent:
            self.name_preview_text.set_text(f"Will save as: {self.selected_parent}_{name}")
            self.status_text.set_visible(False)
        else:
            self.name_preview_text.set_text(f"Will save as: {name}")
            self.status_text.set_visible(False)
        self.fig.canvas.draw_idle()

    # === RENDERING =============================================================

    def _render_section(self, reset_view: bool = False, force_full: bool = False):
        """Render the current image and room overlays."""
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()

        need_full = force_full or reset_view
        if need_full:
            self._do_full_render(xlim, ylim, reset_view)
        else:
            self._update_room_visuals()
            self.fig.canvas.draw_idle()

    def _do_full_render(self, xlim, ylim, reset_view: bool):
        """Complete redraw of image and room polygons."""
        # Disconnect old selector BEFORE ax.clear() destroys its artists,
        # so it can't respond to events in a broken state.
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.disconnect_events()
            self.selector = None
        self.ax.clear()
        self._room_patch_cache.clear()
        self._room_label_cache.clear()
        self._df_text_cache.clear()
        self._edit_vertex_scatter = None
        self._image_handle = None
        self._df_cursor_bg = None  # invalidate blit background after ax.clear()

        # Load and display current image
        path = self.current_variant_path
        if path is not None:
            img = self._load_image(path)
        else:
            img = None

        if img is not None:
            H, W = img.shape[:2]
            self._image_width  = W
            self._image_height = H
            self._reference_view_w = float(W)
            self._image_handle = self.ax.imshow(
                img, origin='upper',
                extent=[0, W, H, 0],
                aspect='equal', zorder=0,
            )
            # Render PDF overlay if visible and available
            self._overlay_handle = None
            if self._overlay_visible and self._overlay_rgba is not None:
                self._render_overlay(W, H)
            self.ax.set_xlim(0, W)
            self.ax.set_ylim(H, 0)
        else:
            # Blank canvas
            self.ax.set_xlim(0, 1000)
            self.ax.set_ylim(1000, 0)

        # Rebuild snap vertex + edge pools from current-HDR room vertices
        all_verts, edge_starts, edge_ends = [], [], []
        for room in self.rooms:
            if self._is_room_on_current_hdr(room):
                verts = room['vertices']
                all_verts.extend(verts)
                n = len(verts)
                for j in range(n):
                    edge_starts.append(verts[j])
                    edge_ends.append(verts[(j + 1) % n])
        self.current_vertices = np.array(all_verts) if all_verts else np.array([])
        self._snap_edge_starts = np.array(edge_starts, dtype=float) if edge_starts else None
        self._snap_edge_ends = np.array(edge_ends, dtype=float) if edge_ends else None

        # Load DF image for current HDR (cached)
        self._load_current_df_image()
        self._compute_all_room_df_results()

        # Draw room polygons
        self._draw_all_room_polygons()

        # Pre-build vectorized hover arrays for fast hit-testing
        self._rebuild_hover_arrays()

        self.ax.set_aspect('equal', adjustable='box')
        self.ax.axis('off')

        # Restore or reset zoom
        if reset_view or not hasattr(self, 'original_xlim') or xlim == (0.0, 1.0):
            if img is not None:
                self.ax.set_xlim(0, self._image_width)
                self.ax.set_ylim(self._image_height, 0)
            else:
                self.ax.autoscale()
        elif img is not None and xlim != (0.0, 1.0):
            self.ax.set_xlim(xlim)
            self.ax.set_ylim(ylim)

        # Reapply zoom-aware font sizes now that limits are restored
        self._apply_zoom_fontsizes()

        self._render_df_legend()

        # Title
        hdr_name = self.current_hdr_name or "(no HDR)"
        variant  = self.current_variant_path
        variant_label = variant.stem if variant else ""
        # Title removed to maximise image area

        # Draw any stamped DF readings for the current HDR
        self._draw_df_stamps()

        # Always recreate the polygon selector after ax.clear() destroyed the old one
        if not self.edit_mode:
            self._create_polygon_selector()

        # Recreate DF cursor readout (destroyed by ax.clear())
        self._df_cursor_text = self.ax.text(
            0, 0, '', transform=self.ax.transData,
            fontsize=8, color='white', va='bottom', ha='left',
            bbox=dict(boxstyle='round,pad=0.2', facecolor='black', alpha=0.55, edgecolor='none'),
            zorder=300, visible=False, animated=True)
        self._df_cursor_bg = None  # will be captured after first draw

        self.fig.canvas.draw_idle()

    def _render_df_legend(self):
        """Show the DF% legend strip in the bottom toolbar."""
        if self.ax_legend is None:
            return
        self.ax_legend.clear()
        legend_path = self._get_legend_for_variant(self.current_variant_path)
        if legend_path is None:
            self.ax_legend.set_visible(False)
            return
        legend_img = self._load_image(legend_path)
        if legend_img is None:
            self.ax_legend.set_visible(False)
            return
        legend_img = np.rot90(legend_img, k=1)
        lH, lW = legend_img.shape[:2]
        self.ax_legend.imshow(legend_img, origin='upper', extent=[0, lW, lH, 0], aspect='auto')
        self.ax_legend.set_xlim(0, lW)
        self.ax_legend.set_ylim(lH, 0)
        self.ax_legend.axis('off')
        self.ax_legend.set_visible(True)

    def _load_current_df_image(self):
        """Load the DF% image for the current HDR file (cached).

        For IESVE .pic files ``pvalue -o`` is used to undo any EXPOSURE
        header adjustment (e.g. from ``pfilt``) so that the raw irradiance
        values are recovered before converting to DF%.
        """
        if self._hdr2wpd is None or not self.hdr_files:
            self._df_image = None
            return
        hdr_path = self.hdr_files[self.current_hdr_idx]['hdr_path']
        key = str(hdr_path)
        if key in self._df_image_cache:
            self._df_image = self._df_image_cache[key]
            return
        if self._iesve_room_data_path is not None:
            self._df_image = self._load_iesve_df_image(hdr_path)
        else:
            self._df_image = Hdr2Wpd.load_df_image(hdr_path)
        if self._df_image is not None:
            self._df_image_cache[key] = self._df_image

    @staticmethod
    def _load_iesve_df_image(hdr_path: Path) -> Optional[np.ndarray]:
        """Load a DF% image from an IESVE .pic file, undoing any EXPOSURE adjustment.

        IESVE .pic files are post-processed with ``pfilt`` which embeds an
        EXPOSURE header that scales pixel values.  Using ``pvalue -o``
        reverses this so the original irradiance values are recovered.
        """
        try:
            width, height = utils.get_hdr_resolution(hdr_path)
            result = subprocess.run(
                ['pvalue', '-h', '-H', '-b', '-o', '-df', str(hdr_path)],
                capture_output=True, check=True,
            )
            raw_image = np.frombuffer(result.stdout, dtype=np.float32).reshape((height, width))
            df_image = raw_image * 1.79  # W/m² → DF%
            return df_image
        except Exception as exc:
            print(f"Warning: could not load IESVE DF image from {hdr_path}: {exc}")
            return None

    @staticmethod
    def _vertices_hash(vertices: list) -> str:
        """Return a compact hash string for a list of vertex coordinates."""
        return str([(round(v[0], 2), round(v[1], 2)) for v in vertices])

    def _effective_room_type(self, room: dict) -> Optional[str]:
        """Return the effective room type for a room (sub-rooms inherit parent's type)."""
        rtype = room.get('room_type')
        if rtype:
            return rtype
        parent_name = room.get('parent')
        if parent_name:
            for r in self.rooms:
                if r.get('name') == parent_name:
                    return r.get('room_type')
        return None

    def _threshold_for_type(self, room_type: Optional[str]) -> Optional[float]:
        """Return the DF threshold for a given room type, or None if untagged."""
        return self.DF_THRESHOLDS.get(room_type)

    def _compute_all_room_df_results(self):
        """Compute DF threshold results for every room on the current HDR floor.

        Each room type has a single fixed threshold (BED=0.5%, LIVING=1.0%).
        Only rooms with a type (or inherited type) get results computed.
        """
        self._room_df_results.clear()
        if self._df_image is None or self._hdr2wpd is None:
            return
        any_new = False
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room['vertices']
            if len(verts) < 3:
                continue
            eff_type = self._effective_room_type(room)
            threshold = self._threshold_for_type(eff_type)
            if threshold is None:
                continue
            had_cache = bool(room.get('df_cache'))
            lines = self._compute_room_df(room, verts, (threshold,))
            self._room_df_results[i] = lines
            if not had_cache and room.get('df_cache'):
                any_new = True
        # Persist newly computed DF caches to the session JSON
        if any_new:
            self._save_session()

    def _compute_room_df(self, room: dict, verts: list, thresholds_key: tuple) -> list:
        """Compute DF results for a single room, using its cached df_cache if valid.

        For parent rooms with children, sub-room areas are subtracted so the
        result represents only the remaining (non-sub-room) area.
        """
        # Build hash including child vertices for parent rooms
        children = self._get_children(room.get('name', ''))
        child_verts_list = [c['vertices'] for c in children if len(c.get('vertices', [])) >= 3]

        verts_hash = self._vertices_hash(verts)
        if child_verts_list:
            child_hashes = tuple(self._vertices_hash(cv) for cv in child_verts_list)
            verts_hash = hash((verts_hash, child_hashes))

        _DF_DISPLAY_FMT = 2  # bump to invalidate cached display_lines
        cache = room.get('df_cache')
        if (cache
                and cache.get('vertices_hash') == verts_hash
                and tuple(cache.get('thresholds', [])) == thresholds_key):
            if cache.get('display_fmt') == _DF_DISPLAY_FMT:
                return cache['display_lines']
            # Geometry unchanged — regenerate display lines from cached raw_result
            if cache.get('raw_result'):
                lines = self._format_df_lines(cache['raw_result'])
                cache['display_lines'] = lines
                cache['display_fmt'] = _DF_DISPLAY_FMT
                return lines

        # Cache miss — recompute
        if child_verts_list:
            result = self._hdr2wpd.compute_df_for_polygon_excluding(
                self._df_image, verts, child_verts_list, thresholds_key)
        else:
            result = self._hdr2wpd.compute_df_for_polygon(
                self._df_image, verts, thresholds_key)
        lines = self._format_df_lines(result)
        room['df_cache'] = {
            'vertices_hash':  verts_hash,
            'thresholds':     list(thresholds_key),
            'display_fmt':    _DF_DISPLAY_FMT,
            'display_lines':  lines,
            'raw_result':     result,
        }
        return lines

    @staticmethod
    def _format_df_lines(result: dict) -> list:
        """Format DF results as display lines: area (%) then threshold."""
        total_area = result['total_area_m2']
        lines = []
        for tr in result['thresholds']:
            pct = (tr['area_m2'] / total_area * 100) if total_area > 0 else 0.0
            lines.append(f"{tr['area_m2']:.2f} m\u00b2 ({pct:.0f}%)")
            lines.append(f"@ {tr['threshold']:g}% DF")
        return lines

    def _invalidate_room_df_cache(self, room_idx: int):
        """Clear the DF cache for a specific room, forcing recomputation.

        Also invalidates the parent room's cache if the edited room is a
        sub-room, since LIVING-type parents depend on child geometry.
        """
        if 0 <= room_idx < len(self.rooms):
            room = self.rooms[room_idx]
            room.pop('df_cache', None)
            # Invalidate parent if this is a sub-room
            parent_name = room.get('parent')
            if parent_name:
                for i, r in enumerate(self.rooms):
                    if r.get('name') == parent_name:
                        r.pop('df_cache', None)
                        break

    def _draw_all_room_polygons(self):
        """Draw all room polygons for the current view."""
        for i, room in enumerate(self.rooms):
            if self._is_room_on_current_hdr(room):
                self._draw_room_polygon(room, i, is_current_floor=True)

    def _draw_room_polygon(self, room: dict, idx: int, is_current_floor: bool):
        """Draw a single room polygon with its label."""
        verts = room['vertices']
        if len(verts) < 3:
            return

        is_selected = (idx == self.selected_room_idx or idx in self.multi_selected_room_idxs)
        is_hover    = (idx == self.hover_room_idx)
        is_editing  = (idx == self.edit_room_idx and self.edit_mode)
        is_subroom  = room.get('parent') is not None
        is_div      = '_DIV' in room.get('name', '')

        if is_editing:
            edge_color, lw, linestyle, alpha = 'cyan',    3, '-',  1.0
        elif is_selected:
            edge_color, lw, linestyle, alpha = 'yellow',  4, '-',  1.0
        elif is_hover and self.edit_mode:
            edge_color, lw, linestyle, alpha = 'magenta', 2, '-',  1.0
        elif is_div and is_current_floor:
            edge_color, lw, linestyle, alpha = 'red',     self._zoom_linewidth(), '--', 0.6
        elif is_current_floor:
            edge_color, lw, linestyle, alpha = 'red',     self._zoom_linewidth(), '-',  1.0
        else:
            edge_color, lw, linestyle, alpha = 'gray',    self._zoom_linewidth(base=1.0), '-', 1.0

        poly = Polygon(verts, closed=True,
                       edgecolor=edge_color, facecolor='none', alpha=alpha, linewidth=lw,
                       linestyle=linestyle, clip_on=True)
        poly._patch_type = 'current' if is_current_floor else 'other'
        self.ax.add_patch(poly)
        self._room_patch_cache[idx] = poly

        # Edit mode: draw vertex handles
        if self.edit_mode and is_current_floor:
            verts_array = np.array(verts)
            xs, ys, colors, sizes = [], [], [], []
            for v_idx, (vx, vy) in enumerate(verts_array):
                is_hovered  = (idx == self.hover_room_idx  and v_idx == self.hover_vertex_idx)
                is_dragging = (idx == self.edit_room_idx   and v_idx == self.edit_vertex_idx)
                if is_dragging:
                    c, s = 'yellow', 9
                elif is_hovered:
                    c, s = 'red', 8
                else:
                    c, s = 'cyan', 5
                xs.append(vx); ys.append(vy); colors.append(c); sizes.append(s ** 2)
            if xs:
                self.ax.scatter(xs, ys, c=colors, s=sizes,
                                edgecolors='black', linewidths=1.0, zorder=100, picker=5)

            # Highlight hovered edge
            if idx == self.hover_edge_room_idx and self.hover_edge_idx is not None:
                n = len(verts)
                j = self.hover_edge_idx
                ex = [verts[j][0], verts[(j + 1) % n][0]]
                ey = [verts[j][1], verts[(j + 1) % n][1]]
                self.ax.plot(ex, ey, '-', color='lime', linewidth=3, zorder=99, alpha=0.8)
                if self.hover_edge_point is not None:
                    self.ax.plot([self.hover_edge_point[0]], [self.hover_edge_point[1]], 'o',
                                 markersize=8, color='lime', markeredgecolor='darkgreen',
                                 markeredgewidth=1.5, zorder=101, alpha=0.9)

        # Label and DF results — suppressed for CIRC rooms (which includes all
        # DIV sub-rooms, auto-typed as CIRC on creation) to reduce visual clutter.
        is_circ = room.get('room_type', '') == 'CIRC'
        if not is_circ:
            centroid = self._polygon_label_point(verts)
            label    = room.get('name', '')
            if not is_current_floor:
                hf = room.get('hdr_file', '')
                label += f"\n({hf})"

            fs_name = self._zoom_fontsize()
            label_text = self.ax.text(
                centroid[0], centroid[1], label,
                color='red', fontsize=fs_name,
                ha='center', va='center', clip_on=True,
                path_effects=[patheffects.withStroke(linewidth=fs_name * 0.06, foreground='black')],
            )
            label_text.set_clip_path(self.ax.patch)
            self._room_label_cache[idx] = label_text

        # DF results as smaller subtext below the room name (only if room type is set,
        # and not suppressed for CIRC rooms)
        df_lines = self._room_df_results.get(idx, [])
        if df_lines and is_current_floor and not is_circ:
            line_step = self._df_line_step()
            for line_i, line in enumerate(df_lines):
                dy = 1.2 * line_step + line_i * line_step * 0.7
                fs_df = self._zoom_fontsize(base=6.5)
                df_text = self.ax.text(
                    centroid[0], centroid[1] + dy, line,
                    color='red', fontsize=fs_df,
                    ha='center', va='center', clip_on=True, alpha=0.85,
                    path_effects=[patheffects.withStroke(linewidth=fs_df * 0.06, foreground='black')],
                )
                df_text.set_clip_path(self.ax.patch)
                # Store centroid + line index for zoom repositioning
                df_text._df_centroid = centroid
                df_text._df_line_i  = line_i
                self._df_text_cache.append(df_text)

    def _update_room_visuals(self):
        """Update room colours without a full redraw (for hover/selection changes)."""
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            patch = self._room_patch_cache.get(i)
            if patch is None:
                continue
            is_selected = (i == self.selected_room_idx or i in self.multi_selected_room_idxs)
            is_hover    = (i == self.hover_room_idx)
            is_editing  = (i == self.edit_room_idx and self.edit_mode)
            is_subroom  = room.get('parent') is not None

            if is_editing:
                patch.set_edgecolor('cyan');    patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(3)
                patch._patch_type = 'special'
            elif is_selected:
                patch.set_edgecolor('yellow');  patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(4)
                patch._patch_type = 'special'
            elif is_hover and self.edit_mode:
                patch.set_edgecolor('magenta'); patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(2)
                patch._patch_type = 'special'
            else:
                is_div = '_DIV' in room.get('name', '')
                patch.set_edgecolor('red');     patch.set_facecolor('none'); patch.set_alpha(1.0 if not is_div else 0.6); patch.set_linewidth(self._zoom_linewidth())
                patch.set_linestyle('--' if is_div else '-')
                patch._patch_type = 'current'

        self.fig.canvas.draw_idle()

    # === UNDO & SNAPPING =======================================================

    def _push_undo(self, room_idx: int):
        """Push a snapshot of a room's vertices onto the undo stack."""
        verts_copy = [list(v) for v in self.rooms[room_idx]['vertices']]
        self._edit_undo_stack.append((room_idx, verts_copy))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)

    def _undo_edit(self):
        """Restore the last undone vertex state (Ctrl+Z in edit mode)."""
        if not self.edit_mode:
            return
        if not self._edit_undo_stack:
            self._update_status("Nothing to undo", 'orange')
            return
        entry = self._edit_undo_stack.pop()

        # Divider undo: full rooms-list snapshot
        if isinstance(entry, tuple) and len(entry) == 2 and entry[0] == 'divider':
            self.rooms = entry[1]
            self._reset_hover_state()
            self._update_status("Undid room division", 'blue')
            self._save_session()
            self._update_room_list()
            self._update_hdr_list()
            self._render_section(force_full=True)
            return

        room_idx, verts_snapshot = entry
        self.rooms[room_idx]['vertices'] = verts_snapshot
        self._invalidate_room_df_cache(room_idx)
        self._reset_hover_state()
        room_name = self.rooms[room_idx].get('name', 'unnamed')
        remaining = len(self._edit_undo_stack)
        self._update_status(
            f"Undid change to '{room_name}' ({remaining} undo{'s' if remaining != 1 else ''} left)", 'blue')
        self._save_session()
        self._render_section(force_full=True)

    def _push_draw_undo(self, entry: tuple):
        """Push an entry onto the draw-mode undo stack."""
        self._draw_undo_stack.append(entry)
        if len(self._draw_undo_stack) > self._draw_undo_max:
            self._draw_undo_stack.pop(0)

    def _undo_draw(self):
        """Undo the last draw-mode operation (Ctrl+Z outside edit mode)."""
        if not self._draw_undo_stack:
            self._update_status("Nothing to undo", 'orange')
            return
        entry = self._draw_undo_stack.pop()
        tag   = entry[0]

        if tag == 'delete':
            _, idx, room = entry
            idx = min(idx, len(self.rooms))
            self.rooms.insert(idx, room)
            name = room.get('name', 'unnamed')
            self._update_status(f"Restored '{name}'", 'blue')
            print(f"Restored deleted room '{name}'")

        elif tag == 'type':
            _, old_types = entry
            for idx, old_type in old_types:
                if idx < len(self.rooms):
                    self.rooms[idx]['room_type'] = old_type
                    self._invalidate_room_df_cache(idx)
            count = len(old_types)
            self._update_status(f"Undid type change ({count} room{'s' if count != 1 else ''})", 'blue')

        elif tag == 'rename':
            _, idx, old_name = entry
            if idx < len(self.rooms):
                self.rooms[idx]['name'] = old_name
                self._update_status(f"Undid rename → '{old_name}'", 'blue')

        elif tag == 'create':
            _, idx, parent_type_changed = entry
            if idx < len(self.rooms):
                name = self.rooms[idx].get('name', 'unnamed')
                self.rooms.pop(idx)
                self.selected_room_idx = None
                # Revert auto-assigned parent type if applicable
                if parent_type_changed is not None:
                    p_idx, p_old_type = parent_type_changed
                    if p_idx < len(self.rooms):
                        self.rooms[p_idx]['room_type'] = p_old_type
                        self._invalidate_room_df_cache(p_idx)
                self._update_status(f"Undid creation of '{name}'", 'blue')

        else:
            self._update_status("Unknown undo entry", 'red')
            return

        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)

    def _snap_to_vertex(self, x: float, y: float) -> tuple:
        """Snap to nearest existing polygon vertex within snap_distance_px pixels."""
        if len(self.current_vertices) == 0:
            return x, y
        dists = np.hypot(self.current_vertices[:, 0] - x, self.current_vertices[:, 1] - y)
        min_idx  = int(np.argmin(dists))
        min_dist = dists[min_idx]
        if min_dist <= self._snap_distance_px:
            return float(self.current_vertices[min_idx, 0]), float(self.current_vertices[min_idx, 1])
        return x, y

    def _snap_to_edge(self, x: float, y: float) -> tuple:
        """Snap to nearest existing polygon edge within snap_distance_px pixels."""
        if self._snap_edge_starts is None or self._snap_edge_ends is None:
            return x, y
        A = self._snap_edge_starts
        B = self._snap_edge_ends
        AB = B - A
        AP = np.array([[x, y]]) - A
        seg_len_sq = np.sum(AB * AB, axis=1)
        safe_len = np.where(seg_len_sq == 0, 1.0, seg_len_sq)
        t = np.clip(np.sum(AP * AB, axis=1) / safe_len, 0.0, 1.0)
        t = np.where(seg_len_sq == 0, 0.0, t)
        proj = A + t[:, None] * AB
        dists = np.hypot(x - proj[:, 0], y - proj[:, 1])
        min_idx = int(np.argmin(dists))
        if dists[min_idx] <= self._snap_distance_px:
            return float(proj[min_idx, 0]), float(proj[min_idx, 1])
        return x, y

    def _point_to_segment_dist(self, px, py, ax, ay, bx, by):
        """Return (distance, proj_x, proj_y) from point P to segment A→B."""
        dx, dy    = bx - ax, by - ay
        seg_len_sq = dx*dx + dy*dy
        if seg_len_sq == 0:
            return np.hypot(px - ax, py - ay), ax, ay
        t = max(0.0, min(1.0, ((px - ax)*dx + (py - ay)*dy) / seg_len_sq))
        proj_x, proj_y = ax + t*dx, ay + t*dy
        return np.hypot(px - proj_x, py - proj_y), proj_x, proj_y

    @staticmethod
    def _polygon_label_point(verts) -> np.ndarray:
        """Return a point guaranteed to be inside the polygon for label placement.

        Uses the true geometric centroid (centre of mass via the shoelace
        formula) rather than the simple average of vertices.  If the centroid
        falls outside a concave polygon, falls back to scanning interior
        candidate points to find one that is inside and as far from the edges
        as possible.
        """
        pts = np.array(verts, dtype=float)
        n = len(pts)
        if n < 3:
            return pts.mean(axis=0)

        # --- Signed area via shoelace formula ---
        x, y = pts[:, 0], pts[:, 1]
        x1, y1 = np.roll(x, -1), np.roll(y, -1)
        cross = x * y1 - x1 * y
        signed_area = cross.sum() / 2.0

        if abs(signed_area) < 1e-10:
            # Degenerate polygon — fall back to vertex average
            return pts.mean(axis=0)

        # --- True centroid (centre of mass) ---
        cx = ((x + x1) * cross).sum() / (6.0 * signed_area)
        cy = ((y + y1) * cross).sum() / (6.0 * signed_area)
        centroid = np.array([cx, cy])

        # Check if centroid is inside the polygon
        path = MplPath(pts)
        if path.contains_point(centroid):
            return centroid

        # --- Fallback: find best interior point for concave polygons ---
        # Sample candidate points on a grid inside the bounding box and pick
        # the one with the greatest distance to the nearest edge (visual
        # "pole of inaccessibility" approximation).
        xmin, ymin = pts.min(axis=0)
        xmax, ymax = pts.max(axis=0)
        w, h = xmax - xmin, ymax - ymin
        # Use ~20 steps along the longer dimension
        steps = 20
        xs = np.linspace(xmin, xmax, max(steps, 4))
        ys = np.linspace(ymin, ymax, max(steps, 4))
        grid_x, grid_y = np.meshgrid(xs, ys)
        candidates = np.column_stack([grid_x.ravel(), grid_y.ravel()])

        inside_mask = path.contains_points(candidates)
        inside_pts = candidates[inside_mask]

        if len(inside_pts) == 0:
            # Extremely unlikely — return vertex average as last resort
            return pts.mean(axis=0)

        # For each inside point, compute min distance to any polygon edge
        best_point = inside_pts[0]
        best_dist  = -1.0
        for cp in inside_pts:
            min_d = np.inf
            for i in range(n):
                j = (i + 1) % n
                a, b = pts[i], pts[j]
                edge = b - a
                edge_len_sq = edge @ edge
                if edge_len_sq < 1e-12:
                    d = np.linalg.norm(cp - a)
                else:
                    t = np.clip((cp - a) @ edge / edge_len_sq, 0, 1)
                    proj = a + t * edge
                    d = np.linalg.norm(cp - proj)
                if d < min_d:
                    min_d = d
            if min_d > best_dist:
                best_dist  = min_d
                best_point = cp

        return best_point

    @staticmethod
    def _snap_to_pixel(x: float, y: float) -> tuple:
        """Snap coordinates to the nearest pixel centre (int + 0.5)."""
        return int(x) + 0.5, int(y) + 0.5

    # === EVENT HANDLERS ========================================================

    def _on_click_with_snap(self, event):
        """Handle mouse clicks; snap left-clicks to existing polygon vertices."""
        if event.inaxes != self.ax:
            return

        # Middle-click: start pan
        if event.button == 2:
            self._pan_active = True
            self._pan_start = (event.x, event.y)
            self._pan_xlim = self.ax.get_xlim()
            self._pan_ylim = self.ax.get_ylim()
            return

        # Right-click in divider mode: undo last placed point
        if event.button == 3 and self.divider_mode:
            self._divider_undo_last_point()
            return

        # Right-click in edit mode: delete hovered vertex
        if event.button == 3 and self.edit_mode:
            if self.hover_vertex_idx is not None and self.hover_room_idx is not None:
                room = self.rooms[self.hover_room_idx]
                if len(room['vertices']) > 3:
                    self._push_undo(self.hover_room_idx)
                    room['vertices'].pop(self.hover_vertex_idx)
                    rname = room.get('name', 'unnamed')
                    self.hover_vertex_idx = None
                    self.hover_room_idx   = None
                    self._update_status(f"Removed vertex from '{rname}' - press 's' to save", 'green')
                    self._save_session()
                    self._render_section(force_full=True)
                else:
                    self._update_status("Cannot remove - polygon must have at least 3 vertices", 'red')
            return

        # Right-click in default mode (not edit, not divider): remove nearest DF stamp
        if event.button == 3 and not self.edit_mode and not self.divider_mode:
            if event.xdata is not None and event.ydata is not None:
                self._remove_df_stamp(event.xdata, event.ydata)
            return

        # Left-click in align mode: collect alignment points
        if event.button == 1 and self._align_mode and event.xdata is not None:
            self._on_align_click(event.xdata, event.ydata)
            return


        # Left-click in divider mode: place division line endpoints
        if event.button == 1 and self.divider_mode and event.xdata is not None and event.ydata is not None:
            self._on_divider_click(event.xdata, event.ydata)
            return

        # Left-click in default mode (not draw, not edit): stamp DF (if placement on) then select room
        if event.button == 1 and not self.draw_mode and not self.edit_mode:
            stamped = False
            if self.placement_mode and event.xdata is not None and event.ydata is not None and self._df_image is not None:
                px, py = int(event.xdata), int(event.ydata)
                h, w = self._df_image.shape[:2]
                if 0 <= py < h and 0 <= px < w:
                    df_val = self._df_image[py, px]
                    hdr = self.current_hdr_name
                    if hdr not in self._df_stamps:
                        self._df_stamps[hdr] = []
                    self._df_stamps[hdr].append((float(event.xdata), float(event.ydata), float(df_val), px, py))
                    self._df_cursor_bg = None
                    stamped = True
            self._select_room_at(event.xdata, event.ydata, ctrl=event.key == 'control')
            if stamped:
                self._save_session()
                # force full render so the new stamp is drawn (select_room_at may not have)
                self._render_section(force_full=True)
            return

        # Left-click in edit mode: drag vertex or insert vertex on edge
        if event.button == 1 and self.edit_mode and event.xdata is not None and event.ydata is not None:
            if self.hover_vertex_idx is not None and self.hover_room_idx is not None:
                # Begin vertex drag
                self._push_undo(self.hover_room_idx)
                self.edit_room_idx   = self.hover_room_idx
                self.edit_vertex_idx = self.hover_vertex_idx
                v = self.rooms[self.hover_room_idx]['vertices'][self.hover_vertex_idx]
                self._edit_drag_origin = (float(v[0]), float(v[1]))
                room_name = self.rooms[self.edit_room_idx].get('name', 'unnamed')
                self._update_status(f"Dragging vertex in '{room_name}'", 'cyan')
                self._render_section(force_full=True)
                self._start_blit_drag(self.edit_room_idx)
                return

            if self.hover_edge_room_idx is not None and self.hover_edge_idx is not None and self.hover_edge_point is not None:
                if event.key == 'shift':
                    # Shift+click: begin edge drag (move both endpoints together)
                    self._push_undo(self.hover_edge_room_idx)
                    self.edit_edge_room_idx = self.hover_edge_room_idx
                    self.edit_edge_idx      = self.hover_edge_idx
                    self.edit_edge_start    = (event.xdata, event.ydata)
                    room_name = self.rooms[self.edit_edge_room_idx].get('name', 'unnamed')
                    self._update_status(f"Dragging edge in '{room_name}'", 'cyan')
                    self._start_blit_drag(self.edit_edge_room_idx)
                    return
                # Click: insert two vertices on edge (straddling the click point)
                self._push_undo(self.hover_edge_room_idx)
                room  = self.rooms[self.hover_edge_room_idx]
                j     = self.hover_edge_idx
                verts = room['vertices']
                ax_, ay_ = verts[j]
                bx_, by_ = verts[(j + 1) % len(verts)]
                cx, cy = self.hover_edge_point

                # Place two points offset along the edge from the click point
                edge_dx, edge_dy = bx_ - ax_, by_ - ay_
                edge_len = (edge_dx ** 2 + edge_dy ** 2) ** 0.5
                offset = min(15.0, edge_len * 0.15)  # 15 px or 15% of edge
                if edge_len > 0:
                    ux, uy = edge_dx / edge_len, edge_dy / edge_len
                else:
                    ux, uy = 1.0, 0.0
                p1 = [float(cx - ux * offset), float(cy - uy * offset)]
                p2 = [float(cx + ux * offset), float(cy + uy * offset)]

                room['vertices'].insert(j + 1, p1)
                room['vertices'].insert(j + 2, p2)
                # Start dragging the second vertex (closer to b) for immediate repositioning
                self.edit_room_idx   = self.hover_edge_room_idx
                self.edit_vertex_idx = j + 2
                self._edit_drag_origin = (float(p2[0]), float(p2[1]))
                self.hover_edge_room_idx = None
                self.hover_edge_idx      = None
                self.hover_edge_point    = None
                self._update_status("Inserted 2 vertices — drag to reposition, right-click to remove", 'cyan')
                self._render_section(force_full=True)
                return
            return  # click on empty space in edit mode

        # Left-click in draw mode: store snapped position for correction on release
        if event.button == 1 and not self.edit_mode and event.xdata is not None and event.ydata is not None:
            x, y = event.xdata, event.ydata

            # Orthogonal constraint: lock to horizontal or vertical from last vertex
            if self.ortho_mode and hasattr(self, 'selector') and self.selector is not None and self.selector.verts:
                last_x, last_y = self.selector.verts[-1]
                dx, dy = abs(x - last_x), abs(y - last_y)
                if dx >= dy:
                    y = last_y   # horizontal line
                else:
                    x = last_x   # vertical line

            x, y = self._snap_to_pixel(x, y)
            snapped_x, snapped_y = self._snap_to_vertex(x, y)
            # If vertex snap didn't fire, try edge snap
            if snapped_x == x and snapped_y == y:
                snapped_x, snapped_y = self._snap_to_edge(x, y)
            self._pending_snap = (snapped_x, snapped_y)

            # Auto-detect parent room on first point of a new polygon
            if (not self.selected_parent
                    and hasattr(self, 'selector') and self.selector is not None
                    and not self.selector.verts):
                for room in self.rooms:
                    if (room.get('parent') is None
                            and self._is_room_on_current_hdr(room)
                            and len(room['vertices']) >= 3):
                        path = MplPath(np.array(room['vertices']))
                        if path.contains_point((snapped_x, snapped_y)):
                            self.selected_parent = room['name']
                            self.btn_parent.label.set_text(room['name'])
                            self._update_status(
                                f"Auto-detected parent: {room['name']}", 'blue')
                            break

    def _on_button_release(self, event):
        """Handle mouse button release (end of drag, pan, or snap correction)."""
        # End pan (check before inaxes guard — mouse may have left the axes)
        if event.button == 2 and self._pan_active:
            self._pan_active = False
            self._pan_start = None
            return


        if event.inaxes != self.ax:
            return

        # Correct the vertex the PolygonSelector just placed with our snapped position
        if (event.button == 1 and not self.edit_mode
                and hasattr(self, '_pending_snap') and self._pending_snap is not None):
            sx, sy = self._pending_snap
            self._pending_snap = None
            # _xys[-1] is the cursor tracking point; the just-added vertex is [-2]
            if self.selector is not None and len(self.selector._xys) >= 2 and not self.selector._selection_completed:
                self.selector._xys[-2] = (sx, sy)
                self.selector._draw_polygon()

        if self.edit_vertex_idx is not None and self.edit_room_idx is not None:
            self._end_blit_drag()
            room = self.rooms[self.edit_room_idx]
            current_pos = room['vertices'][self.edit_vertex_idx]
            # Snap to pixel centre, then to existing vertex if close
            px, py = self._snap_to_pixel(current_pos[0], current_pos[1])
            sx, sy = self._snap_to_vertex(px, py)
            room['vertices'][self.edit_vertex_idx] = [float(sx), float(sy)]
            self.edit_vertex_idx = None
            self._edit_drag_origin = None
            self._save_session()
            room_name = room.get('name', 'unnamed')
            self._update_status(f"Moved vertex in '{room_name}'", 'green')
            self._render_section(force_full=True)
            self._create_polygon_selector()

        if self.edit_edge_room_idx is not None and self.edit_edge_idx is not None:
            self._end_blit_drag()
            room = self.rooms[self.edit_edge_room_idx]
            j  = self.edit_edge_idx
            j2 = (j + 1) % len(room['vertices'])
            # Snap both endpoints to pixel centre
            for vi in (j, j2):
                px, py = self._snap_to_pixel(room['vertices'][vi][0], room['vertices'][vi][1])
                room['vertices'][vi] = [float(px), float(py)]
            self.edit_edge_room_idx = None
            self.edit_edge_idx      = None
            self.edit_edge_start    = None
            self._save_session()
            room_name = room.get('name', 'unnamed')
            self._update_status(f"Moved edge in '{room_name}'", 'green')
            self._render_section(force_full=True)
            self._create_polygon_selector()

    def _on_mouse_motion(self, event):
        """Handle mouse movement — dispatches to drag, pan, or hover handlers."""
        # Pan: use raw pixel coords so it works even when inaxes changes during drag
        if self._pan_active and self._pan_start is not None:
            dx_px = event.x - self._pan_start[0]
            dy_px = event.y - self._pan_start[1]
            # Convert pixel delta to data-coordinate delta
            fig_w, fig_h = self.fig.get_size_inches() * self.fig.dpi
            bbox = self.ax.get_position()
            ax_w_px = bbox.width * fig_w
            ax_h_px = bbox.height * fig_h
            xlim = self._pan_xlim
            ylim = self._pan_ylim
            dx_data = (xlim[1] - xlim[0]) * dx_px / ax_w_px
            dy_data = (ylim[1] - ylim[0]) * dy_px / ax_h_px
            self.ax.set_xlim(xlim[0] - dx_data, xlim[1] - dx_data)
            self.ax.set_ylim(ylim[0] - dy_data, ylim[1] - dy_data)
            self._df_cursor_bg = None  # invalidate stale blit background after pan
            self.fig.canvas.draw_idle()
            return


        if event.inaxes != self.ax:
            if self._df_cursor_text.get_visible():
                self._df_cursor_text.set_visible(False)
                self.fig.canvas.draw_idle()
            return
        x, y = event.xdata, event.ydata
        if x is None or y is None:
            if self._df_cursor_text.get_visible():
                self._df_cursor_text.set_visible(False)
                self.fig.canvas.draw_idle()
            return

        # DF% cursor readout — blit only the text for lag-free updates
        canvas = self.fig.canvas
        supports_blit = getattr(canvas, 'supports_blit', False)
        if self._df_image is not None:
            px, py = int(x), int(y)
            h, w = self._df_image.shape[:2]
            if 0 <= py < h and 0 <= px < w:
                df_val = self._df_image[py, px]
                self._df_cursor_text.set_text(f"DF: {df_val:.2f}%")
                # Fixed 14-pixel offset in display space → convert to data coords
                disp_x, disp_y = self.ax.transData.transform((x, y))
                data_x, data_y = self.ax.transData.inverted().transform(
                    (disp_x + 14, disp_y - 14))
                self._df_cursor_text.set_position((data_x, data_y))
                self._df_cursor_text.set_visible(True)
            else:
                self._df_cursor_text.set_visible(False)
        else:
            self._df_cursor_text.set_visible(False)

        if supports_blit:
            try:
                if self._df_cursor_bg is None:
                    self._df_cursor_text.set_visible(False)
                    canvas.draw()
                    self._df_cursor_bg = canvas.copy_from_bbox(self.ax.bbox)
                canvas.restore_region(self._df_cursor_bg)
                self.ax.draw_artist(self._df_cursor_text)
                canvas.blit(self.ax.bbox)
            except (AttributeError, RuntimeError):
                # Figure or renderer not yet ready (e.g. during ax.clear() / HDR
                # switch / window close). Invalidate the cached background so it
                # is rebuilt on the next motion event and fall back to draw_idle.
                self._df_cursor_bg = None
                canvas.draw_idle()
        else:
            canvas.draw_idle()

        # Active drag operations (vertex or edge)
        if self.edit_vertex_idx is not None and self.edit_room_idx is not None:
            # Ortho constraint: lock to H or V from drag origin
            if self.ortho_mode and self._edit_drag_origin is not None:
                ox, oy = self._edit_drag_origin
                dx, dy = abs(x - ox), abs(y - oy)
                if dx >= dy:
                    y = oy   # horizontal movement
                else:
                    x = ox   # vertical movement
            self.rooms[self.edit_room_idx]['vertices'][self.edit_vertex_idx] = [float(x), float(y)]
            self._update_dragged_patch(self.edit_room_idx)
            return
        if self.edit_edge_room_idx is not None and self.edit_edge_idx is not None:
            self._handle_edge_drag(x, y)
            return

        # Divider mode: update preview line and vertex snap highlight
        if self.divider_mode:
            self._update_divider_preview(x, y)
            return

        if not self.edit_mode:
            return

        # Throttle hover detection (~15 fps)
        now = time.monotonic()
        if now - self._last_hover_check < 0.067:
            return
        self._last_hover_check = now
        self._handle_hover_detection(x, y)

    def _handle_edge_drag(self, x: float, y: float):
        """Move both endpoints of the hovered edge perpendicular to it.

        The drag is constrained to the direction orthogonal to the edge,
        so horizontal edges move only vertically and vice-versa.
        """
        dx = x - self.edit_edge_start[0]
        dy = y - self.edit_edge_start[1]
        room = self.rooms[self.edit_edge_room_idx]
        j  = self.edit_edge_idx
        j2 = (j + 1) % len(room['vertices'])

        # Edge direction vector
        ex = room['vertices'][j2][0] - room['vertices'][j][0]
        ey = room['vertices'][j2][1] - room['vertices'][j][1]
        edge_len_sq = ex * ex + ey * ey

        if edge_len_sq > 0:
            # Normal (perpendicular) to the edge: (-ey, ex)
            nx, ny = -ey, ex
            # Project the drag delta onto the normal direction
            proj = (dx * nx + dy * ny) / edge_len_sq
            dx = proj * nx
            dy = proj * ny

        room['vertices'][j][0]  += dx;  room['vertices'][j][1]  += dy
        room['vertices'][j2][0] += dx;  room['vertices'][j2][1] += dy
        self.edit_edge_start = (self.edit_edge_start[0] + dx, self.edit_edge_start[1] + dy)
        self._update_dragged_patch(self.edit_edge_room_idx)

    def _rebuild_hover_arrays(self):
        """Pre-build combined numpy arrays for vectorized hover hit-testing."""
        all_verts, vert_room, vert_local = [], [], []
        edge_starts, edge_ends, edge_room, edge_local = [], [], [], []
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room['vertices']
            n = len(verts)
            for vi, v in enumerate(verts):
                all_verts.append(v)
                vert_room.append(i)
                vert_local.append(vi)
            for j in range(n):
                edge_starts.append(verts[j])
                edge_ends.append(verts[(j + 1) % n])
                edge_room.append(i)
                edge_local.append(j)
        self._hover_all_verts       = np.array(all_verts, dtype=float) if all_verts else None
        self._hover_vert_room_idx   = np.array(vert_room, dtype=int) if vert_room else None
        self._hover_vert_local_idx  = np.array(vert_local, dtype=int) if vert_local else None
        self._hover_edge_starts     = np.array(edge_starts, dtype=float) if edge_starts else None
        self._hover_edge_ends       = np.array(edge_ends, dtype=float) if edge_ends else None
        self._hover_edge_room_idx_arr = np.array(edge_room, dtype=int) if edge_room else None
        self._hover_edge_local_idx  = np.array(edge_local, dtype=int) if edge_local else None

    def _handle_hover_detection(self, x: float, y: float):
        """Detect vertex or edge hover using pre-built vectorized arrays."""
        hover_threshold = max(5.0, max(self._image_width, self._image_height) * 0.01)

        # --- Vertex hover (single vectorized distance computation) ---
        closest_dist, closest_room, closest_vidx = float('inf'), None, None
        if self._hover_all_verts is not None and len(self._hover_all_verts) > 0:
            dists = np.hypot(self._hover_all_verts[:, 0] - x, self._hover_all_verts[:, 1] - y)
            min_idx = int(np.argmin(dists))
            closest_dist = dists[min_idx]
            closest_room = int(self._hover_vert_room_idx[min_idx])
            closest_vidx = int(self._hover_vert_local_idx[min_idx])

        if closest_dist < hover_threshold:
            if self.hover_room_idx != closest_room or self.hover_vertex_idx != closest_vidx:
                self.hover_room_idx   = closest_room
                self.hover_vertex_idx = closest_vidx
                self.hover_edge_room_idx = self.hover_edge_idx = self.hover_edge_point = None
                name = self.rooms[closest_room].get('name', 'unnamed')
                self._update_status(f"Vertex in '{name}' - drag or right-click to remove", 'cyan')
                self._render_section()
            return

        # Clear stale vertex hover
        vertex_was_hovered = self.hover_vertex_idx is not None or self.hover_room_idx is not None
        if vertex_was_hovered:
            self.hover_vertex_idx = self.hover_room_idx = None

        # --- Edge hover (vectorized point-to-segment distance) ---
        edge_threshold = hover_threshold * 1.3
        best_dist, best_room, best_edge, best_pt = float('inf'), None, None, None
        if self._hover_edge_starts is not None and len(self._hover_edge_starts) > 0:
            A = self._hover_edge_starts
            B = self._hover_edge_ends
            AB = B - A
            P = np.array([[x, y]])
            AP = P - A
            seg_len_sq = np.sum(AB * AB, axis=1)
            # Avoid division by zero for degenerate edges
            safe_len = np.where(seg_len_sq == 0, 1.0, seg_len_sq)
            t = np.clip(np.sum(AP * AB, axis=1) / safe_len, 0.0, 1.0)
            # For zero-length edges, t=0 projects to A
            t = np.where(seg_len_sq == 0, 0.0, t)
            proj = A + t[:, None] * AB
            edge_dists = np.hypot(x - proj[:, 0], y - proj[:, 1])
            min_edge_idx = int(np.argmin(edge_dists))
            best_dist = edge_dists[min_edge_idx]
            best_room = int(self._hover_edge_room_idx_arr[min_edge_idx])
            best_edge = int(self._hover_edge_local_idx[min_edge_idx])
            best_pt   = (float(proj[min_edge_idx, 0]), float(proj[min_edge_idx, 1]))

        if best_dist < edge_threshold:
            changed = (self.hover_edge_room_idx != best_room or self.hover_edge_idx != best_edge)
            self.hover_edge_room_idx = best_room
            self.hover_edge_idx      = best_edge
            self.hover_edge_point    = best_pt
            if changed:
                self._update_status("Click to insert vertex, Shift+click to drag edge", 'blue')
                self._update_cursor()
                self._render_section()
        else:
            need_redraw = self.hover_edge_room_idx is not None or vertex_was_hovered
            self.hover_edge_room_idx = self.hover_edge_idx = self.hover_edge_point = None
            if need_redraw:
                self._update_status("Edit Mode: Hover over any vertex to drag", 'blue')
                self._update_cursor()
                self._render_section()

    def _on_key_press(self, event):
        """Handle keyboard shortcuts."""
        if event.key in ('backspace', 'delete', 'escape'):
            if event.key == 'escape':
                if self._align_mode:
                    self._on_overlay_align_toggle(None)  # exit align mode
                elif self.divider_mode:
                    self._exit_divider_mode(cancelled=True)
                elif self.draw_mode:
                    self._exit_draw_mode()
                elif self.edit_mode:
                    # Exit edit mode without saving
                    self.edit_mode = False
                    self.btn_edit_mode.label.set_text('Edit Mode: OFF (Press E)')
                    self._style_toggle_button(self.btn_edit_mode, False)
                    self._reset_hover_state()
                    self._edit_undo_stack.clear()
                    self._create_polygon_selector()
                    self._update_cursor()
                    self._update_status("Edit Mode cancelled (no save)", 'orange')
                    self._render_section(force_full=True)
                else:
                    self._deselect_room()
            return
        if event.key == 's':
            if self.divider_mode:
                if len(self._divider_points) >= 2:
                    self._finalize_division()
                else:
                    self._update_status(
                        "Need at least 2 points to finalize divider — keep clicking or Esc to cancel",
                        'orange')
            else:
                self._on_save_click(None)
        elif event.key == 'S':
            self._save_session()
        elif event.key == 'r':
            self._on_reset_zoom_click(None)
        elif event.key == 'q':
            plt.close(self.fig)
        elif event.key in ('up', 'down', 'left', 'right'):
            # Arrow keys: move overlay if align mode is ON, otherwise navigate HDR files with Up/Down
            if self._align_mode:
                if (self._overlay_visible and self._overlay_rgba is not None
                        and not self.edit_mode and not self.draw_mode and not self.divider_mode):
                    
                    # Ensure blit background is ready for the first move
                    if self._blit_background is None:
                        self._start_overlay_blit()

                    # Accelerating step: fine (0.05%) at tap, ramps to 40× over ~1.5s hold
                    step_pct = 0.0005
                    now = time.monotonic()
                    if self._align_key_last == event.key and self._align_key_press_start is not None:
                        held = now - self._align_key_press_start
                        accel = min(1.0 + held * 25, 40.0)
                    else:
                        accel = 1.0
                        self._align_key_press_start = now
                    self._align_key_last = event.key
                    step = max(1.0, self._image_width * step_pct * accel)
                    hdr = self.current_hdr_name
                    tf = self._make_overlay_manual(hdr)
                    if event.key == 'left':
                        tf['offset_x'] = tf.get('offset_x', 0.0) - step
                    elif event.key == 'right':
                        tf['offset_x'] = tf.get('offset_x', 0.0) + step
                    elif event.key == 'up':
                        tf['offset_y'] = tf.get('offset_y', 0.0) - step
                    elif event.key == 'down':
                        tf['offset_y'] = tf.get('offset_y', 0.0) + step
                    
                    # Fast update using blitting instead of force_full=True
                    self._update_overlay_blit()

                    # Debounced session save and full render cleanup
                    if hasattr(self, '_blit_save_timer') and self._blit_save_timer:
                        self._blit_save_timer.stop()
                    self._blit_save_timer = self.fig.canvas.new_timer(interval=500)
                    self._blit_save_timer.single_shot = True
                    self._blit_save_timer.add_callback(lambda: self._end_overlay_blit(save=True))
                    self._blit_save_timer.start()
            else:
                if event.key == 'up':
                    self._on_next_hdr_click(None)
                elif event.key == 'down':
                    self._on_prev_hdr_click(None)
                # left/right do nothing in normal mode as requested
        elif event.key == 't':
            self._on_image_toggle_click(None)
        elif event.key == 'e':
            self._on_edit_mode_toggle(None)
        elif event.key == 'o':
            self._on_ortho_toggle(None)
        elif event.key == 'd':
            if self.edit_mode:
                if self.divider_mode:
                    self._exit_divider_mode(cancelled=True)
                else:
                    self._enter_divider_mode()
            else:
                if self.draw_mode:
                    self._exit_draw_mode()
                else:
                    self._enter_draw_mode()
        elif event.key == 'f':
            self._fit_to_selected_room()
        elif event.key == 'ctrl+z':
            if self.edit_mode:
                self._undo_edit()
            else:
                self._undo_draw()
        elif event.key == 'ctrl+a':
            self._select_all_rooms()
        elif event.key == 'p':
            self._on_placement_toggle(None)
        elif event.key == 'ctrl+r':
            self._rotate_overlay_90()

    def _on_key_release(self, event):
        """Reset arrow-key acceleration state when a key is released."""
        if event.key in ('up', 'down', 'left', 'right'):
            self._align_key_press_start = None
            self._align_key_last = None

    def _on_scroll(self, event):
        """Handle scroll wheel for zooming or room list scrolling."""
        if event.inaxes == self.ax_list:
            if event.button == 'down':
                self.room_list_scroll_offset += 1
            else:
                self.room_list_scroll_offset = max(0, self.room_list_scroll_offset - 1)
            now = time.monotonic()
            if now - getattr(self, '_last_list_scroll_draw', 0.0) >= 0.05:
                self._last_list_scroll_draw = now
                self._update_room_list()
            return

        if event.inaxes != self.ax:
            return

        # Shift+scroll: overlay scale (align mode only)
        if (event.key == 'shift' and self._align_mode
                and self._overlay_visible
                and self._overlay_rgba is not None
                and not self.edit_mode and not self.draw_mode):

            # Ensure blit background is ready
            if self._blit_background is None:
                self._start_overlay_blit()

            # Use smaller increments for resizing (1% per scroll tick)
            factor = 1.01 if event.button == 'up' else 1 / 1.01
            hdr = self.current_hdr_name
            tf = self._make_overlay_manual(hdr)

            old_sx = tf.get('scale_x', 1.0)
            old_sy = tf.get('scale_y', 1.0)
            new_sx = old_sx * factor
            new_sy = old_sy * factor

            # Determine rendered pixel dimensions (account for 90° rotation)
            rot = tf.get('rotation_90', 0) % 4
            if rot > 0:
                oh, ow = np.rot90(self._overlay_rgba, k=rot).shape[:2]
            else:
                oh, ow = self._overlay_rgba.shape[:2]

            # Current overlay centre in world coordinates
            img_w = self._image_width
            img_h = self._image_height
            default_ox = (img_w - ow * old_sx) / 2.0
            default_oy = (img_h - oh * old_sy) / 2.0
            ox = tf.get('offset_x', default_ox)
            oy = tf.get('offset_y', default_oy)
            cx = ox + ow * old_sx / 2.0
            cy = oy + oh * old_sy / 2.0

            # Shift offset so the centre stays fixed after scaling
            tf['scale_x'] = new_sx
            tf['scale_y'] = new_sy
            tf['offset_x'] = cx - ow * new_sx / 2.0
            tf['offset_y'] = cy - oh * new_sy / 2.0
            
            # Fast update using blitting
            self._update_overlay_blit()

            # Debounced session save and full render cleanup
            if hasattr(self, '_blit_save_timer') and self._blit_save_timer:
                self._blit_save_timer.stop()
            self._blit_save_timer = self.fig.canvas.new_timer(interval=500)
            self._blit_save_timer.single_shot = True
            self._blit_save_timer.add_callback(lambda: self._end_overlay_blit(save=True))
            self._blit_save_timer.start()
            return

        # Ctrl+scroll: overlay alpha
        if event.key == 'control' and self._overlay_visible:
            step = 0.05 if event.button == 'up' else -0.05
            self._overlay_alpha = max(0.1, min(1.0, self._overlay_alpha + step))
            self._render_section(force_full=True)
            self._update_status(f"Overlay alpha: {self._overlay_alpha:.0%}", 'cyan')
            return

        scale  = 1.2 if event.button == 'down' else 1 / 1.2
        xlim   = self.ax.get_xlim()
        ylim   = self.ax.get_ylim()
        xdata, ydata = event.xdata, event.ydata
        new_w  = (xlim[1] - xlim[0]) * scale
        new_h  = (ylim[1] - ylim[0]) * scale
        relx   = (xdata - xlim[0]) / (xlim[1] - xlim[0])
        rely   = (ydata - ylim[0]) / (ylim[1] - ylim[0])
        self.ax.set_xlim([xdata - new_w * relx,   xdata + new_w * (1 - relx)])
        self.ax.set_ylim([ydata - new_h * rely,   ydata + new_h * (1 - rely)])
        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()
        self._df_cursor_bg = None  # invalidate stale blit background after zoom change

        now = time.monotonic()
        if now - getattr(self, '_last_scroll_draw', 0.0) >= 0.033:
            self._last_scroll_draw = now
            self.fig.canvas.draw_idle()
        else:
            if hasattr(self, '_scroll_draw_timer') and self._scroll_draw_timer:
                self._scroll_draw_timer.stop()
            self._scroll_draw_timer = self.fig.canvas.new_timer(interval=60)
            self._scroll_draw_timer.single_shot = True
            self._scroll_draw_timer.add_callback(lambda: self.fig.canvas.draw_idle())
            self._scroll_draw_timer.start()

    def _on_resize(self, event):
        """Handle window resize events."""
        self._blit_background = None
        if not getattr(self, '_launch_complete', False):
            return
        if event.width > 0 and event.height > 0:
            now = time.monotonic()
            if now - getattr(self, '_last_resize_draw', 0.0) >= 0.1:
                self._last_resize_draw = now
                self.fig.canvas.draw_idle()

    def _force_resize_update(self):
        """Force update after window maximisation."""
        try:
            self._launch_complete = True
            canvas    = self.fig.canvas
            tk_widget = canvas.get_tk_widget()
            tk_widget.update_idletasks()
            width  = tk_widget.winfo_width()
            height = tk_widget.winfo_height()
            if width > 1 and height > 1:
                # set_size_inches is a matplotlib API requirement; values are pixel dimensions
                # divided by the figure's dots-per-inch scaling factor.
                self.fig.set_size_inches(width / self.fig.dpi, height / self.fig.dpi, forward=False)
                canvas.resize(width, height)
            canvas.draw_idle()
        except Exception:
            pass

    # === ROOM SELECTION ========================================================

    def _select_room_at(self, x, y, ctrl=False):
        """Select the room polygon at the given point."""
        if x is None or y is None:
            if not ctrl:
                self._deselect_room()
            return
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = np.array(room['vertices'])
            if MplPath(verts).contains_point((x, y)):
                if ctrl:
                    # Ctrl+click on screen: toggle room in multi-selection
                    if i in self.multi_selected_room_idxs:
                        self.multi_selected_room_idxs.discard(i)
                    else:
                        self.multi_selected_room_idxs.add(i)
                        # Also fold the current primary selection into the multi-set
                        if self.selected_room_idx is not None:
                            self.multi_selected_room_idxs.add(self.selected_room_idx)
                    n = len(self.multi_selected_room_idxs)
                    self._update_status(f"Multi-select: {n} room(s)", 'blue')
                    self._update_room_list()
                    self._render_section()
                else:
                    self._select_room(i)
                return
        if not ctrl:
            self._deselect_room()

    def _select_room(self, idx: int):
        """Select a room by index and populate the name textbox and parent."""
        self._deselect_room()
        self.selected_room_idx = idx
        room = self.rooms[idx]
        self.name_textbox.set_val(room.get('name', ''))
        self.room_type = room.get('room_type')
        self._update_room_type_buttons()

        # Auto-populate parent apartment selector
        parent = room.get('parent')
        if parent:
            # Sub-room: show its parent
            self.selected_parent = parent
            self.btn_parent.label.set_text(parent)
            self.name_label_text.set_text("Room Name:")
        else:
            # This IS a parent apartment — set it as the active parent
            self.selected_parent = room.get('name', '')
            self.btn_parent.label.set_text(self.selected_parent)
            self.name_label_text.set_text("Room Name:")

        self._update_status(f"Selected: {room.get('name', 'unnamed')}", 'orange')
        self._update_room_list()
        self._render_section()

    def _deselect_room(self):
        """Deselect any selected room and clear multi-selection."""
        changed = self.selected_room_idx is not None or self.multi_selected_room_idxs
        self.multi_selected_room_idxs.clear()
        if self.selected_room_idx is not None:
            self.selected_room_idx = None
            self.name_textbox.set_val('')
            self.room_type = None
            self._update_room_type_buttons()
        if changed:
            self._update_status("Ready to draw", 'blue')
            self._update_room_list()
            self.fig.canvas.draw_idle()

    def _select_all_rooms(self):
        """Select all rooms on the current HDR (Ctrl+A)."""
        self.multi_selected_room_idxs = {
            i for i, r in enumerate(self.rooms) if self._is_room_on_current_hdr(r)
        }
        n = len(self.multi_selected_room_idxs)
        self._update_status(f"Multi-select: {n} room(s) on current HDR", 'blue')
        self._update_room_list()
        self.fig.canvas.draw_idle()

    # === DF STAMPS =============================================================

    def _remove_df_stamp(self, x: float, y: float):
        """Remove the nearest DF stamp within 40 pixels of (x, y)."""
        hdr = self.current_hdr_name
        stamps = self._df_stamps.get(hdr, [])
        if not stamps:
            return
        # Convert click to display coords for pixel-accurate distance
        disp_click = self.ax.transData.transform((x, y))
        best_idx, best_dist = None, float('inf')
        for i, stamp in enumerate(stamps):
            sx, sy = stamp[0], stamp[1]
            disp_s = self.ax.transData.transform((sx, sy))
            dist = np.hypot(disp_click[0] - disp_s[0], disp_click[1] - disp_s[1])
            if dist < best_dist:
                best_dist, best_idx = dist, i
        if best_idx is not None and best_dist <= 40:
            stamps.pop(best_idx)
            self._df_cursor_bg = None
            self._save_session()
            self._render_section(force_full=True)

    def _draw_df_stamps(self):
        """Draw all stamped DF readings for the current HDR."""
        hdr = self.current_hdr_name
        stamps = self._df_stamps.get(hdr, [])
        for stamp in stamps:
            x, y, df_val = stamp[0], stamp[1], stamp[2]
            # px, py added in later sessions; fall back to rounded data coords for old stamps
            px = stamp[3] if len(stamp) > 3 else int(round(x))
            py = stamp[4] if len(stamp) > 4 else int(round(y))
            self.ax.text(
                x, y, f"DF: {df_val:.2f}%\npx({px},{py})",
                fontsize=7, color='white', va='bottom', ha='left',
                bbox=dict(boxstyle='round,pad=0.2', facecolor='#222222', alpha=0.8, edgecolor='none'),
                zorder=310,
            )
            # Small dot at the stamped pixel
            self.ax.plot(x, y, 'o', color='cyan', markersize=4, zorder=311)

    # === POLYGON SELECTOR ======================================================

    def _create_polygon_selector(self):
        """Create or recreate the polygon selector."""
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.disconnect_events()
        self.selector = PolygonSelector(
            self.ax,
            self._on_polygon_select,
            useblit=True,
            props=dict(color='cyan', linestyle='-', linewidth=2, alpha=0.5),
            handle_props=dict(markersize=8, markerfacecolor='lime',
                              markeredgecolor='darkgreen', markeredgewidth=1.5),
        )
        # Monkey-patch _onmove to apply ortho constraint to the rubber-band
        # line in real time, so the preview matches the final placement.
        _orig_onmove = self.selector._onmove
        editor = self

        def _ortho_onmove(event):
            _orig_onmove(event)
            sel = editor.selector
            if (editor.ortho_mode
                    and not sel._selection_completed
                    and len(sel._xys) >= 2):
                # _xys[-1] is the cursor tracking point; previous vertex is [-2]
                last_x, last_y = sel._xys[-2]
                cx, cy = sel._xys[-1]
                dx, dy = abs(cx - last_x), abs(cy - last_y)
                if dx >= dy:
                    cy = last_y   # horizontal
                else:
                    cx = last_x   # vertical
                sel._xys[-1] = (cx, cy)
                sel._draw_polygon()

        self.selector._onmove = _ortho_onmove
        # Honour draw_mode: selector is inactive by default until 'd' is pressed
        self.selector.set_active(getattr(self, 'draw_mode', False))

    def _on_polygon_select(self, vertices):
        """Callback when a polygon drawing is completed."""
        if len(vertices) < 3:
            return
        self.current_polygon_vertices = list(vertices)
        if self.selected_room_idx is not None:
            self._deselect_room()
        area = 0.5 * abs(
            np.dot([v[0] for v in vertices], np.roll([v[1] for v in vertices], 1))
            - np.dot([v[1] for v in vertices], np.roll([v[0] for v in vertices], 1))
        )
        self._update_status(f"Polygon ready: {len(vertices)} pts, {area:.0f} px2", 'green')

    # === ROOM OPERATIONS =======================================================

    def _on_save_click(self, event):
        """Save room, update selected room, or save edited boundary."""
        if self.divider_mode:
            self._update_status("Complete or cancel divider mode first (Esc to cancel)", 'orange')
            return
        if self.edit_mode and self.edit_room_idx is not None:
            self._save_edited_room()
        elif self.selected_room_idx is not None:
            self._update_selected_room()
        else:
            self._save_current_room()

    def _save_edited_room(self):
        """Save changes to an edited room boundary (also applies name edit)."""
        if self.edit_room_idx is None:
            return
        room = self.rooms[self.edit_room_idx]

        # Apply name from textbox if provided
        typed_name = self.name_textbox.text.strip().upper()
        if typed_name and typed_name != room.get('name', ''):
            old_name = room['name']
            new_name = self._make_unique_name(typed_name, exclude_idx=self.edit_room_idx)
            self._push_draw_undo(('rename', self.edit_room_idx, old_name))
            room['name'] = new_name

        name = room.get('name', 'unnamed')
        self._save_session()
        self.edit_room_idx    = None
        self.hover_vertex_idx = None
        self._update_status(f"Saved changes for '{name}'", 'green')
        self._render_section(force_full=True)
        self._create_polygon_selector()
        self._update_room_list()
        self._update_hdr_list()
        print(f"Saved edited room '{name}'")

    def _save_current_room(self):
        """Save the currently drawn polygon as a new room."""
        # Auto-close: if polygon isn't completed but has 3+ vertices, grab them
        if (len(self.current_polygon_vertices) < 3
                and hasattr(self, 'selector') and self.selector is not None
                and not self.selector._selection_completed
                and len(self.selector.verts) >= 3):
            self.current_polygon_vertices = list(self.selector.verts)

        if len(self.current_polygon_vertices) < 3:
            self._update_status("No polygon to save - draw one first", 'red')
            return

        name = self.name_textbox.text.strip().upper()
        if not name:
            name = f"ROOM_{len(self.rooms) + 1:03d}"

        if self.selected_parent:
            full_name = f"{self.selected_parent}_{name}"
        else:
            full_name = name

        full_name = self._make_unique_name(full_name)
        vertices  = [[float(x), float(y)] for x, y in self.current_polygon_vertices]

        # Boundary containment check
        warning_msg      = ""
        is_outside_parent = False
        if self.selected_parent:
            parent_room = self._get_parent_room(self.selected_parent)
            if parent_room:
                is_outside_parent = not self._check_boundary_containment(vertices, parent_room['vertices'])
                if is_outside_parent:
                    warning_msg = " (WARNING: extends outside parent boundary!)"
                    print(f"WARNING: Room '{full_name}' extends outside parent '{self.selected_parent}'")
            else:
                print(f"Warning: Parent room '{self.selected_parent}' not found")

        # Auto-assign types: sub-rooms default to BED, parent defaults to LIVING
        room_type = self.room_type
        parent_type_changed = None          # (parent_idx, old_type) for undo
        if self.selected_parent:
            if not room_type:
                room_type = 'BED'
            parent_room = self._get_parent_room(self.selected_parent)
            if parent_room and not parent_room.get('room_type'):
                parent_idx = self.rooms.index(parent_room)
                parent_type_changed = (parent_idx, parent_room.get('room_type'))
                parent_room['room_type'] = 'LIVING'

        room = {
            'name':      full_name,
            'parent':    self.selected_parent,
            'vertices':  vertices,
            'hdr_file':  self.current_hdr_name,
            'room_type': room_type,
        }
        self.rooms.append(room)
        new_idx = len(self.rooms) - 1
        self._push_draw_undo(('create', new_idx, parent_type_changed))

        status_color = 'orange' if is_outside_parent else 'green'
        self._update_status(f"Saved '{full_name}'{warning_msg}", status_color)
        print(f"Saved room '{full_name}' on HDR '{self.current_hdr_name}'")

        # Reset drawing state
        self.current_polygon_vertices = []
        self.name_textbox.set_val('')
        self.room_type = None
        self._update_room_type_buttons()
        self._update_parent_options()
        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)
        self._create_polygon_selector()

    def _update_selected_room(self):
        """Update the name of the selected room."""
        if self.selected_room_idx is None:
            return
        idx      = self.selected_room_idx
        old_name = self.rooms[idx]['name']
        new_name = self.name_textbox.text.strip().upper() or old_name
        new_name = self._make_unique_name(new_name, exclude_idx=idx)
        if new_name == old_name:
            return
        self._push_draw_undo(('rename', idx, old_name))
        self.rooms[idx]['name'] = new_name
        self._update_status(f"Renamed '{old_name}' → '{new_name}'", 'green')
        self._update_room_list()
        self._render_section(force_full=True)
        self._save_session()

    def _on_clear_click(self, event):
        """Clear the current polygon drawing."""
        self._deselect_room()
        self.current_polygon_vertices = []
        if self.selector is not None:
            self.selector.clear()
        self._update_status("Cleared - ready to draw", 'blue')

    def _on_delete_click(self, event):
        """Delete the selected room (Ctrl+Z to undo)."""
        if self.selected_room_idx is None:
            self._update_status("No room selected to delete", 'red')
            return
        idx  = self.selected_room_idx
        room = self.rooms.pop(idx)
        name = room.get('name', 'unnamed')
        self._push_draw_undo(('delete', idx, room))
        self.selected_room_idx = None
        self.current_polygon_vertices = []
        self._update_status(f"Deleted '{name}' (Ctrl+Z to undo)", 'green')
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)
        self._create_polygon_selector()
        print(f"Deleted room '{name}'")
        self._save_session()

    # === ZOOM & DISPLAY ========================================================

    def _view_ratio(self) -> float:
        """Return view_w / image_width. 1.0 = fully zoomed out, <1 = zoomed in."""
        if self._image_width <= 1:
            return 1.0
        view_w = abs(self.ax.get_xlim()[1] - self.ax.get_xlim()[0])
        return max(0.01, view_w / self._image_width)

    def _zoom_fontsize(self, base: float = 8.0) -> float:
        """Return font size scaled relative to the default (full-image) zoom.

        At default zoom returns base. Scales down when zoomed out further,
        scales up when zoomed in. Clamps to [4, 20].
        """
        if self._reference_view_w <= 1:
            return base
        view_w = abs(self.ax.get_xlim()[1] - self.ax.get_xlim()[0])
        scale  = self._reference_view_w / max(view_w, 1.0)
        return max(4.0, min(20.0, base * scale))

    def _df_line_step(self) -> float:
        """Return zoom-aware vertical spacing (in data coords) between DF result lines.

        Converts the actual DF font size from points to data-space units via
        the axes transform. This is correct at any zoom level — including
        fit-to-room — because it tracks the real rendered text height rather
        than a view-ratio approximation that breaks when _zoom_fontsize clamps.
        """
        fs_pt = self._zoom_fontsize(base=6.5)   # font size in points
        # 1 point = 1/72 inch. Convert to display pixels, then to data units.
        dpi = self.fig.dpi
        px_per_pt = dpi / 72.0
        fs_px = fs_pt * px_per_pt              # font height in display pixels
        # Map a vertical span of fs_px display pixels to data coordinates.
        # transData maps data → display; its inverse maps display → data.
        inv = self.ax.transData.inverted()
        # Use a fixed display reference point; only the delta matters.
        _, y0_data = inv.transform((0, 0))
        _, y1_data = inv.transform((0, fs_px))
        fs_data = abs(y1_data - y0_data)       # font height in data units
        return fs_data * 1.6                   # 1.6× line height for comfortable spacing

    def _apply_zoom_fontsizes(self):
        """Reapply zoom-dependent font sizes and stroke widths to all cached text."""
        fs = self._zoom_fontsize()
        stroke_name = [patheffects.withStroke(linewidth=fs * 0.06, foreground='black')]
        for label in self._room_label_cache.values():
            if label is not None:
                label.set_fontsize(fs)
                label.set_path_effects(stroke_name)
        fs_df = self._zoom_fontsize(base=6.5)
        stroke_df = [patheffects.withStroke(linewidth=fs_df * 0.06, foreground='black')]
        line_step = self._df_line_step()
        for dt in self._df_text_cache:
            dt.set_fontsize(fs_df)
            dt.set_path_effects(stroke_df)
            # Reposition based on current zoom-aware line step
            centroid = getattr(dt, '_df_centroid', None)
            line_i   = getattr(dt, '_df_line_i', None)
            if centroid is not None and line_i is not None:
                dt.set_position((centroid[0], centroid[1] + 1.2 * line_step + line_i * line_step * 0.7))

    def _zoom_linewidth(self, base: float = 1.5) -> float:
        """Return linewidth that stays visually constant regardless of zoom.

        Matplotlib linewidths are in *points* (screen units) but the polygon
        edges are in *data* coordinates, so when you zoom in the data-space
        features grow on-screen while the point-width stays fixed — making
        lines appear thinner.  To compensate we scale inversely with the
        view ratio: zooming in (smaller view_ratio) → larger linewidth.
        Clamped to [base * 0.5, base * 4] to stay reasonable at extremes.
        """
        ratio = self._view_ratio()                       # 1.0 = full image, <1 = zoomed in
        lw = base / max(ratio, 0.01)                     # inverse: zoom in → thicker
        return max(base * 0.5, min(base * 4.0, lw))

    def _apply_zoom_linewidths(self):
        """Reapply zoom-dependent linewidths to all cached room patches."""
        lw_current = self._zoom_linewidth()
        lw_other = self._zoom_linewidth(base=1.0)
        for patch in self._room_patch_cache.values():
            ptype = getattr(patch, '_patch_type', None)
            if ptype == 'current':
                patch.set_linewidth(lw_current)
            elif ptype == 'other':
                patch.set_linewidth(lw_other)


    def _on_reset_zoom_click(self, event):
        """Reset zoom to the full image extent."""
        self.ax.set_xlim(0, self._image_width)
        self.ax.set_ylim(self._image_height, 0)
        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()
        self._df_cursor_bg = None  # invalidate stale blit background after zoom change
        self.fig.canvas.draw_idle()

    def _fit_to_selected_room(self):
        """Zoom to fit the selected room boundary on screen (press f)."""
        if self.selected_room_idx is None:
            self._update_status("Select a room first, then press 'f' to fit", 'red')
            return
        room = self.rooms[self.selected_room_idx]
        verts = room['vertices']
        if len(verts) < 3:
            return

        xs = [v[0] for v in verts]
        ys = [v[1] for v in verts]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)

        # Add 10% padding around the bounding box
        pad_x = max((max_x - min_x) * 0.10, 20)
        pad_y = max((max_y - min_y) * 0.10, 20)
        self.ax.set_xlim(min_x - pad_x, max_x + pad_x)
        self.ax.set_ylim(max_y + pad_y, min_y - pad_y)  # y-axis is inverted (origin='upper')

        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()
        self._df_cursor_bg = None  # invalidate stale blit background after zoom change
        room_name = room.get('name', 'unnamed')
        self._update_status(f"Fit to '{room_name}' (r to reset zoom)", 'blue')
        self.fig.canvas.draw_idle()

    # === TOGGLE BUTTONS & ROOM TYPE ============================================

    def _update_cursor(self):
        """Set the canvas cursor based on the active editing mode.

        Matplotlib's TkAgg backend overrides the cursor on every mouse-move
        via its own set_cursor(). We patch that method on the canvas instance
        so our custom cursor is not stomped, then restore normal behaviour
        when no custom cursor is needed.
        """
        try:
            canvas = self.fig.canvas
            tkcanvas = canvas._tkcanvas
        except Exception:
            return

        if self._align_mode:
            tkcanvas.configure(cursor='crosshair')
            canvas.set_cursor = lambda cursor: None          # suppress mpl overrides
        elif self.divider_mode:
            tkcanvas.configure(cursor='crosshair')
            canvas.set_cursor = lambda cursor: None          # suppress mpl overrides
        elif self.edit_mode and self.hover_edge_room_idx is not None:
            tkcanvas.configure(cursor='plus')
            canvas.set_cursor = lambda cursor: None          # suppress mpl overrides
        else:
            # Restore matplotlib's own set_cursor before resetting
            from matplotlib.backends._backend_tk import FigureCanvasTk
            canvas.set_cursor = types.MethodType(FigureCanvasTk.set_cursor, canvas)
            tkcanvas.configure(cursor='arrow')

    def _style_toggle_button(self, btn, is_on: bool):
        """Apply pressed/depressed styling to a toggle button.

        ON  = sunken look: darker background, bold white text.
        OFF = raised look: lighter background, normal dark text.
        """
        body = getattr(btn, '_body', None)
        highlight = getattr(btn, '_highlight', None)
        if is_on:
            btn.color      = self._btn_on_color
            btn.hovercolor = self._btn_on_hover
            if body:
                body.set_facecolor(self._btn_on_color)
                body.set_edgecolor('#606058')
            if highlight:
                highlight.set_facecolor('#FFFFFF15')
            btn.label.set_fontweight('bold')
            btn.label.set_color('#FFFFFF')
        else:
            btn.color      = self._btn_color
            btn.hovercolor = self._btn_hover
            if body:
                body.set_facecolor(self._btn_color)
                body.set_edgecolor('#B0B0A8')
            if highlight:
                highlight.set_facecolor('#FFFFFF30')
            btn.label.set_fontweight('medium')
            btn.label.set_color('#333333')

    def _on_edit_mode_toggle(self, event):
        """Toggle edit mode for modifying existing room boundaries."""
        if self._align_mode:
            self._on_overlay_align_toggle(None)
        if self.divider_mode:
            self._exit_divider_mode(cancelled=True)
        if self.draw_mode:
            self._exit_draw_mode()

        self.edit_mode = not self.edit_mode
        self.btn_edit_mode.label.set_text(
            'Boundary Edit Mode: ON (Press E)' if self.edit_mode else 'Boundary Edit Mode: OFF (Press E)')
        self._style_toggle_button(self.btn_edit_mode, self.edit_mode)

        if self.edit_mode:
            self._edit_undo_stack.clear()
            if hasattr(self, 'selector') and self.selector is not None:
                self.selector.set_active(False)
            self._update_status("Boundary Edit Mode: Hover over any vertex to drag (all rooms editable)", 'cyan')
        else:
            self._reset_hover_state()
            self._edit_undo_stack.clear()
            self._save_session()
            self._create_polygon_selector()   # set_active(False) applied inside via draw_mode=False
            self._update_status("Boundary Edit Mode OFF — press 'd' to draw, left-click to select", 'blue')
            self._update_cursor()

        self._render_section(force_full=True)

    def _on_draw_mode_toggle(self, event):
        """Toggle draw mode via the UI button."""
        if self._align_mode:
            self._on_overlay_align_toggle(None)
        if self.edit_mode:
            return   # draw mode not available in edit mode
        if self.draw_mode:
            self._exit_draw_mode()
        else:
            self._enter_draw_mode()

    def _on_ortho_toggle(self, event):
        """Toggle orthogonal lines mode."""
        self.ortho_mode = not self.ortho_mode
        self.btn_ortho.label.set_text(
            'Ortho Lines: ON (Press O)' if self.ortho_mode else 'Ortho Lines: OFF (Press O)')
        self._style_toggle_button(self.btn_ortho, self.ortho_mode)
        state = "ON" if self.ortho_mode else "OFF"
        self._update_status(f"Orthogonal mode: {state}", 'blue')
        self.fig.canvas.draw_idle()

    def _on_placement_toggle(self, event):
        """Toggle DF% stamping on screen on/off (press P).

        When ON:  left-clicking the floor plan stamps a DF% reading at that point.
        When OFF: left-clicking only selects rooms; no new DF% stamps are placed.
        Existing stamps remain visible regardless of this toggle.
        """
        self.placement_mode = not self.placement_mode
        self.btn_placement.label.set_text(
            'Place DF% Point: ON (Press P)' if self.placement_mode else 'Place DF% Point: OFF (Press P)')
        self._style_toggle_button(self.btn_placement, self.placement_mode)
        state = "ON" if self.placement_mode else "OFF"
        self._update_status(f"DF% placement mode: {state}", 'blue')
        self.fig.canvas.draw_idle()

    # ── PDF Overlay Callbacks ──────────────────────────────────────────────────

    def _on_overlay_toggle(self, event):
        """Toggle PDF overlay visibility."""
        if self._overlay_rgba is None:
            if self._overlay_pdf_path is not None:
                self._rasterize_overlay_page()
            else:
                self._update_status("No PDF loaded — set pdf_path in HdrAoiEditor()", 'orange')
                return

        # Preserve current zoom so toggling visibility never changes the view
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()

        self._overlay_visible = not self._overlay_visible
        label = 'Floor Plan: ON' if self._overlay_visible else 'Floor Plan: OFF'
        self.btn_overlay_toggle.label.set_text(label)
        self._style_toggle_button(self.btn_overlay_toggle, self._overlay_visible)
        self._save_session()
        self._do_full_render(xlim, ylim, reset_view=False)

    def _on_overlay_page_cycle(self, event):
        """Cycle to next PDF page."""
        if self._overlay_pdf_path is None or self._overlay_pdf_info is None:
            self._update_status("No PDF loaded", 'orange')
            return
        n = self._overlay_pdf_info['page_count']
        self._overlay_page_idx = (self._overlay_page_idx + 1) % n
        self._rasterize_overlay_page()
        self._update_overlay_page_label()
        # Persist page assignment per HDR
        hdr = self.current_hdr_name
        tf = self._overlay_transforms.setdefault(hdr, {})
        tf['page_idx'] = self._overlay_page_idx
        self._save_session()
        self._render_section(force_full=True)

    def _on_aoi_level_cycle(self, event):
        """Cycle to the next FFL group and reassign those rooms to the current .pic level."""
        ffl_groups = self._aoi_ffl_groups()
        if not ffl_groups:
            self._update_status("No IESVE AOI rooms loaded", 'orange')
            return
        n = len(ffl_groups)
        self._aoi_level_idx = (self._aoi_level_idx + 1) % n
        self._reassign_aoi_level()
        self._update_aoi_level_label()
        self._save_session()
        self._render_section(force_full=True)

    def _aoi_ffl_groups(self) -> list:
        """Return sorted list of unique FFL values from rooms that have a 'ffl' key."""
        ffls = sorted({r['ffl'] for r in self.rooms if 'ffl' in r})
        return ffls

    def _reassign_aoi_level(self):
        """Re-project rooms in the current FFL group using the current .pic's VIEW header
        and assign them to the current HDR file."""
        ffl_groups = self._aoi_ffl_groups()
        if not ffl_groups or not self.hdr_files:
            return
        target_ffl = ffl_groups[self._aoi_level_idx]
        entry      = self.hdr_files[self.current_hdr_idx]
        view_params = self._read_view_params(entry['hdr_path'])
        if view_params is None:
            self._update_status(f"No VIEW params in {entry['name']}", 'orange')
            return
        vp_x, vp_y, vh_val, vv_val, img_w, img_h = view_params
        for room in self.rooms:
            if room.get('ffl') == target_ffl and 'world_vertices' in room:
                room['vertices'] = self._world_to_pixels(
                    room['world_vertices'], vp_x, vp_y, vh_val, vv_val, img_w, img_h)
                room['hdr_file'] = entry['name']
                room.pop('df_cache', None)
        self._update_status(
            f"Assigned FFL {target_ffl} m rooms → {entry['name']}", 'green')

    def _update_aoi_level_label(self):
        """Update the AOI level cycle button label."""
        if not hasattr(self, 'btn_aoi_level'):
            return
        ffl_groups = self._aoi_ffl_groups()
        if not ffl_groups:
            self.btn_aoi_level.label.set_text('AOI Level: -/-')
        else:
            n   = len(ffl_groups)
            idx = self._aoi_level_idx % n
            ffl = ffl_groups[idx]
            self.btn_aoi_level.label.set_text(f'AOI Level: {idx + 1}/{n} ({ffl} m)')
        self.fig.canvas.draw_idle()

    # === OVERLAY BLITTING OPTIMIZATIONS ========================================

    def _start_overlay_blit(self):
        """Capture a clean background for fast overlay movement via blitting.
        Background includes the HDR image and all room polygons/labels, but
        specifically excludes the PDF overlay itself.
        """
        canvas = self.fig.canvas
        if not getattr(canvas, 'supports_blit', False) or self._overlay_handle is None:
            return

        # Temporarily hide overlay to capture everything underneath
        self._overlay_handle.set_visible(False)
        canvas.draw()
        self._blit_background = canvas.copy_from_bbox(self.ax.bbox)
        self._overlay_handle.set_visible(True)
        self._blit_active = True

    def _update_overlay_blit(self):
        """Update only the PDF overlay's position and redraw using blitting.
        This provides instantaneous feedback during arrow-key alignment.
        """
        if self._overlay_handle is None:
            return

        # Update the artist's extent from stored transforms
        hdr_name = self.current_hdr_name
        tf = self._overlay_transforms.get(hdr_name, {})
        sx = tf.get('scale_x', 1.0)
        sy = tf.get('scale_y', 1.0)
        rot = tf.get('rotation_90', 0) % 4
        
        rgba = self._overlay_rgba
        if rot > 0:
            oh, ow = np.rot90(rgba, k=rot).shape[:2]
        else:
            oh, ow = rgba.shape[:2]
            
        ox = tf.get('offset_x', 0.0)
        oy = tf.get('offset_y', 0.0)
        
        self._overlay_handle.set_extent([
            ox, ox + ow * sx,
            oy + oh * sy, oy,
        ])

        canvas = self.fig.canvas
        if self._blit_active and self._blit_background is not None:
            canvas.restore_region(self._blit_background)
            self.ax.draw_artist(self._overlay_handle)
            canvas.blit(self.ax.bbox)
        else:
            # Fallback for systems without blit support
            canvas.draw_idle()

    def _end_overlay_blit(self, save=True):
        """Finalise movement and save session state."""
        self._blit_active = False
        self._blit_background = None
        if save:
            self._save_session()
        self._render_section(force_full=True) # Full render to restore everything properly

    def _on_overlay_align_toggle(self, event):
        """Enter/exit two-point alignment mode."""
        if self._overlay_rgba is None:
            self._update_status("Load a PDF first", 'orange')
            return
        self._align_mode = not self._align_mode
        
        # Show/hide Reset Level Alignment button
        self.btn_overlay_reset.ax.set_visible(self._align_mode)
        
        if self._align_mode:
            # Exit other modes to prevent conflicts
            if self.edit_mode:
                self._on_edit_mode_toggle(None)
            if self.draw_mode:
                self._on_draw_mode_toggle(None)
            self._align_points_overlay = []
            self._align_points_hdr = []
            self._clear_align_markers()
            self.btn_overlay_align.label.set_text('Align Floor Plan: ON')
            self._style_toggle_button(self.btn_overlay_align, True)
            self._update_status("ALIGN: Click first point on overlay", 'cyan')
        else:
            self._clear_align_markers()
            self.btn_overlay_align.label.set_text('Align Floor Plan: OFF')
            self._style_toggle_button(self.btn_overlay_align, False)
            self._update_status("Alignment finished", 'green')
            self._end_overlay_blit(save=True)
        self.fig.canvas.draw_idle()

    def _on_overlay_reset_click(self, event):
        """Clear manual overrides for the current level and revert to inheritance."""
        hdr = self.current_hdr_name
        if hdr in self._overlay_transforms:
            tf = self._overlay_transforms[hdr]
            # Remove manual flags and spatial overrides
            tf.pop('is_manual', None)
            tf.pop('offset_x', None)
            tf.pop('offset_y', None)
            tf.pop('scale_x', None)
            tf.pop('scale_y', None)
            tf.pop('rotation_90', None)
            
            self._update_status(f"Reset '{hdr}' to default inheritance", 'blue')
            self._save_session()
            self._render_section(force_full=True)
        else:
            self._update_status("No custom alignment to reset on this level", 'orange')

    _DPI_PRESETS = [72, 100, 150, 200, 300]

    def _on_overlay_dpi_click(self, event):
        """Toggle DPI preset dropdown visibility."""
        self._toggle_dpi_dropdown(not self._dpi_dropdown_visible)

    def _on_dpi_radio_select(self, label):
        """Handle selection from DPI RadioButtons dropdown."""
        # label is e.g. "150 DPI"
        new_dpi = int(label.split()[0])
        old_dpi = self._overlay_raster_dpi
        if new_dpi == old_dpi:
            self._toggle_dpi_dropdown(False)
            return

        # Adjust existing transforms to maintain PDF world-space scale and position
        # ow_new = ow_old * (new_dpi / old_dpi)
        # To keep (ow * scale) constant: scale_new = scale_old * (old_dpi / new_dpi)
        k = old_dpi / new_dpi
        for hdr in self._overlay_transforms:
            tf = self._overlay_transforms[hdr]
            if 'scale_x' in tf: tf['scale_x'] *= k
            if 'scale_y' in tf: tf['scale_y'] *= k
            # offset_x/y are in world coordinates (HDR pixels) and don't need adjustment

        self._overlay_raster_dpi = new_dpi
        self.btn_overlay_dpi.label.set_text(f'PDF Res: {self._overlay_raster_dpi} DPI')

        # Hide dropdown after selection
        self._toggle_dpi_dropdown(False)

        self._save_session()
        if self._overlay_pdf_path is not None:
            # Capture current view limits before re-rasterization/full render
            xlim = self.ax.get_xlim()
            ylim = self.ax.get_ylim()
            self._rasterize_overlay_page()
            # force_full=True triggers _do_full_render with these limits
            self._do_full_render(xlim, ylim, reset_view=False)
        else:
            self.fig.canvas.draw_idle()

    def _toggle_dpi_dropdown(self, visible: bool):
        """Thoroughly show or hide the DPI dropdown and all its components."""
        self._dpi_dropdown_visible = visible
        self.ax_dpi_radio.set_visible(visible)
        # Explicitly toggle all children (circles, labels, spines) to ensure visibility
        for artist in self.ax_dpi_radio.get_children():
            artist.set_visible(visible)
        self.fig.canvas.draw_idle()

    def _get_overlay_cache_path(self, page_idx: Optional[int] = None, dpi: Optional[int] = None) -> Optional[Path]:
        """Generate a unique, persistent cache path for a specific PDF page and resolution.
        
        Filename includes the PDF stem, page index, DPI, and a short hash of the 
        absolute PDF path to prevent collisions between different files with the same name.
        """
        if self._overlay_pdf_path is None:
            return None
            
        page = page_idx if page_idx is not None else self._overlay_page_idx
        res  = dpi if dpi is not None else self._overlay_raster_dpi
        pdf_path = self._overlay_pdf_path
        
        # Create a 6-char hash of the absolute path to handle duplicate filenames in different folders
        pdf_hash = hashlib.md5(str(pdf_path.absolute()).encode()).hexdigest()[:6]

        cache_dir = self.project_aoi_dir / ".overlay_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)        
        # Example: FloorPlans_p0_300dpi_a1b2c3.npy
        fname = f"{pdf_path.stem}_p{page}_{res}dpi_{pdf_hash}.npy"
        return cache_dir / fname

    def _start_background_prefetch(self):
        """Prefetch adjacent pages (n+1, n-1) in a background thread to eliminate future lag."""
        if self._overlay_pdf_info is None or self._overlay_pdf_path is None:
            return
            
        current = self._overlay_page_idx
        total = self._overlay_pdf_info['page_count']
        
        # Pages to prefetch: look ahead first, then look behind
        to_fetch = []
        if current + 1 < total:
            to_fetch.append(current + 1)
        if current - 1 >= 0:
            to_fetch.append(current - 1)
            
        # Filter out pages already being prefetched to avoid redundant threads
        to_fetch = [p for p in to_fetch if p not in self._prefetching_pages]
        
        if not to_fetch:
            return

        def _worker():
            pdf_mtime = self._overlay_pdf_path.stat().st_mtime
            for p in to_fetch:
                self._prefetching_pages.add(p)
                try:
                    path = self._get_overlay_cache_path(page_idx=p)
                    # Only rasterize if not already in cache or if cache is stale
                    if not path.exists() or path.stat().st_mtime < pdf_mtime:
                        self._rasterize_overlay_page(page_idx=p, use_bg_thread=True)
                finally:
                    self._prefetching_pages.discard(p)

        threading.Thread(target=_worker, daemon=True).start()

    def _rasterize_overlay_page(self, page_idx: Optional[int] = None, use_bg_thread: bool = False):
        """Rasterize a PDF page and apply transparency, using a persistent multi-page cache.
        
        If page_idx is provided and use_bg_thread is True, it performs a silent 
        background cache generation. Otherwise, it updates the active overlay image.
        """
        if self._overlay_pdf_path is None:
            return

        target_page = page_idx if page_idx is not None else self._overlay_page_idx
        cache_path = self._get_overlay_cache_path(page_idx=target_page)
        pdf_mtime  = self._overlay_pdf_path.stat().st_mtime

        # Cache is valid if it exists and is newer than the source PDF
        cache_hit = (
            cache_path.exists()
            and cache_path.stat().st_mtime >= pdf_mtime
        )

        if cache_hit:
            if not use_bg_thread:
                # Instant load for active page
                try:
                    self._overlay_rgba = np.load(str(cache_path))
                except (ValueError, EOFError):
                    # Corrupt cache (rare due to atomic save, but possible)
                    cache_hit = False

        if not cache_hit:
            # Rasterize and process (the slow part)
            rgba = rasterize_pdf_page(
                self._overlay_pdf_path,
                target_page,
                dpi=self._overlay_raster_dpi,
            )
            processed = make_lines_only(rgba, white_threshold=240)
            
            # Atomic Save: Save to temp file then rename to prevent partial-read crashes
            tmp_path = cache_path.with_suffix(".tmp.npy")
            np.save(str(tmp_path), processed)
            os.replace(str(tmp_path), str(cache_path))
            
            if not use_bg_thread:
                self._overlay_rgba = processed
                self._overlay_cache_pdf = str(self._overlay_pdf_path)
                self._overlay_cache_dpi = self._overlay_raster_dpi
                self._save_session()

        # If we just loaded/created the main active page, trigger pre-fetch for its neighbors
        if not use_bg_thread:
            self._start_background_prefetch()

    def _update_overlay_page_label(self):
        """Update the page cycle button label."""
        if self._overlay_pdf_info is None:
            self.btn_overlay_page.label.set_text('Cycle Floor Plans: -/-')
        else:
            n = self._overlay_pdf_info['page_count']
            self.btn_overlay_page.label.set_text(f'Cycle Floor Plans: {self._overlay_page_idx + 1}/{n}')
        self.fig.canvas.draw_idle()

    def _make_overlay_manual(self, hdr_name: str):
        """Ensure the given HDR level has a manual transform, starting from inherited values.
        
        This prevents 'jumps' when starting to adjust an inherited alignment.
        """
        tf = self._overlay_transforms.get(hdr_name, {})
        if tf.get('is_manual'):
            return tf

        # Find the HDR index for this name
        hdr_idx = -1
        for i, entry in enumerate(self.hdr_files):
            if entry['name'] == hdr_name:
                hdr_idx = i
                break

        # Get the effective values (including inherited ones)
        eff_tf = self._get_effective_overlay_transform(hdr_idx)

        # Populate the current level's transform with these values and mark it manual
        tf = self._overlay_transforms.setdefault(hdr_name, {})
        for k in ['offset_x', 'offset_y', 'scale_x', 'scale_y', 'rotation_90']:
            if k in eff_tf:
                tf[k] = eff_tf[k]
        tf['is_manual'] = True
        return tf

    def _get_effective_overlay_transform(self, hdr_idx: Optional[int] = None) -> dict:
        """Return the transform for the given HDR level, inheriting from below if needed.

        If the current level hasn't been manually adjusted, it searches downwards
        through lower floors to find the first level that has manual alignment
        data and inherits its scale, offset, and rotation.
        """
        if hdr_idx is None:
            hdr_idx = self.current_hdr_idx

        if not (0 <= hdr_idx < len(self.hdr_files)):
            return {}

        hdr_name = self.hdr_files[hdr_idx]['name']
        tf = self._overlay_transforms.get(hdr_name, {})

        # Helper to check if a transform dict contains meaningful spatial data
        def is_actually_manual(t):
            return t.get('is_manual') or 'scale_x' in t or 'offset_x' in t

        # If this level is explicitly or implicitly manual, we're done
        if is_actually_manual(tf):
            return tf

        # Search downwards for a manual level to inherit from
        for i in range(hdr_idx - 1, -1, -1):
            prev_name = self.hdr_files[i]['name']
            prev_tf = self._overlay_transforms.get(prev_name, {})
            if is_actually_manual(prev_tf):
                # Inherit scale, offset, and rotation, but NOT the page index
                inherited = prev_tf.copy()
                inherited['is_manual'] = False # Still not manual for this level
                
                # Keep our current level-specific overrides if we have them
                # page_idx is always level-specific
                if 'page_idx' in tf:
                    inherited['page_idx'] = tf['page_idx']
                # rotation_90 is also preserved if set specifically on this level 
                # even if not manual (though usually it would be manual)
                if 'rotation_90' in tf:
                    inherited['rotation_90'] = tf['rotation_90']
                    
                return inherited

        # No manual level found below, return what we have (or defaults)
        return tf

    def _render_overlay(self, img_w: int, img_h: int):
        """Draw the PDF overlay as a second imshow artist above the base image."""
        tf = self._get_effective_overlay_transform()
        sx = tf.get('scale_x', 1.0)
        sy = tf.get('scale_y', 1.0)
        rot = tf.get('rotation_90', 0) % 4

        rgba = self._overlay_rgba
        if rgba is None:
            return

        if rot > 0:
            rgba = np.rot90(rgba, k=rot)

        oh, ow = rgba.shape[:2]

        # Default offset: centre overlay on HDR image; use saved offset if present
        default_ox = (img_w - ow * sx) / 2.0
        default_oy = (img_h - oh * sy) / 2.0
        ox = tf.get('offset_x', default_ox)
        oy = tf.get('offset_y', default_oy)

        # Build display array with user-controlled alpha
        display = rgba.copy()
        display[:, :, 3] = (display[:, :, 3].astype(float) * self._overlay_alpha).astype(np.uint8)

        extent_left   = ox
        extent_right  = ox + ow * sx
        extent_top    = oy
        extent_bottom = oy + oh * sy

        self._overlay_handle = self.ax.imshow(
            display, origin='upper',
            extent=[extent_left, extent_right, extent_bottom, extent_top],
            aspect='equal', zorder=1, interpolation='bilinear',
        )

    def _update_overlay_extent(self):
        """Update the overlay artist's extent from current transform state (fast path for drag)."""
        if self._overlay_handle is None or self._overlay_rgba is None:
            return
        tf = self._get_effective_overlay_transform()
        sx = tf.get('scale_x', 1.0)
        sy = tf.get('scale_y', 1.0)
        rot = tf.get('rotation_90', 0) % 4

        rgba = self._overlay_rgba
        if rot > 0:
            oh, ow = np.rot90(rgba, k=rot).shape[:2]
        else:
            oh, ow = rgba.shape[:2]

        img_w = getattr(self, '_image_width', 0)
        img_h = getattr(self, '_image_height', 0)
        default_ox = (img_w - ow * sx) / 2.0
        default_oy = (img_h - oh * sy) / 2.0
        ox = tf.get('offset_x', default_ox)
        oy = tf.get('offset_y', default_oy)

        self._overlay_handle.set_extent([
            ox, ox + ow * sx,
            oy + oh * sy, oy,
        ])

    def _clear_align_markers(self):
        """Remove alignment marker artists from the canvas."""
        for m in self._align_markers:
            try:
                m.remove()
            except ValueError:
                pass
        self._align_markers = []

    def _on_align_click(self, x, y):
        """Handle a click during two-point alignment mode."""
        n_overlay = len(self._align_points_overlay)
        n_hdr = len(self._align_points_hdr)

        if n_overlay == n_hdr:
            # Collecting overlay point
            self._align_points_overlay.append((x, y))
            marker = self.ax.scatter(x, y, c='lime', s=100, zorder=500, marker='+', linewidths=2)
            self._align_markers.append(marker)
            if n_overlay == 0:
                self._update_status("ALIGN: Now click matching point on HDR image", 'cyan')
            else:
                self._update_status("ALIGN: Now click matching point #2 on HDR image", 'cyan')
            self.fig.canvas.draw_idle()
        else:
            # Collecting HDR point
            self._align_points_hdr.append((x, y))
            marker = self.ax.scatter(x, y, c='red', s=100, zorder=500, marker='+', linewidths=2)
            self._align_markers.append(marker)
            self.fig.canvas.draw_idle()

            if len(self._align_points_hdr) == 1:
                self._update_status("ALIGN: Click second point on overlay", 'cyan')
            elif len(self._align_points_hdr) == 2:
                self._compute_two_point_alignment()

    def _compute_two_point_alignment(self):
        """Compute translation + uniform scale from two point pairs."""
        o1, o2 = self._align_points_overlay
        h1, h2 = self._align_points_hdr

        hdr = self.current_hdr_name
        # Use effective (inherited) values as base for the computation
        eff_tf = self._get_effective_overlay_transform()
        old_sx = eff_tf.get('scale_x', 1.0)
        old_sy = eff_tf.get('scale_y', 1.0)
        old_ox = eff_tf.get('offset_x', 0.0)
        old_oy = eff_tf.get('offset_y', 0.0)

        # Convert overlay data-coord clicks to overlay-image-space
        img_x1 = (o1[0] - old_ox) / old_sx
        img_y1 = (o1[1] - old_oy) / old_sy
        img_x2 = (o2[0] - old_ox) / old_sx
        img_y2 = (o2[1] - old_oy) / old_sy

        d_img = np.sqrt((img_x2 - img_x1)**2 + (img_y2 - img_y1)**2)
        if d_img < 1e-6:
            self._update_status("Overlay points too close together", 'red')
            return

        h1_arr, h2_arr = np.array(h1), np.array(h2)
        d_hdr = np.linalg.norm(h2_arr - h1_arr)

        new_scale = d_hdr / d_img

        # Offset so image point 1 maps to HDR point 1
        new_ox = h1[0] - img_x1 * new_scale
        new_oy = h1[1] - img_y1 * new_scale

        tf = self._overlay_transforms.setdefault(hdr, {})
        tf['scale_x'] = new_scale
        tf['scale_y'] = new_scale
        tf['offset_x'] = new_ox
        tf['offset_y'] = new_oy
        tf['is_manual'] = True

        # Exit align mode
        self._align_mode = False
        self._clear_align_markers()
        self.btn_overlay_align.label.set_text('Align Floor Plan: OFF')
        self._style_toggle_button(self.btn_overlay_align, False)
        self.btn_overlay_reset.ax.set_visible(False)
        self._save_session()
        self._render_section(force_full=True)
        self._update_status("Alignment applied (2-point)", 'green')

    def _rotate_overlay_90(self):
        """Rotate overlay by 90 degrees clockwise."""
        if self._overlay_rgba is None:
            return
        hdr = self.current_hdr_name
        tf = self._overlay_transforms.setdefault(hdr, {})
        tf['rotation_90'] = (tf.get('rotation_90', 0) + 1) % 4
        self._save_session()
        self._render_section(force_full=True)
        self._update_status(f"Overlay rotated to {tf['rotation_90'] * 90}\u00b0", 'cyan')

    def _on_room_type_toggle(self, room_type: str, **_kw):
        """Toggle a room type on/off.  Supports bulk tagging when multi-selected."""
        # In multi-select mode: always set (no toggle-off), apply to all
        if self.multi_selected_room_idxs:
            self.room_type = room_type
            self._apply_room_type_change()
            return

        # Single-select: toggle behaviour
        if self.room_type == room_type:
            self.room_type = None
            self._apply_room_type_change()
            return
        self.room_type = room_type
        self._apply_room_type_change()

    def _apply_room_type_change(self):
        """Update UI and persist room type change.

        When ``multi_selected_room_idxs`` is non-empty the change is applied
        to every room in the set.
        """
        self._update_room_type_buttons()

        # Collect target indices (multi-select takes priority)
        if self.multi_selected_room_idxs:
            target_idxs = list(self.multi_selected_room_idxs)
        elif self.selected_room_idx is not None:
            target_idxs = [self.selected_room_idx]
        else:
            target_idxs = []

        if not target_idxs:
            self.fig.canvas.draw_idle()
            return

        # Snapshot old types for undo
        old_types = [(idx, self.rooms[idx].get('room_type')) for idx in target_idxs]
        self._push_draw_undo(('type', old_types))

        for idx in target_idxs:
            self.rooms[idx]['room_type'] = self.room_type
            self._invalidate_room_df_cache(idx)

        self._save_session()
        self._update_room_list()
        self._render_section(force_full=True)

        # Status feedback
        tag = self.room_type or 'None'
        applied = len(target_idxs)
        if applied > 1:
            self._update_status(f"Set {applied} rooms to {tag}", 'green')
        elif applied == 1:
            name = self.rooms[target_idxs[0]].get('name', '')
            self._update_status(f"{name} → {tag}", 'green')

        self.fig.canvas.draw_idle()

    def _update_room_type_buttons(self):
        """Style BED/LIVING/CIRC buttons to reflect current room_type."""
        self._style_toggle_button(self.btn_room_type_bed, self.room_type == 'BED')
        self._style_toggle_button(self.btn_room_type_living, self.room_type == 'LIVING')
        self._style_toggle_button(self.btn_room_type_nonresi, self.room_type == 'NON-RESI')
        self._style_toggle_button(self.btn_room_type_circ, self.room_type == 'CIRC')

    # === ROOM DIVIDER ==========================================================

    @staticmethod
    def _points_close(a, b, tol=1.0) -> bool:
        """Return True if two points are within *tol* pixels of each other."""
        return abs(a[0] - b[0]) < tol and abs(a[1] - b[1]) < tol

    @staticmethod
    def _point_on_segment(pt, a, b, tol=2.0) -> bool:
        """Return True if *pt* lies on segment a→b within pixel tolerance."""
        px, py = float(pt[0]), float(pt[1])
        ax_, ay_ = float(a[0]), float(a[1])
        bx_, by_ = float(b[0]), float(b[1])
        # Collinearity via cross product
        cross = abs((bx_ - ax_) * (py - ay_) - (by_ - ay_) * (px - ax_))
        seg_len = ((bx_ - ax_) ** 2 + (by_ - ay_) ** 2) ** 0.5
        if seg_len > 0 and cross / seg_len > tol:
            return False
        # Bounding box check
        min_x, max_x = min(ax_, bx_) - tol, max(ax_, bx_) + tol
        min_y, max_y = min(ay_, by_) - tol, max(ay_, by_) + tol
        return min_x <= px <= max_x and min_y <= py <= max_y

    @staticmethod
    def _line_polygon_intersections(line_start, line_end, polygon_verts):
        """Find all points where an infinite line (through *line_start* and
        *line_end*) intersects the edges of *polygon_verts*.

        Returns a list of (x, y) tuples sorted by distance from *line_start*.
        Near-duplicate points (within 1 px) are filtered out.
        """
        intersections = []
        sx, sy = float(line_start[0]), float(line_start[1])
        ex, ey = float(line_end[0]),   float(line_end[1])
        dx, dy = ex - sx, ey - sy

        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]),           float(polygon_verts[i][1])
            bx_, by_ = float(polygon_verts[(i + 1) % n][0]), float(polygon_verts[(i + 1) % n][1])
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = dx * edge_dy - dy * edge_dx
            if abs(denom) < 1e-10:
                continue  # parallel / collinear
            u = ((ax_ - sx) * dy - (ay_ - sy) * dx) / denom
            if -1e-10 <= u <= 1.0 + 1e-10:
                ix = ax_ + u * edge_dx
                iy = ay_ + u * edge_dy
                intersections.append((float(ix), float(iy)))

        # Deduplicate (adjacent edges sharing a vertex produce two hits)
        if len(intersections) > 1:
            unique = [intersections[0]]
            for pt in intersections[1:]:
                if not any(abs(pt[0] - up[0]) < 1.0 and abs(pt[1] - up[1]) < 1.0 for up in unique):
                    unique.append(pt)
            intersections = unique

        intersections.sort(key=lambda p: (p[0] - sx) ** 2 + (p[1] - sy) ** 2)
        return intersections

    def _split_polygon_by_line(self, verts, p1, p2):
        """Split a closed polygon into two sub-polygons along segment p1→p2.

        *p1* and *p2* must lie on the polygon boundary.
        Returns ``(poly_a, poly_b)`` as lists of ``[x, y]``, or
        ``(None, None)`` on failure.
        """
        n = len(verts)
        augmented = []
        p1_idx = None
        p2_idx = None

        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1

            # Check if p1 lies on edge a→b
            if p1_idx is None and self._point_on_segment(p1, a, b):
                if self._points_close(p1, a):
                    p1_idx = cur
                elif self._points_close(p1, b):
                    pass  # will be picked up when b is appended
                else:
                    augmented.append([float(p1[0]), float(p1[1])])
                    p1_idx = len(augmented) - 1

            # Check if p2 lies on edge a→b
            if p2_idx is None and self._point_on_segment(p2, a, b):
                if self._points_close(p2, a):
                    p2_idx = cur
                elif self._points_close(p2, b):
                    pass  # will be picked up when b is appended
                else:
                    augmented.append([float(p2[0]), float(p2[1])])
                    p2_idx = len(augmented) - 1

        # Handle the deferred "close to b" cases
        if p1_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p1, v):
                    p1_idx = i
                    break
        if p2_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p2, v):
                    p2_idx = i
                    break

        if p1_idx is None or p2_idx is None:
            return None, None

        # Walk forward from p1→p2 to collect poly_a
        m = len(augmented)
        poly_a = []
        idx = p1_idx
        while True:
            poly_a.append(augmented[idx])
            if idx == p2_idx and len(poly_a) > 1:
                break
            idx = (idx + 1) % m

        # Walk forward from p2→p1 to collect poly_b
        poly_b = []
        idx = p2_idx
        while True:
            poly_b.append(augmented[idx])
            if idx == p1_idx and len(poly_b) > 1:
                break
            idx = (idx + 1) % m

        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    # --- draw mode lifecycle ---------------------------------------------------

    def _enter_draw_mode(self):
        """Activate polygon drawing mode."""
        self.draw_mode = True
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.set_active(True)
        if hasattr(self, 'btn_draw_mode'):
            self.btn_draw_mode.label.set_text('Draw Mode: ON (Press D)')
            self._style_toggle_button(self.btn_draw_mode, True)
        self._update_status("Draw Mode: left-click to place vertices, s=save, Esc=exit", 'cyan')
        self.fig.canvas.draw_idle()

    def _exit_draw_mode(self):
        """Deactivate polygon drawing mode and clear any in-progress polygon."""
        self.draw_mode = False
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.set_active(False)
            self.selector.clear()
        self.current_polygon_vertices = []
        self._pending_snap = None
        if hasattr(self, 'btn_draw_mode'):
            self.btn_draw_mode.label.set_text('Draw Mode: OFF (Press D)')
            self._style_toggle_button(self.btn_draw_mode, False)
        self._update_status("Draw Mode OFF — press 'd' to draw, left-click to select", 'blue')
        self.fig.canvas.draw_idle()

    # --- divider mode lifecycle ------------------------------------------------

    def _enter_divider_mode(self):
        """Enter divider mode: draw an ortho polyline to split the selected room."""
        if not self.edit_mode:
            self._update_status("Must be in edit mode to use divider", 'red')
            return
        if self.selected_room_idx is None:
            self._update_status("Select a room first (right-click), then press 'd'", 'red')
            return

        room = self.rooms[self.selected_room_idx]
        if len(room['vertices']) < 3:
            self._update_status("Selected room has too few vertices to divide", 'red')
            return

        self.divider_mode          = True
        self._divider_room_idx     = self.selected_room_idx
        self._divider_points       = []
        self._divider_markers      = []
        self._divider_segments     = []
        self._divider_preview_line = None

        self._reset_hover_state()
        self._update_cursor()

        room_name = room.get('name', 'unnamed')
        self._update_status(
            f"DIVIDER: Click points to build ortho path on '{room_name}' "
            f"(s=finish, right-click=undo, Esc=cancel)", 'magenta')
        self.fig.canvas.draw_idle()

    def _exit_divider_mode(self, cancelled: bool = False):
        """Exit divider mode, cleaning up all preview artists."""
        self.divider_mode      = False
        self._divider_room_idx = None
        self._divider_points   = []
        self._update_cursor()

        for artist in self._divider_markers:
            if artist is not None:
                try: artist.remove()
                except ValueError: pass
        self._divider_markers = []

        for artist in self._divider_segments:
            if artist is not None:
                try: artist.remove()
                except ValueError: pass
        self._divider_segments = []

        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        if self._divider_snap_marker is not None:
            try: self._divider_snap_marker.remove()
            except ValueError: pass
            self._divider_snap_marker = None
        self._divider_snap_pt = None

        if cancelled:
            self._update_status("Divider mode cancelled", 'blue')
        self._render_section(force_full=True)

    # --- divider click / preview / undo ----------------------------------------

    def _on_divider_click(self, x: float, y: float):
        """Handle a left-click during divider mode — accumulate polyline points."""
        if self._divider_room_idx is None:
            return

        # Use highlighted snap vertex if one is active, else normal snap pipeline
        if self._divider_snap_pt is not None:
            x, y = self._divider_snap_pt
        else:
            x, y = self._snap_to_pixel(x, y)
            x, y = self._snap_to_vertex(x, y)

        # Ortho constraint from last placed point
        if self._divider_points:
            lx, ly = self._divider_points[-1]
            dx, dy = abs(x - lx), abs(y - ly)
            if dx >= dy:
                y = ly   # horizontal
            else:
                x = lx   # vertical

        # Reject zero-length segment
        if self._divider_points:
            lx, ly = self._divider_points[-1]
            if abs(x - lx) < 0.5 and abs(y - ly) < 0.5:
                return

        self._divider_points.append((x, y))

        # Draw marker dot
        marker = self.ax.plot(
            x, y, 'o', color='magenta', markersize=8,
            markeredgecolor='white', markeredgewidth=1.5, zorder=200)[0]
        self._divider_markers.append(marker)

        # Draw solid segment from previous point
        if len(self._divider_points) >= 2:
            px, py = self._divider_points[-2]
            seg, = self.ax.plot(
                [px, x], [py, y],
                '-', color='magenta', linewidth=2, alpha=0.9, zorder=195)
            self._divider_segments.append(seg)

        # Clear old preview line (becomes solid segment)
        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        n = len(self._divider_points)
        self._update_status(
            f"DIVIDER: {n} pt{'s' if n != 1 else ''} — "
            f"click=add, s=finish, right-click=undo last, Esc=cancel", 'magenta')
        self.fig.canvas.draw_idle()

    def _divider_undo_last_point(self):
        """Remove the last placed divider point (right-click during divider mode)."""
        if not self._divider_points:
            self._update_status("No points to undo — Esc to cancel divider", 'orange')
            return

        self._divider_points.pop()

        if self._divider_markers:
            artist = self._divider_markers.pop()
            if artist is not None:
                try: artist.remove()
                except ValueError: pass

        if self._divider_segments:
            artist = self._divider_segments.pop()
            if artist is not None:
                try: artist.remove()
                except ValueError: pass

        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        n = len(self._divider_points)
        if n == 0:
            self._update_status(
                "DIVIDER: All points removed — click to start again, Esc to cancel", 'magenta')
        else:
            self._update_status(
                f"DIVIDER: Undid last point, {n} remaining", 'magenta')
        self.fig.canvas.draw_idle()

    def _update_divider_preview(self, x: float, y: float):
        """Draw / update dashed preview from last placed point to cursor.

        Also detects nearby vertices from any room and shows a highlight ring
        so the user can snap the next divider point to an existing vertex.
        """
        # --- Vertex snap detection (active regardless of whether first point is placed) ---
        snap_x, snap_y = self._snap_to_vertex(x, y)
        snapped = (snap_x != x or snap_y != y)
        self._divider_snap_pt = (snap_x, snap_y) if snapped else None

        if snapped:
            if self._divider_snap_marker is None:
                self._divider_snap_marker, = self.ax.plot(
                    snap_x, snap_y, 'o',
                    color='none', markersize=14,
                    markeredgecolor='yellow', markeredgewidth=2.5,
                    zorder=300)
            else:
                self._divider_snap_marker.set_data([snap_x], [snap_y])
                self._divider_snap_marker.set_visible(True)
        else:
            if self._divider_snap_marker is not None:
                self._divider_snap_marker.set_visible(False)

        # --- Preview line (only once at least one point is placed) ---
        if not self._divider_points:
            now = time.monotonic()
            if now - self._last_drag_draw >= 0.033:
                self._last_drag_draw = now
                self.fig.canvas.draw_idle()
            return

        lx, ly = self._divider_points[-1]

        # Ortho constraint applied to the (possibly snapped) target
        tx, ty = (snap_x, snap_y) if snapped else (x, y)
        dx, dy = abs(tx - lx), abs(ty - ly)
        if dx >= dy:
            ty = ly
        else:
            tx = lx

        if self._divider_preview_line is not None:
            self._divider_preview_line.set_data([lx, tx], [ly, ty])
        else:
            line, = self.ax.plot(
                [lx, tx], [ly, ty],
                '--', color='magenta', linewidth=2, alpha=0.8, zorder=150)
            self._divider_preview_line = line

        now = time.monotonic()
        if now - self._last_drag_draw >= 0.033:
            self._last_drag_draw = now
            self.fig.canvas.draw_idle()

    # --- ray intersection helper -----------------------------------------------

    @staticmethod
    def _ray_polygon_intersection(origin, direction, polygon_verts):
        """Find the first intersection of a ray with the polygon boundary.

        The ray starts at *origin* and extends in *direction*.
        Returns ``(x, y)`` of the nearest hit, or ``None``.
        """
        ox, oy   = float(origin[0]),    float(origin[1])
        rdx, rdy = float(direction[0]), float(direction[1])
        if abs(rdx) < 1e-12 and abs(rdy) < 1e-12:
            return None

        best_t  = float('inf')
        best_pt = None

        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]),           float(polygon_verts[i][1])
            bx_, by_ = float(polygon_verts[(i + 1) % n][0]), float(polygon_verts[(i + 1) % n][1])
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = rdx * edge_dy - rdy * edge_dx
            if abs(denom) < 1e-10:
                continue
            t = ((ax_ - ox) * edge_dy - (ay_ - oy) * edge_dx) / denom
            u = ((ax_ - ox) * rdy   - (ay_ - oy) * rdx)       / denom
            if t > 1e-6 and -1e-10 <= u <= 1.0 + 1e-10:
                if t < best_t:
                    best_t  = t
                    best_pt = (float(ax_ + u * edge_dx), float(ay_ + u * edge_dy))

        return best_pt

    # --- finalize & split along polyline ---------------------------------------

    def _finalize_division(self):
        """Finalize the multi-segment divider: find boundary intersections
        for the first and last segments, then split the room polygon."""
        if self._divider_room_idx is None:
            return
        if len(self._divider_points) < 2:
            self._update_status("Need at least 2 points to finalize — keep clicking", 'orange')
            return

        room  = self.rooms[self._divider_room_idx]
        verts = room['vertices']
        pts   = list(self._divider_points)

        # --- boundary intersection for FIRST segment (extend pts[0] outward) ---
        boundary_start = self._find_boundary_hit(
            pts[0], pts[1], verts, outward=True)

        # --- boundary intersection for LAST segment (extend pts[-1] outward) ---
        boundary_end = self._find_boundary_hit(
            pts[-1], pts[-2], verts, outward=True)

        if boundary_start is None:
            self._update_status("Could not find boundary intersection for first segment", 'red')
            return
        if boundary_end is None:
            self._update_status("Could not find boundary intersection for last segment", 'red')
            return

        # Build full polyline: boundary_start → user pts → boundary_end
        polyline = list(pts)
        if self._points_close(boundary_start, polyline[0], tol=2.0):
            polyline[0] = boundary_start
        else:
            polyline.insert(0, boundary_start)

        if self._points_close(boundary_end, polyline[-1], tol=2.0):
            polyline[-1] = boundary_end
        else:
            polyline.append(boundary_end)

        polyline = [(float(p[0]), float(p[1])) for p in polyline]
        if len(polyline) < 2:
            self._update_status("Divider polyline too short", 'red')
            return

        poly_a, poly_b = self._split_polygon_by_polyline(verts, polyline)
        if poly_a is None or poly_b is None:
            self._update_status("Division failed — could not split polygon along polyline", 'red')
            return

        self._apply_division_with_polys(self._divider_room_idx, poly_a, poly_b)

    def _find_boundary_hit(self, tip, anchor, polygon_verts, outward=True):
        """Find where the segment anchor→tip (extended) crosses the polygon boundary.

        First checks if *tip* is already on the boundary.  If not, casts a ray
        from *tip* outward (direction tip−anchor) and returns the first hit.
        Falls back to the reverse direction if the outward ray misses.
        """
        # Check if tip is already on boundary
        n = len(polygon_verts)
        for i in range(n):
            a = polygon_verts[i]
            b = polygon_verts[(i + 1) % n]
            if self._point_on_segment(tip, a, b, tol=3.0):
                return (float(tip[0]), float(tip[1]))

        # Ray outward from tip (direction = tip − anchor)
        direction = (float(tip[0]) - float(anchor[0]),
                     float(tip[1]) - float(anchor[1]))
        hit = self._ray_polygon_intersection(tip, direction, polygon_verts)
        if hit is not None:
            return hit

        # Fallback: try inward direction (point may be outside the room)
        inward = (-direction[0], -direction[1])
        hit = self._ray_polygon_intersection(tip, inward, polygon_verts)
        return hit

    def _split_polygon_by_polyline(self, verts, polyline):
        """Split a closed polygon along a multi-segment polyline.

        *polyline[0]* and *polyline[-1]* must lie on the polygon boundary.
        Interior points may be inside the polygon.

        Returns ``(poly_a, poly_b)`` or ``(None, None)`` on failure.
        """
        p_entry  = polyline[0]
        p_exit   = polyline[-1]
        interior = polyline[1:-1]

        n = len(verts)
        augmented = []
        entry_idx = None
        exit_idx  = None

        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1

            if entry_idx is None and self._point_on_segment(p_entry, a, b):
                if self._points_close(p_entry, a):
                    entry_idx = cur
                elif self._points_close(p_entry, b):
                    pass
                else:
                    augmented.append([float(p_entry[0]), float(p_entry[1])])
                    entry_idx = len(augmented) - 1

            if exit_idx is None and self._point_on_segment(p_exit, a, b):
                if self._points_close(p_exit, a):
                    exit_idx = cur
                elif self._points_close(p_exit, b):
                    pass
                else:
                    augmented.append([float(p_exit[0]), float(p_exit[1])])
                    exit_idx = len(augmented) - 1

        # Deferred "close to b" resolution
        if entry_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_entry, v):
                    entry_idx = i
                    break
        if exit_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_exit, v):
                    exit_idx = i
                    break

        if entry_idx is None or exit_idx is None:
            return None, None

        m = len(augmented)

        # Walk boundary entry → exit
        bnd_a = []
        idx = entry_idx
        while True:
            bnd_a.append(augmented[idx])
            if idx == exit_idx and len(bnd_a) > 1:
                break
            idx = (idx + 1) % m

        # Walk boundary exit → entry
        bnd_b = []
        idx = exit_idx
        while True:
            bnd_b.append(augmented[idx])
            if idx == entry_idx and len(bnd_b) > 1:
                break
            idx = (idx + 1) % m

        interior_fwd = [[float(p[0]), float(p[1])] for p in interior]
        interior_rev = list(reversed(interior_fwd))

        # poly_a: boundary entry→exit, then interior reversed (exit→entry)
        poly_a = bnd_a + interior_rev
        # poly_b: boundary exit→entry, then interior forward (entry→exit)
        poly_b = bnd_b + interior_fwd

        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    # --- apply division --------------------------------------------------------

    def _apply_division_with_polys(self, room_idx: int, poly_a: list, poly_b: list):
        """Apply a division using pre-computed sub-polygons.

        The parent room boundary is left completely unchanged. The smaller of
        the two split polygons is created as a new DIV sub-room drawn *within*
        the parent boundary. Deleting the child later therefore leaves the
        parent untouched.
        Exits divider mode **and** edit mode, then saves the session.
        """
        room            = self.rooms[room_idx]
        original_name   = room.get('name', 'unnamed')
        original_parent = room.get('parent')
        hdr_file        = room.get('hdr_file', self.current_hdr_name)

        if original_parent is None:
            division_parent = original_name
        else:
            division_parent = original_parent

        # Snapshot entire rooms list for undo
        rooms_snapshot = [dict(r, vertices=[list(v) for v in r['vertices']]) for r in self.rooms]
        self._edit_undo_stack.append(('divider', rooms_snapshot))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)

        # Use the smaller polygon as the new sub-room drawn inside the parent
        def _shoelace(poly):
            n = len(poly)
            if n < 3:
                return 0.0
            s = 0.0
            for i in range(n):
                x0, y0 = poly[i]
                x1, y1 = poly[(i + 1) % n]
                s += float(x0) * float(y1) - float(x1) * float(y0)
            return abs(s) / 2.0

        small_poly = poly_a if _shoelace(poly_a) <= _shoelace(poly_b) else poly_b

        # Create the smaller piece as a new DIV sub-room
        base_name = f"{division_parent}_DIV"
        div_name = self._make_unique_name(f"{base_name}1")

        div_room = {
            'name':      div_name,
            'parent':    division_parent,
            'vertices':  [[float(x), float(y)] for x, y in small_poly],
            'hdr_file':  hdr_file,
            'room_type': 'CIRC',
        }

        # Parent room vertices are intentionally left unchanged

        # Insert the new DIV sub-room after the original
        self.rooms.insert(room_idx + 1, div_room)

        self.selected_room_idx = None
        self._exit_divider_mode(cancelled=False)

        # Exit edit mode and save (per requirement)
        self.edit_mode = False
        self.btn_edit_mode.label.set_text('Edit Mode: OFF (Press E)')
        self._style_toggle_button(self.btn_edit_mode, False)
        self._reset_hover_state()
        self._create_polygon_selector()

        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)

        self._update_status(
            f"Divided '{original_name}' → '{div_name}' (sub-room inside parent)", 'green')
        print(f"Room divider: '{original_name}' -> parent unchanged, created '{div_name}'")

    # === STATUS & ROOM LIST ====================================================

    def _update_status(self, message: str, color: str = 'blue'):
        """Update the status text in the side panel."""
        if hasattr(self, 'status_text') and self.status_text is not None:
            self.status_text.set_text(f"Status: {message}")
            self.status_text.set_color(color)
            self.status_text.set_visible(True)
            # Clear the name preview so they don't overlap
            if hasattr(self, 'name_preview_text') and self.name_preview_text is not None:
                self.name_preview_text.set_text("")
            self.fig.canvas.draw_idle()

    def _update_room_list(self):
        """Refresh the saved rooms list as a scrollable, click-to-select panel."""
        self.ax_list.clear()
        self.ax_list.set_facecolor('#FAFAF8')
        self.ax_list.set_xlim(0, 1)
        self.ax_list.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
        for spine in self.ax_list.spines.values():
            spine.set_edgecolor('#CCCCCC'); spine.set_linewidth(0.5)
        self._room_list_hit_boxes = []

        hdr_rooms = [(i, r) for i, r in enumerate(self.rooms) if self._is_room_on_current_hdr(r)]
        if not hdr_rooms:
            self.ax_list.set_ylim(0, 1)
            self.ax_list.text(0.05, 0.5, "(no rooms on this HDR)", fontsize=7,
                              style='italic', color='gray', va='center')
            self.fig.canvas.draw_idle()
            return

        flat_items, children_by_parent = self._build_room_tree(hdr_rooms)
        self._room_list_flat_items = flat_items
        self._render_room_list_rows(flat_items, children_by_parent)
        self.fig.canvas.draw_idle()

    def _build_room_tree(self, hdr_rooms):
        """Build a flat parent→children tree for the room list. Returns (items, children_map).
        Each item is (room_idx, indent, is_last_child)."""
        apartments         = [(i, r) for i, r in hdr_rooms if r.get('parent') is None]
        children_by_parent = {}
        for i, room in hdr_rooms:
            parent = room.get('parent')
            if parent is not None:
                children_by_parent.setdefault(parent, []).append((i, room))
        flat_items = []
        for apt_idx, apt in apartments:
            flat_items.append((apt_idx, 0, False))
            kids = children_by_parent.get(apt.get('name', ''), [])
            for ki, (child_idx, _) in enumerate(kids):
                is_last = (ki == len(kids) - 1)
                flat_items.append((child_idx, 1, is_last))
        return flat_items, children_by_parent

    def _render_room_list_rows(self, flat_items, children_by_parent):
        """Render visible rows in two columns with scrollbar into ax_list."""
        # Row heights: parents get 1.0 unit, children get 0.5 unit
        pad_top, pad_bot = 0.01, 0.03
        usable = 1.0 - pad_top - pad_bot
        total = len(flat_items)

        # Compute total weight for scroll capacity (parents=1, children=0.7)
        weights = [0.7 if indent > 0 else 1.0 for (_, indent, _) in flat_items]

        # Determine how many items fit in ONE column
        unit_h = usable / 9.0
        cumulative = 0.0
        col_capacity = 0
        for w in weights:
            cumulative += w
            if cumulative * unit_h > usable:
                break
            col_capacity += 1
        col_capacity = max(1, col_capacity)

        # Two columns = 2x capacity
        visible_count = min(total, col_capacity * 2)

        # Scroll offset
        max_offset = max(0, total - visible_count)
        self.room_list_scroll_offset = max(0, min(self.room_list_scroll_offset, max_offset))
        self.ax_list.set_ylim(0, 1)

        # Column layout constants
        col_w     = 0.48          # each column width in axes fraction
        col_x     = [0.01, 0.51]  # left edges for col 0 and col 1
        divider_x = 0.50          # vertical divider between columns

        # Split visible items into two columns
        visible_slice = flat_items[self.room_list_scroll_offset:self.room_list_scroll_offset + visible_count]
        col0_items = visible_slice[:col_capacity]
        col1_items = visible_slice[col_capacity:]

        # Draw vertical divider between columns (if there are items in col 1)
        if col1_items:
            self.ax_list.plot(
                [divider_x, divider_x], [pad_bot, 1.0 - pad_top],
                color='#DDDDDD', linewidth=0.5,
                transform=self.ax_list.transAxes, clip_on=True)

        # Render each column
        for col_idx, col_items in enumerate([col0_items, col1_items]):
            cx = col_x[col_idx]
            y_cursor = 1.0 - pad_top

            for row_i, (room_idx, indent, is_last_child) in enumerate(col_items):
                room   = self.rooms[room_idx]
                name   = room.get('name', 'unnamed')
                is_sel   = (room_idx == self.selected_room_idx)
                is_multi = (room_idx in self.multi_selected_room_idxs)
                row_h   = unit_h * (0.7 if indent > 0 else 1.0)
                row_top = y_cursor
                row_bot = y_cursor - row_h
                row_mid = (row_top + row_bot) / 2
                y_cursor = row_bot

                box_pad = 0.005 if indent > 0 else 0.01
                box_inset = 0.001 if indent > 0 else 0.002
                if is_sel:
                    self.ax_list.add_patch(FancyBboxPatch(
                        (cx, row_bot + box_inset), col_w, row_h - box_inset * 2,
                        boxstyle=f'round,pad={box_pad}', facecolor='#FFE082', edgecolor='orange',
                        linewidth=1.0, transform=self.ax_list.transAxes, clip_on=True))
                elif is_multi:
                    self.ax_list.add_patch(FancyBboxPatch(
                        (cx, row_bot + box_inset), col_w, row_h - box_inset * 2,
                        boxstyle=f'round,pad={box_pad}', facecolor='#BBDEFB', edgecolor='#42A5F5',
                        linewidth=0.8, transform=self.ax_list.transAxes, clip_on=True))

                rtype    = room.get('room_type', '')
                type_tag = f" : {rtype}" if rtype else ""
                if indent > 0:
                    parent_name = room.get('parent', '')
                    short = name[len(parent_name) + 1:] if name.startswith(f"{parent_name}_") else name
                    text   = f"{short}{type_tag}"
                    color, fs, fw = ('#E65100' if is_sel else '#0D47A1'), 6.5, 'normal'

                    # Draw vertical connector line from this row up to the previous row
                    line_x = cx + 0.02 + 0.02
                    line_bot = row_mid
                    line_top = row_top + row_h * 0.5
                    self.ax_list.plot(
                        [line_x, line_x], [line_bot, line_top],
                        color='#90A4AE', linewidth=0.7, solid_capstyle='round',
                        transform=self.ax_list.transAxes, clip_on=True)
                    # Horizontal tick from vertical line to the label
                    tick_end = line_x + 0.02
                    self.ax_list.plot(
                        [line_x, tick_end], [row_mid, row_mid],
                        color='#90A4AE', linewidth=0.7, solid_capstyle='round',
                        transform=self.ax_list.transAxes, clip_on=True)

                    text_x = tick_end + 0.01
                else:
                    n_kids = len(children_by_parent.get(name, []))
                    suffix = f" ({n_kids})" if n_kids else ""
                    text   = f"{name}{suffix}{type_tag}"
                    color, fs, fw = ('#E65100' if is_sel else '#1B5E20'), 7, 'bold'
                    text_x = cx + 0.02

                self.ax_list.text(text_x, row_mid, text, fontsize=fs,
                                  fontweight=fw, color=color, va='center',
                                  transform=self.ax_list.transAxes, clip_on=True)
                self._room_list_hit_boxes.append((row_bot, row_top, cx, cx + col_w, room_idx))

        # Scrollbar
        if total > visible_count:
            pct = self.room_list_scroll_offset / max(1, max_offset)
            ind_h = visible_count / total
            ind_y = (1.0 - ind_h) * (1.0 - pct)
            self.ax_list.add_patch(FancyBboxPatch(
                (0.965, ind_y), 0.025, ind_h, boxstyle='round,pad=0.005',
                facecolor='#AAAAAA', edgecolor='none',
                transform=self.ax_list.transAxes, clip_on=True))
            self.ax_list.text(
                0.5, 0.01,
                f"\u2191\u2193 scroll  ({self.room_list_scroll_offset + 1}"
                f"-{min(self.room_list_scroll_offset + visible_count, total)} of {total})",
                fontsize=6, color='#888888', ha='center', va='bottom',
                transform=self.ax_list.transAxes)

    def _on_list_click(self, event):
        """Handle clicks on the saved rooms list to select a room.

        Ctrl+click toggles rooms in/out of the multi-selection set,
        allowing bulk room-type tagging.  Shift+click selects all rooms
        between the last clicked room and the current one.  Plain click
        reverts to single-select behaviour.
        """
        if event.inaxes != self.ax_list:
            return
        if event.xdata is None or event.ydata is None:
            return
        x = event.xdata
        y = event.ydata
        ctrl  = event.key == 'control'
        shift = event.key == 'shift'
        for hit in self._room_list_hit_boxes:
            if len(hit) == 5:
                y_min, y_max, x_min, x_max, room_idx = hit
            else:
                y_min, y_max, room_idx = hit
                x_min, x_max = 0, 1
            if y_min <= y <= y_max and x_min <= x <= x_max:
                if shift and self._last_list_click_idx is not None:
                    # Shift+click: select range between last click and this one
                    flat_idxs = [item[0] for item in self._room_list_flat_items]
                    try:
                        a = flat_idxs.index(self._last_list_click_idx)
                        b = flat_idxs.index(room_idx)
                    except ValueError:
                        a, b = 0, len(flat_idxs) - 1
                    lo, hi = min(a, b), max(a, b)
                    self.multi_selected_room_idxs = set(flat_idxs[lo:hi + 1])
                    n = len(self.multi_selected_room_idxs)
                    self._update_status(f"Multi-select: {n} room(s)", 'blue')
                elif ctrl:
                    # Ctrl+click: toggle room in multi-selection
                    if room_idx in self.multi_selected_room_idxs:
                        self.multi_selected_room_idxs.discard(room_idx)
                    else:
                        self.multi_selected_room_idxs.add(room_idx)
                    # Also add current primary selection to multi-set
                    if self.selected_room_idx is not None:
                        self.multi_selected_room_idxs.add(self.selected_room_idx)
                    n = len(self.multi_selected_room_idxs)
                    self._update_status(f"Multi-select: {n} room(s)", 'blue')
                    self._last_list_click_idx = room_idx
                else:
                    # Plain click: clear multi-selection, single-select
                    self.multi_selected_room_idxs.clear()
                    self._last_list_click_idx = room_idx
                    if self.selected_room_idx == room_idx:
                        self._deselect_room()
                    else:
                        self._select_room(room_idx)
                self._update_room_list()
                return

    # === SESSION PERSISTENCE ===================================================

    def _save_session(self):
        """Save all room boundaries, DF stamps, and DF cache to JSON."""
        self.session_path.parent.mkdir(parents=True, exist_ok=True)
        # Convert stamp tuples to lists for JSON serialisation
        stamps_json = {hdr: [list(s) for s in stamps]
                       for hdr, stamps in self._df_stamps.items() if stamps}
        data = {
            'image_dir':      str(self.image_dir),
            'current_hdr_idx':     self.current_hdr_idx,
            'current_variant_idx': self.current_variant_idx,
            'df_thresholds':  self.DF_THRESHOLDS,
            'rooms':          self.rooms,
            'df_stamps':      stamps_json,
            'overlay_pdf_path':    str(self._overlay_pdf_path) if self._overlay_pdf_path else None,
            'overlay_page_idx':    self._overlay_page_idx,
            'aoi_level_idx':       self._aoi_level_idx,
            'overlay_visible':     self._overlay_visible,
            'overlay_transforms':  self._overlay_transforms,
            'overlay_alpha':       self._overlay_alpha,
            'overlay_raster_dpi':  self._overlay_raster_dpi,
            'overlay_cache_pdf':   self._overlay_cache_pdf,
            'overlay_cache_dpi':   self._overlay_cache_dpi,
            'window_settings':     self.window_settings,
        }
        # Write to a temp file alongside the target, then atomically rename it
        # over the real path. This guarantees the session is never left in a
        # partially-written (corrupt) state if the process is interrupted mid-save.
        tmp_path = self.session_path.with_suffix('.json.tmp')
        with open(tmp_path, 'w') as f:
            json.dump(data, f, indent=2)
        os.replace(tmp_path, self.session_path)
        print(f"Session saved to {self.session_path}")
        # Remove stale CSV so the JSON is the single source of truth
        if self.csv_path.exists():
            self.csv_path.unlink()
            print(f"Removed stale CSV: {self.csv_path}")

    def _load_session(self):
        """Load room boundaries, DF stamps, and cached DF results from JSON session or AOI files."""
        if self.session_path.exists():
            with open(self.session_path, 'r') as f:
                data = json.load(f)
            self.rooms = data.get('rooms', [])
            # Restore stamped DF readings (lists back to tuples).
            # Old sessions have 3-element stamps (x, y, df_val); new ones have 5
            # (x, y, df_val, px, py). Both are handled transparently by _draw_df_stamps.
            stamps_raw = data.get('df_stamps', {})
            self._df_stamps = {hdr: [tuple(s) for s in stamps]
                               for hdr, stamps in stamps_raw.items()}
            
            # Restore window settings
            self.window_settings = data.get('window_settings', {})
            
            # Restore HDR and variant selection
            self.current_hdr_idx = data.get('current_hdr_idx', 0)
            self.current_variant_idx = data.get('current_variant_idx', 0)
            if not (0 <= self.current_hdr_idx < len(self.hdr_files)):
                self.current_hdr_idx = 0
            self._rebuild_image_variants()

            # Restore overlay state
            self._overlay_visible = data.get('overlay_visible', False)
            self._overlay_transforms = data.get('overlay_transforms', {})
            self._overlay_alpha = data.get('overlay_alpha', 0.6)
            self._overlay_raster_dpi = data.get('overlay_raster_dpi', 150)
            self._overlay_cache_pdf  = data.get('overlay_cache_pdf')
            self._overlay_cache_dpi  = data.get('overlay_cache_dpi')

            # Fix DPI mismatch: if the cached raster was built at a different DPI
            # than the current raster DPI, compensate all saved scale values so the
            # overlay renders at the correct world-space size after rasterization.
            if (self._overlay_cache_dpi and self._overlay_raster_dpi
                    and self._overlay_cache_dpi != self._overlay_raster_dpi):
                k = self._overlay_cache_dpi / self._overlay_raster_dpi
                for tf in self._overlay_transforms.values():
                    if 'scale_x' in tf:
                        tf['scale_x'] *= k
                    if 'scale_y' in tf:
                        tf['scale_y'] *= k
                self._overlay_cache_dpi = self._overlay_raster_dpi

            pdf_path_str = data.get('overlay_pdf_path')
            if pdf_path_str and Path(pdf_path_str).exists():
                self._overlay_pdf_path = Path(pdf_path_str)
                self._overlay_page_idx = data.get('overlay_page_idx', 0)
                self._overlay_needs_rasterize = True
            self._aoi_level_idx = data.get('aoi_level_idx', 0)
            # If iesve_room_data is set but session has no rooms or rooms lack 'ffl'
            # keys, the session predates IESVE AOI support or was saved before rooms
            # were loaded — (re-)load from IESVE AOI files.
            if (self._iesve_room_data_path is not None
                    and (not self.rooms
                         or not any('ffl' in r for r in self.rooms))):
                print("Session has no IESVE rooms — loading from IESVE AOI files.")
                self.rooms = []
                n = self._load_from_iesve_aoi()
                if n == 0:
                    return
                self._save_session()
            # df_thresholds from old sessions are ignored — now fixed per room type
            source = "session"
            cached = sum(1 for r in self.rooms if r.get('df_cache'))
            if cached:
                print(f"Restored DF cache for {cached}/{len(self.rooms)} rooms")
        elif self.aoi_dir.exists() and list(self.aoi_dir.glob('*.aoi')):
            # Distinguish IESVE .aoi format (world X/Y only, 'AoI Points File :' header)
            # from modern Archilume format (pixel coords included, 'AOI Points File:' header).
            # IESVE files are only processed when iesve_room_data is provided.
            first_aoi = next(self.aoi_dir.glob('*.aoi'), None)
            is_iesve = False
            if first_aoi is not None and self._iesve_room_data_path is not None:
                try:
                    with open(first_aoi, 'r') as _f:
                        is_iesve = _f.readline().startswith('AoI Points File :')
                except Exception:
                    pass
            if is_iesve:
                n = self._load_from_iesve_aoi()
                if n == 0:
                    return
                source = "IESVE AOI files"
                self._save_session()  # persist so conversion never runs again
            else:
                self._load_from_aoi_files(self.aoi_dir)
                source = "AOI files"
        else:
            return

        renamed_count = self._enforce_unique_names()
        if renamed_count > 0:
            print(f"Renamed {renamed_count} rooms to ensure uniqueness")
            self._save_session()

        self._update_status(f"Loaded {len(self.rooms)} rooms from {source}", 'green')
        if hasattr(self, 'ax'):
            self._render_section(force_full=True)
        print(f"Loaded {len(self.rooms)} rooms from {source}")


    # === ARCHIVE ===============================================================

    def _on_close(self, event):
        """Handle editor window close. Capture window state before saving session."""
        if self._closing:
            return
        self._closing = True

        try:
            # Access the window via the specific figure manager to be safe
            if hasattr(self, 'fig') and self.fig.canvas.manager:
                window = self.fig.canvas.manager.window
                
                # Capture current window state
                is_maximized = False
                try:
                    if sys.platform == "win32":
                        is_maximized = window.state() == 'zoomed'
                    else:
                        is_maximized = window.attributes('-zoomed')
                except Exception:
                    # Window might already be partially destroyed
                    pass

                # Get geometry using winfo for more robustness than string parsing
                try:
                    # winfo_x/y includes the window decorations (title bar, borders)
                    # winfo_width/height is the client area (the actual canvas)
                    # We want the values we can pass back to window.geometry()
                    # window.geometry() on Tk returns "WIDTHxHEIGHT+X+Y"
                    # We'll use the current geometry string for the most accurate restore
                    geom_str = window.geometry()
                    # Geometry string is WxH+X+Y or WxH-X-Y
                    match = re.match(r'(\d+)x(\d+)([+-]-?\d+)([+-]-?\d+)', geom_str)
                    if match:
                        w, h, x, y = match.groups()
                        self.window_settings = {
                            'x': int(x.lstrip('+')),
                            'y': int(y.lstrip('+')),
                            'width': int(w),
                            'height': int(h),
                            'maximized': bool(is_maximized)
                        }
                        print(f"Captured window state: {self.window_settings}")
                except Exception as e:
                    print(f"Note: could not capture geometry details: {e}")
        except Exception as exc:
            print(f"Warning: could not capture window state on close: {exc}")

        self._save_session()

    # === EXPORT ================================================================

    def _show_progress(self, fraction: float, label: str = ""):
        """Update the export progress bar (0.0–1.0) and optional label."""
        if not hasattr(self, 'ax_progress'):
            return
        self.ax_progress.clear()
        self.ax_progress.set_xlim(0, 1)
        self.ax_progress.set_ylim(0, 1)
        self.ax_progress.axis('off')

        # Background track
        self.ax_progress.add_patch(FancyBboxPatch(
            (0, 0.1), 1.0, 0.8, boxstyle='round,pad=0.02',
            facecolor='#E0E0E0', edgecolor='#CCCCCC', linewidth=0.5,
        ))
        # Filled portion
        if fraction > 0:
            self.ax_progress.add_patch(FancyBboxPatch(
                (0, 0.1), max(0.01, min(fraction, 1.0)), 0.8,
                boxstyle='round,pad=0.02',
                facecolor='#4CAF50', edgecolor='none',
            ))
        # Label
        display = label or f"{fraction * 100:.0f}%"
        self.ax_progress.text(
            0.5, 0.5, display, fontsize=7, color='#333333',
            ha='center', va='center', fontweight='bold',
        )
        self.ax_progress.set_visible(True)

    def _hide_progress(self):
        """Hide the export progress bar."""
        if hasattr(self, 'ax_progress'):
            self.ax_progress.set_visible(False)

    def _on_export_report(self, event):
        """Export per-pixel illuminance and DF data for every room to Excel.

        The heavy work (HDR loading, rasterisation, Excel writing) runs in a
        background thread so the UI stays responsive.  A matplotlib timer
        polls for completion and updates the progress bar.
        """
        if self._hdr2wpd is None:
            self._update_status("Export unavailable - DF analysis not initialised", 'red')
            return
        if not self.rooms:
            self._update_status("No rooms to export", 'red')
            return
        if getattr(self, '_export_thread', None) and self._export_thread.is_alive():
            self._update_status("Export already in progress...", 'orange')
            return

        n_rooms = len(self.rooms)
        self._update_status(f"Exporting {n_rooms} rooms...", 'orange')
        self.btn_export.label.set_text('Exporting...')
        self._show_progress(0.0, f"0 / {n_rooms} rooms")
        self.fig.canvas.draw_idle()

        progress = {'done': 0, 'total': n_rooms, 'phase': 'pixels', 'error': None}
        self._export_thread = threading.Thread(
            target=self._export_worker, args=(progress,), daemon=True)
        self._export_thread.start()
        self._start_export_progress_poll(progress, n_rooms)

    @staticmethod
    def _compute_summary_for_hdr(hdr_name, hdr_path, rooms_for_hdr, iesve_mode=False):
        """Load one HDR's DF image and compute per-room summary statistics.

        Designed to run in a ThreadPoolExecutor — the heavy work is the
        subprocess call to `pvalue` (I/O-bound, releases the GIL) and numpy
        array operations (also release the GIL).

        Args:
            rooms_for_hdr: list of (parent, name, room_type, verts, threshold_or_None,
                           child_verts_list).
                           child_verts_list is a list of vertex lists for sub-rooms to
                           exclude from this room's pixel set (mirrors the UI behaviour).
                           Pass an empty list for rooms that have no children.
            iesve_mode: When True, use ``pvalue -o`` to undo EXPOSURE header
                        adjustments (e.g. from ``pfilt``).

        Returns:
            (summary_rows, pixel_chunks) where summary_rows is a list of dicts
            (one per room) and pixel_chunks is a list of
            (room_name, lux_array, df_pct_array).
            Child polygon pixels are excluded from parent room results in both outputs,
            so no pixel appears in both a parent room and its sub-rooms.
        """
        from skimage.draw import polygon as skimage_polygon

        if iesve_mode:
            df_img = HdrAoiEditor._load_iesve_df_image(hdr_path)
        else:
            df_img = Hdr2Wpd.load_df_image(hdr_path)
        if df_img is None:
            return [], []

        h, w = df_img.shape
        summary_rows = []
        pixel_chunks = []   # list of (room_name, lux_array, df_pct_array)
        for parent, name, room_type, verts, threshold, child_verts_list in rooms_for_hdr:
            xs = [int(round(v[0])) for v in verts]
            ys = [int(round(v[1])) for v in verts]
            rr, cc = skimage_polygon(ys, xs, shape=(h, w))
            if len(rr) == 0:
                continue

            if child_verts_list:
                # Build a boolean mask for the parent polygon then punch out each child,
                # so child pixels are never counted in the parent's results.
                mask = np.zeros((h, w), dtype=bool)
                mask[rr, cc] = True
                for child_verts in child_verts_list:
                    cxs = [int(round(v[0])) for v in child_verts]
                    cys = [int(round(v[1])) for v in child_verts]
                    crr, ccc = skimage_polygon(cys, cxs, shape=(h, w))
                    mask[crr, ccc] = False
                rr, cc = np.where(mask)

            n = len(rr)
            if n == 0:
                continue

            df_vals  = df_img[rr, cc]   # already DF% (e.g. 0.98 = 0.98%)
            lux_vals = df_vals * 100.0  # DF% → lux  (df% = lux/10000*100 → lux = df%*100)

            if threshold is not None:
                passing     = int((df_vals >= threshold).sum())
                passing_pct = round(passing / n * 100, 1)
            else:
                passing     = None
                passing_pct = None

            summary_rows.append({
                'HDR File':            hdr_name,
                'Parent':              parent,
                'Room':                name,
                'Room Type':           room_type or '',
                'Total Pixels':        n,
                'DF Threshold (%)':    threshold,
                'Pixels >= Threshold': passing,
                '% Area >= Threshold': passing_pct,
            })
            pixel_chunks.append((name, np.round(lux_vals, 2), np.round(df_vals, 4)))

        return summary_rows, pixel_chunks

    def _export_worker(self, progress: dict):
        """Background thread: compute per-room DF summary stats and write outputs.

        Groups rooms by HDR file and uses ThreadPoolExecutor to parallelise
        DF image loading (subprocess I/O) and rasterisation (numpy, GIL-free).

        Outputs written to WPD_DIR:
          - aoi_report_daylight.xlsx  — one summary row per room
          - <room_name>_pixels.csv    — one file per room with per-pixel lux + DF%
        """
        try:
            # Build HDR name → path lookup
            hdr_lookup = {entry['name']: entry['hdr_path'] for entry in self.hdr_files}

            # Build a lookup of room name → child vertex lists for exclusion.
            # Only rooms with children (i.e. LIVING/parent rooms) will have entries.
            children_verts_by_name = {}  # room_name → [child_verts, ...]
            for room in self.rooms:
                p = room.get('parent', '')
                if p and len(room.get('vertices', [])) >= 3:
                    children_verts_by_name.setdefault(p, []).append(room['vertices'])

            # Group rooms by HDR file; carry per-room threshold and child verts for exclusion
            hdr_room_groups = {}  # hdr_name → list of (parent, name, room_type, verts, threshold, child_verts_list)
            for room in self.rooms:
                verts    = room.get('vertices', [])
                hdr_name = room.get('hdr_file', '')
                if len(verts) < 3 or hdr_name not in hdr_lookup:
                    continue
                parent    = room.get('parent', '')
                name      = room.get('name', 'unnamed')
                room_type = self._effective_room_type(room)
                threshold = self._threshold_for_type(room_type)   # None if untyped
                # Child verts to subtract: only applicable to parent rooms
                child_verts_list = children_verts_by_name.get(name, [])
                hdr_room_groups.setdefault(hdr_name, []).append(
                    (parent, name, room_type or '', verts, threshold, child_verts_list))

            # Phase 1: Parallel extraction (pvalue + numpy are GIL-free)
            all_summary_rows = []
            all_pixel_chunks = []   # list of (room_name, lux_array, df_pct_array)
            max_workers = min(len(hdr_room_groups), 4) if hdr_room_groups else 1
            completed_rooms = 0
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                iesve = self._iesve_room_data_path is not None
                futures = {
                    pool.submit(
                        self._compute_summary_for_hdr,
                        hdr_name, hdr_lookup[hdr_name], rooms_list, iesve
                    ): rooms_list
                    for hdr_name, rooms_list in hdr_room_groups.items()
                }
                for future in as_completed(futures):
                    summary_rows, pixel_chunks = future.result()
                    all_summary_rows.extend(summary_rows)
                    all_pixel_chunks.extend(pixel_chunks)
                    completed_rooms += len(futures[future])
                    progress['done'] = completed_rooms

            # Phase 2: Write outputs
            progress['phase'] = 'writing'
            output_dir = config.WPD_DIR
            output_dir.mkdir(parents=True, exist_ok=True)

            # Excel summary (one row per room)
            df = pd.DataFrame(all_summary_rows) if all_summary_rows else pd.DataFrame()
            output_path = output_dir / 'aoi_report_daylight.xlsx'
            df.to_excel(output_path, sheet_name='Room Summary', index=False)
            wb = load_workbook(output_path)
            ws = wb.active
            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4
            # Convert to a structured Excel table with filters
            if ws.max_row > 1:
                last_col = get_column_letter(ws.max_column)
                table_ref = f"A1:{last_col}{ws.max_row}"
                table = Table(displayName="RoomSummary", ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
            wb.save(output_path)
            print(f"Report saved to {output_path}")

            # Per-room CSVs (one file per AOI, all pixel values)
            csv_subdir = output_dir / 'aoi_pixel_data'
            csv_subdir.mkdir(parents=True, exist_ok=True)
            for room_name, lux_vals, df_pct_vals in all_pixel_chunks:
                safe_name = room_name.replace('/', '_').replace('\\', '_')
                csv_path = csv_subdir / f"{safe_name}_pixels.csv"
                with open(csv_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Room', 'Illuminance (Lux)', 'Daylight Factor (%)'])
                    for lux, df_pct in zip(lux_vals, df_pct_vals):
                        writer.writerow([room_name, lux, df_pct])
            print(f"Per-room CSVs saved to {csv_subdir}")

            progress['phase'] = 'overlays'
            self._export_overlay_images(progress)
            self._export_pdf_underlay_images()
            progress['phase'] = 'done'
        except Exception as exc:
            progress['error'] = exc

    def _start_export_progress_poll(self, progress: dict, n_rooms: int):
        """Poll the background export thread and update the progress bar."""
        def _check():
            if self._export_thread.is_alive():
                if progress['phase'] == 'overlays':
                    self._show_progress(0.97, "Rendering overlay images...")
                elif progress['phase'] == 'writing':
                    self._show_progress(0.95, "Writing Excel file...")
                else:
                    frac = progress['done'] / max(progress['total'], 1) * 0.9
                    self._show_progress(frac, f"{progress['done']} / {progress['total']} rooms")
                self.fig.canvas.draw_idle()
                return
            self._export_poll_timer.stop()
            if progress['error']:
                self._update_status(f"Export failed: {progress['error']}", 'red')
                self._show_progress(1.0, "Export failed")
            else:
                self._show_progress(0.98, "Archiving outputs...")
                self.fig.canvas.draw_idle()
                archive_path = self._run_archive()
                if archive_path:
                    self._update_status(f"Export & archive complete → {archive_path.name}", 'green')
                else:
                    self._update_status("Export complete (archive failed)", 'orange')
                self._show_progress(1.0, "Complete")
            self.btn_export.label.set_text('Export & Archive')
            self.fig.canvas.draw_idle()
            hide = self.fig.canvas.new_timer(interval=4000)
            hide.single_shot = True
            hide.add_callback(lambda: (self._hide_progress(), self.fig.canvas.draw_idle()))
            hide.start()

        self._export_poll_timer = self.fig.canvas.new_timer(interval=300)
        self._export_poll_timer.add_callback(_check)
        self._export_poll_timer.start()

    # === ARCHIVE / EXTRACT =====================================================

    def _run_archive(self) -> Optional[Path]:
        """Zip the outputs directory into the archive folder. Returns the zip path or None."""
        from archilume.config import OUTPUTS_DIR, ARCHIVE_DIR
        try:
            ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            project_suffix = f"_{self.project}" if self.project else ""
            zip_name  = f"archilume_export{project_suffix}_{timestamp}"
            zip_path  = ARCHIVE_DIR / f"{zip_name}.zip"
            shutil.make_archive(str(ARCHIVE_DIR / zip_name), 'zip', str(OUTPUTS_DIR.parent), OUTPUTS_DIR.name)
            print(f"Archive created: {zip_path}")
            return zip_path
        except Exception as e:
            print(f"Archive failed: {e}")
            return None

    def _on_extract_click(self, event):
        """Open a file picker to select a zip archive, extract it to outputs, and reload."""
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        initial_dir = str(ARCHIVE_DIR) if ARCHIVE_DIR.exists() else str(Path.cwd())
        zip_path = filedialog.askopenfilename(
            title="Select archive to extract",
            initialdir=initial_dir,
            filetypes=[("Zip files", "*.zip"), ("All files", "*.*")],
        )
        if not zip_path:
            root.destroy()
            return
        confirmed = messagebox.askyesno(
            title="Overwrite outputs?",
            message=(
                f"Extract '{Path(zip_path).name}' and overwrite all files in:\n\n"
                f"{OUTPUTS_DIR}\n\n"
                "This will replace all current output data. Continue?"
            ),
            parent=root,
        )
        root.destroy()
        if not confirmed:
            return
        zip_path = Path(zip_path)
        try:
            # Clear outputs directory; wipe first to avoid stale files
            if OUTPUTS_DIR.exists():
                shutil.rmtree(OUTPUTS_DIR)
            OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)

            # Extract, stripping a leading "outputs/" prefix if present (old archives
            # have no prefix; new archives created after the make_archive fix do).
            with zipfile.ZipFile(zip_path, 'r') as zf:
                for member in zf.infolist():
                    member_path = member.filename
                    if member_path.startswith('outputs/'):
                        member_path = member_path[len('outputs/'):]
                    if not member_path:
                        continue
                    dest = OUTPUTS_DIR / member_path
                    if member.is_dir():
                        dest.mkdir(parents=True, exist_ok=True)
                    else:
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        with zf.open(member) as src, open(dest, 'wb') as dst:
                            shutil.copyfileobj(src, dst)
            print(f"Extracted archive: {zip_path} → {OUTPUTS_DIR}")

            # Full editor reload
            self.rooms = []
            self.hdr_files        = self._scan_hdr_files()
            self.current_hdr_idx  = 0
            self._rebuild_image_variants()
            self._load_session()
            self._reset_hover_state()
            self._update_hdr_list()
            self._update_room_list()
            self._render_section(reset_view=True, force_full=True)
            self._create_polygon_selector()
            self._update_status(f"Extracted and reloaded: {zip_path.name}", 'green')
        except Exception as e:
            self._update_status(f"Extract failed: {e}", 'red')
            print(f"Extract failed: {e}")
        self.fig.canvas.draw_idle()

    @staticmethod
    def _render_single_overlay(tiff_path, rooms_data, output_dir, stamps=None):
        """Render room boundary overlays onto a single TIFF and save.

        Designed to run in a ThreadPoolExecutor (I/O + PIL drawing).

        Args:
            tiff_path: Path to source TIFF image.
            rooms_data: list of (name, vertices, df_display_lines) tuples.
            output_dir: directory to save the overlay TIFF.
            stamps: list of (x, y, df_val[, px, py]) stamped DF readings.
        """
        if not tiff_path.exists():
            return
        img = Image.open(tiff_path).convert('RGB')
        draw = ImageDraw.Draw(img)

        font_size = max(12, int(img.height * 0.012))
        font_size_small = max(10, int(font_size * 0.8))
        font = _load_pil_font(font_size, bold=True)
        font_sm = _load_pil_font(font_size_small, bold=False)

        red = (255, 0, 0)
        black = (0, 0, 0)
        outline_w = max(1, font_size // 12)

        def _outlined_text(x, y, text, fnt):
            for ox in range(-outline_w, outline_w + 1):
                for oy in range(-outline_w, outline_w + 1):
                    if ox or oy:
                        draw.text((x + ox, y + oy), text, fill=black, font=fnt)
            draw.text((x, y), text, fill=red, font=fnt)

        def _dashed_polygon(pts_closed, fill, dash=8, gap=6):
            """Draw a dashed polyline along pts_closed using alternating draw/skip segments."""
            for i in range(len(pts_closed) - 1):
                x0, y0 = pts_closed[i]
                x1, y1 = pts_closed[i + 1]
                seg_len = ((x1 - x0) ** 2 + (y1 - y0) ** 2) ** 0.5
                if seg_len == 0:
                    continue
                dx, dy = (x1 - x0) / seg_len, (y1 - y0) / seg_len
                pos, drawing = 0.0, True
                while pos < seg_len:
                    step = dash if drawing else gap
                    end = min(pos + step, seg_len)
                    if drawing:
                        draw.line(
                            [(int(x0 + dx * pos), int(y0 + dy * pos)),
                             (int(x0 + dx * end), int(y0 + dy * end))],
                            fill=fill, width=1)
                    pos, drawing = end, not drawing

        for name, verts, df_lines, is_circ in rooms_data:
            pts = [(int(round(v[0])), int(round(v[1]))) for v in verts]
            pts.append(pts[0])
            if is_circ:
                _dashed_polygon(pts, fill=red)
            else:
                draw.line(pts, fill=red, width=1)

            if is_circ:
                continue

            centroid = HdrAoiEditor._polygon_label_point(verts)
            cx, cy = int(round(centroid[0])), int(round(centroid[1]))

            bbox = draw.textbbox((0, 0), name, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            _outlined_text(cx - tw // 2, cy - th // 2, name, font)

            y_offset = cy + th
            for line in df_lines:
                bbox_s = draw.textbbox((0, 0), line, font=font_sm)
                tw_s = bbox_s[2] - bbox_s[0]
                th_s = bbox_s[3] - bbox_s[1]
                _outlined_text(cx - tw_s // 2, y_offset, line, font_sm)
                y_offset += th_s + 2

        # Draw stamped DF pixel readings (cyan dot + label), matching on-screen style
        for stamp in (stamps or []):
            sx, sy, df_val = stamp[0], stamp[1], stamp[2]
            px = stamp[3] if len(stamp) > 3 else int(round(sx))
            py = stamp[4] if len(stamp) > 4 else int(round(sy))
            ix, iy = int(round(sx)), int(round(sy))
            r = max(3, font_size // 4)
            draw.ellipse((ix - r, iy - r, ix + r, iy + r), fill=(0, 255, 255))
            label = f"DF: {df_val:.2f}%\npx({px},{py})"
            _outlined_text(ix + r + 2, iy - font_size_small // 2, label, font_sm)

        out_path = output_dir / f"{tiff_path.stem}_aoi_overlay.png"
        if out_path.exists():
            out_path.unlink()
        img.save(out_path)
        print(f"Overlay saved: {out_path}")

    def _export_overlay_images(self, progress: dict):
        """Render room boundary overlays onto each TIFF using ThreadPoolExecutor."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        output_dir = self.image_dir

        # Collect all (tiff_path, rooms_data, stamps) jobs
        jobs = []
        for entry in self.hdr_files:
            hdr_name = entry['name']
            rooms_on_hdr = [
                (r.get('name', ''), r['vertices'],
                 r.get('df_cache', {}).get('display_lines', []),
                 r.get('room_type', '') == 'CIRC')
                for r in self.rooms
                if r.get('hdr_file') == hdr_name
                and len(r.get('vertices', [])) >= 3
            ]
            if not rooms_on_hdr:
                continue
            stamps = self._df_stamps.get(hdr_name, [])
            for tiff_path in entry.get('tiff_paths', []):
                jobs.append((tiff_path, rooms_on_hdr, output_dir, stamps))

        max_workers = min(len(jobs), 4) if jobs else 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(self._render_single_overlay, tiff_path, rooms_data, out_dir, stamps)
                       for tiff_path, rooms_data, out_dir, stamps in jobs]
            for future in as_completed(futures):
                future.result()  # propagate exceptions

    @staticmethod
    def _render_pdf_underlay(
        df_false_path: Path,
        rooms_data: list,
        stamps: list,
        pdf_path,           # Path or None
        pdf_transform: dict,
        session_dpi: int,
        output_dir: Path,
        pdf_opacity: float = 0.6,
    ):
        """Render one level's PDF underlay composite and save.

        Layers (bottom to top):
          1. df_false PNG — fully opaque falsecolour base (HDR pixel dimensions).
          2. PDF floor plan — composited at pdf_opacity (matching the viewer's
             overlay alpha) so the falsecolour remains visible beneath.
          3. AOI room boundaries and labels drawn directly onto the canvas.

        Designed to run in a ThreadPoolExecutor (PIL + numpy, GIL-free I/O).

        Args:
            df_false_path:  Path to the df_false PNG (base canvas, HDR pixel dims).
            rooms_data:     list of (name, vertices, df_display_lines, is_circ).
            stamps:         list of (x, y, df_val[, px, py]) stamped DF readings.
            pdf_path:       Path to PDF, or None if no overlay configured.
            pdf_transform:  dict with scale_x/y, offset_x/y, rotation_90, page_idx.
            session_dpi:    DPI the PDF was rasterized at in the editor session.
            output_dir:     Directory to write the output PNG.
            pdf_opacity:    Opacity of the PDF layer (0–1). Should match the
                            session's _overlay_alpha so the export looks identical
                            to the on-screen viewer. Default 0.6.
        """
        EXPORT_DPI = 300

        base_img = Image.open(df_false_path).convert('RGB')
        hdr_w, hdr_h = base_img.size

        # --- Composite PDF onto df_false base ---
        if pdf_path is not None and pdf_transform and 'scale_x' in pdf_transform:
            page_idx = pdf_transform.get('page_idx', 0)
            sx = pdf_transform.get('scale_x', 1.0)
            sy = pdf_transform.get('scale_y', 1.0)
            ox = pdf_transform.get('offset_x', 0.0)
            oy = pdf_transform.get('offset_y', 0.0)
            rot = pdf_transform.get('rotation_90', 0) % 4

            pdf_rgba = rasterize_pdf_page(pdf_path, page_idx, dpi=session_dpi)
            if rot > 0:
                pdf_rgba = np.rot90(pdf_rgba, k=rot)
            pdf_h_px, pdf_w_px = pdf_rgba.shape[:2]

            pdf_canvas = Image.new('RGBA', (hdr_w, hdr_h), (255, 255, 255, 0))
            pdf_pil = Image.fromarray(pdf_rgba[:, :, :3]).convert('RGBA')

            dest_w = max(1, int(round(pdf_w_px * sx)))
            dest_h = max(1, int(round(pdf_h_px * sy)))
            if dest_w != pdf_w_px or dest_h != pdf_h_px:
                pdf_pil = pdf_pil.resize((dest_w, dest_h), Image.LANCZOS)

            pdf_canvas.paste(pdf_pil, (int(round(ox)), int(round(oy))))

            alpha_val = int(round(pdf_opacity * 255))
            pdf_arr = np.array(pdf_canvas)
            pdf_arr[:, :, 3] = np.where(pdf_arr[:, :, 3] > 0, alpha_val, 0).astype(np.uint8)
            pdf_canvas = Image.fromarray(pdf_arr, 'RGBA')
            composited = Image.alpha_composite(base_img.convert('RGBA'), pdf_canvas).convert('RGB')
        else:
            composited = base_img

        # --- Scale up to 300 DPI ---
        scale_up = EXPORT_DPI / session_dpi
        if abs(scale_up - 1.0) > 0.01:
            new_w = max(1, int(round(hdr_w * scale_up)))
            new_h = max(1, int(round(hdr_h * scale_up)))
            composited = composited.resize((new_w, new_h), Image.LANCZOS)
        else:
            scale_up = 1.0

        draw = ImageDraw.Draw(composited)
        out_h = composited.size[1]

        font_size = max(16, int(out_h * 0.018))
        font_size_small = max(13, int(font_size * 0.8))
        font = _load_pil_font(font_size, bold=True)
        font_sm = _load_pil_font(font_size_small, bold=False)

        red = (255, 0, 0)
        black = (0, 0, 0)
        outline_w = max(1, font_size // 10)
        line_w = max(2, int(round(font_size // 7)))

        def _sv(hdr_x, hdr_y):
            return int(round(hdr_x * scale_up)), int(round(hdr_y * scale_up))

        def _outlined_text(x, y, text, fnt):
            for ddx in range(-outline_w, outline_w + 1):
                for ddy in range(-outline_w, outline_w + 1):
                    if ddx or ddy:
                        draw.text((x + ddx, y + ddy), text, fill=black, font=fnt)
            draw.text((x, y), text, fill=red, font=fnt)

        def _dashed_polygon(pts_closed, fill, dash=8, gap=6):
            for i in range(len(pts_closed) - 1):
                x0, y0 = pts_closed[i]
                x1, y1 = pts_closed[i + 1]
                seg_len = ((x1 - x0) ** 2 + (y1 - y0) ** 2) ** 0.5
                if seg_len == 0:
                    continue
                ddx, ddy = (x1 - x0) / seg_len, (y1 - y0) / seg_len
                pos, drawing = 0.0, True
                while pos < seg_len:
                    step = dash if drawing else gap
                    end = min(pos + step, seg_len)
                    if drawing:
                        draw.line(
                            [(int(x0 + ddx * pos), int(y0 + ddy * pos)),
                             (int(x0 + ddx * end), int(y0 + ddy * end))],
                            fill=fill, width=line_w)
                    pos, drawing = end, not drawing

        for name, verts, df_lines, is_circ in rooms_data:
            pts = [_sv(v[0], v[1]) for v in verts]
            pts.append(pts[0])
            if is_circ:
                _dashed_polygon(pts, fill=red)
            else:
                draw.line(pts, fill=red, width=line_w)

            if is_circ:
                continue

            centroid = HdrAoiEditor._polygon_label_point(verts)
            cx, cy = _sv(centroid[0], centroid[1])

            bbox = draw.textbbox((0, 0), name, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            _outlined_text(cx - tw // 2, cy - th // 2, name, font)

            y_offset = cy + th
            for line in df_lines:
                bbox_s = draw.textbbox((0, 0), line, font=font_sm)
                tw_s = bbox_s[2] - bbox_s[0]
                th_s = bbox_s[3] - bbox_s[1]
                _outlined_text(cx - tw_s // 2, y_offset, line, font_sm)
                y_offset += th_s + 2

        for stamp in (stamps or []):
            sx_stamp, sy_stamp, df_val = stamp[0], stamp[1], stamp[2]
            ix, iy = _sv(sx_stamp, sy_stamp)
            r = max(3, font_size // 4)
            draw.ellipse((ix - r, iy - r, ix + r, iy + r), fill=(0, 255, 255))
            px_lbl = stamp[3] if len(stamp) > 3 else int(round(sx_stamp))
            py_lbl = stamp[4] if len(stamp) > 4 else int(round(sy_stamp))
            _outlined_text(ix + r + 2, iy - font_size_small // 2,
                           f"DF: {df_val:.2f}%\npx({px_lbl},{py_lbl})", font_sm)

        out_path = output_dir / f"{df_false_path.stem.split('_df_false')[0]}_pdf_aoi_overlay.png"
        if out_path.exists():
            out_path.unlink()
        composited.save(out_path, dpi=(EXPORT_DPI, EXPORT_DPI))
        print(f"PDF underlay overlay saved: {out_path}")

    def _export_pdf_underlay_images(self):
        """Dispatch per-level PDF underlay renders concurrently via ThreadPoolExecutor."""
        output_dir = self.image_dir
        session_dpi = self._overlay_raster_dpi
        pdf_opacity = self._overlay_alpha  # match the viewer's current overlay transparency
        pdf_path = self._overlay_pdf_path if (
            self._overlay_pdf_path is not None and self._overlay_pdf_path.exists()
        ) else None

        jobs = []
        for idx, entry in enumerate(self.hdr_files):
            hdr_name = entry['name']

            rooms_on_hdr = [
                (r.get('name', ''), r['vertices'],
                 r.get('df_cache', {}).get('display_lines', []),
                 r.get('room_type', '') == 'CIRC')
                for r in self.rooms
                if r.get('hdr_file') == hdr_name
                and len(r.get('vertices', [])) >= 3
            ]
            if not rooms_on_hdr:
                continue

            df_false_path = next(
                (p for p in entry.get('tiff_paths', []) if '_df_false' in p.stem),
                None
            )
            if df_false_path is None or not df_false_path.exists():
                print(f"PDF underlay export: no df_false image found for {hdr_name}, skipping")
                continue

            tf = self._get_effective_overlay_transform(hdr_idx=idx)
            # Pass a copy so the thread holds no reference into mutable editor state
            pdf_transform = dict(tf) if (tf and 'scale_x' in tf) else {}
            stamps = list(self._df_stamps.get(hdr_name, []))

            jobs.append((df_false_path, rooms_on_hdr, stamps, pdf_path,
                         pdf_transform, session_dpi, output_dir, pdf_opacity))

        max_workers = min(len(jobs), 4) if jobs else 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [
                pool.submit(self._render_pdf_underlay, *job)
                for job in jobs
            ]
            for future in as_completed(futures):
                future.result()  # propagate exceptions

    def _export_rooms_as_aoi_files(self) -> int:
        """Write editor rooms as .aoi files compatible with Hdr2Wpd.

        Each .aoi file follows the standard format with headers that
        Hdr2Wpd._group_aoi_by_view() and _process_daylight_view_groups()
        expect. Pixel→world conversion uses the coordinate map header.

        Returns:
            Number of .aoi files written.
        """
        # Parse pixel→world mapping parameters from the coordinate map header
        coord_map_path = self.aoi_dir / "pixel_to_world_coordinate_map.txt"
        vp_x = vp_y = vh = vv = 0.0
        img_w = img_h = 1
        if coord_map_path.exists():
            with open(coord_map_path, 'r') as f:
                header_lines = [f.readline() for _ in range(4)]
            # Line 0: # VIEW: VIEW= -vtl v -vp X Y Z ... -vh VH -vv VV
            view_line = header_lines[0]
            vp_match = re.search(r'-vp\s+([\d.-]+)\s+([\d.-]+)\s+([\d.-]+)', view_line)
            if vp_match:
                vp_x, vp_y = float(vp_match.group(1)), float(vp_match.group(2))
            vh_match = re.search(r'-vh\s+([\d.-]+)', view_line)
            vv_match = re.search(r'-vv\s+([\d.-]+)', view_line)
            if vh_match:
                vh = float(vh_match.group(1))
            if vv_match:
                vv = float(vv_match.group(1))
            # Line 1: # Image dimensions in pixels: width=W, height=H
            dim_line = header_lines[1]
            img_w = int(dim_line.split('width=')[1].split(',')[0])
            img_h = int(dim_line.split('height=')[1])

        self.aoi_dir.mkdir(parents=True, exist_ok=True)
        count = 0

        for room in self.rooms:
            name     = room.get('name', 'unnamed')
            hdr_name = room.get('hdr_file', '')
            verts    = room['vertices']
            if len(verts) < 3:
                continue

            # Derive associated view file from HDR name (e.g. model_plan_ffl_25300 → plan_ffl_25300.vp)
            ffl_match = re.search(r'plan_ffl_(\d+)', hdr_name)
            associated_vp = f"plan_ffl_{ffl_match.group(1)}.vp" if ffl_match else f"{hdr_name}.vp"

            # Compute centroid in pixel space (true centre of mass)
            centroid_pt = self._polygon_label_point(verts)
            cx, cy = centroid_pt[0], centroid_pt[1]

            # Build vertex lines: world_x world_y pixel_x pixel_y
            vertex_lines = []
            for px, py in verts:
                if vh > 0 and vv > 0:
                    world_x = vp_x + (px - img_w / 2) * (vh / img_w)
                    world_y = vp_y + (img_h / 2 - py) * (vv / img_h)
                else:
                    world_x, world_y = px, py
                vertex_lines.append(f"{world_x:.4f} {world_y:.4f} {int(round(px))} {int(round(py))}")

            content = "\n".join([
                f"AOI Points File: {name}",
                f"ASSOCIATED VIEW FILE: {associated_vp}",
                f"FFL z height(m): 0.0",
                f"CENTRAL x,y: {cx:.4f} {cy:.4f}",
                f"NO. PERIMETER POINTS {len(verts)}: x,y pixel_x pixel_y positions",
            ] + vertex_lines)

            clean_name = re.sub(r'[^\w\s-]', '', name).strip()
            clean_name = re.sub(r'[-\s]+', '_', clean_name)
            filepath = self.aoi_dir / f"{clean_name}.aoi"
            with open(filepath, 'w') as f:
                f.write(content)
            count += 1

        print(f"Wrote {count} .aoi files to {self.aoi_dir}")
        return count

    def export_room_boundaries_csv(self, output_path: Optional[Path] = None):
        """Export room boundaries as CSV.

        Format: name, parent, hdr_file, X_px Y_px, X_px Y_px, ...
        """
        if not self.rooms:
            return

        output_path = Path(output_path) if output_path else self.csv_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        rows = []
        for room in self.rooms:
            coord_strings = [f"X_{x:.3f} Y_{y:.3f}" for x, y in room['vertices']]
            parent   = room.get('parent') or ''
            hdr_file = room.get('hdr_file', '')
            rows.append([room['name'], parent, hdr_file] + coord_strings)

        max_cols = max(len(r) for r in rows)
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            for row in rows:
                row += [''] * (max_cols - len(row))
                writer.writerow(row)

        self._update_status(f"Exported CSV ({len(self.rooms)} rooms)", 'green')
        print(f"Exported room boundaries CSV to {output_path}")
