# archilume/apps/dash_editor.py
#
# Dash-based HDR AOI Room Boundary Editor.
# Replaces the matplotlib-based HdrAoiEditor with a web UI.
#
# Run:  python -m archilume.apps.dash_editor
#       http://127.0.0.1:8050/
#
# Or:   python -m archilume.apps.dash_editor --project 527DP

from __future__ import annotations

import base64
import csv
import datetime
import io
import json
import logging
import os
import re
import shutil
import subprocess
import threading
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Optional, Tuple, Union

import archilume.utils as archilume_utils

import dash
import dash_bootstrap_components as dbc
import imageio.v2 as imageio
import numpy as np
import plotly.graph_objects as go
from dash import ALL, MATCH, ctx, dcc, html, Input, Output, State, no_update
from dash_iconify import DashIconify
from PIL import Image

from archilume import config
from archilume.apps.project_config import (
    get_last_project,
    list_projects,
    load_project_toml,
    save_project_toml,
    set_last_project,
)

log = logging.getLogger("archilume.dash_editor")

# ═══════════════════════════════════════════════════════════════════════════════
# EditorState — server-side singleton holding all room data & business logic
# ═══════════════════════════════════════════════════════════════════════════════


class EditorState:
    """Pure-data editor state. No UI/rendering imports (matplotlib-free)."""

    def __init__(
        self,
        project: Optional[str] = None,
        image_dir: Optional[Union[Path, str]] = None,
        pdf_path: Optional[Union[Path, str]] = None,
        iesve_room_data: Optional[Union[Path, str]] = None,
    ):
        if not project:
            last = get_last_project()
            available = list_projects()
            if last and last in available:
                project = last
            elif len(available) == 1:
                project = available[0]

        self.project = project
        self._iesve_room_data_path: Optional[Path] = None

        _placeholder = config.PROJECTS_DIR / "_no_project"
        self.project_input_dir: Path = _placeholder
        self.project_aoi_dir: Path = _placeholder / "aoi"
        self.archive_dir: Path = _placeholder / "archive"
        self.wpd_dir: Path = _placeholder / "wpd"
        self.image_dir: Path = _placeholder / "image"
        self._overlay_pdf_path: Optional[Path] = None

        if project:
            self._init_project_paths(project, image_dir, pdf_path, iesve_room_data)
        else:
            self._set_blank_state()

        if project:
            toml_cfg = load_project_toml(project)
            mode = toml_cfg.get("project", {}).get("mode", "archilume")
            if mode in ("archilume", "hdr"):
                self.aoi_dir = config.get_project_paths(project).aoi_dir
            else:
                self.aoi_dir = self.project_aoi_dir
        else:
            self.aoi_dir = self.project_aoi_dir

        self.session_path = self.project_aoi_dir / "aoi_session.json"
        self.csv_path = self.project_aoi_dir / "aoi_boundaries.csv"

        self.hdr_files: List[dict] = self._scan_hdr_files()
        self.current_hdr_idx: int = 0

        self.image_variants: List[Path] = []
        self.current_variant_idx: int = 0
        self._rebuild_image_variants()

        # Room storage
        self.rooms: List[dict] = []
        self.current_polygon_vertices: list = []
        self.selected_room_idx: Optional[int] = None

        # Mode flags
        self.draw_mode: bool = False
        self.edit_mode: bool = False
        self.divider_mode: bool = False
        self.ortho_mode: bool = True
        self.placement_mode: bool = False

        # Divider state
        self._divider_room_idx: Optional[int] = None
        self._divider_points: list = []

        # Edit-mode vertex selection (two-click move)
        self._edit_selected_vertex: Optional[tuple] = None  # (room_idx, vertex_idx)

        # Edge snap arrays (built by _rebuild_snap_arrays)
        self._snap_edge_starts: np.ndarray = np.array([])
        self._snap_edge_ends: np.ndarray = np.array([])

        # DF% placement
        self._df_image: Optional[np.ndarray] = None   # (H, W) float array of DF%
        self._df_image_cache: dict = {}               # hdr_path_str → np.ndarray
        self._df_stamps: dict = {}                    # hdr_name → [(x, y, df_val), ...]

        # Snap
        self._snap_distance_px: float = 10.0
        self.current_vertices: np.ndarray = np.array([])

        # Undo stacks
        self._edit_undo_stack: list = []
        self._edit_undo_max: int = 50
        self._draw_undo_stack: list = []
        self._draw_undo_max: int = 50

        # Image cache
        self._image_cache: dict = {}
        self._image_cache_lock = threading.Lock()
        self._image_cache_limit: int = 15

        # Parent
        self.selected_parent: Optional[str] = None
        self.parent_options: List[str] = []

        # Tree
        self._tree_collapsed: set = set()

        # DF stamps
        self._df_stamps: dict = {}
        self.DF_THRESHOLDS = {"BED": 0.5, "LIVING": 1.0, "NON-RESI": 2.0}
        self._annotation_scale: float = 1.0

        # IESVE
        self._aoi_level_idx: int = 0
        self._aoi_level_map: dict = {}

        # Overlay
        self._overlay_visible: bool = False
        self._overlay_alpha: float = 0.6
        self._overlay_raster_dpi: int = 150
        self._overlay_transforms: dict = {}
        self._overlay_page_idx: int = 0
        self._overlay_cache_pdf: Optional[str] = None
        self._overlay_cache_dpi: Optional[int] = None
        # In-memory base64 cache keyed by (pdf_path_str, page_idx, dpi)
        self._overlay_b64_cache: dict = {}
        # Overlay align mode (user is adjusting position/scale)
        self.overlay_align_mode: bool = False

        # Multi-select
        self.multi_selected_room_idxs: set = set()

        # Status message
        self.status_message: str = ""
        self.status_color: str = "blue"

        # Loading flag
        self._loading: bool = True

        # Double-d timing
        self._last_d_press_time: float = 0.0

        # Window settings (unused in Dash but kept for session compat)
        self.window_settings: dict = {}

        # Load session
        self._load_session()
        self._loading = False
        self._rebuild_snap_arrays()

    # === PROJECT PATHS ========================================================

    def _init_project_paths(self, project, image_dir=None, pdf_path=None, iesve_room_data=None):
        paths = config.get_project_paths(project)
        self.project_input_dir = paths.inputs_dir
        self.project_aoi_dir = paths.aoi_inputs_dir
        self.archive_dir = paths.archive_dir
        self.wpd_dir = paths.wpd_dir
        paths.create_dirs()

        toml_cfg = load_project_toml(project)
        toml_paths = toml_cfg.get("paths", {})
        if pdf_path is None and toml_paths.get("pdf_path"):
            pdf_path = paths.inputs_dir / toml_paths["pdf_path"]
        if iesve_room_data is None and toml_paths.get("iesve_room_data"):
            iesve_room_data = paths.inputs_dir / toml_paths["iesve_room_data"]
        project_mode = toml_cfg.get("project", {}).get("mode", "hdr")
        if image_dir is None and toml_paths.get("image_dir"):
            p = Path(toml_paths["image_dir"])
            image_dir = p if p.is_absolute() else paths.project_dir / p

        if project_mode == "iesve":
            self.image_dir = paths.pic_dir
        elif image_dir is None:
            self.image_dir = paths.image_dir
        else:
            image_dir = Path(image_dir)
            self.image_dir = (
                self.project_input_dir / image_dir
                if not image_dir.is_absolute()
                else image_dir
            )

        if pdf_path is not None:
            pdf_path = Path(pdf_path)
            pdf_path = (
                self.project_input_dir / pdf_path
                if not pdf_path.is_absolute()
                else pdf_path
            )
            if not pdf_path.exists():
                pdf_path = None
        self._overlay_pdf_path = pdf_path

        if iesve_room_data is not None:
            iesve_room_data = Path(iesve_room_data)
            iesve_room_data = (
                self.project_input_dir / iesve_room_data
                if not iesve_room_data.is_absolute()
                else iesve_room_data
            )
            if not iesve_room_data.exists():
                iesve_room_data = None
        self._iesve_room_data_path = iesve_room_data

    def _set_blank_state(self):
        pass  # placeholder paths already set in __init__

    # === HDR SCANNING =========================================================

    def _scan_hdr_files(self) -> List[dict]:
        if not self.image_dir.exists():
            return []
        hdr_paths = sorted(
            [*self.image_dir.glob("*.hdr"), *self.image_dir.glob("*.pic")]
        )
        result = []
        for hdr_path in hdr_paths:
            stem = hdr_path.stem
            tiff_paths = sorted(
                p
                for p in self.image_dir.glob("*.png")
                if p.stem.startswith(stem + "_")
                and not p.stem.endswith("_aoi_overlay")
            )
            result.append(
                {
                    "hdr_path": hdr_path,
                    "tiff_paths": tiff_paths,
                    "name": stem,
                    "suffix": hdr_path.suffix,
                }
            )
        self.legend_map: dict = {}
        for legend_path in sorted(self.image_dir.glob("*_legend.png")):
            key = legend_path.stem[: -len("_legend")]
            self.legend_map[key] = legend_path
        return result

    def _rebuild_image_variants(self):
        if not self.hdr_files:
            self.image_variants = []
            self.current_variant_idx = 0
            return
        active_suffix = None
        if self.image_variants and 0 < self.current_variant_idx < len(self.image_variants):
            old_path = self.image_variants[self.current_variant_idx]
            if old_path.suffix.lower() != ".hdr":
                old_hdr_stem = self.hdr_files[self.current_hdr_idx]["name"]
                active_suffix = old_path.stem[len(old_hdr_stem):]
        entry = self.hdr_files[self.current_hdr_idx]
        self.image_variants = [entry["hdr_path"]] + list(entry["tiff_paths"])
        if active_suffix:
            new_hdr_stem = entry["name"]
            for i, path in enumerate(self.image_variants):
                if (
                    path.suffix.lower() != ".hdr"
                    and path.stem[len(new_hdr_stem):] == active_suffix
                ):
                    self.current_variant_idx = i
                    return
        self.current_variant_idx = 0

    @property
    def current_hdr_name(self) -> str:
        if not self.hdr_files:
            return ""
        return self.hdr_files[self.current_hdr_idx]["name"]

    @property
    def current_variant_path(self) -> Optional[Path]:
        if not self.image_variants:
            return None
        idx = self.current_variant_idx % len(self.image_variants)
        return self.image_variants[idx]

    # === IMAGE LOADING ========================================================

    def _load_image(self, path: Path) -> Optional[np.ndarray]:
        key = str(path)
        with self._image_cache_lock:
            if key in self._image_cache:
                return self._image_cache[key]
        try:
            if path.suffix.lower() in (".hdr", ".pic"):
                img = imageio.imread(str(path)).astype(np.float32)
                if img.ndim == 2:
                    img = np.stack([img, img, img], axis=-1)
                p99 = np.percentile(img, 99)
                if p99 > 0:
                    img = img / p99
                img = np.clip(img ** (1.0 / 2.2), 0.0, 1.0)
            else:
                pil_img = Image.open(path).convert("RGB")
                img = np.array(pil_img, dtype=np.float32) / 255.0
            with self._image_cache_lock:
                if len(self._image_cache) >= self._image_cache_limit:
                    first_key = next(iter(self._image_cache))
                    self._image_cache.pop(first_key)
                self._image_cache[key] = img
            return img
        except Exception as exc:
            print(f"Warning: could not load image {path}: {exc}")
            return None

    def get_image_base64(self) -> Optional[str]:
        """Load current image variant and return as base64-encoded PNG."""
        path = self.current_variant_path
        if path is None:
            return None
        img = self._load_image(path)
        if img is None:
            return None
        img_uint8 = (np.clip(img, 0, 1) * 255).astype(np.uint8)
        pil_img = Image.fromarray(img_uint8)
        buf = io.BytesIO()
        pil_img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode("ascii")
        return f"data:image/png;base64,{b64}"

    def get_image_dimensions(self) -> Tuple[int, int]:
        """Return (width, height) of current image."""
        path = self.current_variant_path
        if path is None:
            return 1, 1
        img = self._load_image(path)
        if img is None:
            return 1, 1
        return img.shape[1], img.shape[0]

    # === AOI LOADING ==========================================================

    def _load_from_aoi_files(self, aoi_dir: Path):
        aoi_files = sorted(aoi_dir.glob("*.aoi"))
        for aoi_path in aoi_files:
            with open(aoi_path, "r") as f:
                lines = [l.strip() for l in f.readlines()]
            if len(lines) < 6:
                continue
            name_match = re.match(r"AOI Points File:\s*(.+)", lines[0])
            name = name_match.group(1).strip() if name_match else aoi_path.stem
            vp_match = re.search(r"plan_ffl_(\d+)", lines[1])
            hdr_file = self.current_hdr_name
            if vp_match:
                ffl_val = vp_match.group(1)
                for entry in self.hdr_files:
                    if ffl_val in entry["name"]:
                        hdr_file = entry["name"]
                        break
            vertices = []
            for line in lines[5:]:
                parts = line.split()
                if len(parts) >= 4:
                    px, py = float(parts[2]), float(parts[3])
                    vertices.append([px, py])
            if len(vertices) >= 3:
                self.rooms.append(
                    {
                        "name": name,
                        "parent": None,
                        "vertices": vertices,
                        "hdr_file": hdr_file,
                    }
                )

    @staticmethod
    def _read_view_params(pic_path: Path):
        try:
            with open(pic_path, "r", encoding="utf-8", errors="ignore") as f:
                view_line = next(
                    (l.strip() for l in f if l.startswith("VIEW=")), None
                )
            if not view_line:
                return None
            vp = re.search(r"-vp\s+([\d.-]+)\s+([\d.-]+)", view_line)
            vh = re.search(r"-vh\s+([\d.-]+)", view_line)
            vv = re.search(r"-vv\s+([\d.-]+)", view_line)
            if not (vp and vh and vv):
                return None
            vp_x, vp_y = float(vp.group(1)), float(vp.group(2))
            vh_val = float(vh.group(1))
            vv_val = float(vv.group(1))
            img = imageio.imread(str(pic_path))
            img_h, img_w = img.shape[:2]
            return vp_x, vp_y, vh_val, vv_val, img_w, img_h
        except Exception as exc:
            print(f"Warning: could not read VIEW params from {pic_path}: {exc}")
            return None

    @staticmethod
    def _world_to_pixels(world_verts, vp_x, vp_y, vh_val, vv_val, img_w, img_h):
        pixels = []
        for wx, wy in world_verts:
            px = (wx - vp_x) / (vh_val / img_w) + img_w / 2
            py = img_h / 2 - (wy - vp_y) / (vv_val / img_h)
            pixels.append([px, py])
        return pixels

    def _load_from_iesve_aoi(self) -> int:
        csv_path = self._iesve_room_data_path
        if not self.hdr_files:
            return 0
        ffl_lookup: dict = {}
        if csv_path is not None and csv_path.exists():
            try:
                import pandas as pd

                try:
                    df_csv = pd.read_csv(csv_path, encoding="utf-8")
                except (UnicodeDecodeError, pd.errors.ParserError, pd.errors.EmptyDataError):
                    try:
                        df_csv = pd.read_excel(csv_path)
                    except Exception:
                        df_csv = pd.read_csv(csv_path, encoding="cp1252")
                id_col = "Space ID"
                ffl_col = "Min. Height (m) (Real)"
                if id_col in df_csv.columns and ffl_col in df_csv.columns:
                    ffl_lookup = dict(zip(df_csv[id_col].astype(str), df_csv[ffl_col]))
            except Exception as exc:
                print(f"Warning: could not read iesve_room_data: {exc}")

        ffl_mm_to_entry: dict = {}
        for entry in self.hdr_files:
            m = re.search(r"plan_ffl_(\d+)", entry["name"])
            if m:
                ffl_mm_to_entry[int(m.group(1))] = entry

        view_cache: dict = {}
        for ffl_mm, entry in ffl_mm_to_entry.items():
            vp = self._read_view_params(entry["hdr_path"])
            if vp is not None:
                view_cache[ffl_mm] = (entry, vp)

        use_ffl_filter = bool(ffl_mm_to_entry)
        if not use_ffl_filter:
            first_entry = self.hdr_files[0]
            fallback_vp = self._read_view_params(first_entry["hdr_path"])
            if fallback_vp is None:
                return 0

        aoi_files = sorted(self.aoi_dir.glob("*.aoi"))
        count = 0
        for aoi_path in aoi_files:
            with open(aoi_path, "r") as f:
                lines = [l.strip() for l in f.readlines()]
            if len(lines) < 4:
                continue
            zone_match = re.match(r"ZONE\s+(\S+)\s+(.*)", lines[1])
            if not zone_match:
                continue
            space_id = zone_match.group(1)
            room_name = zone_match.group(2).strip()
            ffl = ffl_lookup.get(space_id, 0.0)

            if use_ffl_filter:
                ffl_mm = int(round(ffl * 1000))
                cached = view_cache.get(ffl_mm)
                if cached is None:
                    continue
                entry, (vp_x, vp_y, vh_val, vv_val, img_w, img_h) = cached
            else:
                entry = first_entry
                vp_x, vp_y, vh_val, vv_val, img_w, img_h = fallback_vp

            world_verts = []
            for line in lines[3:]:
                parts = line.split()
                if len(parts) >= 2:
                    try:
                        world_verts.append([float(parts[0]), float(parts[1])])
                    except ValueError:
                        continue
            if len(world_verts) < 3:
                continue
            pixels = self._world_to_pixels(
                world_verts, vp_x, vp_y, vh_val, vv_val, img_w, img_h
            )
            self.rooms.append(
                {
                    "name": f"{space_id} {room_name}",
                    "parent": None,
                    "vertices": pixels,
                    "world_vertices": world_verts,
                    "ffl": ffl,
                    "hdr_file": entry["name"],
                }
            )
            count += 1
        return count

    # === ROOM HELPERS =========================================================

    @staticmethod
    def _room_matches_hdr(room: dict, hdr_name: str) -> bool:
        room_hdr = room.get("hdr_file", "")
        if not room_hdr:
            return True
        return room_hdr == hdr_name

    def _is_room_on_current_hdr(self, room: dict) -> bool:
        if not self.hdr_files:
            return False
        hdr_name = self.current_hdr_name
        if "ffl" in room and hdr_name in self._aoi_level_map:
            return room["ffl"] == self._aoi_level_map[hdr_name]
        return self._room_matches_hdr(room, hdr_name)

    def _get_apartments_for_hdr(self, hdr_name: str) -> List[str]:
        return [
            room["name"]
            for room in self.rooms
            if room.get("parent") is None and self._room_matches_hdr(room, hdr_name)
        ]

    def _get_children(self, parent_name: str) -> List[dict]:
        return [r for r in self.rooms if r.get("parent") == parent_name]

    def _get_parent_room(self, parent_name: str) -> Optional[dict]:
        for r in self.rooms:
            if r.get("name") == parent_name and r.get("parent") is None:
                return r
        return None

    def _make_unique_name(self, base_name: str, exclude_idx: Optional[int] = None) -> str:
        existing = {r["name"] for i, r in enumerate(self.rooms) if i != exclude_idx}
        if base_name not in existing:
            return base_name
        match = re.match(r"^(.*?)(\d+)$", base_name)
        root = match.group(1) if match else base_name
        counter = 1
        while f"{root}{counter}" in existing:
            counter += 1
        return f"{root}{counter}"

    def _enforce_unique_names(self) -> int:
        seen = set()
        renamed = 0
        for room in self.rooms:
            name = room["name"]
            if name not in seen:
                seen.add(name)
                continue
            match = re.match(r"^(.*?)(\d+)$", name)
            root = match.group(1) if match else name
            counter = 1
            while f"{root}{counter}" in seen:
                counter += 1
            new_name = f"{root}{counter}"
            room["name"] = new_name
            seen.add(new_name)
            renamed += 1
        return renamed

    def _check_boundary_containment(self, child_verts, parent_verts) -> bool:
        if not parent_verts or len(parent_verts) < 3:
            return True
        if not child_verts or len(child_verts) < 3:
            return True
        from matplotlib.path import Path as MplPath

        arr = np.array(parent_verts)
        if not np.allclose(arr[0], arr[-1]):
            arr = np.vstack([arr, arr[0]])
        path = MplPath(arr)
        return all(path.contains_point(v) for v in child_verts)

    def _update_parent_options(self):
        self.parent_options = self._get_apartments_for_hdr(self.current_hdr_name)

    # === SNAP =================================================================

    def _rebuild_snap_arrays(self):
        """Build flat vertex and edge arrays for snap detection."""
        all_verts = []
        edge_starts = []
        edge_ends = []
        for room in self.rooms:
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room["vertices"]
            all_verts.extend(verts)
            n = len(verts)
            for i in range(n):
                edge_starts.append(verts[i])
                edge_ends.append(verts[(i + 1) % n])
        self.current_vertices = np.array(all_verts) if all_verts else np.array([]).reshape(0, 2)
        self._snap_edge_starts = np.array(edge_starts) if edge_starts else np.array([]).reshape(0, 2)
        self._snap_edge_ends = np.array(edge_ends) if edge_ends else np.array([]).reshape(0, 2)

    def _snap_to_vertex(self, x: float, y: float) -> tuple:
        if len(self.current_vertices) == 0:
            return x, y
        dists = np.hypot(
            self.current_vertices[:, 0] - x, self.current_vertices[:, 1] - y
        )
        min_idx = int(np.argmin(dists))
        if dists[min_idx] <= self._snap_distance_px:
            return (
                float(self.current_vertices[min_idx, 0]),
                float(self.current_vertices[min_idx, 1]),
            )
        return x, y

    def _snap_to_edge(self, x: float, y: float) -> tuple:
        """Snap (x, y) to the nearest point on any existing polygon edge.

        Returns the snapped (x, y) if within snap distance, otherwise the
        original (x, y). Only runs when there are edges to snap to.
        """
        if len(self._snap_edge_starts) == 0:
            return x, y
        p = np.array([x, y])
        a = self._snap_edge_starts          # (N, 2)
        b = self._snap_edge_ends            # (N, 2)
        ab = b - a                          # edge vectors
        ap = p - a                          # vectors from edge starts to p
        # Parameter t in [0, 1] along each edge
        ab_len2 = np.einsum("ij,ij->i", ab, ab)
        # Avoid division by zero for degenerate edges
        valid = ab_len2 > 1e-10
        t = np.zeros(len(a))
        t[valid] = np.clip(
            np.einsum("ij,ij->i", ap[valid], ab[valid]) / ab_len2[valid], 0.0, 1.0
        )
        closest = a + t[:, None] * ab       # (N, 2) closest points on edges
        dists = np.hypot(closest[:, 0] - x, closest[:, 1] - y)
        min_idx = int(np.argmin(dists))
        if dists[min_idx] <= self._snap_distance_px:
            return float(closest[min_idx, 0]), float(closest[min_idx, 1])
        return x, y

    @staticmethod
    def _snap_to_pixel(x: float, y: float) -> tuple:
        return int(x) + 0.5, int(y) + 0.5

    # === GEOMETRY =============================================================

    @staticmethod
    def _points_close(a, b, tol=1.0) -> bool:
        return abs(a[0] - b[0]) < tol and abs(a[1] - b[1]) < tol

    @staticmethod
    def _point_on_segment(pt, a, b, tol=2.0) -> bool:
        px, py = float(pt[0]), float(pt[1])
        ax_, ay_ = float(a[0]), float(a[1])
        bx_, by_ = float(b[0]), float(b[1])
        cross = abs((bx_ - ax_) * (py - ay_) - (by_ - ay_) * (px - ax_))
        seg_len = ((bx_ - ax_) ** 2 + (by_ - ay_) ** 2) ** 0.5
        if seg_len > 0 and cross / seg_len > tol:
            return False
        min_x, max_x = min(ax_, bx_) - tol, max(ax_, bx_) + tol
        min_y, max_y = min(ay_, by_) - tol, max(ay_, by_) + tol
        return min_x <= px <= max_x and min_y <= py <= max_y

    @staticmethod
    def _line_polygon_intersections(line_start, line_end, polygon_verts):
        intersections = []
        sx, sy = float(line_start[0]), float(line_start[1])
        ex, ey = float(line_end[0]), float(line_end[1])
        dx, dy = ex - sx, ey - sy
        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]), float(polygon_verts[i][1])
            bx_, by_ = (
                float(polygon_verts[(i + 1) % n][0]),
                float(polygon_verts[(i + 1) % n][1]),
            )
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = dx * edge_dy - dy * edge_dx
            if abs(denom) < 1e-10:
                continue
            u = ((ax_ - sx) * dy - (ay_ - sy) * dx) / denom
            if -1e-10 <= u <= 1.0 + 1e-10:
                ix = ax_ + u * edge_dx
                iy = ay_ + u * edge_dy
                intersections.append((float(ix), float(iy)))
        if len(intersections) > 1:
            unique = [intersections[0]]
            for pt in intersections[1:]:
                if not any(
                    abs(pt[0] - up[0]) < 1.0 and abs(pt[1] - up[1]) < 1.0
                    for up in unique
                ):
                    unique.append(pt)
            intersections = unique
        intersections.sort(key=lambda p: (p[0] - sx) ** 2 + (p[1] - sy) ** 2)
        return intersections

    @staticmethod
    def _ray_polygon_intersection(origin, direction, polygon_verts):
        ox, oy = float(origin[0]), float(origin[1])
        rdx, rdy = float(direction[0]), float(direction[1])
        if abs(rdx) < 1e-12 and abs(rdy) < 1e-12:
            return None
        best_t = float("inf")
        best_pt = None
        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]), float(polygon_verts[i][1])
            bx_, by_ = (
                float(polygon_verts[(i + 1) % n][0]),
                float(polygon_verts[(i + 1) % n][1]),
            )
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = rdx * edge_dy - rdy * edge_dx
            if abs(denom) < 1e-10:
                continue
            t = ((ax_ - ox) * edge_dy - (ay_ - oy) * edge_dx) / denom
            u = ((ax_ - ox) * rdy - (ay_ - oy) * rdx) / denom
            if t > 1e-6 and -1e-10 <= u <= 1.0 + 1e-10:
                if t < best_t:
                    best_t = t
                    u_clamped = max(0.0, min(1.0, u))
                    if u_clamped < 1e-6:
                        best_pt = (ax_, ay_)
                    elif u_clamped > 1.0 - 1e-6:
                        best_pt = (bx_, by_)
                    else:
                        best_pt = (
                            float(ax_ + u_clamped * edge_dx),
                            float(ay_ + u_clamped * edge_dy),
                        )
        return best_pt

    def _split_polygon_by_line(self, verts, p1, p2):
        n = len(verts)
        augmented = []
        p1_idx = None
        p2_idx = None
        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1
            if p1_idx is None and self._point_on_segment(p1, a, b):
                if self._points_close(p1, a):
                    p1_idx = cur
                elif self._points_close(p1, b):
                    pass
                else:
                    augmented.append([float(p1[0]), float(p1[1])])
                    p1_idx = len(augmented) - 1
            if p2_idx is None and self._point_on_segment(p2, a, b):
                if self._points_close(p2, a):
                    p2_idx = cur
                elif self._points_close(p2, b):
                    pass
                else:
                    augmented.append([float(p2[0]), float(p2[1])])
                    p2_idx = len(augmented) - 1
        if p1_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p1, v):
                    p1_idx = i
                    break
        if p2_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p2, v):
                    p2_idx = i
                    break
        if p1_idx is None or p2_idx is None:
            return None, None
        m = len(augmented)
        poly_a = []
        idx = p1_idx
        while True:
            poly_a.append(augmented[idx])
            if idx == p2_idx and len(poly_a) > 1:
                break
            idx = (idx + 1) % m
        poly_b = []
        idx = p2_idx
        while True:
            poly_b.append(augmented[idx])
            if idx == p1_idx and len(poly_b) > 1:
                break
            idx = (idx + 1) % m
        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    def _split_polygon_by_polyline(self, verts, polyline):
        p_entry = polyline[0]
        p_exit = polyline[-1]
        interior = polyline[1:-1]
        n = len(verts)
        augmented = []
        entry_idx = None
        exit_idx = None
        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1
            if entry_idx is None and self._point_on_segment(p_entry, a, b):
                if self._points_close(p_entry, a):
                    entry_idx = cur
                elif self._points_close(p_entry, b):
                    pass
                else:
                    augmented.append([float(p_entry[0]), float(p_entry[1])])
                    entry_idx = len(augmented) - 1
            if exit_idx is None and self._point_on_segment(p_exit, a, b):
                if self._points_close(p_exit, a):
                    exit_idx = cur
                elif self._points_close(p_exit, b):
                    pass
                else:
                    augmented.append([float(p_exit[0]), float(p_exit[1])])
                    exit_idx = len(augmented) - 1
        if entry_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_entry, v):
                    entry_idx = i
                    break
        if exit_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_exit, v):
                    exit_idx = i
                    break
        if entry_idx is None or exit_idx is None:
            return None, None
        m = len(augmented)
        bnd_a = []
        idx = entry_idx
        while True:
            bnd_a.append(augmented[idx])
            if idx == exit_idx and len(bnd_a) > 1:
                break
            idx = (idx + 1) % m
        bnd_b = []
        idx = exit_idx
        while True:
            bnd_b.append(augmented[idx])
            if idx == entry_idx and len(bnd_b) > 1:
                break
            idx = (idx + 1) % m
        interior_fwd = [[float(p[0]), float(p[1])] for p in interior]
        interior_rev = list(reversed(interior_fwd))
        poly_a = bnd_a + interior_rev
        poly_b = bnd_b + interior_fwd
        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    def _find_boundary_hit(self, tip, anchor, polygon_verts, outward=True):
        n = len(polygon_verts)
        for i in range(n):
            a = polygon_verts[i]
            b = polygon_verts[(i + 1) % n]
            if self._point_on_segment(tip, a, b, tol=3.0):
                ax_, ay_ = float(a[0]), float(a[1])
                bx_, by_ = float(b[0]), float(b[1])
                edge_dx, edge_dy = bx_ - ax_, by_ - ay_
                edge_len_sq = edge_dx**2 + edge_dy**2
                if edge_len_sq < 1e-20:
                    return (ax_, ay_)
                u = (
                    (float(tip[0]) - ax_) * edge_dx + (float(tip[1]) - ay_) * edge_dy
                ) / edge_len_sq
                u = max(0.0, min(1.0, u))
                if u < 1e-6:
                    return (ax_, ay_)
                elif u > 1.0 - 1e-6:
                    return (bx_, by_)
                return (ax_ + u * edge_dx, ay_ + u * edge_dy)
        direction = (
            float(tip[0]) - float(anchor[0]),
            float(tip[1]) - float(anchor[1]),
        )
        hit = self._ray_polygon_intersection(tip, direction, polygon_verts)
        if hit is not None:
            return hit
        inward = (-direction[0], -direction[1])
        return self._ray_polygon_intersection(tip, inward, polygon_verts)

    # === DIVIDER ==============================================================

    def add_divider_point(self, x: float, y: float) -> str:
        """Add a divider point. Returns status message."""
        if self._divider_room_idx is None:
            return "No room selected for division"
        x, y = self._snap_to_pixel(x, y)
        x, y = self._snap_to_vertex(x, y)
        if self._divider_points:
            lx, ly = self._divider_points[-1]
            dx, dy = abs(x - lx), abs(y - ly)
            if dx >= dy:
                y = ly
            else:
                x = lx
            if abs(x - lx) < 0.5 and abs(y - ly) < 0.5:
                return "Point too close to previous"
        self._divider_points.append((x, y))
        n = len(self._divider_points)
        return f"DIVIDER: {n} pt{'s' if n != 1 else ''} placed"

    def undo_divider_point(self) -> str:
        if not self._divider_points:
            return "No points to undo"
        self._divider_points.pop()
        n = len(self._divider_points)
        return f"DIVIDER: Undid last point, {n} remaining" if n else "All points removed"

    def finalize_division(self) -> str:
        """Finalize division. Returns status message."""
        if self._divider_room_idx is None:
            return "No room selected"
        if len(self._divider_points) < 2:
            return "Need at least 2 points"
        room = self.rooms[self._divider_room_idx]
        verts = room["vertices"]
        pts = list(self._divider_points)
        boundary_start = self._find_boundary_hit(pts[0], pts[1], verts, outward=True)
        boundary_end = self._find_boundary_hit(pts[-1], pts[-2], verts, outward=True)
        if boundary_start is None:
            return "Could not find boundary intersection for first segment"
        if boundary_end is None:
            return "Could not find boundary intersection for last segment"
        polyline = list(pts)
        if self._points_close(boundary_start, polyline[0], tol=2.0):
            polyline[0] = boundary_start
        else:
            polyline.insert(0, boundary_start)
        if self._points_close(boundary_end, polyline[-1], tol=2.0):
            polyline[-1] = boundary_end
        else:
            polyline.append(boundary_end)
        polyline = [(float(p[0]), float(p[1])) for p in polyline]
        if len(polyline) < 2:
            return "Divider polyline too short"
        poly_a, poly_b = self._split_polygon_by_polyline(verts, polyline)
        if poly_a is None or poly_b is None:
            return "Division failed — could not split polygon"
        return self._apply_division_with_polys(self._divider_room_idx, poly_a, poly_b)

    def _apply_division_with_polys(self, room_idx: int, poly_a: list, poly_b: list) -> str:
        room = self.rooms[room_idx]
        original_name = room.get("name", "unnamed")
        original_parent = room.get("parent")
        hdr_file = room.get("hdr_file", self.current_hdr_name)
        ffl = room.get("ffl")
        division_parent = original_name if original_parent is None else original_parent
        rooms_snapshot = [
            dict(r, vertices=[list(v) for v in r["vertices"]]) for r in self.rooms
        ]
        self._edit_undo_stack.append(("divider", rooms_snapshot))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)

        def _shoelace(poly):
            n = len(poly)
            if n < 3:
                return 0.0
            s = 0.0
            for i in range(n):
                x0, y0 = poly[i]
                x1, y1 = poly[(i + 1) % n]
                s += float(x0) * float(y1) - float(x1) * float(y0)
            return abs(s) / 2.0

        small_poly = poly_a if _shoelace(poly_a) <= _shoelace(poly_b) else poly_b
        base_name = f"{division_parent}_DIV"
        div_name = self._make_unique_name(f"{base_name}1")
        div_room = {
            "name": div_name,
            "parent": division_parent,
            "vertices": [[float(x), float(y)] for x, y in small_poly],
            "hdr_file": hdr_file,
            "room_type": "CIRC",
        }
        if ffl is not None:
            div_room["ffl"] = ffl
        self.rooms.insert(room_idx + 1, div_room)
        self.selected_room_idx = None
        self.divider_mode = False
        self._divider_room_idx = None
        self._divider_points = []
        self.edit_mode = False
        self._save_session()
        self._rebuild_snap_arrays()
        return f"Divided '{original_name}' → '{div_name}'"

    # === UNDO =================================================================

    def undo_edit(self) -> str:
        if not self._edit_undo_stack:
            return "Nothing to undo"
        entry = self._edit_undo_stack.pop()
        if entry[0] == "divider":
            self.rooms = entry[1]
            self._rebuild_snap_arrays()
            self._save_session()
            return "Undid division"
        room_idx, old_verts = entry
        if 0 <= room_idx < len(self.rooms):
            self.rooms[room_idx]["vertices"] = old_verts
            self._rebuild_snap_arrays()
            self._save_session()
            return "Undid vertex edit"
        return "Undo failed"

    def move_vertex(self, room_idx: int, vertex_idx: int, x: float, y: float) -> str:
        """Move a single vertex of a room polygon, pushing the old state to the undo stack."""
        if not (0 <= room_idx < len(self.rooms)):
            return "Invalid room index"
        verts = self.rooms[room_idx]["vertices"]
        if not (0 <= vertex_idx < len(verts)):
            return "Invalid vertex index"
        # Push current state for undo
        self._edit_undo_stack.append((room_idx, [list(v) for v in verts]))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)
        verts[vertex_idx] = [x, y]
        self._edit_selected_vertex = None
        self._rebuild_snap_arrays()
        self._save_session()
        return f"Moved vertex {vertex_idx} of room '{self.rooms[room_idx]['name']}'"

    def delete_vertex(self, room_idx: int, vertex_idx: int) -> str:
        """Delete a vertex from a room polygon (minimum 3 vertices enforced)."""
        if not (0 <= room_idx < len(self.rooms)):
            return "Invalid room index"
        verts = self.rooms[room_idx]["vertices"]
        if len(verts) <= 3:
            return "Cannot delete: polygon must have at least 3 vertices"
        self._edit_undo_stack.append((room_idx, [list(v) for v in verts]))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)
        verts.pop(vertex_idx)
        self._edit_selected_vertex = None
        self._rebuild_snap_arrays()
        self._save_session()
        return f"Deleted vertex {vertex_idx} of room '{self.rooms[room_idx]['name']}'"

    # ── PDF overlay helpers ───────────────────────────────────────────────────

    def get_overlay_page_count(self) -> int:
        """Return the number of pages in the overlay PDF, or 0 if none loaded."""
        if self._overlay_pdf_path is None or not self._overlay_pdf_path.exists():
            return 0
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(str(self._overlay_pdf_path))
            n = len(doc)
            doc.close()
            return n
        except Exception:
            return 0

    def get_overlay_base64(self, page_idx: Optional[int] = None, dpi: Optional[int] = None) -> Optional[str]:
        """Rasterize a PDF page to a base64 PNG string, with in-memory caching.

        Returns a data URI string suitable for use as a Plotly layout image
        source, or None if no PDF is available.
        """
        if self._overlay_pdf_path is None or not self._overlay_pdf_path.exists():
            return None
        if page_idx is None:
            page_idx = self._overlay_page_idx
        if dpi is None:
            dpi = self._overlay_raster_dpi

        cache_key = (str(self._overlay_pdf_path), page_idx, dpi)
        if cache_key in self._overlay_b64_cache:
            return self._overlay_b64_cache[cache_key]

        try:
            from archilume.utils import rasterize_pdf_page, make_lines_only
            import base64
            import io as _io
            from PIL import Image as _Image

            rgba = rasterize_pdf_page(self._overlay_pdf_path, page_idx, dpi=dpi)
            # Make white pixels transparent so only linework shows
            rgba = make_lines_only(rgba)
            pil_img = _Image.fromarray(rgba, "RGBA")
            buf = _io.BytesIO()
            pil_img.save(buf, format="PNG")
            b64 = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode()
            self._overlay_b64_cache[cache_key] = b64
            return b64
        except Exception as exc:
            log.warning("Overlay rasterize failed: %s", exc)
            return None

    def get_overlay_transform(self) -> dict:
        """Return the per-HDR overlay transform for the current HDR, with defaults."""
        hdr_name = self.current_hdr_name or ""
        return self._overlay_transforms.get(hdr_name, {
            "offset_x": 0.0,
            "offset_y": 0.0,
            "scale_x": 1.0,
            "scale_y": 1.0,
        })

    def set_overlay_transform(self, offset_x: float, offset_y: float,
                               scale_x: float, scale_y: float) -> None:
        """Persist the per-HDR overlay transform for the current HDR."""
        hdr_name = self.current_hdr_name or ""
        self._overlay_transforms[hdr_name] = {
            "offset_x": offset_x,
            "offset_y": offset_y,
            "scale_x": scale_x,
            "scale_y": scale_y,
        }
        self._save_session()

    def nudge_overlay(self, direction: str, step: float = 2.0) -> None:
        """Nudge the overlay position by *step* pixels in the given direction.

        direction is one of: 'left', 'right', 'up', 'down'.
        step defaults to 2 px; callers may increase this for acceleration.
        """
        tf = self.get_overlay_transform()
        ox = tf.get("offset_x", 0.0)
        oy = tf.get("offset_y", 0.0)
        sx = tf.get("scale_x", 1.0)
        sy = tf.get("scale_y", 1.0)
        if direction == "left":
            ox -= step
        elif direction == "right":
            ox += step
        elif direction == "up":
            oy -= step
        elif direction == "down":
            oy += step
        self.set_overlay_transform(ox, oy, sx, sy)

    def rotate_overlay_90(self) -> None:
        """Swap the overlay's scale_x and scale_y to simulate a 90° rotation.

        Also clears the base64 cache since the rasterised image orientation
        changes.
        """
        tf = self.get_overlay_transform()
        sx = tf.get("scale_x", 1.0)
        sy = tf.get("scale_y", 1.0)
        ox = tf.get("offset_x", 0.0)
        oy = tf.get("offset_y", 0.0)
        self.set_overlay_transform(ox, oy, sy, sx)  # swap sx/sy
        self._overlay_b64_cache.clear()

    def undo_draw(self) -> str:
        if not self._draw_undo_stack:
            return "Nothing to undo"
        entry = self._draw_undo_stack.pop()
        tag = entry[0]
        if tag == "delete":
            idx, room_dict = entry[1], entry[2]
            self.rooms.insert(idx, room_dict)
            self._rebuild_snap_arrays()
            self._save_session()
            return f"Restored deleted room '{room_dict['name']}'"
        elif tag == "create":
            idx = entry[1]
            if 0 <= idx < len(self.rooms):
                name = self.rooms[idx]["name"]
                self.rooms.pop(idx)
                self._rebuild_snap_arrays()
                self._save_session()
                return f"Undid creation of '{name}'"
        elif tag == "type":
            for idx, old_type in entry[1]:
                if 0 <= idx < len(self.rooms):
                    self.rooms[idx]["room_type"] = old_type
            self._save_session()
            return "Undid type change"
        elif tag == "rename":
            idx, old_name = entry[1], entry[2]
            if 0 <= idx < len(self.rooms):
                self.rooms[idx]["name"] = old_name
                self._save_session()
                return f"Undid rename (restored '{old_name}')"
        return "Undo failed"

    # === POLYGON LABEL ========================================================

    @staticmethod
    def _polygon_label_point(verts) -> list:
        pts = np.array(verts, dtype=float)
        n = len(pts)
        if n < 3:
            return pts.mean(axis=0).tolist()
        x, y = pts[:, 0], pts[:, 1]
        x1, y1 = np.roll(x, -1), np.roll(y, -1)
        cross = x * y1 - x1 * y
        signed_area = cross.sum() / 2.0
        if abs(signed_area) < 1e-10:
            return pts.mean(axis=0).tolist()
        cx = ((x + x1) * cross).sum() / (6.0 * signed_area)
        cy = ((y + y1) * cross).sum() / (6.0 * signed_area)
        centroid = [cx, cy]
        # Check containment using ray-casting
        inside = False
        px, py = cx, cy
        for i in range(n):
            j = (i + 1) % n
            yi, yj = pts[i, 1], pts[j, 1]
            xi, xj = pts[i, 0], pts[j, 0]
            if ((yi > py) != (yj > py)) and (px < (xj - xi) * (py - yi) / (yj - yi) + xi):
                inside = not inside
        if inside:
            return centroid
        return pts.mean(axis=0).tolist()

    # === SESSION ==============================================================

    def _save_session(self):
        if getattr(self, "_loading", False):
            return
        self.session_path.parent.mkdir(parents=True, exist_ok=True)
        stamps_json = {
            hdr: [list(s) for s in stamps]
            for hdr, stamps in self._df_stamps.items()
            if stamps
        }
        data = {
            "current_hdr_idx": self.current_hdr_idx,
            "current_variant_idx": self.current_variant_idx,
            "df_thresholds": self.DF_THRESHOLDS,
            "rooms": self.rooms,
            "df_stamps": stamps_json,
            "overlay_pdf_path": (
                str(self._overlay_pdf_path.relative_to(self.project_input_dir))
                if self._overlay_pdf_path
                and self._overlay_pdf_path.is_relative_to(self.project_input_dir)
                else (str(self._overlay_pdf_path) if self._overlay_pdf_path else None)
            ),
            "overlay_page_idx": self._overlay_page_idx,
            "aoi_level_idx": self._aoi_level_idx,
            "aoi_level_map": self._aoi_level_map,
            "overlay_visible": self._overlay_visible,
            "overlay_align_mode": self.overlay_align_mode,
            "overlay_transforms": self._overlay_transforms,
            "overlay_alpha": self._overlay_alpha,
            "overlay_raster_dpi": self._overlay_raster_dpi,
            "overlay_cache_pdf": self._overlay_cache_pdf,
            "overlay_cache_dpi": self._overlay_cache_dpi,
            "annotation_scale": self._annotation_scale,
            "window_settings": self.window_settings,
        }
        tmp_path = self.session_path.with_suffix(".json.tmp")
        with open(tmp_path, "w") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp_path, self.session_path)
        if self.csv_path.exists():
            self.csv_path.unlink()

    def _load_session(self):
        if self.session_path.exists():
            with open(self.session_path, "r") as f:
                data = json.load(f)
            if "rooms" in data:
                self.rooms = data["rooms"]
            stamps_raw = data.get("df_stamps", {})
            self._df_stamps = {
                hdr: [tuple(s) for s in stamps]
                for hdr, stamps in stamps_raw.items()
            }
            self.current_hdr_idx = data.get("current_hdr_idx", 0)
            self.current_variant_idx = data.get("current_variant_idx", 0)
            if not (0 <= self.current_hdr_idx < len(self.hdr_files)):
                self.current_hdr_idx = 0
            self._rebuild_image_variants()
            self._overlay_visible = data.get("overlay_visible", False)
            self.overlay_align_mode = data.get("overlay_align_mode", False)
            self._overlay_transforms = data.get("overlay_transforms", {})
            self._overlay_alpha = data.get("overlay_alpha", 0.6)
            self._overlay_raster_dpi = data.get("overlay_raster_dpi", 150)
            self._overlay_cache_pdf = data.get("overlay_cache_pdf")
            self._overlay_cache_dpi = data.get("overlay_cache_dpi")
            self._annotation_scale = data.get("annotation_scale", 1.0)
            self._aoi_level_idx = data.get("aoi_level_idx", 0)
            self._aoi_level_map = data.get("aoi_level_map", {})
            self.window_settings = data.get("window_settings", {})
            # Handle IESVE reload if needed
            if self._iesve_room_data_path is not None and (
                not self.rooms or not any("ffl" in r for r in self.rooms)
            ):
                self.rooms = []
                n = self._load_from_iesve_aoi()
                if n > 0:
                    self._aoi_level_idx = 0
                    self._reassign_aoi_level()
                    self._save_session()
        elif self.aoi_dir.exists() and list(self.aoi_dir.glob("*.aoi")):
            first_aoi = next(self.aoi_dir.glob("*.aoi"), None)
            is_iesve = False
            if first_aoi is not None and self._iesve_room_data_path is not None:
                try:
                    with open(first_aoi, "r") as _f:
                        is_iesve = _f.readline().startswith("AoI Points File :")
                except Exception:
                    pass
            if is_iesve:
                n = self._load_from_iesve_aoi()
                if n > 0:
                    self._aoi_level_idx = 0
                    self._reassign_aoi_level()
                    self._save_session()
            else:
                self._load_from_aoi_files(self.aoi_dir)
        self._enforce_unique_names()
        self._update_parent_options()

    def _aoi_ffl_groups(self) -> list:
        return sorted({r["ffl"] for r in self.rooms if "ffl" in r})

    def _reassign_aoi_level(self):
        ffl_groups = self._aoi_ffl_groups()
        if not ffl_groups or not self.hdr_files:
            return
        target_ffl = ffl_groups[self._aoi_level_idx]
        entry = self.hdr_files[self.current_hdr_idx]
        hdr_name = entry["name"]
        view_params = self._read_view_params(entry["hdr_path"])
        if view_params is None:
            return
        vp_x, vp_y, vh_val, vv_val, img_w, img_h = view_params
        self._aoi_level_map[hdr_name] = target_ffl
        for room in self.rooms:
            if "ffl" not in room or "world_vertices" not in room:
                continue
            if room["ffl"] == target_ffl:
                room["vertices"] = self._world_to_pixels(
                    room["world_vertices"], vp_x, vp_y, vh_val, vv_val, img_w, img_h
                )
                room["hdr_file"] = hdr_name
                room.pop("df_cache", None)

    # === TREE =================================================================

    def build_layer_tree(self) -> list:
        """Build flat tree item list for rendering."""
        collapsed = self._tree_collapsed
        flat = []
        for hdr_i, entry in enumerate(self.hdr_files):
            hdr_name = entry["name"]
            is_current = hdr_i == self.current_hdr_idx
            hdr_key = ("hdr", hdr_name)
            flat.append(
                {
                    "type": "hdr",
                    "indent": 0,
                    "label": f"{hdr_name}{entry.get('suffix', '.hdr')}",
                    "data": {"hdr_idx": hdr_i},
                    "node_key": str(hdr_key),
                    "has_children": True,
                    "is_current": is_current,
                }
            )
            if not is_current or hdr_key in collapsed:
                continue
            # Children
            hdr_rooms = [
                (i, r)
                for i, r in enumerate(self.rooms)
                if self._is_room_on_current_hdr(r)
            ]
            apartments = [(i, r) for i, r in hdr_rooms if r.get("parent") is None]
            children_by_parent = {}
            for i, room in hdr_rooms:
                parent = room.get("parent")
                if parent is not None:
                    children_by_parent.setdefault(parent, []).append((i, room))

            room_group_key = ("room_group", hdr_name)
            flat.append(
                {
                    "type": "room_group",
                    "indent": 1,
                    "label": f"Room Boundaries ({len(hdr_rooms)})",
                    "data": {},
                    "node_key": str(room_group_key),
                    "has_children": bool(apartments),
                }
            )
            if room_group_key in collapsed:
                continue
            for apt_idx, apt in apartments:
                apt_name = apt.get("name", "")
                kids = children_by_parent.get(apt_name, [])
                rtype = apt.get("room_type", "")
                type_tag = f" : {rtype}" if rtype else ""
                n_kids = len(kids)
                suffix = f" ({n_kids})" if n_kids else ""
                apt_key = ("room_parent", apt_name)
                flat.append(
                    {
                        "type": "room_parent",
                        "indent": 2,
                        "label": f"{apt_name}{suffix}{type_tag}",
                        "data": {"room_idx": apt_idx},
                        "node_key": str(apt_key),
                        "has_children": bool(kids),
                    }
                )
                if apt_key in collapsed:
                    continue
                for child_idx, child in kids:
                    child_name = child.get("name", "")
                    parent_name = child.get("parent", "")
                    short = (
                        child_name[len(parent_name) + 1 :]
                        if child_name.startswith(f"{parent_name}_")
                        else child_name
                    )
                    crtype = child.get("room_type", "")
                    ctag = f" : {crtype}" if crtype else ""
                    flat.append(
                        {
                            "type": "room_child",
                            "indent": 3,
                            "label": f"{short}{ctag}",
                            "data": {"room_idx": child_idx},
                            "node_key": None,
                            "has_children": False,
                        }
                    )
        return flat

    # === POINT IN POLYGON =====================================================

    def find_parent_at(self, x: float, y: float) -> Optional[str]:
        """Return the name of the parent-level room containing (x, y), or None."""
        for room in self.rooms:
            if not self._is_room_on_current_hdr(room):
                continue
            if room.get("parent") is not None:
                continue  # skip child rooms
            verts = room["vertices"]
            if len(verts) >= 3 and self._point_in_polygon(x, y, verts):
                return room["name"]
        return None

    def find_room_at(self, x: float, y: float) -> Optional[int]:
        """Return index of room containing point (x, y) on current HDR, or None."""
        # Check smaller (child) rooms first for better hit detection
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            if room.get("parent") is not None:
                verts = room["vertices"]
                if len(verts) >= 3 and self._point_in_polygon(x, y, verts):
                    return i
        # Then check parents
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            if room.get("parent") is None:
                verts = room["vertices"]
                if len(verts) >= 3 and self._point_in_polygon(x, y, verts):
                    return i
        return None

    @staticmethod
    def _point_in_polygon(px: float, py: float, verts) -> bool:
        """Ray-casting point-in-polygon test."""
        n = len(verts)
        inside = False
        for i in range(n):
            j = (i + 1) % n
            yi, yj = float(verts[i][1]), float(verts[j][1])
            xi, xj = float(verts[i][0]), float(verts[j][0])
            if ((yi > py) != (yj > py)) and (
                px < (xj - xi) * (py - yi) / (yj - yi) + xi
            ):
                inside = not inside
        return inside

    # === SAVE / DELETE ROOM ===================================================

    def save_room(self, name: str, room_type: Optional[str] = None) -> str:
        """Save the current polygon as a new room. Returns status message."""
        if len(self.current_polygon_vertices) < 3:
            return "Need at least 3 vertices"
        verts = [[float(x), float(y)] for x, y in self.current_polygon_vertices]
        # Auto-detect parent
        first_x, first_y = verts[0]
        parent = None
        for room in self.rooms:
            if not self._is_room_on_current_hdr(room):
                continue
            if room.get("parent") is not None:
                continue
            if self._point_in_polygon(first_x, first_y, room["vertices"]):
                parent = room["name"]
                break
        if parent and self.selected_parent is None:
            self.selected_parent = parent
        if self.selected_parent:
            if not name.startswith(self.selected_parent + "_"):
                name = f"{self.selected_parent}_{name}"
        name = self._make_unique_name(name)
        new_room = {
            "name": name,
            "parent": self.selected_parent,
            "vertices": verts,
            "hdr_file": self.current_hdr_name,
            "room_type": room_type,
        }
        self.rooms.append(new_room)
        idx = len(self.rooms) - 1
        self._draw_undo_stack.append(("create", idx))
        if len(self._draw_undo_stack) > self._draw_undo_max:
            self._draw_undo_stack.pop(0)
        self.current_polygon_vertices = []
        self._rebuild_snap_arrays()
        self._save_session()
        return f"Saved room '{name}'"

    def delete_room(self, idx: int) -> str:
        if 0 <= idx < len(self.rooms):
            room = self.rooms.pop(idx)
            self._draw_undo_stack.append(("delete", idx, room))
            if len(self._draw_undo_stack) > self._draw_undo_max:
                self._draw_undo_stack.pop(0)
            if self.selected_room_idx == idx:
                self.selected_room_idx = None
            elif self.selected_room_idx is not None and self.selected_room_idx > idx:
                self.selected_room_idx -= 1
            self._rebuild_snap_arrays()
            self._save_session()
            return f"Deleted room '{room['name']}'"
        return "Invalid room index"

    def set_room_type(self, idx: int, room_type: str) -> str:
        if 0 <= idx < len(self.rooms):
            old_type = self.rooms[idx].get("room_type")
            self._draw_undo_stack.append(("type", [(idx, old_type)]))
            self.rooms[idx]["room_type"] = room_type
            self._save_session()
            return f"Set '{self.rooms[idx]['name']}' type to {room_type}"
        return "Invalid room index"

    # === EXPORT ===============================================================

    def export_room_boundaries_csv(self, output_path: Optional[Path] = None):
        if not self.rooms:
            return
        output_path = Path(output_path) if output_path else self.csv_path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        rows = []
        for room in self.rooms:
            coord_strings = [f"X_{x:.3f} Y_{y:.3f}" for x, y in room["vertices"]]
            parent = room.get("parent") or ""
            hdr_file = room.get("hdr_file", "")
            rows.append([room["name"], parent, hdr_file] + coord_strings)
        max_cols = max(len(r) for r in rows)
        with open(output_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in rows:
                row += [""] * (max_cols - len(row))
                writer.writerow(row)

    def _export_rooms_as_aoi_files(self):
        aoi_dir = self.project_aoi_dir
        aoi_dir.mkdir(parents=True, exist_ok=True)
        for aoi_path in aoi_dir.glob("*.aoi"):
            aoi_path.unlink()
        for room in self.rooms:
            name = room["name"]
            verts = room["vertices"]
            hdr_file = room.get("hdr_file", "")
            safe_name = re.sub(r'[<>:"/\\|?*]', "_", name)
            aoi_path = aoi_dir / f"{safe_name}.aoi"
            with open(aoi_path, "w") as f:
                f.write(f"AOI Points File: {name}\n")
                f.write(f"ASSOCIATED VIEW FILE: {hdr_file}.vp\n")
                f.write(f"FFL z height(m): 0.000\n")
                f.write(f"CENTRAL x,y: 0.000 0.000\n")
                f.write(f"NO. PERIMETER POINTS {len(verts)}: x,y pixel_x pixel_y positions\n")
                for px, py in verts:
                    f.write(f"0.000 0.000 {px:.3f} {py:.3f}\n")

    # ── DF compliance results ─────────────────────────────────────────────────

    def _compute_all_room_df_results(self) -> None:
        """Compute per-room DF% compliance stats for all rooms on the current HDR.

        Results stored in each room dict under 'df_result':
            {'pass_pct': float|None, 'mean_df': float, 'threshold': float|None}
        Run in a background thread to avoid blocking the Dash worker.
        """
        if self._df_image is None:
            return
        img = self._df_image
        h, w = img.shape[:2]
        try:
            from skimage.draw import polygon as skimage_polygon
        except ImportError:
            return

        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room.get("vertices", [])
            if len(verts) < 3:
                continue
            xs = [int(round(v[0])) for v in verts]
            ys = [int(round(v[1])) for v in verts]
            try:
                rr, cc = skimage_polygon(ys, xs, shape=(h, w))
            except Exception:
                continue
            if len(rr) == 0:
                continue
            df_vals = img[rr, cc]
            mean_df = float(np.mean(df_vals))
            rtype = self._effective_room_type(room)
            threshold = self._threshold_for_type(rtype)
            if threshold is not None and len(df_vals) > 0:
                pass_pct = float((df_vals >= threshold).sum() / len(df_vals) * 100)
            else:
                pass_pct = None
            room["df_result"] = {
                "pass_pct": pass_pct,
                "mean_df": round(mean_df, 2),
                "threshold": threshold,
            }

    # ── DF% image loading ─────────────────────────────────────────────────────

    def load_df_image(self) -> None:
        """Load the DF% image for the current HDR file into self._df_image (cached).

        Runs synchronously — call from a background thread to avoid blocking the
        Dash worker. Sets self._df_image to None if loading fails.
        """
        if not self.hdr_files:
            self._df_image = None
            return
        hdr_entry = self.hdr_files[self.current_hdr_idx]
        hdr_path: Path = hdr_entry["hdr_path"]
        key = str(hdr_path)
        if key in self._df_image_cache:
            self._df_image = self._df_image_cache[key]
            return
        try:
            # subprocess imported at top
            # archilume_utils imported at top
            width, height = archilume_utils.get_hdr_resolution(hdr_path)
            result = subprocess.run(
                ["pvalue", "-h", "-H", "-b", "-o", "-df", str(hdr_path)],
                capture_output=True, check=True,
            )
            raw = np.frombuffer(result.stdout, dtype=np.float32).reshape((height, width))
            df_arr = raw * 1.79  # W/m² → DF%
            self._df_image = df_arr
            self._df_image_cache[key] = df_arr
        except Exception as exc:
            log.warning("load_df_image failed for %s: %s", hdr_path.name, exc)
            self._df_image = None

    def stamp_df(self, x: float, y: float) -> Optional[float]:
        """Read DF% at pixel (x, y) and add a stamp for the current HDR.

        Returns the DF% value, or None if the DF image is not loaded or coords
        are out of bounds.
        """
        if self._df_image is None:
            return None
        py, px = int(round(y)), int(round(x))
        h, w = self._df_image.shape[:2]
        if not (0 <= py < h and 0 <= px < w):
            return None
        df_val = float(self._df_image[py, px])
        hdr_name = self.current_hdr_name or ""
        self._df_stamps.setdefault(hdr_name, []).append((x, y, df_val))
        self._save_session()
        return df_val

    def remove_nearest_stamp(self, x: float, y: float, threshold_px: float = 20.0) -> bool:
        """Remove the DF% stamp nearest to (x, y) within threshold_px.

        Returns True if a stamp was removed.
        """
        hdr_name = self.current_hdr_name or ""
        stamps = self._df_stamps.get(hdr_name, [])
        if not stamps:
            return False
        dists = [(abs(sx - x) ** 2 + abs(sy - y) ** 2) for sx, sy, _ in stamps]
        idx = dists.index(min(dists))
        if dists[idx] ** 0.5 <= threshold_px:
            stamps.pop(idx)
            self._df_stamps[hdr_name] = stamps
            self._save_session()
            return True
        return False

    # ── AOI level cycling ─────────────────────────────────────────────────────

    def cycle_aoi_level(self) -> str:
        """Cycle to the next IESVE FFL / AOI level and reassign rooms."""
        ffl_groups = self._aoi_ffl_groups()
        n = len(ffl_groups)
        if n == 0:
            return "No AOI levels available"
        self._aoi_level_idx = (self._aoi_level_idx + 1) % n
        self._reassign_aoi_level()
        return f"AOI level → {self._aoi_level_idx + 1} / {n}"

    def get_aoi_level_label(self) -> str:
        ffl_groups = self._aoi_ffl_groups()
        n = len(ffl_groups)
        if n == 0:
            return "—"
        return f"{self._aoi_level_idx + 1} / {n}"

    # ── DF report helpers ─────────────────────────────────────────────────────

    def _effective_room_type(self, room: dict) -> Optional[str]:
        """Return the room's type, inheriting from its parent if unset."""
        rtype = room.get("room_type") or ""
        if rtype:
            return rtype
        parent_name = room.get("parent")
        if parent_name:
            for r in self.rooms:
                if r.get("name") == parent_name:
                    return r.get("room_type") or ""
        return ""

    def _threshold_for_type(self, room_type: Optional[str]) -> Optional[float]:
        """Return the DF% compliance threshold for a room type, or None."""
        if not room_type:
            return None
        return self.DF_THRESHOLDS.get(room_type.upper())

    @staticmethod
    def _compute_summary_for_hdr(hdr_name, hdr_path, rooms_for_hdr, iesve_mode=False):
        """Load one HDR's DF image and compute per-room summary statistics.

        rooms_for_hdr: list of (parent, name, room_type, verts, threshold, child_verts_list).
        Returns (summary_rows, pixel_chunks).
        """
        try:
            from skimage.draw import polygon as skimage_polygon
        except ImportError:
            log.warning("scikit-image not available — skipping DF computation for %s", hdr_name)
            return [], []

        try:
            # subprocess imported at top
            # archilume_utils imported at top
            if iesve_mode:
                width, height = archilume_utils.get_hdr_resolution(hdr_path)
                result = subprocess.run(
                    ["pvalue", "-h", "-H", "-b", "-o", "-df", str(hdr_path)],
                    capture_output=True, check=True,
                )
                raw = np.frombuffer(result.stdout, dtype=np.float32).reshape((height, width))
                df_img = raw * 1.79
            else:
                from archilume.post.hdr2wpd import Hdr2Wpd as _Hdr2Wpd
                df_img = _Hdr2Wpd().load_df_image(hdr_path)
        except Exception as exc:
            log.warning("DF image load failed for %s: %s", hdr_name, exc)
            return [], []

        if df_img is None:
            return [], []

        h, w = df_img.shape
        summary_rows = []
        pixel_chunks = []
        for parent, name, room_type, verts, threshold, child_verts_list in rooms_for_hdr:
            xs = [int(round(v[0])) for v in verts]
            ys = [int(round(v[1])) for v in verts]
            rr, cc = skimage_polygon(ys, xs, shape=(h, w))
            if len(rr) == 0:
                continue
            if child_verts_list:
                mask = np.zeros((h, w), dtype=bool)
                mask[rr, cc] = True
                for child_verts in child_verts_list:
                    cxs = [int(round(v[0])) for v in child_verts]
                    cys = [int(round(v[1])) for v in child_verts]
                    crr, ccc = skimage_polygon(cys, cxs, shape=(h, w))
                    mask[crr, ccc] = False
                rr, cc = np.where(mask)
            n = len(rr)
            if n == 0:
                continue
            df_vals = df_img[rr, cc]
            lux_vals = df_vals * 100.0
            if threshold is not None:
                passing = int((df_vals >= threshold).sum())
                passing_pct = round(passing / n * 100, 1)
            else:
                passing = None
                passing_pct = None
            summary_rows.append({
                "HDR File": hdr_name,
                "Parent": parent,
                "Room": name,
                "Room Type": room_type or "",
                "Total Pixels": n,
                "DF Threshold (%)": threshold,
                "Pixels >= Threshold": passing,
                "% Area >= Threshold": passing_pct,
            })
            pixel_chunks.append((name, np.round(lux_vals, 2), np.round(df_vals, 4)))
        return summary_rows, pixel_chunks

    # ── Export worker ─────────────────────────────────────────────────────────

    def run_export(self, progress: dict) -> None:
        """Background export: write AOI files, CSV, Excel DF report, and archive.

        *progress* keys: 'phase' (str), 'message' (str), 'pct' (int 0–100).
        """
        # zipfile imported at top
        # datetime imported at top
        # ThreadPoolExecutor imported at top
        try:
            progress.update({"phase": "aoi", "message": "Writing AOI files…", "pct": 5})
            self._export_rooms_as_aoi_files()

            progress.update({"phase": "csv", "message": "Writing boundary CSV…", "pct": 15})
            self.export_room_boundaries_csv()

            # ── DF computation ────────────────────────────────────────────────
            progress.update({"phase": "df", "message": "Computing DF statistics…", "pct": 20})
            hdr_lookup = {entry["name"]: entry["hdr_path"] for entry in self.hdr_files}
            children_verts_by_name: dict = {}
            for room in self.rooms:
                p = room.get("parent", "")
                if p and len(room.get("vertices", [])) >= 3:
                    children_verts_by_name.setdefault(p, []).append(room["vertices"])

            hdr_room_groups: dict = {}
            for room in self.rooms:
                verts = room.get("vertices", [])
                hdr_name = room.get("hdr_file", "")
                if len(verts) < 3 or hdr_name not in hdr_lookup:
                    continue
                parent = room.get("parent", "")
                name = room.get("name", "unnamed")
                room_type = self._effective_room_type(room)
                threshold = self._threshold_for_type(room_type)
                child_verts_list = children_verts_by_name.get(name, [])
                hdr_room_groups.setdefault(hdr_name, []).append(
                    (parent, name, room_type or "", verts, threshold, child_verts_list)
                )

            all_summary_rows = []
            all_pixel_chunks = []
            iesve = self._iesve_room_data_path is not None
            total_hdrs = len(hdr_room_groups)
            max_workers = max(1, min(total_hdrs, 4))
            completed = 0
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {
                    pool.submit(
                        self._compute_summary_for_hdr,
                        hdr_name, hdr_lookup[hdr_name], rooms_list, iesve
                    ): hdr_name
                    for hdr_name, rooms_list in hdr_room_groups.items()
                }
                for future in as_completed(futures):
                    rows, chunks = future.result()
                    all_summary_rows.extend(rows)
                    all_pixel_chunks.extend(chunks)
                    completed += 1
                    pct = 20 + int(completed / max(total_hdrs, 1) * 50)
                    progress.update({"pct": pct, "message": f"DF computed {completed}/{total_hdrs} HDRs…"})

            # ── Write Excel report ────────────────────────────────────────────
            if all_summary_rows:
                progress.update({"phase": "excel", "message": "Writing Excel report…", "pct": 72})
                try:
                    import pandas as _pd
                    from openpyxl import load_workbook as _lwb
                    from openpyxl.utils import get_column_letter as _gcl
                    from openpyxl.worksheet.table import Table as _Table, TableStyleInfo as _TSI

                    output_dir = self.wpd_dir
                    output_dir.mkdir(parents=True, exist_ok=True)
                    df_frame = _pd.DataFrame(all_summary_rows)
                    xlsx_path = output_dir / "aoi_report_daylight.xlsx"
                    df_frame.to_excel(xlsx_path, sheet_name="Room Summary", index=False)
                    wb = _lwb(xlsx_path)
                    ws = wb.active
                    for col in ws.columns:
                        max_len = max(
                            len(str(cell.value)) if cell.value is not None else 0
                            for cell in col
                        )
                        ws.column_dimensions[col[0].column_letter].width = max_len + 4
                    if ws.max_row > 1:
                        last_col = _gcl(ws.max_column)
                        tbl = _Table(
                            displayName="RoomSummary",
                            ref=f"A1:{last_col}{ws.max_row}",
                        )
                        tbl.tableStyleInfo = _TSI(
                            name="TableStyleMedium9",
                            showFirstColumn=False, showLastColumn=False,
                            showRowStripes=True, showColumnStripes=False,
                        )
                        ws.add_table(tbl)
                    wb.save(xlsx_path)
                    log.info("Excel report saved to %s", xlsx_path)

                    # Per-room pixel CSVs
                    progress.update({"phase": "pixelcsv", "message": "Writing pixel CSVs…", "pct": 80})
                    csv_subdir = output_dir / "aoi_pixel_data"
                    csv_subdir.mkdir(parents=True, exist_ok=True)
                    for room_name, lux_vals, df_pct_vals in all_pixel_chunks:
                        safe = re.sub(r'[<>:"/\\|?*]', "_", room_name)
                        with open(csv_subdir / f"{safe}_pixels.csv", "w", newline="") as f:
                            writer = csv.writer(f)
                            writer.writerow(["Room", "Illuminance (Lux)", "Daylight Factor (%)"])
                            for lux, df_pct in zip(lux_vals, df_pct_vals):
                                writer.writerow([room_name, lux, df_pct])
                except Exception as exc:
                    log.warning("Excel report failed (pandas/openpyxl missing?): %s", exc)

            # ── Archive ───────────────────────────────────────────────────────
            progress.update({"phase": "archive", "message": "Archiving…", "pct": 88})
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            if self.project:
                archive_dir = self.archive_dir
                archive_dir.mkdir(parents=True, exist_ok=True)
                zip_path = archive_dir / f"aoi_archive_{ts}.zip"
                with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                    if self.session_path.exists():
                        zf.write(self.session_path, self.session_path.name)
                    for aoi_f in self.project_aoi_dir.glob("*.aoi"):
                        zf.write(aoi_f, aoi_f.name)
                    if self.csv_path.exists():
                        zf.write(self.csv_path, self.csv_path.name)
                    xlsx_path = self.wpd_dir / "aoi_report_daylight.xlsx"
                    if xlsx_path.exists():
                        zf.write(xlsx_path, xlsx_path.name)

            progress.update({"phase": "done", "message": "Export complete.", "pct": 100})
        except Exception as exc:
            log.error("Export failed: %s", exc)
            progress.update({"phase": "error", "message": f"Export failed: {exc}", "pct": 0})

    # ── Extract archive ───────────────────────────────────────────────────────

    def list_archives(self) -> list:
        """Return a list of archive zip paths sorted newest-first."""
        if not self.archive_dir.exists():
            return []
        return sorted(self.archive_dir.glob("*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)

    def extract_and_reload(self, zip_path: Path) -> str:
        """Extract a zip archive into the project output area and reload state.

        Mirrors matplotlib_app._on_extract_click extraction logic.
        Returns a status message string.
        """
        # zipfile imported at top
        # shutil imported at top
        try:
            # The archive sits alongside the session; AOI files restore to project_aoi_dir.
            # Clear existing AOI files first so stale files don't persist.
            if self.project_aoi_dir.exists():
                shutil.rmtree(self.project_aoi_dir)
            self.project_aoi_dir.mkdir(parents=True, exist_ok=True)

            with zipfile.ZipFile(zip_path, "r") as zf:
                for member in zf.infolist():
                    member_path = member.filename
                    # Strip leading path prefixes from old archives
                    for prefix in ("outputs/", "aoi/"):
                        if member_path.startswith(prefix):
                            member_path = member_path[len(prefix):]
                            break
                    if not member_path:
                        continue
                    dest = self.project_aoi_dir / member_path
                    if member.is_dir():
                        dest.mkdir(parents=True, exist_ok=True)
                    else:
                        dest.parent.mkdir(parents=True, exist_ok=True)
                        with zf.open(member) as src, open(dest, "wb") as dst:
                            shutil.copyfileobj(src, dst)

            # Full state reload
            self.rooms = []
            self.hdr_files = self._scan_hdr_files()
            self.current_hdr_idx = 0
            self._rebuild_image_variants()
            self._rebuild_snap_arrays()
            self._update_parent_options()
            self._load_session()
            self.selected_room_idx = None
            self._edit_selected_vertex = None
            log.info("Extracted and reloaded: %s", zip_path.name)
            return f"Extracted: {zip_path.name}"
        except Exception as exc:
            log.error("Extract failed: %s", exc)
            return f"Extract failed: {exc}"


# ═══════════════════════════════════════════════════════════════════════════════
# Dash App — Layout & Callbacks
# ═══════════════════════════════════════════════════════════════════════════════

# Colour tokens
C = {
    "sidebar": "#f5f6f7",
    "sidebar_act": "#e8eaed",
    "panel_bg": "#ffffff",
    "panel_bdr": "#e2e5e9",
    "viewport": "#1a1a2e",
    "text_pri": "#1a1f27",
    "text_sec": "#5a6472",
    "text_dim": "#9ba6b2",
    "accent": "#0d9488",
    "hover": "#eef0f3",
    "btn_on": "#ccfbf1",
    "danger": "#dc2626",
    "success": "#059669",
    "deep": "#f0f2f4",
}

FONT = "'DM Mono', monospace"

# Global state
STATE: Optional[EditorState] = None


def _init_state(project: Optional[str] = None):
    global STATE
    STATE = EditorState(project=project)


# ── Helpers ──────────────────────────────────────────────────────────────────


def sb_btn(icon, btn_id, tip, active=False):
    color = C["accent"] if active else C["text_sec"]
    bg = C["sidebar_act"] if active else "transparent"
    return html.Div(
        [
            dbc.Button(
                DashIconify(icon=icon, width=20, color=color),
                id=btn_id,
                color="link",
                style={
                    "padding": "9px 0",
                    "width": "44px",
                    "display": "flex",
                    "justifyContent": "center",
                    "borderRadius": "6px",
                    "backgroundColor": bg,
                    "border": "none",
                },
            ),
            dbc.Tooltip(tip, target=btn_id, placement="right"),
        ],
        style={"marginBottom": "3px"},
    )


def divider_line():
    return html.Hr(
        style={
            "margin": "6px 8px",
            "opacity": "0.3",
            "borderColor": C["panel_bdr"],
        }
    )


def room_type_btn(label, btn_id, active=False):
    bg = C["btn_on"] if active else C["deep"]
    border = C["accent"] if active else C["panel_bdr"]
    color = C["accent"] if active else C["text_sec"]
    return dbc.Button(
        label,
        id=btn_id,
        size="sm",
        color="secondary",
        style={
            "fontFamily": FONT,
            "fontSize": "10px",
            "padding": "3px 7px",
            "backgroundColor": bg,
            "border": f"1px solid {border}",
            "color": color,
            "marginRight": "3px",
            "borderRadius": "3px",
        },
    )


# ── Layout ───────────────────────────────────────────────────────────────────


def _menu_item(label, btn_id):
    """A plain text menu item styled for the top menubar dropdowns."""
    return dbc.DropdownMenuItem(
        label,
        id=btn_id,
        style={"fontFamily": FONT, "fontSize": "12px"},
    )


def make_menubar():
    item_style = {
        "fontFamily": FONT,
        "fontSize": "12px",
        "color": C["text_pri"],
        "background": "none",
        "border": "none",
        "padding": "4px 10px",
        "cursor": "pointer",
    }
    menu_style = {
        "backgroundColor": C["sidebar"],
        "height": "28px",
        "display": "flex",
        "alignItems": "center",
        "borderBottom": f"1px solid {C['panel_bdr']}",
        "padding": "0 4px",
        "gap": "2px",
        "flexShrink": "0",
    }
    toggle_style = {
        **item_style,
        "padding": "2px 8px",
        "borderRadius": "3px",
        "textDecoration": "none",
    }
    return html.Div(
        [
            dbc.DropdownMenu(
                label="File",
                children=[
                    _menu_item("Open Project",    "menu-open-project"),
                    _menu_item("Create Project",  "menu-create-project"),
                    dbc.DropdownMenuItem(divider=True),
                    _menu_item("Export & Archive","menu-export"),
                    _menu_item("Import ZIP",      "menu-import"),
                ],
                toggle_style=toggle_style,
                color="link",
                size="sm",
                style={"fontFamily": FONT},
            ),
            dbc.DropdownMenu(
                label="View",
                children=[
                    _menu_item("Toggle Room Browser",   "menu-toggle-browser"),
                    _menu_item("Toggle Image Layer",    "menu-toggle-image"),
                    _menu_item("Reset Zoom",            "menu-reset-zoom"),
                    _menu_item("Toggle Grid",           "menu-toggle-grid"),
                ],
                toggle_style=toggle_style,
                color="link",
                size="sm",
                style={"fontFamily": FONT},
            ),
            dbc.DropdownMenu(
                label="Tools",
                children=[
                    _menu_item("Keyboard Shortcuts", "menu-shortcuts"),
                ],
                toggle_style=toggle_style,
                color="link",
                size="sm",
                style={"fontFamily": FONT},
            ),
        ],
        id="top-menubar",
        style=menu_style,
    )


def make_layout():
    left_sidebar = html.Div(
        [
            sb_btn("lucide:menu", "sb-menu", "Toggle Room Browser"),
            html.Div(style={"height": "4px"}),
            divider_line(),
            sb_btn("lucide:pen-line", "sb-edit-mode", "Edit Mode [E]"),
            sb_btn("lucide:corner-down-right", "sb-ortho", "Ortho Lines [O]", active=True),
            sb_btn("lucide:scissors", "sb-divider", "Room Divider [DD]"),
            divider_line(),
            sb_btn("lucide:layers", "sb-image-toggle", "Toggle Image Layer [T]"),
            sb_btn("lucide:zoom-in", "sb-reset-zoom", "Reset Zoom [R]"),
            divider_line(),
            sb_btn("lucide:save", "sb-export", "Export & Archive"),
            divider_line(),
            sb_btn("lucide:grid", "sb-grid-toggle", "Toggle Grid [G]", active=True),
            html.Div(
                [
                    dbc.Input(
                        id="input-grid-spacing",
                        type="number",
                        value=100,
                        min=10,
                        max=10000,
                        step=10,
                        size="sm",
                        debounce=True,
                        style={
                            "width": "44px",
                            "fontFamily": FONT,
                            "fontSize": "9px",
                            "padding": "3px 4px",
                            "textAlign": "center",
                            "backgroundColor": C["sidebar_act"],
                            "border": f"1px solid {C['panel_bdr']}",
                            "borderRadius": "4px",
                            "color": C["text_sec"],
                        },
                    ),
                    dbc.Tooltip("Grid spacing (mm)", target="input-grid-spacing", placement="right"),
                ],
                style={"display": "flex", "justifyContent": "center", "marginBottom": "3px"},
            ),
            html.Div(style={"flex": "1"}),
            sb_btn("lucide:keyboard", "sb-shortcuts", "Keyboard Shortcuts"),
        ],
        style={
            "width": "52px",
            "height": "100vh",
            "backgroundColor": C["sidebar"],
            "display": "flex",
            "flexDirection": "column",
            "alignItems": "center",
            "padding": "10px 0",
            "borderRight": f"1px solid {C['panel_bdr']}",
            "flexShrink": "0",
            "overflowY": "auto",
        },
        className="sidebar-scroll",
    )

    # Right panel
    right_panel = html.Div(
        [
            # HDR Navigation
            html.Div(
                [
                    html.Div(
                        "HDR FILE",
                        style={
                            "fontSize": "9px",
                            "fontFamily": FONT,
                            "color": C["text_dim"],
                            "letterSpacing": "0.1em",
                            "marginBottom": "4px",
                        },
                    ),
                    html.Div(
                        [
                            dbc.Button(
                                DashIconify(icon="lucide:chevron-up", width=14),
                                id="btn-hdr-prev",
                                size="sm",
                                color="secondary",
                                style={"padding": "2px 6px"},
                            ),
                            html.Span(
                                id="hdr-name-display",
                                style={
                                    "fontFamily": FONT,
                                    "fontSize": "11px",
                                    "color": C["text_pri"],
                                    "flex": "1",
                                    "textAlign": "center",
                                    "overflow": "hidden",
                                    "textOverflow": "ellipsis",
                                    "whiteSpace": "nowrap",
                                },
                            ),
                            dbc.Button(
                                DashIconify(icon="lucide:chevron-down", width=14),
                                id="btn-hdr-next",
                                size="sm",
                                color="secondary",
                                style={"padding": "2px 6px"},
                            ),
                        ],
                        style={
                            "display": "flex",
                            "alignItems": "center",
                            "gap": "4px",
                        },
                    ),
                ],
                style={
                    "padding": "8px 10px",
                    "borderBottom": f"1px solid {C['panel_bdr']}",
                },
            ),
            # Room browser header — Collapse / Expand all
            html.Div(
                [
                    html.Div(
                        "ROOMS",
                        style={
                            "fontSize": "9px",
                            "fontFamily": FONT,
                            "color": C["text_dim"],
                            "letterSpacing": "0.1em",
                            "flex": "1",
                        },
                    ),
                    dbc.Button(
                        DashIconify(icon="lucide:chevrons-down-up", width=12),
                        id="btn-collapse-all",
                        color="link",
                        size="sm",
                        style={"padding": "0 4px", "color": C["text_dim"]},
                    ),
                    dbc.Tooltip("Collapse all", target="btn-collapse-all", placement="top"),
                    dbc.Button(
                        DashIconify(icon="lucide:chevrons-up-down", width=12),
                        id="btn-expand-all",
                        color="link",
                        size="sm",
                        style={"padding": "0 4px", "color": C["text_dim"]},
                    ),
                    dbc.Tooltip("Expand all", target="btn-expand-all", placement="top"),
                ],
                style={
                    "display": "flex",
                    "alignItems": "center",
                    "padding": "4px 10px",
                    "borderBottom": f"1px solid {C['panel_bdr']}",
                },
            ),
            # Room tree
            html.Div(
                id="room-tree-container",
                style={
                    "flex": "1",
                    "overflowY": "auto",
                    "padding": "4px 0",
                },
            ),
            # Room input panel
            html.Div(
                [
                    html.Div(
                        "PARENT",
                        style={
                            "fontSize": "9px",
                            "fontFamily": FONT,
                            "color": C["text_dim"],
                            "letterSpacing": "0.1em",
                        },
                    ),
                    html.Div(
                        id="parent-display",
                        style={
                            "fontFamily": FONT,
                            "fontSize": "11px",
                            "color": C["accent"],
                            "marginBottom": "6px",
                        },
                    ),
                    html.Div(
                        "ROOM NAME",
                        style={
                            "fontSize": "9px",
                            "fontFamily": FONT,
                            "color": C["text_dim"],
                            "letterSpacing": "0.1em",
                        },
                    ),
                    dbc.Input(
                        id="input-room-name",
                        placeholder="e.g. BED1",
                        size="sm",
                        style={
                            "fontFamily": FONT,
                            "fontSize": "11px",
                            "marginBottom": "6px",
                        },
                    ),
                    html.Div(
                        "ROOM TYPE",
                        style={
                            "fontSize": "9px",
                            "fontFamily": FONT,
                            "color": C["text_dim"],
                            "letterSpacing": "0.1em",
                            "marginBottom": "4px",
                        },
                    ),
                    html.Div(
                        [
                            room_type_btn("BED", "btn-type-bed"),
                            room_type_btn("LIVING", "btn-type-living"),
                            room_type_btn("NON-RESI", "btn-type-nonresi"),
                            room_type_btn("CIRC", "btn-type-circ"),
                        ],
                        style={"display": "flex", "flexWrap": "wrap", "marginBottom": "8px"},
                    ),
                    html.Div(
                        [
                            dbc.Button(
                                [
                                    DashIconify(icon="lucide:save", width=13, style={"marginRight": "4px"}),
                                    "Save",
                                ],
                                id="btn-save-room",
                                color="success",
                                size="sm",
                                style={"flex": "1", "fontFamily": FONT, "fontSize": "11px"},
                            ),
                            dbc.Button(
                                [
                                    DashIconify(icon="lucide:trash-2", width=13, style={"marginRight": "4px"}),
                                    "Delete",
                                ],
                                id="btn-delete-room",
                                color="danger",
                                size="sm",
                                style={"flex": "1", "fontFamily": FONT, "fontSize": "11px"},
                            ),
                        ],
                        style={"display": "flex", "gap": "4px"},
                    ),
                ],
                style={
                    "padding": "8px 10px",
                    "borderTop": f"1px solid {C['panel_bdr']}",
                },
            ),
        ],
        id="right-panel",
        style={
            "width": "240px",
            "height": "100vh",
            "backgroundColor": C["panel_bg"],
            "borderRight": f"1px solid {C['panel_bdr']}",
            "display": "flex",
            "flexDirection": "column",
            "flexShrink": "0",
        },
    )

    # ── Project panel (shown in place of right-panel) ────────────────────────
    _panel_style = {
        "width": "260px",
        "height": "100%",
        "backgroundColor": C["panel_bg"],
        "borderRight": f"1px solid {C['panel_bdr']}",
        "display": "none",          # hidden by default; toggled by callback
        "flexDirection": "column",
        "flexShrink": "0",
        "overflowY": "auto",
    }

    def _field_label(text):
        return html.Div(
            text,
            style={
                "fontSize": "9px",
                "fontFamily": FONT,
                "color": C["text_dim"],
                "letterSpacing": "0.1em",
                "marginBottom": "3px",
                "marginTop": "8px",
            },
        )

    def _panel_btn(label, btn_id, color="secondary"):
        return dbc.Button(
            label,
            id=btn_id,
            color=color,
            size="sm",
            style={"fontFamily": FONT, "fontSize": "11px", "width": "100%", "marginTop": "6px"},
        )

    # ── Open panel content ────────────────────────────────────────────────────
    open_panel_content = html.Div(
        [
            html.Div(
                "OPEN PROJECT",
                style={
                    "fontSize": "9px", "fontFamily": FONT, "color": C["text_dim"],
                    "letterSpacing": "0.1em", "marginBottom": "10px",
                },
            ),
            _field_label("SELECT PROJECT"),
            dcc.Dropdown(
                id="proj-dropdown",
                options=[],
                value=None,
                clearable=False,
                style={"fontFamily": FONT, "fontSize": "11px"},
            ),
            _field_label("MODE"),
            html.Div(id="proj-info-mode", style={"fontFamily": FONT, "fontSize": "11px", "color": C["text_pri"]}),
            _field_label("PDF PATH"),
            html.Div(id="proj-info-pdf",  style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "wordBreak": "break-all"}),
            _field_label("IMAGE DIR"),
            html.Div(id="proj-info-img",  style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "wordBreak": "break-all"}),
            _field_label("IESVE CSV"),
            html.Div(id="proj-info-csv",  style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "wordBreak": "break-all"}),
            html.Hr(style={"margin": "12px 0", "opacity": "0.3"}),
            _panel_btn("Load Project",         "btn-proj-load",   color="success"),
            _panel_btn("Reveal in Explorer",   "btn-proj-reveal"),
            _panel_btn("Cancel",               "btn-proj-cancel"),
        ],
        id="open-panel-content",
        style={"padding": "12px 10px"},
    )

    # ── Create panel content ──────────────────────────────────────────────────
    create_panel_content = html.Div(
        [
            html.Div(
                "CREATE PROJECT",
                style={
                    "fontSize": "9px", "fontFamily": FONT, "color": C["text_dim"],
                    "letterSpacing": "0.1em", "marginBottom": "10px",
                },
            ),
            _field_label("PROJECT NAME"),
            dbc.Input(
                id="proj-name-input",
                placeholder="e.g. MyProject",
                size="sm",
                style={"fontFamily": FONT, "fontSize": "11px"},
            ),
            _field_label("RESULTS SOURCE"),
            dbc.RadioItems(
                id="proj-mode-radio",
                options=[
                    {"label": "Archilume simulation", "value": "hdr"},
                    {"label": "IESVE simulation",      "value": "iesve"},
                ],
                value="hdr",
                inline=False,
                style={"fontFamily": FONT, "fontSize": "11px"},
            ),
            _field_label("PDF PLAN (optional)"),
            dcc.Upload(
                id="proj-pdf-upload",
                children=html.Div(
                    "Drag & drop or click to browse…",
                    style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_dim"]},
                ),
                style={
                    "border": f"1px dashed {C['panel_bdr']}",
                    "borderRadius": "4px",
                    "padding": "6px 8px",
                    "textAlign": "center",
                    "cursor": "pointer",
                },
                multiple=False,
            ),
            html.Div(id="proj-pdf-name",  style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "marginTop": "2px"}),
            _field_label("AOI FILES"),
            dcc.Upload(
                id="proj-aoi-upload",
                children=html.Div(
                    "Drag & drop or click to browse…",
                    style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_dim"]},
                ),
                style={
                    "border": f"1px dashed {C['panel_bdr']}",
                    "borderRadius": "4px",
                    "padding": "6px 8px",
                    "textAlign": "center",
                    "cursor": "pointer",
                },
                multiple=True,
            ),
            html.Div(id="proj-aoi-names", style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "marginTop": "2px"}),
            # IESVE-only fields
            html.Div(
                [
                    _field_label("IESVE PIC DIR"),
                    dcc.Upload(
                        id="proj-imgdir-upload",
                        children=html.Div(
                            "Drag & drop folder or files…",
                            style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_dim"]},
                        ),
                        style={
                            "border": f"1px dashed {C['panel_bdr']}",
                            "borderRadius": "4px",
                            "padding": "6px 8px",
                            "textAlign": "center",
                            "cursor": "pointer",
                        },
                        multiple=True,
                    ),
                    html.Div(id="proj-imgdir-name", style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "marginTop": "2px"}),
                    _field_label("ROOM BOUNDARIES CSV"),
                    dcc.Upload(
                        id="proj-csv-upload",
                        children=html.Div(
                            "Drag & drop or click to browse…",
                            style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_dim"]},
                        ),
                        style={
                            "border": f"1px dashed {C['panel_bdr']}",
                            "borderRadius": "4px",
                            "padding": "6px 8px",
                            "textAlign": "center",
                            "cursor": "pointer",
                        },
                        multiple=False,
                    ),
                    html.Div(id="proj-csv-name", style={"fontFamily": FONT, "fontSize": "10px", "color": C["text_sec"], "marginTop": "2px"}),
                ],
                id="proj-iesve-fields",
                style={"display": "none"},
            ),
            html.Hr(style={"margin": "12px 0", "opacity": "0.3"}),
            _panel_btn("Create & Load",      "btn-proj-create", color="success"),
            _panel_btn("Cancel",             "btn-proj-cancel-create"),
        ],
        id="create-panel-content",
        style={"padding": "12px 10px"},
    )

    project_panel = html.Div(
        [open_panel_content, create_panel_content],
        id="project-panel",
        style=_panel_style,
    )

    # Viewport
    viewport = html.Div(
        [
            dcc.Graph(
                id="viewport-graph",
                config={
                    "scrollZoom": True,
                    "displayModeBar": False,
                    "doubleClick": "reset",
                },
                style={"height": "100%", "width": "100%"},
            ),
        ],
        id="viewport-container",
        style={
            "flex": "1",
            "backgroundColor": C["viewport"],
            "position": "relative",
            "overflow": "hidden",
        },
    )

    # Status bar
    status_bar = html.Div(
        id="status-bar",
        style={
            "height": "24px",
            "backgroundColor": C["sidebar"],
            "borderTop": f"1px solid {C['panel_bdr']}",
            "display": "flex",
            "alignItems": "center",
            "padding": "0 12px",
            "fontFamily": FONT,
            "fontSize": "10px",
            "color": C["text_sec"],
        },
        className="status-bar",
    )

    # Stores
    stores = html.Div(
        [
            dcc.Store(id="store-draw-vertices", data=[]),
            dcc.Store(id="store-divider-points", data=[]),
            dcc.Store(id="store-trigger", data=0),
            dcc.Store(id="keyboard-event", data=""),
            dcc.Store(id="store-grid-spacing", data=100),
            dcc.Store(id="store-grid-visible", data=True),
            dcc.Store(id="store-project-panel", data={"open": False, "mode": "open"}),
            dcc.Interval(id="keyboard-poll", interval=150, n_intervals=0),
        ]
    )

    # Shortcuts modal
    shortcuts_modal = dbc.Modal(
        [
            dbc.ModalHeader("Keyboard Shortcuts"),
            dbc.ModalBody(
                html.Table(
                    [
                        html.Tr([html.Td("D", style={"fontFamily": FONT, "paddingRight": "16px"}), html.Td("Draw mode")]),
                        html.Tr([html.Td("E"), html.Td("Edit mode")]),
                        html.Tr([html.Td("DD"), html.Td("Room divider")]),
                        html.Tr([html.Td("S"), html.Td("Save / Finalize")]),
                        html.Tr([html.Td("O"), html.Td("Toggle ortho")]),
                        html.Tr([html.Td("T"), html.Td("Toggle image layer")]),
                        html.Tr([html.Td("R"), html.Td("Reset zoom")]),
                        html.Tr([html.Td("F"), html.Td("Fit to selected room")]),
                        html.Tr([html.Td("Esc"), html.Td("Exit current mode")]),
                        html.Tr([html.Td("Ctrl+Z"), html.Td("Undo")]),
                        html.Tr([html.Td("G"), html.Td("Toggle grid")]),
                        html.Tr([html.Td("Up/Down"), html.Td("Navigate HDR files")]),
                        html.Tr([html.Td("Click"), html.Td("Select room / Place vertex")]),
                    ],
                    style={"fontFamily": FONT, "fontSize": "12px"},
                )
            ),
        ],
        id="shortcuts-modal",
        is_open=False,
    )

    layout = html.Div(
        [
            stores,
            shortcuts_modal,
            html.Div(
                [
                    make_menubar(),
                    html.Div(
                        [
                            left_sidebar,
                            right_panel,
                            project_panel,
                            html.Div(
                                [viewport, status_bar],
                                style={
                                    "flex": "1",
                                    "display": "flex",
                                    "flexDirection": "column",
                                    "minWidth": "0",
                                },
                            ),
                        ],
                        style={
                            "display": "flex",
                            "flex": "1",
                            "overflow": "hidden",
                        },
                    ),
                ],
                style={
                    "display": "flex",
                    "flexDirection": "column",
                    "height": "100vh",
                    "overflow": "hidden",
                },
            ),
        ]
    )
    return layout


# ── Figure Builder ───────────────────────────────────────────────────────────


def build_figure(
    grid_spacing: int = 100,
    grid_visible: bool = True,
    state: Optional["EditorState"] = None,
) -> go.Figure:
    """Build the Plotly figure with image and room polygons.

    *state* defaults to the module-level STATE singleton so existing callers
    are unaffected. Pass an explicit state to use a different EditorState
    instance (e.g. from dash_app.py).
    """
    s = state if state is not None else STATE
    if s is None:
        return go.Figure()

    fig = go.Figure()
    img_w, img_h = s.get_image_dimensions()

    # Add image as a go.Image trace (not layout_image) so it responds to clickData.
    path = s.current_variant_path
    if path is not None:
        img = s._load_image(path)
        if img is not None:
            img_uint8 = (np.clip(img, 0, 1) * 255).astype(np.uint8)
            fig.add_trace(
                go.Image(
                    z=img_uint8,
                    hoverinfo="x+y",
                    name="_image",
                )
            )

    # Room polygon colours — no fill, red border for all types
    _ROOM_LINE = "rgba(220, 38, 38, 0.8)"
    TYPE_COLORS = {
        "BED": "rgba(0, 0, 0, 0)",
        "LIVING": "rgba(0, 0, 0, 0)",
        "NON-RESI": "rgba(0, 0, 0, 0)",
        "CIRC": "rgba(0, 0, 0, 0)",
    }
    TYPE_LINE_COLORS = {
        "BED": _ROOM_LINE,
        "LIVING": _ROOM_LINE,
        "NON-RESI": _ROOM_LINE,
        "CIRC": _ROOM_LINE,
    }
    DEFAULT_FILL = "rgba(0, 0, 0, 0)"
    DEFAULT_LINE = _ROOM_LINE
    SELECTED_FILL = "rgba(255, 255, 0, 0.15)"
    SELECTED_LINE = "rgba(255, 255, 0, 1.0)"

    # Draw rooms
    for i, room in enumerate(s.rooms):
        if not s._is_room_on_current_hdr(room):
            continue
        verts = room["vertices"]
        if len(verts) < 3:
            continue

        is_selected = i == s.selected_room_idx
        rtype = room.get("room_type", "")

        if is_selected:
            fill_color = SELECTED_FILL
            line_color = SELECTED_LINE
            line_width = 3
        else:
            fill_color = TYPE_COLORS.get(rtype, DEFAULT_FILL)
            line_color = TYPE_LINE_COLORS.get(rtype, DEFAULT_LINE)
            line_width = 2

        xs = [v[0] for v in verts] + [verts[0][0]]
        ys = [v[1] for v in verts] + [verts[0][1]]

        room_name = room.get("name", "")
        line_dash = "dash" if "_DIV" in room_name else "solid"

        fig.add_trace(
            go.Scatter(
                x=xs,
                y=ys,
                fill="toself",
                fillcolor=fill_color,
                line=dict(color=line_color, width=line_width, dash=line_dash),
                mode="lines",
                name=room.get("name", ""),
                hoverinfo="text",
                text=room.get("name", ""),
                showlegend=False,
                customdata=[{"room_idx": i}] * len(xs),
            )
        )

        # Label — room name + optional DF compliance annotation
        label_pt = s._polygon_label_point(verts)
        label_color = "white" if not is_selected else "cyan"
        label_size = max(6, int(10 * s._annotation_scale))
        label_text = room.get("name", "")

        df_result = room.get("df_result")
        if df_result is not None:
            pass_pct = df_result.get("pass_pct")
            mean_df = df_result.get("mean_df", 0.0)
            threshold = df_result.get("threshold")
            if pass_pct is not None:
                # Colour: green ≥90%, amber ≥50%, red <50%
                if pass_pct >= 90:
                    compliance_color = "#22c55e"
                elif pass_pct >= 50:
                    compliance_color = "#f59e0b"
                else:
                    compliance_color = "#ef4444"
                df_line = f"{pass_pct:.0f}% ≥ {threshold}%DF"
                # Add DF annotation as a separate trace slightly below the name
                fig.add_trace(
                    go.Scatter(
                        x=[label_pt[0]],
                        y=[label_pt[1] - label_size * 1.4],
                        mode="text",
                        text=[df_line],
                        textfont=dict(
                            size=max(5, int(8 * s._annotation_scale)),
                            color=compliance_color,
                            family=FONT,
                        ),
                        showlegend=False,
                        hoverinfo="skip",
                    )
                )
            else:
                # No threshold — just show mean DF
                fig.add_trace(
                    go.Scatter(
                        x=[label_pt[0]],
                        y=[label_pt[1] - label_size * 1.4],
                        mode="text",
                        text=[f"DF {mean_df:.2f}%"],
                        textfont=dict(
                            size=max(5, int(8 * s._annotation_scale)),
                            color="#94a3b8",
                            family=FONT,
                        ),
                        showlegend=False,
                        hoverinfo="skip",
                    )
                )

        fig.add_trace(
            go.Scatter(
                x=[label_pt[0]],
                y=[label_pt[1]],
                mode="text",
                text=[label_text],
                textfont=dict(
                    size=label_size,
                    color=label_color,
                    family=FONT,
                ),
                showlegend=False,
                hoverinfo="skip",
            )
        )

    # Draw in-progress polygon
    if s.draw_mode and s.current_polygon_vertices:
        draw_verts = s.current_polygon_vertices
        xs = [v[0] for v in draw_verts]
        ys = [v[1] for v in draw_verts]
        fig.add_trace(
            go.Scatter(
                x=xs,
                y=ys,
                mode="lines+markers",
                line=dict(color="cyan", width=2, dash="dash"),
                marker=dict(color="cyan", size=8),
                showlegend=False,
                hoverinfo="skip",
            )
        )

    # Draw divider points
    if s.divider_mode and s._divider_points:
        pts = s._divider_points
        xs = [p[0] for p in pts]
        ys = [p[1] for p in pts]
        fig.add_trace(
            go.Scatter(
                x=xs,
                y=ys,
                mode="lines+markers",
                line=dict(color="magenta", width=2),
                marker=dict(color="magenta", size=8, line=dict(color="white", width=1.5)),
                showlegend=False,
                hoverinfo="skip",
            )
        )

    # Edit mode vertex handles
    if s.edit_mode:
        pending = s._edit_selected_vertex  # (room_idx, vertex_idx) or None
        for i, room in enumerate(s.rooms):
            if not s._is_room_on_current_hdr(room):
                continue
            verts = room["vertices"]
            xs = [v[0] for v in verts]
            ys = [v[1] for v in verts]
            # Per-vertex colour: yellow/larger for the pending-selected vertex
            colors = []
            sizes = []
            for j in range(len(verts)):
                if pending and pending == (i, j):
                    colors.append("yellow")
                    sizes.append(12)
                else:
                    colors.append("cyan")
                    sizes.append(7)
            fig.add_trace(
                go.Scatter(
                    x=xs,
                    y=ys,
                    mode="markers",
                    marker=dict(color=colors, size=sizes, line=dict(color="white", width=1)),
                    showlegend=False,
                    hoverinfo="text",
                    text=[f"Room {room.get('name','')}, vertex {j}" for j in range(len(verts))],
                    customdata=[{"room_idx": i, "vertex_idx": j} for j in range(len(verts))],
                )
            )

    # DF% stamps
    hdr_name = s.current_hdr_name or ""
    df_stamp_list = s._df_stamps.get(hdr_name, [])
    if df_stamp_list:
        sx_list = [st[0] for st in df_stamp_list]
        sy_list = [st[1] for st in df_stamp_list]
        df_vals = [st[2] for st in df_stamp_list]
        fig.add_trace(
            go.Scatter(
                x=sx_list,
                y=sy_list,
                mode="markers+text",
                marker=dict(
                    color=[v for v in df_vals],
                    colorscale=[[0, "#dc2626"], [0.5, "#d97706"], [1, "#059669"]],
                    cmin=0,
                    cmax=5,
                    size=14,
                    line=dict(color="white", width=1.5),
                    symbol="circle",
                ),
                text=[f"{v:.1f}%" for v in df_vals],
                textposition="top center",
                textfont=dict(size=max(6, int(9 * s._annotation_scale)), color="white", family=FONT),
                showlegend=False,
                hoverinfo="text",
                hovertext=[f"DF: {v:.2f}%" for v in df_vals],
                customdata=[{"stamp_idx": i} for i in range(len(df_stamp_list))],
                name="_df_stamps",
            )
        )

    # PDF overlay
    layout_images = []
    if s._overlay_visible and img_w > 0 and img_h > 0:
        b64 = s.get_overlay_base64()
        if b64 is not None:
            tf = s.get_overlay_transform()
            ox = tf.get("offset_x", 0.0)
            oy = tf.get("offset_y", 0.0)
            sx = tf.get("scale_x", 1.0)
            sy = tf.get("scale_y", 1.0)
            # Estimate overlay pixel dimensions from original PDF raster size
            # Plotly layout images: x/y is top-left in data coords,
            # sizex/sizey is width/height in data coords.
            # We don't know exact PDF pixel size here without re-rasterizing,
            # so we use the full image extent scaled by transform.
            overlay_w = img_w * sx
            overlay_h = img_h * sy
            layout_images.append(dict(
                source=b64,
                xref="x", yref="y",
                x=ox,
                y=oy + overlay_h,   # Plotly y is top of image in data coords (image y=0 is top)
                sizex=overlay_w,
                sizey=overlay_h,
                sizing="stretch",
                opacity=s._overlay_alpha,
                layer="above" if s.overlay_align_mode else "below",
            ))

    # Grid lines — grid_spacing is in mm; convert to pixels using VIEW params
    grid_shapes = []
    if grid_visible and grid_spacing and grid_spacing > 0 and img_w > 0 and img_h > 0:
        # Read only the VIEW= text line — no imageio needed, img_w already known
        step_px = None
        if s is not None:
            hdr_entry = s.hdr_files[s.current_hdr_idx] if s.hdr_files else None
            if hdr_entry:
                try:
                    with open(hdr_entry["hdr_path"], "r", encoding="utf-8", errors="ignore") as _f:
                        _view_line = next(
                            (l.strip() for l in _f if l.startswith("VIEW=")), None
                        )
                    if _view_line:
                        _vh = re.search(r"-vh\s+([\d.]+)", _view_line)
                        if _vh:
                            vh_metres = float(_vh.group(1))
                            if vh_metres > 0:
                                pixels_per_mm = img_w / (vh_metres * 1000.0)
                                step_px = grid_spacing * pixels_per_mm
                except Exception:
                    pass
        if step_px is None:
            step_px = float(grid_spacing)  # fallback: treat mm value as raw pixels
        step = max(int(round(step_px)), 1)
        grid_color = "rgba(255, 255, 255, 0.15)"
        # Vertical lines
        x = 0
        while x <= img_w:
            grid_shapes.append(dict(
                type="line",
                x0=x, y0=0, x1=x, y1=img_h,
                line=dict(color=grid_color, width=1),
                layer="above",
            ))
            x += step
        # Horizontal lines
        y = 0
        while y <= img_h:
            grid_shapes.append(dict(
                type="line",
                x0=0, y0=y, x1=img_w, y1=y,
                line=dict(color=grid_color, width=1),
                layer="above",
            ))
            y += step

    fig.update_layout(
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            showticklabels=False,
            fixedrange=False,
            constrain="domain",
        ),
        yaxis=dict(
            showgrid=False,
            zeroline=False,
            showticklabels=False,
            fixedrange=False,
            scaleanchor="x",
            constrain="domain",
        ),
        margin=dict(l=0, r=0, t=0, b=0),
        plot_bgcolor=C["viewport"],
        paper_bgcolor=C["viewport"],
        dragmode="pan",
        hovermode="closest",
        shapes=grid_shapes,
        images=layout_images,
    )

    return fig


# ── Tree Renderer ────────────────────────────────────────────────────────────


def render_tree(state: Optional["EditorState"] = None) -> list:
    """Render the room tree as Dash components.

    *state* defaults to the module-level STATE singleton so existing callers
    are unaffected.
    """
    s = state if state is not None else STATE
    if s is None:
        return []
    items = s.build_layer_tree()
    rows = []
    for item in items:
        indent = item["indent"] * 14
        is_sel = (
            item.get("data", {}).get("room_idx") is not None
            and item["data"]["room_idx"] == s.selected_room_idx
        )
        bg = C["btn_on"] if is_sel else "transparent"
        font_weight = "600" if item["type"] == "hdr" else "normal"
        font_size = "11px" if item["type"] == "hdr" else "10px"
        color = C["accent"] if item.get("is_current") else C["text_pri"]
        if item["type"] in ("room_child",):
            color = C["text_sec"]

        chevron = None
        if item.get("has_children"):
            node_key = item.get("node_key")
            is_collapsed = node_key and eval(node_key) in s._tree_collapsed if node_key else False
            chevron_icon = "lucide:chevron-right" if is_collapsed else "lucide:chevron-down"
            chevron = DashIconify(
                icon=chevron_icon,
                width=11,
                color=C["text_dim"],
                style={"marginRight": "3px", "flexShrink": "0", "cursor": "pointer"},
            )

        room_idx = item.get("data", {}).get("room_idx")
        row_id = {"type": "tree-row", "index": room_idx if room_idx is not None else f"node-{item.get('node_key', '')}"}

        row = html.Div(
            [
                html.Div(style={"width": f"{indent}px", "flexShrink": "0"}),
                chevron if chevron else html.Div(style={"width": "14px", "flexShrink": "0"}),
                html.Span(
                    item["label"],
                    style={
                        "fontFamily": FONT,
                        "fontSize": font_size,
                        "fontWeight": font_weight,
                        "color": color,
                        "overflow": "hidden",
                        "textOverflow": "ellipsis",
                        "whiteSpace": "nowrap",
                    },
                ),
            ],
            id=row_id,
            className=f"tree-row {'tree-row-selected' if is_sel else ''}",
            style={
                "display": "flex",
                "alignItems": "center",
                "padding": "3px 8px",
                "cursor": "pointer",
                "backgroundColor": bg,
                "borderRadius": "3px",
                "margin": "0 4px",
            },
            n_clicks=0,
        )
        rows.append(row)
    return rows


# ── App Creation ─────────────────────────────────────────────────────────────


def create_app(project: Optional[str] = None) -> dash.Dash:
    _init_state(project)

    app = dash.Dash(
        __name__,
        external_stylesheets=[
            dbc.themes.BOOTSTRAP,
            "https://fonts.googleapis.com/css2?family=DM+Mono:wght@300;400;500&display=swap",
        ],
        suppress_callback_exceptions=True,
        assets_folder=str(Path(__file__).parent / "assets"),
    )
    app.layout = make_layout()

    # ── Callbacks ────────────────────────────────────────────────────────────

    # Initial render
    @app.callback(
        Output("viewport-graph", "figure"),
        Output("hdr-name-display", "children"),
        Output("room-tree-container", "children"),
        Output("status-bar", "children"),
        Output("parent-display", "children"),
        Input("store-trigger", "data"),
        State("store-grid-spacing", "data"),
        State("store-grid-visible", "data"),
        prevent_initial_call=False,
    )
    def update_all(_trigger, grid_spacing, grid_visible):
        if STATE is None:
            return go.Figure(), "No project", [], "No project loaded", "—"
        fig = build_figure(
            grid_spacing=grid_spacing or 100,
            grid_visible=grid_visible if grid_visible is not None else True,
        )
        hdr_name = STATE.current_hdr_name or "No images"
        tree = render_tree()
        n_rooms = sum(1 for r in STATE.rooms if STATE._is_room_on_current_hdr(r))
        status = f"{hdr_name} | {n_rooms} rooms | " + (
            "DRAW" if STATE.draw_mode else "DIVIDER" if STATE.divider_mode else "EDIT" if STATE.edit_mode else "SELECT"
        )
        parent = STATE.selected_parent or "—"
        return fig, hdr_name, tree, status, parent

    # HDR navigation
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("btn-hdr-prev", "n_clicks"),
        prevent_initial_call=True,
    )
    def hdr_prev(_):
        if STATE is None or STATE.current_hdr_idx <= 0:
            return no_update
        STATE.current_hdr_idx -= 1
        STATE._rebuild_image_variants()
        STATE._rebuild_snap_arrays()
        STATE._update_parent_options()
        STATE.selected_room_idx = None
        return time.time()

    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("btn-hdr-next", "n_clicks"),
        prevent_initial_call=True,
    )
    def hdr_next(_):
        if STATE is None or STATE.current_hdr_idx >= len(STATE.hdr_files) - 1:
            return no_update
        STATE.current_hdr_idx += 1
        STATE._rebuild_image_variants()
        STATE._rebuild_snap_arrays()
        STATE._update_parent_options()
        STATE.selected_room_idx = None
        return time.time()

    # Image variant toggle
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("sb-image-toggle", "n_clicks"),
        prevent_initial_call=True,
    )
    def toggle_variant(_):
        if STATE is None or not STATE.image_variants:
            return no_update
        STATE.current_variant_idx = (STATE.current_variant_idx + 1) % len(STATE.image_variants)
        return time.time()

    # Canvas click → select room or place vertex
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Output("store-draw-vertices", "data"),
        Input("viewport-graph", "clickData"),
        State("store-draw-vertices", "data"),
        prevent_initial_call=True,
    )
    def on_canvas_click(click_data, draw_verts):
        if STATE is None or click_data is None:
            return no_update, no_update
        pt = click_data.get("points", [{}])[0]
        x, y = pt.get("x"), pt.get("y")

        # Check for room_idx in customdata (from polygon/label trace clicks)
        customdata = pt.get("customdata")
        if isinstance(customdata, list) and customdata:
            customdata = customdata[0]
        if isinstance(customdata, dict) and "room_idx" in customdata:
            room_idx = customdata["room_idx"]
            if not STATE.divider_mode and not STATE.draw_mode:
                if 0 <= room_idx < len(STATE.rooms):
                    STATE.selected_room_idx = room_idx
                    STATE.selected_parent = STATE.rooms[room_idx].get("parent")
                    return time.time(), no_update
            # For divider/draw modes, try to get coords from bbox center
            if x is None and "bbox" in pt:
                bbox = pt["bbox"]
                x = (bbox.get("x0", 0) + bbox.get("x1", 0)) / 2
                y = (bbox.get("y0", 0) + bbox.get("y1", 0)) / 2

        if x is None or y is None:
            return no_update, no_update

        # Divider mode
        if STATE.divider_mode:
            STATE.add_divider_point(x, y)
            return time.time(), no_update

        # Draw mode
        if STATE.draw_mode:
            sx, sy = STATE._snap_to_pixel(x, y)
            sx, sy = STATE._snap_to_vertex(sx, sy)
            if STATE.ortho_mode and STATE.current_polygon_vertices:
                lx, ly = STATE.current_polygon_vertices[-1]
                dx, dy = abs(sx - lx), abs(sy - ly)
                if dx >= dy:
                    sy = ly
                else:
                    sx = lx
            STATE.current_polygon_vertices.append([sx, sy])
            draw_verts = [[v[0], v[1]] for v in STATE.current_polygon_vertices]
            return time.time(), draw_verts

        # Select mode: find room at click point
        room_idx = STATE.find_room_at(x, y)
        if room_idx is not None:
            STATE.selected_room_idx = room_idx
            STATE.selected_parent = STATE.rooms[room_idx].get("parent")
        else:
            STATE.selected_room_idx = None
            STATE.selected_parent = None
        return time.time(), no_update

    # Tree row clicks
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input({"type": "tree-row", "index": ALL}, "n_clicks"),
        prevent_initial_call=True,
    )
    def on_tree_click(n_clicks_list):
        if STATE is None:
            return no_update
        triggered = ctx.triggered_id
        if triggered is None:
            return no_update
        idx = triggered.get("index") if isinstance(triggered, dict) else None
        if isinstance(idx, int):
            STATE.selected_room_idx = idx
            room = STATE.rooms[idx]
            STATE.selected_parent = room.get("parent")
        elif isinstance(idx, str) and idx.startswith("node-"):
            # Toggle collapse
            node_key_str = idx[5:]
            try:
                node_key = eval(node_key_str)
                if node_key in STATE._tree_collapsed:
                    STATE._tree_collapsed.discard(node_key)
                else:
                    STATE._tree_collapsed.add(node_key)
            except Exception:
                pass
        return time.time()

    # Collapse / Expand all tree nodes
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("btn-collapse-all", "n_clicks"),
        Input("btn-expand-all",   "n_clicks"),
        prevent_initial_call=True,
    )
    def on_collapse_expand_all(*_):
        if STATE is None:
            return no_update
        triggered = ctx.triggered_id
        if triggered == "btn-collapse-all":
            # Build all possible node keys across every HDR, not just the current one
            keys = set()
            for entry in STATE.hdr_files:
                hdr_name = entry["name"]
                keys.add(("hdr", hdr_name))
                room_groups = {}
                for room in STATE.rooms:
                    if room.get("hdr_name") == hdr_name:
                        grp = room.get("parent") or room.get("name", "")
                        room_groups.setdefault(grp, [])
                for grp in room_groups:
                    keys.add(("room_group", hdr_name, grp))
                    apts = {r.get("apartment") for r in STATE.rooms
                            if r.get("hdr_name") == hdr_name and r.get("parent") == grp and r.get("apartment")}
                    for apt in apts:
                        keys.add(("apt", hdr_name, grp, apt))
            STATE._tree_collapsed = keys
        elif triggered == "btn-expand-all":
            STATE._tree_collapsed = set()
        return time.time()

    # Save room
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Output("store-draw-vertices", "data", allow_duplicate=True),
        Input("btn-save-room", "n_clicks"),
        State("input-room-name", "value"),
        prevent_initial_call=True,
    )
    def on_save(_n, name):
        if STATE is None:
            return no_update, no_update
        # Finalize divider
        if STATE.divider_mode:
            msg = STATE.finalize_division()
            print(msg)
            return time.time(), []
        # Save polygon
        if STATE.draw_mode and STATE.current_polygon_vertices:
            name = name or "ROOM"
            msg = STATE.save_room(name)
            print(msg)
            return time.time(), []
        return no_update, no_update

    # Delete room
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("btn-delete-room", "n_clicks"),
        prevent_initial_call=True,
    )
    def on_delete(_n):
        if STATE is None or STATE.selected_room_idx is None:
            return no_update
        msg = STATE.delete_room(STATE.selected_room_idx)
        print(msg)
        return time.time()

    # Room type buttons
    for type_id, type_val in [
        ("btn-type-bed", "BED"),
        ("btn-type-living", "LIVING"),
        ("btn-type-nonresi", "NON-RESI"),
        ("btn-type-circ", "CIRC"),
    ]:
        @app.callback(
            Output("store-trigger", "data", allow_duplicate=True),
            Input(type_id, "n_clicks"),
            prevent_initial_call=True,
        )
        def on_type_click(_n, _tv=type_val):
            if STATE is None or STATE.selected_room_idx is None:
                return no_update
            STATE.set_room_type(STATE.selected_room_idx, _tv)
            return time.time()

    # Edit mode toggle
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("sb-edit-mode", "n_clicks"),
        prevent_initial_call=True,
    )
    def toggle_edit(_):
        if STATE is None:
            return no_update
        STATE.edit_mode = not STATE.edit_mode
        if not STATE.edit_mode:
            STATE.divider_mode = False
            STATE._divider_points = []
        return time.time()

    # Ortho toggle
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("sb-ortho", "n_clicks"),
        prevent_initial_call=True,
    )
    def toggle_ortho(_):
        if STATE is None:
            return no_update
        STATE.ortho_mode = not STATE.ortho_mode
        return time.time()

    # Divider mode
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("sb-divider", "n_clicks"),
        prevent_initial_call=True,
    )
    def toggle_divider(_):
        if STATE is None:
            return no_update
        if STATE.divider_mode:
            STATE.divider_mode = False
            STATE._divider_points = []
            STATE._divider_room_idx = None
        else:
            if not STATE.edit_mode:
                STATE.edit_mode = True
            if STATE.selected_room_idx is not None:
                STATE.divider_mode = True
                STATE._divider_room_idx = STATE.selected_room_idx
                STATE._divider_points = []
        return time.time()

    # Shortcuts modal
    @app.callback(
        Output("shortcuts-modal", "is_open", allow_duplicate=True),
        Input("sb-shortcuts", "n_clicks"),
        State("shortcuts-modal", "is_open"),
        prevent_initial_call=True,
    )
    def toggle_shortcuts(n, is_open):
        return not is_open

    # Right panel toggle — update store so toggle_project_panel re-renders styles
    @app.callback(
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("sb-menu", "n_clicks"),
        Input("menu-toggle-browser", "n_clicks"),
        State("store-project-panel", "data"),
        prevent_initial_call=True,
    )
    def toggle_right_panel(*args):
        panel_data = args[-1] or {"open": False, "mode": "open", "browser_visible": True}
        visible = panel_data.get("browser_visible", True)
        return {**panel_data, "browser_visible": not visible}

    # Grid toggle
    @app.callback(
        Output("store-grid-visible", "data", allow_duplicate=True),
        Output("sb-grid-toggle", "style"),
        Input("sb-grid-toggle", "n_clicks"),
        State("store-grid-visible", "data"),
        prevent_initial_call=True,
    )
    def toggle_grid(_, grid_visible):
        new_visible = not (grid_visible if grid_visible is not None else True)
        color = C["accent"] if new_visible else C["text_sec"]
        bg = C["sidebar_act"] if new_visible else "transparent"
        btn_style = {
            "padding": "9px 0",
            "width": "44px",
            "display": "flex",
            "justifyContent": "center",
            "borderRadius": "6px",
            "backgroundColor": bg,
            "border": "none",
        }
        return new_visible, btn_style

    # Grid spacing input
    @app.callback(
        Output("store-grid-spacing", "data"),
        Input("input-grid-spacing", "value"),
        prevent_initial_call=True,
    )
    def update_grid_spacing(value):
        if value is None or value < 1:
            return no_update
        return int(value)

    # Rebuild figure when grid settings change
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("store-grid-spacing", "data"),
        Input("store-grid-visible", "data"),
        prevent_initial_call=True,
    )
    def grid_changed(_, __):
        return time.time()

    # Reset zoom
    @app.callback(
        Output("viewport-graph", "figure", allow_duplicate=True),
        Input("sb-reset-zoom", "n_clicks"),
        State("store-grid-spacing", "data"),
        State("store-grid-visible", "data"),
        prevent_initial_call=True,
    )
    def reset_zoom(_, grid_spacing, grid_visible):
        return build_figure(
            grid_spacing=grid_spacing or 100,
            grid_visible=grid_visible if grid_visible is not None else True,
        )

    # ── Project panel: open/create buttons + menubar items ───────────────────

    @app.callback(
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("menu-open-project",   "n_clicks"),
        prevent_initial_call=True,
    )
    def on_open_project(_):
        return {"open": True, "mode": "open"}

    @app.callback(
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("menu-create-project",  "n_clicks"),
        prevent_initial_call=True,
    )
    def on_create_project(_):
        return {"open": True, "mode": "create"}

    # Menubar items that mirror existing sidebar buttons
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("menu-toggle-image", "n_clicks"),
        prevent_initial_call=True,
    )
    def menu_toggle_image(_):
        if STATE is None:
            return no_update
        STATE.image_visible = not getattr(STATE, "image_visible", True)
        return time.time()

    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("menu-reset-zoom", "n_clicks"),
        prevent_initial_call=True,
    )
    def menu_reset_zoom(_):
        if STATE is None:
            return no_update
        STATE.zoom_reset = True
        return time.time()

    @app.callback(
        Output("store-grid-visible", "data", allow_duplicate=True),
        Input("menu-toggle-grid", "n_clicks"),
        State("store-grid-visible", "data"),
        prevent_initial_call=True,
    )
    def menu_toggle_grid(_, visible):
        return not visible if visible is not None else False

    @app.callback(
        Output("shortcuts-modal", "is_open", allow_duplicate=True),
        Input("menu-shortcuts", "n_clicks"),
        prevent_initial_call=True,
    )
    def menu_shortcuts(_):
        return True

    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("menu-export", "n_clicks"),
        prevent_initial_call=True,
    )
    def menu_export(_):
        if STATE is None:
            return no_update
        STATE._trigger_export = True
        return time.time()

    # ── Project panel: visibility + content rendering ─────────────────────────

    @app.callback(
        Output("project-panel",        "style"),
        Output("right-panel",          "style"),
        Output("open-panel-content",   "style"),
        Output("create-panel-content", "style"),
        Output("proj-dropdown",        "options"),
        Output("proj-dropdown",        "value"),
        Input("store-project-panel",   "data"),
    )
    def toggle_project_panel(panel_data):
        _panel_base = {
            "width": "260px",
            "height": "100%",
            "backgroundColor": C["panel_bg"],
            "borderRight": f"1px solid {C['panel_bdr']}",
            "flexDirection": "column",
            "flexShrink": "0",
            "overflowY": "auto",
        }
        _right_base = {
            "width": "240px",
            "height": "100%",
            "backgroundColor": C["panel_bg"],
            "borderRight": f"1px solid {C['panel_bdr']}",
            "flexDirection": "column",
            "flexShrink": "0",
        }
        browser_visible = (panel_data or {}).get("browser_visible", True)
        if not panel_data or not panel_data.get("open"):
            right_display = "flex" if browser_visible else "none"
            return (
                {**_panel_base, "display": "none"},
                {**_right_base, "display": right_display},
                {"padding": "12px 10px"},
                {"padding": "12px 10px", "display": "none"},
                [],
                None,
            )
        mode = panel_data.get("mode", "open")
        projects = list_projects()
        options = [{"label": p, "value": p} for p in projects]
        current = STATE.project if STATE else None
        open_style   = {"padding": "12px 10px"} if mode == "open"   else {"padding": "12px 10px", "display": "none"}
        create_style = {"padding": "12px 10px"} if mode == "create" else {"padding": "12px 10px", "display": "none"}
        return (
            {**_panel_base, "display": "flex"},
            {**_right_base, "display": "none"},
            open_style,
            create_style,
            options,
            current,
        )

    # ── proj-dropdown → populate info fields ─────────────────────────────────

    @app.callback(
        Output("proj-info-mode", "children"),
        Output("proj-info-pdf",  "children"),
        Output("proj-info-img",  "children"),
        Output("proj-info-csv",  "children"),
        Input("proj-dropdown",   "value"),
        prevent_initial_call=True,
    )
    def populate_project_info(name):
        if not name:
            return "—", "—", "—", "—"
        cfg = load_project_toml(name)
        paths_cfg = cfg.get("paths", {})
        mode = cfg.get("project", {}).get("mode", "hdr")
        return (
            mode,
            paths_cfg.get("pdf_path") or "—",
            paths_cfg.get("image_dir") or "—",
            paths_cfg.get("iesve_room_data") or "—",
        )

    # ── Upload filename display ───────────────────────────────────────────────

    @app.callback(
        Output("proj-pdf-name",   "children"),
        Input("proj-pdf-upload",  "filename"),
        prevent_initial_call=True,
    )
    def show_pdf_name(fn):
        return fn or ""

    @app.callback(
        Output("proj-aoi-names",  "children"),
        Input("proj-aoi-upload",  "filename"),
        prevent_initial_call=True,
    )
    def show_aoi_names(fns):
        if not fns:
            return ""
        return ", ".join(fns) if isinstance(fns, list) else fns

    @app.callback(
        Output("proj-imgdir-name","children"),
        Input("proj-imgdir-upload","filename"),
        prevent_initial_call=True,
    )
    def show_imgdir_name(fns):
        if not fns:
            return ""
        return f"{len(fns)} file(s)" if isinstance(fns, list) else fns

    @app.callback(
        Output("proj-csv-name",   "children"),
        Input("proj-csv-upload",  "filename"),
        prevent_initial_call=True,
    )
    def show_csv_name(fn):
        return fn or ""

    # ── Mode radio → show/hide IESVE fields ──────────────────────────────────

    @app.callback(
        Output("proj-iesve-fields", "style"),
        Input("proj-mode-radio",    "value"),
        prevent_initial_call=True,
    )
    def toggle_iesve_fields(mode):
        return {"display": "block"} if mode == "iesve" else {"display": "none"}

    # ── Cancel buttons ────────────────────────────────────────────────────────

    @app.callback(
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("btn-proj-cancel",        "n_clicks"),
        Input("btn-proj-cancel-create", "n_clicks"),
        prevent_initial_call=True,
    )
    def cancel_panel(*_):
        return {"open": False, "mode": "open"}

    # ── Load Project ──────────────────────────────────────────────────────────

    @app.callback(
        Output("store-trigger",       "data", allow_duplicate=True),
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("btn-proj-load",        "n_clicks"),
        State("proj-dropdown",        "value"),
        prevent_initial_call=True,
    )
    def load_project(_, name):
        if not name:
            return no_update, no_update
        _init_state(name)
        set_last_project(name)
        return time.time(), {"open": False, "mode": "open"}

    # ── Reveal in Explorer ────────────────────────────────────────────────────

    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Input("btn-proj-reveal", "n_clicks"),
        State("proj-dropdown",   "value"),
        prevent_initial_call=True,
    )
    def reveal_project(_, name):
        if not name:
            return no_update
        project_dir = config.PROJECTS_DIR / name
        if project_dir.exists():
            os.startfile(str(project_dir))
        return no_update

    # ── Create & Load ─────────────────────────────────────────────────────────

    @app.callback(
        Output("store-trigger",       "data", allow_duplicate=True),
        Output("store-project-panel", "data", allow_duplicate=True),
        Input("btn-proj-create",      "n_clicks"),
        State("proj-name-input",      "value"),
        State("proj-mode-radio",      "value"),
        State("proj-pdf-upload",      "contents"),
        State("proj-pdf-upload",      "filename"),
        State("proj-aoi-upload",      "contents"),
        State("proj-aoi-upload",      "filename"),
        State("proj-imgdir-upload",   "contents"),
        State("proj-imgdir-upload",   "filename"),
        State("proj-csv-upload",      "contents"),
        State("proj-csv-upload",      "filename"),
        prevent_initial_call=True,
    )
    def create_project(
        _,
        name, mode,
        pdf_content, pdf_filename,
        aoi_contents, aoi_filenames,
        imgdir_contents, imgdir_filenames,
        csv_content, csv_filename,
    ):
        if not name or not name.strip():
            return no_update, no_update
        name = name.strip()
        paths = config.get_project_paths(name)
        paths.create_dirs()

        def _save_upload(content, filename, dest_dir):
            """Decode a dcc.Upload base64 payload and write to dest_dir."""
            if not content or not filename:
                return None
            import base64 as _b64
            header, data = content.split(",", 1)
            raw = _b64.b64decode(data)
            dest = Path(dest_dir) / filename
            dest.write_bytes(raw)
            return dest

        rel_pdf = None
        if pdf_content and pdf_filename:
            dest = _save_upload(pdf_content, pdf_filename, paths.inputs_dir)
            if dest:
                rel_pdf = str(dest.relative_to(paths.project_dir))

        rel_aoi_files = []
        if aoi_contents and aoi_filenames:
            pairs = zip(aoi_contents, aoi_filenames) if isinstance(aoi_contents, list) else [(aoi_contents, aoi_filenames)]
            for c, fn in pairs:
                dest = _save_upload(c, fn, paths.aoi_inputs_dir)
                if dest:
                    rel_aoi_files.append(str(dest.relative_to(paths.project_dir)))

        rel_csv = None
        if csv_content and csv_filename:
            dest = _save_upload(csv_content, csv_filename, paths.inputs_dir)
            if dest:
                rel_csv = str(dest.relative_to(paths.project_dir))

        rel_imgdir = None
        if imgdir_contents and imgdir_filenames:
            pairs = zip(imgdir_contents, imgdir_filenames) if isinstance(imgdir_contents, list) else [(imgdir_contents, imgdir_filenames)]
            for c, fn in pairs:
                _save_upload(c, fn, paths.pic_dir)
            rel_imgdir = str(paths.pic_dir.relative_to(paths.project_dir))

        cfg = {
            "project": {"name": name, "mode": mode},
            "paths": {
                "pdf_path":       rel_pdf,
                "image_dir":      rel_imgdir,
                "iesve_room_data": rel_csv,
                "aoi_files":      rel_aoi_files,
            },
        }
        save_project_toml(name, cfg)
        _init_state(name)
        set_last_project(name)
        return time.time(), {"open": False, "mode": "open"}

    # Clientside callback: poll window._lastKeyEvent into dcc.Store
    app.clientside_callback(
        """
        function(n) {
            var evt = window._lastKeyEvent || '';
            window._lastKeyEvent = '';
            return evt;
        }
        """,
        Output("keyboard-event", "data"),
        Input("keyboard-poll", "n_intervals"),
    )

    # Server-side keyboard handler
    @app.callback(
        Output("store-trigger", "data", allow_duplicate=True),
        Output("store-draw-vertices", "data", allow_duplicate=True),
        Input("keyboard-event", "data"),
        State("store-draw-vertices", "data"),
        prevent_initial_call=True,
    )
    def on_keyboard(key_json, draw_verts):
        if not key_json or STATE is None:
            return no_update, no_update
        try:
            evt = json.loads(key_json)
        except (json.JSONDecodeError, TypeError):
            return no_update, no_update
        key = evt.get("key", "")
        ctrl = evt.get("ctrl", False)
        shift = evt.get("shift", False)
        trigger = no_update
        dv = no_update
        if key == "Escape":
            STATE.draw_mode = False
            STATE.edit_mode = False
            STATE.divider_mode = False
            STATE._divider_points = []
            STATE._divider_room_idx = None
            STATE.current_polygon_vertices = []
            trigger = time.time()
            dv = []
        elif key.lower() == "d" and not ctrl and not shift:
            STATE.draw_mode = not STATE.draw_mode
            if STATE.draw_mode:
                STATE.edit_mode = False
                STATE.divider_mode = False
            STATE.current_polygon_vertices = []
            trigger = time.time()
            dv = []
        elif key.lower() == "e" and not ctrl:
            STATE.edit_mode = not STATE.edit_mode
            if not STATE.edit_mode:
                STATE.divider_mode = False
                STATE._divider_points = []
            STATE.draw_mode = False
            trigger = time.time()
        elif key.lower() == "s" and not ctrl:
            if STATE.divider_mode and STATE._divider_points:
                msg = STATE.finalize_division()
                print(msg)
                trigger = time.time()
                dv = []
            elif STATE.draw_mode and STATE.current_polygon_vertices:
                msg = STATE.save_room("ROOM")
                print(msg)
                trigger = time.time()
                dv = []
        elif key.lower() == "o":
            STATE.ortho_mode = not STATE.ortho_mode
            trigger = time.time()
        elif key.lower() == "z" and ctrl:
            msg = STATE.undo_edit()
            print(msg)
            trigger = time.time()
        elif key == "ArrowUp":
            if STATE.current_hdr_idx > 0:
                STATE.current_hdr_idx -= 1
                STATE._rebuild_image_variants()
                STATE._rebuild_snap_arrays()
                STATE._update_parent_options()
                STATE.selected_room_idx = None
                trigger = time.time()
        elif key == "ArrowDown":
            if STATE.hdr_files and STATE.current_hdr_idx < len(STATE.hdr_files) - 1:
                STATE.current_hdr_idx += 1
                STATE._rebuild_image_variants()
                STATE._rebuild_snap_arrays()
                STATE._update_parent_options()
                STATE.selected_room_idx = None
                trigger = time.time()
        elif key.lower() == "t":
            if STATE.image_variants:
                STATE.current_variant_idx = (STATE.current_variant_idx + 1) % len(STATE.image_variants)
                trigger = time.time()
        return trigger, dv

    # Keyboard shortcut: G → toggle grid
    @app.callback(
        Output("sb-grid-toggle", "n_clicks"),
        Input("keyboard-event", "data"),
        State("sb-grid-toggle", "n_clicks"),
        prevent_initial_call=True,
    )
    def keyboard_grid_toggle(key_json, n_clicks):
        if not key_json:
            return no_update
        try:
            evt = json.loads(key_json)
        except (json.JSONDecodeError, TypeError):
            return no_update
        if evt.get("key", "").lower() == "g" and not evt.get("ctrl") and not evt.get("shift"):
            return (n_clicks or 0) + 1
        return no_update

    return app


# ── Entry Point ──────────────────────────────────────────────────────────────


def launch(project: Optional[str] = None, debug: bool = False):
    if debug:
        logging.basicConfig(
            level=logging.DEBUG, format="%(name)s %(levelname)s: %(message)s"
        )
    app = create_app(project)
    import webbrowser
    webbrowser.open("http://127.0.0.1:8050/")
    app.run(debug=True, port=8050)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Archilume Dash AOI Editor")
    parser.add_argument("--project", help="Project name")
    parser.add_argument("--debug", action="store_true", help="Enable debug logging")
    args = parser.parse_args()
    launch(project=args.project, debug=args.debug)
