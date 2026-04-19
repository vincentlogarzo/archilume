"""Export pipeline — DF analysis → WPD report → annotated overlays → ZIP archive.

Mirrors :mod:`archilume.apps.matplotlib_app`'s export outputs so both editors
produce identical artefacts for downstream WPD consumers.
"""

from __future__ import annotations

import csv
import re
import sys
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Optional

import numpy as np
from PIL import Image, ImageDraw, ImageFont
from skimage.draw import polygon as sk_polygon

from .df_analysis import DF_THRESHOLDS, compute_room_df, load_df_image
from .geometry import polygon_label_point

ProgressFn = Callable[[int, str], None]

_FALSECOLOUR_SUFFIX = "_df_false.png"
_CONTOUR_SUFFIX = "_df_cntr.png"


# =========================================================================
# Public API
# =========================================================================

def export_report(
    rooms: list[dict],
    hdr_files: list[dict],
    hdr_view_params: dict[str, list[float]],
    image_dir: Path,
    wpd_dir: Path,
    archive_dir: Path,
    outputs_dir: Path,
    project_name: str,
    inputs_dir: Optional[Path] = None,
    iesve_mode: bool = False,
    on_progress: Optional[ProgressFn] = None,
) -> Optional[Path]:
    """Full export pipeline: DF stats → Excel + CSVs → annotated PNGs → ZIP.

    Writes to:
        wpd_dir / aoi_report_daylight.xlsx
        wpd_dir / aoi_pixel_data / <room>_pixels.csv
        image_dir / <hdr>_df_false_aoi_annotated.png
        image_dir / <hdr>_df_cntr_aoi_annotated.png
        archive_dir / <project>_<timestamp>.zip
    """
    _progress(on_progress, 0, "Starting export...")
    wpd_dir.mkdir(parents=True, exist_ok=True)
    image_dir.mkdir(parents=True, exist_ok=True)
    archive_dir.mkdir(parents=True, exist_ok=True)

    hdr_groups = _group_rooms_by_hdr(rooms, hdr_files, hdr_view_params)
    if not hdr_groups:
        _progress(on_progress, 100, "Nothing to export (no rooms on any HDR).")
        return _create_archive(outputs_dir, archive_dir, project_name, inputs_dir)

    summary_rows: list[dict] = []
    pixel_chunks: list[tuple[str, np.ndarray, np.ndarray]] = []
    stats_by_hdr: dict[str, dict[str, dict]] = {}

    total_groups = len(hdr_groups)
    max_workers = min(4, total_groups)
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {
            pool.submit(_process_hdr, group, iesve_mode): group["hdr_name"]
            for group in hdr_groups.values()
        }
        done = 0
        for fut in as_completed(futures):
            hdr_name = futures[fut]
            rows, pixels, stats_by_room = fut.result()
            summary_rows.extend(rows)
            pixel_chunks.extend(pixels)
            stats_by_hdr[hdr_name] = stats_by_room
            done += 1
            pct = int(done / total_groups * 60)
            _progress(on_progress, pct, f"Processed {hdr_name}")

    _progress(on_progress, 65, "Writing Excel report...")
    _write_excel(summary_rows, wpd_dir)

    _progress(on_progress, 75, "Writing per-room CSVs...")
    _write_pixel_csvs(pixel_chunks, wpd_dir / "aoi_pixel_data")

    _progress(on_progress, 85, "Rendering annotated overlays...")
    _render_annotated_overlays(hdr_groups, stats_by_hdr, image_dir)

    _progress(on_progress, 95, "Creating archive...")
    zip_path = _create_archive(outputs_dir, archive_dir, project_name, inputs_dir)

    _progress(on_progress, 100, "Export complete.")
    return zip_path


def extract_archive(zip_path: Path, target_dir: Path) -> bool:
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            zf.extractall(target_dir)
        return True
    except Exception:
        return False


def list_archives(archive_dir: Path) -> list[str]:
    if not archive_dir.exists():
        return []
    return sorted([p.name for p in archive_dir.glob("*.zip")], reverse=True)


# =========================================================================
# Phase 1 — room grouping + HDR-native reprojection
# =========================================================================

def _group_rooms_by_hdr(
    rooms: list[dict],
    hdr_files: list[dict],
    hdr_view_params: dict[str, list[float]],
) -> dict[str, dict]:
    """Group valid rooms by HDR, reproject to HDR-native pixel coords.

    Returns ``{hdr_name: {hdr_name, hdr_path, area_per_pixel_m2, rooms: [...]}}``
    where each room entry has the fields ``_process_hdr`` + ``_draw_annotations``
    need. Rooms with <3 vertices or no matching HDR are skipped.
    """
    hdr_lookup = {h["name"]: h for h in hdr_files}

    child_verts_by_parent: dict[str, list[list[list[float]]]] = {}
    for room in rooms:
        parent = room.get("parent") or ""
        verts = room.get("vertices") or []
        if parent and len(verts) >= 3:
            child_verts_by_parent.setdefault(parent, [])

    groups: dict[str, dict] = {}
    for room in rooms:
        verts = room.get("vertices") or []
        hdr_name = room.get("hdr_file") or ""
        if len(verts) < 3 or hdr_name not in hdr_lookup:
            continue

        hdr_entry = hdr_lookup[hdr_name]
        group = groups.get(hdr_name)
        if group is None:
            vp_params = hdr_view_params.get(hdr_name)
            area_per_pixel_m2 = 0.0
            if vp_params and len(vp_params) >= 6:
                _vpx, _vpy, vh, vv, iw, ih = vp_params[:6]
                if iw > 0 and ih > 0:
                    area_per_pixel_m2 = (vh / iw) * (vv / ih)
            group = {
                "hdr_name": hdr_name,
                "hdr_path": Path(hdr_entry["hdr_path"]),
                "area_per_pixel_m2": area_per_pixel_m2,
                "vp_params": vp_params,
                "rooms": [],
            }
            groups[hdr_name] = group

        hdr_verts = _hdr_native_verts(room, group["vp_params"])
        group["rooms"].append({
            "name": room.get("name") or "unnamed",
            "parent": room.get("parent") or "",
            "room_type": (room.get("room_type") or "NONE") or "NONE",
            "hdr_verts": hdr_verts,
        })

    # Second pass: collect child hdr_verts grouped by parent name per HDR.
    for hdr_name, group in groups.items():
        by_parent: dict[str, list[list[list[float]]]] = {}
        for r in group["rooms"]:
            if r["parent"]:
                by_parent.setdefault(r["parent"], []).append(r["hdr_verts"])
        for r in group["rooms"]:
            r["child_hdr_verts"] = by_parent.get(r["name"], [])

    return groups


def _hdr_native_verts(
    room: dict, vp_params: Optional[list[float]]
) -> list[list[float]]:
    """Reproject room world_vertices → HDR pixel coords.

    Falls back to stored pixel ``vertices`` when ``world_vertices`` or view
    params are missing. Mirrors [editor_state.py:4057-4066].
    """
    world_verts = room.get("world_vertices") or []
    if vp_params and len(world_verts) >= 3 and len(vp_params) >= 6:
        vpx, vpy, vh, vv, iw, ih = vp_params[:6]
        if iw > 0 and ih > 0:
            return [
                [(wx - vpx) / (vh / iw) + iw / 2,
                 ih / 2 - (wy - vpy) / (vv / ih)]
                for wx, wy in world_verts
            ]
    return [list(v) for v in room.get("vertices") or []]


# =========================================================================
# Phase 2 — per-HDR DF compute
# =========================================================================

def _process_hdr(
    group: dict, iesve_mode: bool,
) -> tuple[
    list[dict],
    list[tuple[str, np.ndarray, np.ndarray]],
    dict[str, dict],
]:
    """Compute per-room DF stats + extract raw pixel arrays for one HDR.

    Returns ``(summary_rows, pixel_chunks, stats_by_room)``.
    Summary rows and CSV chunks are only produced for rooms whose room_type
    has a DF threshold in ``DF_THRESHOLDS`` (BED/LIVING/NON-RESI). Full stats
    dicts are returned for every room that successfully computes DF so the
    overlay renderer can access ``area_num``/``area_den``/``area_pct``.
    """
    hdr_name = group["hdr_name"]
    hdr_path = group["hdr_path"]
    area_per_pixel_m2 = group["area_per_pixel_m2"]

    df_image = load_df_image(hdr_path)
    if df_image is None:
        return [], [], {}

    summary_rows: list[dict] = []
    pixel_chunks: list[tuple[str, np.ndarray, np.ndarray]] = []
    stats_by_room: dict[str, dict] = {}

    for room in group["rooms"]:
        name = room["name"]
        room_type = room["room_type"]
        verts = room["hdr_verts"]
        child_verts = room["child_hdr_verts"]

        stats = compute_room_df(
            df_image,
            verts,
            room_type=room_type,
            area_per_pixel_m2=area_per_pixel_m2,
            exclude_polygons=child_verts,
        )
        if stats is None:
            continue
        stats_by_room[name] = stats

        threshold = stats.get("threshold")
        if threshold is None:
            # Untyped / CIRC / NONE — no WPD output per user spec
            continue

        df_vals, lux_vals = _extract_room_pixels(df_image, verts, child_verts)
        if df_vals.size == 0:
            continue

        total_pixels = int(df_vals.size)
        passing = int(np.sum(df_vals >= threshold))
        passing_pct = round(passing / total_pixels * 100.0, 1)

        summary_rows.append({
            "HDR File": hdr_name,
            "Parent": room["parent"],
            "Room": name,
            "Room Type": room_type if room_type != "NONE" else "",
            "Total Pixels": total_pixels,
            "DF Threshold (%)": threshold,
            "Pixels >= Threshold": passing,
            "% Area >= Threshold": passing_pct,
        })
        pixel_chunks.append(
            (name, np.round(lux_vals, 2), np.round(df_vals, 4))
        )

    return summary_rows, pixel_chunks, stats_by_room


def _extract_room_pixels(
    df_image: np.ndarray,
    vertices: list[list[float]],
    exclude_polygons: list[list[list[float]]],
) -> tuple[np.ndarray, np.ndarray]:
    """Return (df_vals, lux_vals) for pixels inside ``vertices`` minus children.

    ``df_image`` is already in DF% units (see :func:`load_df_image`).
    Lux = DF% × 100 (CIE overcast reference 10 000 lux).
    """
    h, w = df_image.shape[:2]
    xs = [int(round(v[0])) for v in vertices]
    ys = [int(round(v[1])) for v in vertices]
    rr, cc = sk_polygon(ys, xs, shape=(h, w))
    if len(rr) == 0:
        return np.array([], dtype=np.float32), np.array([], dtype=np.float32)

    if exclude_polygons:
        mask = np.zeros((h, w), dtype=bool)
        mask[rr, cc] = True
        for child in exclude_polygons:
            if len(child) < 3:
                continue
            cxs = [int(round(v[0])) for v in child]
            cys = [int(round(v[1])) for v in child]
            crr, ccc = sk_polygon(cys, cxs, shape=(h, w))
            mask[crr, ccc] = False
        rr, cc = np.where(mask)
        if len(rr) == 0:
            return (
                np.array([], dtype=np.float32),
                np.array([], dtype=np.float32),
            )

    df_vals = df_image[rr, cc]
    lux_vals = df_vals * 100.0
    return df_vals, lux_vals


# =========================================================================
# Phase 3 — Excel + CSVs
# =========================================================================

def _write_excel(summary_rows: list[dict], wpd_dir: Path) -> Path:
    """Write the per-room summary Excel with an auto-fitted structured table."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    out_path = wpd_dir / "aoi_report_daylight.xlsx"
    df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame()
    df.to_excel(out_path, sheet_name="Room Summary", index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    for col in ws.columns:
        max_len = max(
            len(str(c.value)) if c.value is not None else 0 for c in col
        )
        ws.column_dimensions[col[0].column_letter].width = max_len + 4
    if ws.max_row > 1:
        last_col = get_column_letter(ws.max_column)
        table = Table(
            displayName="RoomSummary", ref=f"A1:{last_col}{ws.max_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    wb.save(out_path)
    return out_path


def _write_pixel_csvs(
    pixel_chunks: list[tuple[str, np.ndarray, np.ndarray]],
    csv_subdir: Path,
) -> None:
    """Write one CSV per applicable room with [Room, Lux, DF%] columns."""
    if not pixel_chunks:
        return
    csv_subdir.mkdir(parents=True, exist_ok=True)
    for room_name, lux_vals, df_vals in pixel_chunks:
        safe = re.sub(r"[\\/:*?\"<>|]", "_", room_name)
        out_path = csv_subdir / f"{safe}_pixels.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Room", "Illuminance (Lux)", "Daylight Factor (%)"])
            for lux, dfp in zip(lux_vals, df_vals):
                writer.writerow([room_name, float(lux), float(dfp)])


# =========================================================================
# Phase 4 — annotated overlays
# =========================================================================

def _render_annotated_overlays(
    hdr_groups: dict[str, dict],
    stats_by_hdr: dict[str, dict[str, dict]],
    image_dir: Path,
) -> None:
    """Render annotated falsecolour + contour PNGs for every HDR with rooms.

    Uses existing ``<hdr>_df_false.png`` / ``<hdr>_df_cntr.png`` as the base
    canvas — we do not regenerate falsecolour/contour content here.
    """
    jobs: list[tuple[Path, list[dict], dict[str, dict], Path]] = []
    for hdr_name, group in hdr_groups.items():
        hdr_stem = group["hdr_path"].stem
        rooms_data = [
            {
                "name": r["name"],
                "verts": r["hdr_verts"],
                "is_circ": r["room_type"] == "CIRC",
            }
            for r in group["rooms"]
        ]
        stats_map = stats_by_hdr.get(hdr_name, {})
        for suffix in (_FALSECOLOUR_SUFFIX, _CONTOUR_SUFFIX):
            base_png = image_dir / f"{hdr_stem}{suffix}"
            if not base_png.exists():
                continue
            out_stem = base_png.stem + "_aoi_annotated"
            out_path = image_dir / f"{out_stem}.png"
            jobs.append((base_png, rooms_data, stats_map, out_path))

    if not jobs:
        return
    max_workers = min(4, len(jobs))
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [
            pool.submit(_draw_annotations, *job) for job in jobs
        ]
        for fut in as_completed(futures):
            fut.result()


_EXPORT_UPSCALE = 2.5


def _draw_annotations(
    base_png_path: Path,
    rooms_data: list[dict],
    stats_by_room: dict[str, dict],
    out_path: Path,
) -> None:
    """Draw room boundaries + DF annotations onto a falsecolour/contour PNG.

    Matches the on-screen Reflex viewport layout:
      * Non-CIRC rooms: solid red outline + stacked area fraction (num / den)
        beside "m² (pct)", coloured by pass/marginal/fail thresholds, followed
        by the `≥ N% DF` target line and room name.
      * CIRC rooms: diagonal red hatch fill + dashed red outline (no labels),
        mirroring the SVG ``#hatch-circ`` pattern in ``viewport.py``.

    The base PNG is upscaled (Lanczos) before drawing so annotations render at
    enough pixel density for the superscript ``²`` and ``≥`` glyphs to read
    cleanly instead of collapsing into tofu boxes.
    """
    base = Image.open(base_png_path).convert("RGBA")
    base_w, base_h = base.size
    scale = _EXPORT_UPSCALE
    up_w = max(1, int(round(base_w * scale)))
    up_h = max(1, int(round(base_h * scale)))
    if (up_w, up_h) != (base_w, base_h):
        img = base.resize((up_w, up_h), Image.LANCZOS)
    else:
        img = base
    w, h = img.size

    fs_large = max(14, int(h * 0.025))
    fs_small = max(10, int(fs_large * 6.5 / 8.5))
    font_large = _load_font(fs_large, bold=True)
    font_large_norm = _load_font(fs_large, bold=False)
    font_small = _load_font(fs_small, bold=False)

    red = (255, 0, 0)
    white = (255, 255, 255)
    black = (0, 0, 0)
    line_w = max(1, int(round(h * 0.003)))

    # Scale room polygons from base PNG coords → upscaled canvas coords.
    scaled_rooms = [
        {
            **room,
            "verts": [[v[0] * scale, v[1] * scale] for v in room["verts"]],
        }
        for room in rooms_data
    ]

    # CIRC hatch pass first so outlines draw on top.
    for room in scaled_rooms:
        if not room["is_circ"]:
            continue
        _hatch_polygon(img, room["verts"], red, spacing=max(6, fs_small),
                       line_w=max(1, line_w // 2))

    # Outlines + annotations.
    draw = ImageDraw.Draw(img)
    for room in scaled_rooms:
        name = room["name"]
        verts = room["verts"]
        is_circ = room["is_circ"]

        pts = [(int(round(v[0])), int(round(v[1]))) for v in verts]
        pts.append(pts[0])

        if is_circ:
            _dashed_polygon(draw, pts, fill=red, width=line_w)
            continue
        draw.line(pts, fill=red, width=line_w)

        stats = stats_by_room.get(name) or {}
        _draw_room_annotation(
            draw, verts, name, stats,
            fs_large, fs_small,
            font_large, font_large_norm, font_small,
            white, black,
        )

    if out_path.exists():
        out_path.unlink()
    img.convert("RGB").save(out_path, "PNG")


def _draw_room_annotation(
    draw: ImageDraw.ImageDraw,
    verts: list[list[float]],
    name: str,
    stats: dict,
    fs_large: int,
    fs_small: int,
    font_large: Any,
    font_large_norm: Any,
    font_small: Any,
    white: tuple,
    black: tuple,
) -> None:
    """Render stacked-fraction DF annotation + threshold + room name for one room."""
    area_num = str(stats.get("area_num") or "")
    area_den = str(stats.get("area_den") or "")
    area_pct = str(stats.get("area_pct") or "")
    result_lines = list(stats.get("result_lines") or [])
    threshold_line = result_lines[1] if len(result_lines) >= 2 else ""

    # Colour — derived from pct_above so untyped rooms fall back to black.
    colour = _pct_colour_from_stats(stats, black)
    no_outline = colour == black
    area_font = font_large if no_outline else font_large_norm
    area_stroke_fg = white if no_outline else black
    area_stroke_lw = fs_large * 0.015 if no_outline else fs_large * 0.03

    pole = polygon_label_point(verts)

    # Measure stacked fraction + unit/pct.
    bar_gap = max(2, int(fs_large * 0.22))
    bar_thick = max(1, int(fs_large * 0.08))
    num_w, num_h = _text_size(draw, area_num, area_font)
    den_w, den_h = _text_size(draw, area_den, area_font)
    unit_pct_text = f" m\u00b2  {area_pct}".rstrip() if area_pct else " m\u00b2"
    has_fraction = bool(area_num and area_den)

    if has_fraction:
        col_w = max(num_w, den_w, 1)
        unit_w, _unit_h = _text_size(draw, unit_pct_text, area_font)
        line0_w = col_w + int(fs_large * 0.35) + unit_w
        line0_h = num_h + bar_gap + bar_thick + bar_gap + den_h
    else:
        # Fallback: render whatever result_lines[0] gave us.
        fallback = result_lines[0] if result_lines else ""
        line0_w, line0_h = _text_size(draw, fallback, area_font)

    line_gap = max(1, int(fs_small * 0.4))
    small_line_h = fs_small
    has_threshold = bool(threshold_line)
    name_w, _name_h = _text_size(draw, name, font_small)
    threshold_w, _t_h = (
        _text_size(draw, threshold_line, font_small) if has_threshold else (0, 0)
    )

    block_w = max(line0_w, threshold_w, name_w)
    block_h = line0_h
    if has_threshold:
        block_h += line_gap + small_line_h
    block_h += line_gap + small_line_h  # room name

    ax, ay = _clamp_block_to_bbox(pole, verts, block_w, block_h)

    # --- Line 0 ---
    cursor_y = ay
    if has_fraction:
        col_w = max(num_w, den_w, 1)
        unit_w, unit_h = _text_size(draw, unit_pct_text, area_font)
        block_total = col_w + int(fs_large * 0.35) + unit_w
        start_x = ax + (block_w - block_total) / 2
        col_cx = start_x + col_w / 2
        # Numerator (top)
        _stroked_text(
            draw, int(round(col_cx - num_w / 2)), int(round(cursor_y)),
            area_num, area_font, colour, area_stroke_fg, area_stroke_lw,
        )
        bar_y = cursor_y + num_h + bar_gap
        bar_x0 = int(round(start_x))
        bar_x1 = int(round(start_x + col_w))
        # Bar — stroked so it stays visible on light backgrounds
        draw.rectangle(
            [(bar_x0, int(round(bar_y))),
             (bar_x1, int(round(bar_y + bar_thick)))],
            fill=colour,
        )
        # Denominator
        den_y = bar_y + bar_thick + bar_gap
        _stroked_text(
            draw, int(round(col_cx - den_w / 2)), int(round(den_y)),
            area_den, area_font, colour, area_stroke_fg, area_stroke_lw,
        )
        # "m² (pct)" — vertically centred on bar
        unit_x = start_x + col_w + int(fs_large * 0.35)
        unit_y = bar_y + bar_thick / 2 - unit_h / 2
        _stroked_text(
            draw, int(round(unit_x)), int(round(unit_y)),
            unit_pct_text, area_font, colour, area_stroke_fg, area_stroke_lw,
        )
    elif result_lines:
        fallback = result_lines[0]
        fb_w, _ = _text_size(draw, fallback, area_font)
        _stroked_text(
            draw,
            int(round(ax + (block_w - fb_w) / 2)), int(round(cursor_y)),
            fallback, area_font, colour, area_stroke_fg, area_stroke_lw,
        )
    cursor_y += line0_h

    # --- Threshold line ---
    if has_threshold:
        cursor_y += line_gap
        _stroked_text(
            draw,
            int(round(ax + (block_w - threshold_w) / 2)), int(round(cursor_y)),
            threshold_line, font_small, white, black, fs_small * 0.03,
        )
        cursor_y += small_line_h

    # --- Room name ---
    cursor_y += line_gap
    _stroked_text(
        draw,
        int(round(ax + (block_w - name_w) / 2)), int(round(cursor_y)),
        name, font_small, white, black, fs_small * 0.03,
    )


def _hatch_polygon(
    img: Image.Image,
    verts: list[list[float]],
    colour: tuple,
    spacing: int,
    line_w: int,
) -> None:
    """Draw diagonal hatch lines (x + y = k) clipped to a polygon.

    Mirrors the SVG ``#hatch-circ`` pattern in ``viewport.py``: a 6×6 cell
    with a single line from (0,6)→(6,0), tiled across the polygon.

    The paste mask is ``polygon_mask ∧ line_alpha`` so only the line pixels
    inside the polygon replace base pixels — off-line interior pixels stay
    untouched (no black fill behind the hatch).
    """
    from PIL import ImageChops

    w, h = img.size
    poly_mask = Image.new("L", (w, h), 0)
    poly = [(int(round(v[0])), int(round(v[1]))) for v in verts]
    ImageDraw.Draw(poly_mask).polygon(poly, fill=255)

    hatch = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    hd = ImageDraw.Draw(hatch)
    colour_rgba = colour if len(colour) == 4 else colour + (255,)
    for k in range(0, w + h + 1, spacing):
        x0 = min(k, w)
        y0 = max(0, k - w)
        x1 = max(0, k - h)
        y1 = min(k, h)
        hd.line([(x0, y0), (x1, y1)], fill=colour_rgba, width=line_w)

    line_alpha = hatch.split()[3]
    combined_mask = ImageChops.multiply(line_alpha, poly_mask)
    img.paste(hatch, (0, 0), combined_mask)


def _clamp_block_to_bbox(
    pole: tuple[float, float],
    verts: list[list[float]],
    block_w: float,
    block_h: float,
) -> tuple[float, float]:
    """Return top-left (ax, ay) centred on pole, clamped inside the polygon bbox."""
    xs = [v[0] for v in verts]
    ys = [v[1] for v in verts]
    xmin, xmax = min(xs), max(xs)
    ymin, ymax = min(ys), max(ys)
    ax = pole[0] - block_w / 2.0
    ay = pole[1] - block_h / 2.0
    ax = max(xmin, min(ax, xmax - block_w))
    ay = max(ymin, min(ay, ymax - block_h))
    return ax, ay


def _text_size(
    draw: ImageDraw.ImageDraw, text: str, font: Any
) -> tuple[int, int]:
    """Return (width, height) of rendered text using anchor='lt'.

    With the left-top anchor the bbox starts at (0, 0), so ``bottom`` is the
    true pixel height of the glyph box — critical for aligning stacked
    numerator/denominator with the divider bar.
    """
    if not text:
        return 0, 0
    try:
        left, top, right, bottom = draw.textbbox(
            (0, 0), text, font=font, anchor="lt",
        )
        return right - left, bottom - top
    except (AttributeError, TypeError, ValueError):
        try:
            left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
            return right - left, bottom - top
        except AttributeError:
            w, h = draw.textsize(text, font=font)  # type: ignore[attr-defined]
            return w, h


def _pct_colour_from_stats(stats: dict, default: tuple) -> tuple:
    """Colour-code annotation by pct_above (matches Reflex viewport logic)."""
    pct = stats.get("pct_above")
    threshold = stats.get("threshold")
    if pct is None or threshold is None:
        return default
    if pct >= 90:
        return (0, 0, 0)
    if pct >= 50:
        return (233, 113, 50)
    return (238, 0, 0)


# =========================================================================
# PIL helpers (ported from matplotlib_app._render_pdf_underlay)
# =========================================================================

def _stroked_text(
    draw: ImageDraw.ImageDraw,
    x: int,
    y: int,
    text: str,
    font: ImageFont.ImageFont,
    fill: tuple,
    stroke_fg: tuple,
    stroke_lw: float,
) -> None:
    """Draw text with a stroke outline. ``y`` is the top pixel of the glyph box.

    Uses ``anchor="lt"`` (left-top) so the caller's y-coordinate is the actual
    top pixel of the rendered text — removes the ascender drift from PIL's
    default ``anchor="la"`` which otherwise pushes digits up and causes them
    to collide with the stacked-fraction divider bar.
    """
    sw = max(1, int(round(stroke_lw)))
    kwargs = {"font": font}
    try:
        draw.text((0, 0), "", fill=fill, font=font, anchor="lt")
        kwargs["anchor"] = "lt"
    except (TypeError, ValueError):
        # Bitmap default font — anchor unsupported; fall back silently.
        pass
    for dx in range(-sw, sw + 1):
        for dy in range(-sw, sw + 1):
            if dx or dy:
                draw.text((x + dx, y + dy), text, fill=stroke_fg, **kwargs)
    draw.text((x, y), text, fill=fill, **kwargs)


def _dashed_polygon(
    draw: ImageDraw.ImageDraw,
    pts_closed: list[tuple[int, int]],
    fill: tuple,
    width: int,
    dash: int = 8,
    gap: int = 6,
) -> None:
    for i in range(len(pts_closed) - 1):
        x0, y0 = pts_closed[i]
        x1, y1 = pts_closed[i + 1]
        seg_len = ((x1 - x0) ** 2 + (y1 - y0) ** 2) ** 0.5
        if seg_len == 0:
            continue
        dx, dy = (x1 - x0) / seg_len, (y1 - y0) / seg_len
        pos, drawing = 0.0, True
        while pos < seg_len:
            step = dash if drawing else gap
            end = min(pos + step, seg_len)
            if drawing:
                draw.line(
                    [(int(x0 + dx * pos), int(y0 + dy * pos)),
                     (int(x0 + dx * end), int(y0 + dy * end))],
                    fill=fill, width=width,
                )
            pos, drawing = end, not drawing


# =========================================================================
# Font loading — prefer a scalable TTF so font sizes honour image height
# =========================================================================

_FONT_CANDIDATES: dict[str, list[str]] = {
    "regular": [
        "arial.ttf",
        "DejaVuSans.ttf",
        "Arial.ttf",
        "tahoma.ttf",
    ],
    "bold": [
        "arialbd.ttf",
        "DejaVuSans-Bold.ttf",
        "Arial-Bold.ttf",
        "tahomabd.ttf",
    ],
}


def _matplotlib_font_dir() -> Optional[Path]:
    """Path to matplotlib's bundled TTF fonts, or None if unavailable.

    Reliable cross-platform source for DejaVu Sans in containerised deploys
    (e.g. ``python:3.12-slim`` which ships no system fonts). matplotlib is
    already a project dependency and its ``mpl-data/fonts/ttf`` directory
    ships ``DejaVuSans.ttf`` / ``DejaVuSans-Bold.ttf`` — both cover ``²``
    (U+00B2) and ``≥`` (U+2265), so annotations render correctly with no
    apt-get / Dockerfile change.
    """
    try:
        import matplotlib  # type: ignore
        d = Path(matplotlib.__file__).parent / "mpl-data" / "fonts" / "ttf"
        return d if d.exists() else None
    except Exception:
        return None


def _platform_font_dirs() -> list[Path]:
    dirs: list[Path] = []
    if sys.platform == "win32":
        import os
        windir = os.environ.get("WINDIR", "C:/Windows")
        dirs.append(Path(windir) / "Fonts")
    elif sys.platform == "darwin":
        dirs += [Path("/System/Library/Fonts"), Path("/Library/Fonts")]
    else:
        dirs += [
            Path("/usr/share/fonts/truetype/dejavu"),
            Path("/usr/share/fonts/truetype"),
            Path("/usr/share/fonts"),
        ]
    mpl_dir = _matplotlib_font_dir()
    if mpl_dir is not None:
        dirs.append(mpl_dir)
    return dirs


_FONT_CACHE: dict[tuple[int, bool], Any] = {}


def _load_font(size: int, bold: bool = False) -> Any:
    """Load a scalable TTF at ``size`` px. Falls back to PIL's default font.

    Tries PIL's short-name resolution first (freetype auto-searches the
    platform's font directory), then explicit paths, then bitmap fallback.
    """
    key = (size, bold)
    cached = _FONT_CACHE.get(key)
    if cached is not None:
        return cached

    candidates = _FONT_CANDIDATES["bold" if bold else "regular"]

    # Short-name lookup first — PIL/freetype searches Windows/Linux/mac font dirs.
    for name in candidates:
        try:
            f = ImageFont.truetype(name, size)
            _FONT_CACHE[key] = f
            return f
        except (OSError, IOError):
            continue

    # Explicit platform paths as a defensive fallback.
    for font_dir in _platform_font_dirs():
        if not font_dir.exists():
            continue
        for name in candidates:
            p = font_dir / name
            if p.exists():
                try:
                    f = ImageFont.truetype(str(p), size)
                    _FONT_CACHE[key] = f
                    return f
                except Exception:
                    continue

    # Bitmap fallback — PIL ≥10 supports size on load_default; older ignore it.
    try:
        f = ImageFont.load_default(size=size)  # type: ignore[call-arg]
    except TypeError:
        f = ImageFont.load_default()
    _FONT_CACHE[key] = f
    return f


# =========================================================================
# Archive
# =========================================================================

def _create_archive(
    outputs_dir: Path,
    archive_dir: Path,
    project_name: str,
    inputs_dir: Optional[Path] = None,
) -> Optional[Path]:
    """Archive both ``inputs/`` and ``outputs/`` so a user can restore state.

    Stored paths are relative to the project root (``outputs_dir.parent``),
    so extraction back into the project directory repopulates both trees in
    place.
    """
    if not outputs_dir.exists() and (inputs_dir is None or not inputs_dir.exists()):
        return None
    archive_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = f"_{project_name}" if project_name else ""
    stem = f"archilume_export{suffix}_{timestamp}"
    zip_path = archive_dir / f"{stem}.zip"
    project_root = outputs_dir.parent

    sources: list[Path] = []
    if inputs_dir is not None and inputs_dir.exists():
        sources.append(inputs_dir)
    if outputs_dir.exists():
        sources.append(outputs_dir)

    try:
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for src in sources:
                for p in src.rglob("*"):
                    if p.is_file():
                        zf.write(p, p.relative_to(project_root))
        return zip_path
    except Exception:
        if zip_path.exists():
            try:
                zip_path.unlink()
            except OSError:
                pass
        return None


# =========================================================================
# Misc
# =========================================================================

def _progress(fn: Optional[ProgressFn], pct: int, msg: str) -> None:
    if fn is not None:
        try:
            fn(pct, msg)
        except Exception:
            pass
