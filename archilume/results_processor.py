"""Results Processor for Working Plan Data (WPD) Generation.

This module handles the extraction and processing of illuminance data from HDR images,
filtering points within Areas of Interest (AOI), and generating compliance reports.
"""

# Archilume imports
from archilume.utils import get_hdr_resolution

# Standard library imports
import subprocess
import sys
import math
import numpy as np
from matplotlib.path import Path as MplPath
from pathlib import Path
import pandas as pd
from concurrent.futures import ProcessPoolExecutor, as_completed
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict
import os
import logging

# Third-party imports
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import ColorScaleRule

logger = logging.getLogger(__name__)

@dataclass
class ResultsProcessor:
    """Main processor for generating Working Plan Data from HDR images.

    This class handles:
    - Extracting point data from HDR images using pvalue
    - Filtering non-zero illuminance points
    - Matching points to AOI polygons using ray casting
    - Writing individual .wpd files per timestep
    - Aggregating results into Excel reports
    """

    # Input parameters
    image_dir: Path
    aoi_dir: Path
    wpd_dir: Path
    pixel_to_world_map: Path
    pixel_threshold_value: float = 0.0
    max_workers: int = 12
    

    def __post_init__(self):
        """Initialize output directories."""
        self.wpd_dir.mkdir(parents=True, exist_ok=True)

        # Calculate area per pixel and pixel increments from the pixel_to_world_map file
        self.area_per_pixel, self.pixel_increment_x, self.pixel_increment_y = self._calculate_area_per_pixel()

    def nsw_adg_sunlight_sequence_wpd_extraction(self) -> None:
        """Execute the complete WPD generation pipeline.

        Phases:
        1. Scan directories for HDR and AOI files
        2. Group files by view
        3. Process view groups in parallel
        4. Generate Excel report from .wpd files
        """
        print("=" * 80 + "\nPARALLEL BATCH PROCESSING (ProcessPoolExecutor)\n" + "=" * 80)

        # Scan directories
        hdr_files_all, aoi_files_all = self._scan_directories()

        if not hdr_files_all or not aoi_files_all:
            print("No files found to process.")
            sys.exit(0)

        # Group files by view
        view_groups = self._group_files_by_view(aoi_files_all, hdr_files_all)

        # Process view groups in parallel
        self._process_view_groups_parallel(view_groups)

        # Generate Excel report
        self._generate_excel_report()

        print("\n" + "=" * 80 + "\nSunlight sequence wpd extraction completed successfully.\n" + "=" * 80)

    def _calculate_area_per_pixel(self) -> tuple[float, float, float]:
        """Calculate the area represented by a single pixel from the pixel_to_world_map file.

        Reads the header metadata from the pixel_to_world_map file to extract:
        - Image dimensions in pixels (width, height)
        - World dimensions in meters (width, height)

        Then calculates:
        - pixel_width = world_width / image_width
        - pixel_height = world_height / image_height
        - area_per_pixel = pixel_width × pixel_height

        Returns:
            Tuple of (area_per_pixel, pixel_increment_x, pixel_increment_y) in m² and mm²
        """
        with open(self.pixel_to_world_map, 'r') as f:
            lines = f.readlines()

            # Parse image dimensions from line 2
            # Format: # Image dimensions in pixels: width=2048, height=1778
            image_dims_line = lines[1].strip()
            image_width = int(image_dims_line.split('width=')[1].split(',')[0])
            image_height = int(image_dims_line.split('height=')[1])

            # Parse world dimensions from line 3
            # Format: # World dimensions in meters: width=29.480000, height=25.590000
            world_dims_line = lines[2].strip()
            world_width = float(world_dims_line.split('width=')[1].split(',')[0])
            world_height = float(world_dims_line.split('height=')[1])

        # Calculate pixel spacing in each direction
        pixel_width_meters = world_width / image_width
        pixel_height_meters = world_height / image_height

        # Calculate area per pixel
        area_per_pixel = pixel_width_meters * pixel_height_meters

        # Round to 4 decimal places
        area_per_pixel_rounded = round(area_per_pixel, 4)

        # Convert to mm (1 m = 1,000 mm) and mm² (1 m² = 1,000,000 mm²)
        pixel_width_mm = round(pixel_width_meters * 1_000)
        pixel_height_mm = round(pixel_height_meters * 1_000)
        area_per_pixel_mm2 = area_per_pixel_rounded * 1_000_000

        print(f"\nPixel-to-World Mapping:")
        print(f"  Image dimensions: {image_width} × {image_height} pixels")
        print(f"  World dimensions: {world_width} × {world_height} meters")
        print(f"  Pixel increment X: {pixel_width_meters:.6f} m ({pixel_width_mm} mm)")
        print(f"  Pixel increment Y: {pixel_height_meters:.6f} m ({pixel_height_mm} mm)")
        print(f"  Area per pixel (unrounded): {area_per_pixel:.10f} m²")
        print(f"  Area per pixel (rounded): {area_per_pixel_rounded} m² ({area_per_pixel_mm2:.2f} mm²)")
        print(f"  Verification: {image_width} × {image_height} = {image_width * image_height} pixels")
        print(f"  Total world area: {world_width * world_height:.4f} m²")
        print(f"  Calculated area: {area_per_pixel_rounded * image_width * image_height:.4f} m²")
        print(f"  Source file: {self.pixel_to_world_map}\n")

        return area_per_pixel_rounded, pixel_width_meters, pixel_height_meters

    def _scan_directories(self) -> Tuple[List[Path], List[Path]]:
        """Scan directories for HDR and AOI files.

        Returns:
            Tuple of (hdr_files, aoi_files)
        """
        print("Scanning directories...")
        hdr_files_all = sorted(self.image_dir.glob("*.hdr"))
        aoi_files_all = sorted(self.aoi_dir.glob("*.aoi"))

        print(f"Found {len(hdr_files_all)} HDR files")
        print(f"Found {len(aoi_files_all)} AOI files\n")

        return hdr_files_all, aoi_files_all

    def _get_associated_view_file(self, aoi_file: Path) -> Optional[str]:
        """Extract the associated view file name from an AOI file.

        Args:
            aoi_file: Path to AOI file

        Returns:
            View name (e.g., "plan_L02") or None if not found
        """
        with open(aoi_file, 'r') as f:
            lines = f.readlines()
            if len(lines) >= 2:
                view_line = lines[1].strip()
                if 'ASSOCIATED VIEW FILE:' in view_line:
                    view_file = view_line.split('ASSOCIATED VIEW FILE:')[1].strip()
                    return view_file.replace('.vp', '')
        return None

    def _group_files_by_view(self, aoi_files: List[Path], hdr_files: List[Path]) -> Dict:
        """Group AOI files and HDR files by their associated view identifier.

        Skips AOI files that already have .wpd files.

        Args:
            aoi_files: List of AOI file paths to group
            hdr_files: List of HDR file paths to match against groups

        Returns:
            Dictionary mapping view names to {'aoi_files': [...], 'hdr_files': [...]}
        """
        groups = defaultdict(lambda: {'aoi_files': [], 'hdr_files': []})

        # Group AOI files by view, skipping existing .wpd files
        print("Grouping AOI files by associated view...")
        skipped_count = 0

        for aoi_file in aoi_files:
            # Check if .wpd file already exists
            wpd_file_path = self.wpd_dir / f"{aoi_file.stem}.wpd"
            if wpd_file_path.exists():
                print(f"  Skipping {aoi_file.name} (already has .wpd file)")
                skipped_count += 1
                continue

            view_name = self._get_associated_view_file(aoi_file)
            if view_name:
                groups[view_name]['aoi_files'].append(aoi_file)
                print(f"  {aoi_file.name} -> {view_name}")
            else:
                print(f"  Warning: Could not extract view file from {aoi_file.name}")

        if skipped_count > 0:
            print(f"\nSkipped {skipped_count} AOI files that already have .wpd files")

        # Match HDR files to groups
        print("\nMatching HDR files to view groups...")
        for hdr_file in hdr_files:
            matched = False
            for view_name in groups.keys():
                if view_name in hdr_file.name:
                    groups[view_name]['hdr_files'].append(hdr_file)
                    matched = True
                    break

            if matched:
                print(f"  {hdr_file.name} -> {view_name}")
            else:
                print(f"  Warning: {hdr_file.name} does not match any view group")

        # Print summary
        print("\n" + "=" * 80 + "\nGROUPING SUMMARY\n" + "=" * 80)
        for view_name, group in sorted(groups.items()):
            print(f"{view_name}:")
            print(f"  AOI files: {len(group['aoi_files'])}")
            print(f"  HDR files: {len(group['hdr_files'])}")
            print(f"  Total operations: {len(group['aoi_files'])} × {len(group['hdr_files'])} = "
                  f"{len(group['aoi_files']) * len(group['hdr_files'])}")

        return dict(groups)

    def _process_view_groups_parallel(self, view_groups: Dict) -> None:
        """Process view groups in parallel using ProcessPoolExecutor.

        Args:
            view_groups: Dictionary of view groups from _group_files_by_view
        """
        print("\n" + "=" * 80 + f"\nPROCESSING {len(view_groups)} VIEW GROUPS IN PARALLEL\n"
              f"Using {self.max_workers} worker processes\nMain PID: {os.getpid()}\n" + "=" * 80)

        # Prepare tasks
        tasks = [
            (view_name, group, self.pixel_threshold_value, self.wpd_dir)
            for view_name, group in view_groups.items()
        ]

        print(f"Submitting {len(tasks)} tasks to ProcessPoolExecutor...")

        all_results = []

        with ProcessPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {
                executor.submit(_process_view_group_worker, task): task[0]
                for task in tasks
            }

            print(f"Submitted {len(futures)} tasks\nProcessing in parallel...\n" + "=" * 80 + "\n")

            # Collect results as they complete
            for future in as_completed(futures):
                view_name = futures[future]
                try:
                    view_name_result, result_df = future.result()
                    all_results.append(result_df)

                    print(f"\n{'='*80}\nCompleted: {view_name_result}\n"
                          f"Progress: {len(all_results)}/{len(futures)} view groups completed\n{'='*80}\n")

                except Exception as e:
                    print(f"\nError processing {view_name}: {e}")
                    import traceback
                    traceback.print_exc()

    def _generate_excel_report(self) -> None:
        """Generate Excel report from all .wpd files."""
        print("\n" + "=" * 80 + "\nGENERATING EXCEL FILE FROM .WPD FILES\n" + "=" * 80)

        # Read all .wpd files
        wpd_files = sorted(self.wpd_dir.glob("*.wpd"))
        print(f"\nFound {len(wpd_files)} .wpd files in {self.wpd_dir}")

        if not wpd_files:
            print("\nNo .wpd files found to generate Excel file")
            return

        # Parse .wpd files
        all_wpd_data = []
        for wpd_file in wpd_files:
            aoi_name = wpd_file.stem + ".aoi"

            with open(wpd_file, 'r') as f:
                lines = f.readlines()

                # Skip header lines (first 2 lines)
                for line in lines[2:]:
                    parts = line.strip().split()
                    if len(parts) >= 2:
                        hdr_file = parts[0]
                        passing_pixels = int(parts[1])
                        total_pixels = int(lines[0].split(':')[1].strip())

                        all_wpd_data.append({
                            'aoi_file': aoi_name,
                            'hdr_file': hdr_file,
                            'total_pixels': total_pixels,
                            'passing_pixels': passing_pixels
                        })

        # Create DataFrame
        combined_results = pd.DataFrame(all_wpd_data)
        combined_results = combined_results.sort_values(['aoi_file', 'hdr_file'])

        # Add area calculation column (passing_pixels × area_per_pixel)
        combined_results['passing_area_m2'] = combined_results['passing_pixels'] * self.area_per_pixel

        # Display summary
        print(f"\nTotal combinations: {len(combined_results)}")
        print(f"Unique AOI files: {combined_results['aoi_file'].nunique()}")
        print(f"Unique HDR files: {combined_results['hdr_file'].nunique()}")
        print(f"Area per pixel: {self.area_per_pixel} m²")

        # Write Excel file
        self._write_excel_file(combined_results)

    def _write_excel_file(self, combined_results: pd.DataFrame) -> None:
        """Write formatted Excel file with raw data and pivot table.

        Args:
            combined_results: DataFrame with all results
        """
        output_excel = self.wpd_dir / "sunlight_analysis_results.xlsx"

        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # Write metadata header and raw data to first sheet
            # Metadata at B1, data starts at B4
            combined_results.to_excel(writer, sheet_name='Raw Data', index=False, startrow=3, startcol=1)

            # Add metadata to Raw Data sheet
            ws_raw = writer.sheets['Raw Data']
            area_per_pixel_mm2 = self.area_per_pixel * 1_000_000
            pixel_increment_x_mm = round(self.pixel_increment_x * 1_000)
            pixel_increment_y_mm = round(self.pixel_increment_y * 1_000)
            ws_raw['B1'] = (f"{pixel_increment_x_mm} mm × {pixel_increment_y_mm} mm grid with an area per pixel of {self.area_per_pixel} m² ({area_per_pixel_mm2:.2f} mm²) | "
                           f"Source: {self.pixel_to_world_map}")

            # Transform HDR filenames for pivot
            combined_results['hdr_file'] = combined_results['hdr_file'].str.split('_SS_').str[1]

            # Create pivot table with passing_area_m2
            pivot_data = combined_results.pivot_table(
                values='passing_area_m2',
                index='aoi_file',
                columns='hdr_file',
                aggfunc='sum',
                fill_value=0
            )

            # Calculate timestep duration from column names (HDR filenames)
            # Assumes format like "0621_0900.hdr", "0621_1000.hdr" etc.
            if len(pivot_data.columns) >= 2:
                # Parse times from first two timesteps
                col1 = pivot_data.columns[0]
                col2 = pivot_data.columns[1]

                # Extract time portion (last 4 digits before .hdr)
                time1_str = col1.split('_')[-1].replace('.hdr', '')
                time2_str = col2.split('_')[-1].replace('.hdr', '')

                # Convert to hours
                hour1 = int(time1_str[:2]) + int(time1_str[2:]) / 60.0
                hour2 = int(time2_str[:2]) + int(time2_str[2:]) / 60.0

                timestep_hours = abs(hour2 - hour1)
                print(f"\nDetected timestep interval: {timestep_hours} hours")
            else:
                timestep_hours = 1.0  # Default to 1 hour if can't determine
                print(f"\nUsing default timestep interval: {timestep_hours} hours")

            # Calculate consecutive timesteps with at least 1 m² of sunlight for each AOI
            consecutive_counts = []
            consecutive_hours = []

            for aoi_name in pivot_data.index:
                row_values = pivot_data.loc[aoi_name].values

                # Count maximum consecutive timesteps >= 1.0 m²
                max_consecutive = 0
                current_consecutive = 0

                for value in row_values:
                    if value >= 1.0:
                        current_consecutive += 1
                        max_consecutive = max(max_consecutive, current_consecutive)
                    else:
                        current_consecutive = 0

                consecutive_counts.append(max_consecutive)
                # Calculate hours: consecutive steps × timestep duration, rounded down to 1 decimal place
                hours_value = max_consecutive * timestep_hours
                hours_rounded = math.floor(hours_value * 10) / 10
                consecutive_hours.append(hours_rounded)

            # Reorder columns: insert consecutive steps and hours after the index (AOI file)
            # First, reset index to make aoi_file a column
            pivot_data_reset = pivot_data.reset_index()

            # Insert consecutive columns at position 1 (after aoi_file)
            pivot_data_reset.insert(1, 'Consecutive Timesteps ≥1m²', consecutive_counts)
            pivot_data_reset.insert(2, 'Hours of Direct Sun', consecutive_hours)

            # Set aoi_file back as index for Excel output
            pivot_data_reordered = pivot_data_reset.set_index('aoi_file')

            # Create aggregation summary by splitting AOI file names
            # Split by first "_" to get apartment and sub-space
            summary_data = pivot_data_reset.copy()
            summary_data['Apartment'] = summary_data['aoi_file'].str.split('_', n=1).str[0]
            summary_data['Sub-Space'] = summary_data['aoi_file'].str.split('_', n=1).str[1]

            # Remove .aoi extension from sub-space if present
            summary_data['Sub-Space'] = summary_data['Sub-Space'].str.replace('.aoi', '', regex=False)

            # Create pivot: Apartment (rows) x Sub-Space (columns) with Hours of Direct Sun values
            summary_pivot = summary_data.pivot_table(
                values='Hours of Direct Sun',
                index='Apartment',
                columns='Sub-Space',
                aggfunc='first',  # Use first since each apartment-subspace combo is unique
                fill_value=0
            )

            # Write aggregation summary at the top left (starting at row 1, column B)
            summary_start_col = 1  # Column B in 0-indexed
            summary_pivot.to_excel(writer, sheet_name='Pivot - Passing Area (m²)', startrow=1, startcol=summary_start_col)

            # Calculate the number of rows the summary takes (including headers and data)
            summary_rows = len(summary_pivot) + 2  # +2 for header row and index label row

            # Calculate detailed pivot starting column: summary start + index column + number of data columns + 1 space
            # summary_pivot.shape[1] gives the number of Sub-Space columns
            # +1 for the index column (Apartment), +1 for spacing
            detail_start_row = 1
            detail_start_col = summary_start_col + 1 + summary_pivot.shape[1] + 1
            pivot_data_reordered.to_excel(writer, sheet_name='Pivot - Passing Area (m²)', startrow=detail_start_row, startcol=detail_start_col)

            # Format Raw Data sheet
            self._format_raw_data_sheet(writer.sheets['Raw Data'])

            # Format Pivot sheet (need to pass summary and detail positions for proper formatting)
            self._format_pivot_sheet(writer.sheets['Pivot - Passing Area (m²)'], summary_start_col, summary_pivot.shape[1], summary_rows, detail_start_row, detail_start_col)

        print(f"\nResults saved to: {output_excel}")
        print(f"  Sheet 1: 'Raw Data' - All combinations with passing_area_m2")
        print(f"  Sheet 2: 'Pivot - Passing Area (m²)' - AOI x HDR matrix")
        print(f"  Metadata: Area per pixel = {self.area_per_pixel} m²")
        print("\n" + "=" * 80 + "\nSAMPLE RESULTS (first 10 rows)\n" + "=" * 80)
        print(combined_results.head(10).to_string(index=False))

    def _format_raw_data_sheet(self, worksheet) -> None:
        """Format Raw Data sheet with autofit and no gridlines.

        Args:
            worksheet: openpyxl worksheet object
        """
        worksheet.sheet_view.showGridLines = False

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _format_pivot_sheet(self, worksheet, summary_start_col: int, summary_num_cols: int, summary_rows: int, detail_start_row: int, detail_start_col: int) -> None:
        """Format Pivot sheet with rotated headers, fixed width columns, highlighting, and gridlines.

        Args:
            worksheet: openpyxl worksheet object
            summary_start_col: Column where summary pivot starts (0-indexed)
            summary_num_cols: Number of data columns in summary pivot
            summary_rows: Number of rows in the summary section
            detail_start_row: Row where detailed pivot data starts (1-indexed for Excel)
            detail_start_col: Column where detailed pivot data starts (0-indexed)
        """
        worksheet.sheet_view.showGridLines = False

        # Define styles
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Get the actual used range
        max_col = worksheet.max_column
        max_row = worksheet.max_row

        # Calculate summary section column range (1-indexed for Excel)
        summary_index_col = summary_start_col + 1  # Apartment column
        summary_first_data_col = summary_index_col + 1  # First Sub-Space column
        summary_last_data_col = summary_first_data_col + summary_num_cols - 1  # Last Sub-Space column
        summary_end_row = summary_rows  # End row of summary data (0-indexed row + headers)

        # Rotate summary section column headers (Sub-Space columns) to vertical (90 degrees)
        summary_header_row = 2  # Row 2 contains the Sub-Space column headers
        for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
            cell = worksheet.cell(row=summary_header_row, column=col_idx)
            if cell.value is not None:
                cell.alignment = Alignment(textRotation=90)

        # Apply green highlighting to summary section for values >= 2
        for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
            for row_idx in range(3, summary_end_row + 1):  # Start at row 3 (data rows)
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    if cell.value >= 2.0:
                        cell.fill = green_fill

        # Autofit summary index column (Apartment)
        max_length = 0
        for row_idx in range(2, summary_end_row + 1):
            cell = worksheet.cell(row=row_idx, column=summary_index_col)
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        if max_length > 0:
            column_letter = worksheet.cell(row=1, column=summary_index_col).column_letter
            worksheet.column_dimensions[column_letter].width = max_length + 2

        # Format summary data columns with autofit
        for col_idx in range(summary_first_data_col, summary_last_data_col + 1):
            max_length = 0
            for row_idx in range(2, summary_end_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            if max_length > 0:
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                worksheet.column_dimensions[column_letter].width = max(max_length + 2, 8)  # Min width 8

        # Format detailed section headers (row detail_start_row + 1 for detailed pivot)
        # Rotate header row text to vertical (90 degrees) for detailed section
        # Column layout: A=empty, B-C=summary, D=empty (detailed index col), E=aoi_file, F=Consecutive Steps, G=Hours of Direct Sun, H onwards=timestep data
        detail_header_row = detail_start_row + 1
        # Calculate the first timestep column index (1-indexed: detail_start_col + 4 + 1)
        first_timestep_col = detail_start_col + 4 + 1
        for idx, cell in enumerate(worksheet[detail_header_row]):
            if cell.value is not None:
                # Only rotate timestep column headers (columns at first_timestep_col onwards)
                # Don't rotate: empty col, summary cols, detailed index, aoi_file, Consecutive Steps, Hours of Direct Sun
                if idx + 1 >= first_timestep_col:
                    cell.alignment = Alignment(textRotation=90)

        # Set column widths and apply formatting
        # Calculate detailed section column indices (1-indexed for Excel)
        detail_aoi_col = detail_start_col + 1 + 1  # +1 for index column, +1 for 1-indexed
        detail_consecutive_col = detail_start_col + 2 + 1
        detail_hours_col = detail_start_col + 3 + 1

        for col_idx in range(1, max_col + 1):
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter

            # Column A is empty (since startcol=1, data starts at B)
            if col_idx == 1:
                worksheet.column_dimensions[column_letter].width = 2

            # AOI file column in detailed section - fit to content with black borders for detailed section only
            elif col_idx == detail_aoi_col:
                max_length = 0
                for row_idx in range(2, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    # Apply black borders to detailed section data rows only (detail_start_row + 2 onwards)
                    if row_idx >= detail_start_row + 2:
                        cell.border = black_border
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Consecutive Steps column in detailed section - fit to content with black borders for detailed section only
            elif col_idx == detail_consecutive_col:
                max_length = 0
                for row_idx in range(2, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    # Apply black borders to detailed section data rows only (detail_start_row + 2 onwards)
                    if row_idx >= detail_start_row + 2:
                        cell.border = black_border
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Hours of Direct Sun column in detailed section - fit to content with black borders for detailed section only
            elif col_idx == detail_hours_col:
                max_length = 0
                for row_idx in range(2, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    # Apply black borders to detailed section data rows only (detail_start_row + 2 onwards)
                    if row_idx >= detail_start_row + 2:
                        cell.border = black_border
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Timestep data columns - fixed width 7.29 and apply green highlighting
            elif col_idx >= first_timestep_col:
                worksheet.column_dimensions[column_letter].width = 7.29

                # Apply green highlighting to cells >= 1.0 m²
                for row_idx in range(3, max_row + 1):  # Skip header rows
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if cell.value >= 1.0:
                            cell.fill = green_fill

        # Apply conditional formatting color scale to Hours of Direct Sun column in detailed section only
        # Range starts at detail_start_row + 2 (data rows of detailed section)
        detail_hours_start = detail_start_row + 2
        hours_col_letter = worksheet.cell(row=1, column=detail_hours_col).column_letter
        hours_range = f"{hours_col_letter}{detail_hours_start}:{hours_col_letter}{max_row}"

        # Create 3-color scale: white (min=2) -> yellow (midpoint=3) -> light red (max=highest value)
        color_scale_rule = ColorScaleRule(
            start_type='num',
            start_value=2,
            start_color='FFFFFF',  # White
            mid_type='num',
            mid_value=3,
            mid_color='FFFF99',  # Light yellow
            end_type='max',
            end_color='FFB6C1'  # Light red
        )

        worksheet.conditional_formatting.add(hours_range, color_scale_rule)

# Module-level worker function (required for multiprocessing pickle)
def _process_view_group_worker(args):
    """Worker function for ProcessPoolExecutor.

    Must be top-level for pickling.

    Args:
        args: tuple of (view_name, group, pixel_threshold_value, wpd_output_dir)

    Returns:
        tuple: (view_name, result_df)
    """
    view_name, group, pixel_threshold_value, wpd_output_dir = args

    print(f"\n{'='*80}\nProcessing view group: {view_name} (PID: {os.getpid()})\n{'='*80}")

    result_df = _process_batch_with_caching(
        hdr_files=group['hdr_files'],
        aoi_files=group['aoi_files'],
        pixel_threshold_value=pixel_threshold_value,
        wpd_output_dir=wpd_output_dir
    )

    return (view_name, result_df)

def _process_single_aoi_with_cached_hdr(hdr_file: Path, aoi_file: Path, data_2d: np.ndarray,
                                       width: int, height: int, pixel_threshold_value: float = 0.0):
    """Process a single AOI file using pre-loaded HDR data.

    Args:
        hdr_file: Path to HDR file (for record keeping)
        aoi_file: Path to AOI file containing polygon coordinates
        data_2d: Pre-loaded 2D numpy array of HDR brightness values
        width: Image width in pixels
        height: Image height in pixels
        pixel_threshold_value: Brightness threshold value

    Returns:
        tuple: (hdr_file, aoi_file, total_in_poly, count_pass)
    """
    # Read polygon coordinates from AOI file
    polygon_coords = []
    with open(aoi_file, 'r') as f:
        lines = f.readlines()
        for line in lines[5:]:  # Skip header lines
            parts = line.strip().split()
            if len(parts) >= 4:
                pixel_x = int(parts[2])
                pixel_y = int(parts[3])
                polygon_coords.append((pixel_x, pixel_y))

    # Generate coordinate grid
    x, y = np.meshgrid(np.arange(width), np.arange(height))
    x_flat = x.flatten()
    y_flat = y.flatten()
    points = np.vstack((x_flat, y_flat)).T

    # Create polygon mask
    path = MplPath(polygon_coords)
    mask_flat = path.contains_points(points, radius=0)
    mask_2d = mask_flat.reshape((height, width))

    # Apply filter
    pixels_of_interest = data_2d[mask_2d & (data_2d > pixel_threshold_value)]

    # Calculate stats
    count_pass = pixels_of_interest.size
    total_in_poly = np.count_nonzero(mask_2d)

    if total_in_poly == 0:
        print(f"Warning: {aoi_file.name} polygon is outside image bounds")
        return (hdr_file, aoi_file, 0, 0)

    return (hdr_file, aoi_file, total_in_poly, count_pass)

def _process_batch_with_caching(hdr_files: List[Path], aoi_files: List[Path],
                               pixel_threshold_value: float = 0.0,
                               wpd_output_dir: Path = None) -> pd.DataFrame:
    """Efficiently process multiple HDR and AOI files by caching HDR data.

    This function loads each HDR file once and processes all AOI files against it,
    avoiding redundant HDR file reads and pvalue operations.

    Args:
        hdr_files: List of HDR file paths to process
        aoi_files: List of AOI file paths to process
        pixel_threshold_value: Brightness threshold value
        wpd_output_dir: Directory to save .wpd files

    Returns:
        pandas.DataFrame with columns: hdr_file, aoi_file, total_pixels, passing_pixels
    """
    if wpd_output_dir is None:
        wpd_output_dir = Path(__file__).parent.parent / "outputs" / "wpd"
    wpd_output_dir.mkdir(parents=True, exist_ok=True)

    # Dictionary to store results grouped by AOI file
    aoi_results = {aoi_file: [] for aoi_file in aoi_files}
    results = []
    total_operations = len(hdr_files) * len(aoi_files)
    current_operation = 0

    print(f"Processing {len(hdr_files)} HDR files × {len(aoi_files)} AOI files = {total_operations} operations\n" + "=" * 80)

    for hdr_file in hdr_files:
        print(f"\nLoading HDR file: {hdr_file.name}")

        # Check if file is empty
        if hdr_file.stat().st_size == 0:
            print(f"  Skipping: {hdr_file.name} (file is empty)")
            continue

        # Load HDR data once
        try:
            width, height = get_hdr_resolution(hdr_file)
        except (ValueError, FileNotFoundError) as e:
            print(f"  Skipping: {hdr_file.name} (error reading resolution: {e})")
            continue

        # Run pvalue to get binary data
        cmd = ['pvalue', '-h', '-H', '-b', '-df', str(hdr_file)]
        try:
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            raw_data = process.stdout.read()
            stderr_data = process.stderr.read()
            process.wait()

            if process.returncode != 0:
                print(f"  Skipping: {hdr_file.name} (pvalue failed with code {process.returncode})")
                if stderr_data:
                    print(f"    Error: {stderr_data.decode('utf-8', errors='ignore').strip()}")
                continue

        except FileNotFoundError:
            print("Error: 'pvalue' command not found.")
            sys.exit(1)

        # Convert to 2D array
        try:
            data = np.frombuffer(raw_data, dtype=np.float32)
            data_2d = data.reshape((height, width))
        except ValueError as e:
            print(f"  Skipping: {hdr_file.name} (error reshaping data: {e})")
            continue

        print(f"  Processing {len(aoi_files)} AOI files with cached HDR data...")

        # Process all AOI files with this cached HDR data
        for aoi_file in aoi_files:
            current_operation += 1

            hdr, aoi, total, passing = _process_single_aoi_with_cached_hdr(
                hdr_file, aoi_file, data_2d, width, height, pixel_threshold_value
            )

            # Strip everything before "plan_" from HDR filename
            hdr_name = hdr.name
            if 'plan_' in hdr_name:
                hdr_name = 'plan_' + hdr_name.split('plan_')[1]

            # Store results for this AOI
            aoi_results[aoi_file].append({
                'hdr_file': hdr_name,
                'total_pixels': total,
                'passing_pixels': passing
            })

            results.append({
                'aoi_file': aoi.name,
                'hdr_file': hdr_name,
                'total_pixels': total,
                'passing_pixels': passing,
            })

            print(f"    [{current_operation}/{total_operations}] {aoi.name}: {passing}/{total} pixels")

    # Write .wpd files for each AOI
    print("\n" + "=" * 80 + "\nWriting .wpd files...\n" + "=" * 80)

    for aoi_file, aoi_data in aoi_results.items():
        # Sort by HDR filename in ascending order
        aoi_data_sorted = sorted(aoi_data, key=lambda x: x['hdr_file'])

        # Get total_pixels (should be same for all HDR files for this AOI)
        total_pixels_in_polygon = aoi_data_sorted[0]['total_pixels'] if aoi_data_sorted else 0

        # Create .wpd filename from AOI filename
        wpd_filename = aoi_file.stem + '.wpd'
        wpd_path = wpd_output_dir / wpd_filename

        # Write .wpd file
        with open(wpd_path, 'w') as f:
            f.write(f"total_pixels_in_polygon: {total_pixels_in_polygon}\n")
            f.write("hdr_file passing_pixels\n")

            for row in aoi_data_sorted:
                f.write(f"{row['hdr_file']} {row['passing_pixels']}\n")

        print(f"  Written: {wpd_filename} ({len(aoi_data_sorted)} HDR files)")

    print("\n" + "=" * 80 + f"\nProcessing complete!\nWPD files saved to: {wpd_output_dir}")

    return pd.DataFrame(results)


if __name__ == "__main__":
    # Initialize processor
    processor = ResultsProcessor(
        image_dir=Path(r"C:/Projects/archilume/outputs/images"),
        aoi_dir=Path(r"C:/Projects/archilume/outputs/aoi"),
        wpd_dir=Path(r"C:/Projects/archilume/outputs/wpd"),
        pixel_to_world_map=Path(r"C:/Projects/archilume/outputs/aoi/pixel_to_world_coordinate_map.txt"),
        pixel_threshold_value=0,
        max_workers=12
    )

    # Run the pipeline
    processor.nsw_adg_sunlight_sequence_wpd_extraction()
