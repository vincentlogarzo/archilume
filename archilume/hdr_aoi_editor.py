"""
Archilume: Interactive Room Boundary Editor for HDR/TIFF Floor Plan Images

NOTE: daylight_workflow_iesve.py must be run before use of this editor.
The workflow generates the HDR/TIFF images and .aoi boundaries this
editor depends on.

Draw apartment and sub-room boundaries on top of HDR or associated TIFF
rendered floor plan images. JSON and CSV are saved automatically to the
image_dir on every save/delete.

Naming convention:
    U101          apartment boundary
    U101_BED1     sub-room (auto-prefixed when parent is selected)
    U101_DIV1     division sub-room (created by the room divider tool)

Controls:
    Up/Down       Navigate HDR files
    t             Toggle image variant (HDR / TIFFs)
    Left-click    Select room (default) / place vertex (draw mode) / drag vertex (edit mode)
    Shift+click   Drag entire edge in edit mode (moves both endpoints together)
    Right-click   Delete hovered vertex (edit mode) / undo divider point (divider mode)
    Scroll        Zoom centred on cursor
    s             Save room / confirm edit
    e             Toggle Edit Mode
    d             Toggle Draw Mode (default) / room divider — multi-segment ortho split (edit mode)
    Esc           Exit draw mode / exit divider mode / deselect room
    Ctrl+Z        Undo — edit: vertex/edge/division; draw: delete/type/rename/create
    o             Toggle orthogonal lines (H/V snap)
    Ctrl+click    Multi-select rooms in list (bulk type tagging)
    Ctrl+A        Select all rooms on current HDR
    f             Fit zoom to selected room
    r             Reset zoom
    q             Quit

Workflow:
    1. Navigate to the desired HDR file with Up/Down
    2. Draw apartment boundary, name "U101", press s to save
    3. Select "U101" as parent
    4. Draw sub-rooms (e.g. "BED1" auto-saved as "U101_BED1")
    5. Repeat for each HDR file / floor
    6. Export and archive results as CSV for analysis and reporting

Functionality:

    Room Divider:
        Splits a room into two sub-rooms along a multi-segment ortho
        polyline. Enter edit mode (e), select a room, press d, then click
        to place points (each segment H or V). Press s to finalize. Lines
        auto-extend to the boundary when points are placed outside the
        room. Right-click undoes the last point. Ctrl+Z undoes the split.

    Edit Mode:
        Toggle with e. All rooms on the current HDR become editable
        simultaneously. Available actions while in edit mode:
        - Drag a vertex to reposition it (ortho-constrained when o is on)
        - Shift+drag an edge to move both its endpoints together
        - Click on an edge to insert two new vertices straddling the
          click point (edit mode only)
        - Right-click a vertex to delete it
        - Press d to enter room divider mode (edit mode only)
        - Press f to fit-zoom to the selected room boundary
        - Ctrl+Z undoes vertex edits (up to 50 levels)
        Press s or toggle e off to save and exit edit mode.

    Ortho Mode:
        Toggle with o. Constrains drawn lines to horizontal or vertical
        only. The live preview snaps to the nearest axis in real time.
        Useful for architectural floor plans with axis-aligned walls.

    Vertex & Edge Snapping:
        Automatically snaps new vertices to nearby existing vertices or
        edges within 10 pixels. Always active during drawing. Helps
        align room boundaries precisely without manual pixel placement.

    Parent Auto-Detection:
        When the first vertex of a new polygon falls inside an existing
        parent room, that parent is auto-selected. Sub-rooms are then
        auto-prefixed (e.g. drawing "BED1" inside U101 saves as
        U101_BED1). Boundary containment is checked on save.

    Room Type Tagging:
        Assign BED, LIVING, or CIRCULATION types via buttons. Sub-rooms
        default to BED; parents default to LIVING when children exist.
        Ctrl+click rooms for multi-select, then tag in bulk. Types
        drive daylight factor threshold evaluation.

    Export & Archive:
        Exports an Excel summary, per-room pixel CSVs, and overlay TIFFs
        with room boundaries rendered. Everything is zipped into a
        timestamped archive. Use "Extract Archive" to restore a previous
        export. Runs in a background thread with progress bar.

    Image Variant Toggle:
        Press t to cycle between the HDR and associated TIFF images for
        the current floor. Useful for comparing raw HDR data against
        processed TIFF renderings side by side.

    Multi-Select:
        Ctrl+click rooms in the list to toggle multi-selection. Ctrl+A
        selects all rooms on the current HDR. Enables bulk room type
        tagging across multiple rooms at once.

    Undo:
        Ctrl+Z in edit mode undoes vertex/edge edits and room divisions
        (50 levels). In draw mode, Ctrl+Z undoes room deletion, room
        type changes, room renames, and new room creation (including
        any auto-assigned parent type). Both stacks hold up to 50
        entries.
"""

# fmt: off
# autopep8: off

# Standard library imports
import csv
import json
import re
import shutil
import threading
import time
import zipfile
from pathlib import Path
from typing import List, Optional, Tuple, Union

# Third-party imports
import matplotlib.pyplot as plt
from matplotlib.widgets import PolygonSelector, TextBox, Button
from matplotlib.patches import Polygon, FancyBboxPatch
from matplotlib.path import Path as MplPath
import matplotlib.patheffects as patheffects
import numpy as np

# Archilume imports
from archilume import config, utils
from archilume.hdr2wpd import Hdr2Wpd


class HdrAoiEditor:
    """Interactive room boundary drawing tool for HDR/TIFF floor plan images.

    Displays an HDR or associated TIFF image and provides a PolygonSelector
    for drawing room boundary polygons. Supports hierarchical parent/child
    relationships (e.g. U101 → U101_BED1). Auto-saves JSON and CSV on every
    save or delete action.

    Args:
        image_dir: Directory containing .hdr files and associated .tiff files.
        aoi_dir: Directory containing .aoi files with pixel-mapped room boundaries.
        session_path: Optional path for saving/loading editor sessions (JSON).
    """

    def __init__(
        self,
        image_dir:              Union[Path, str]            = config.IMAGE_DIR,
        aoi_dir:                Union[Path, str]            = config.AOI_DIR,
        session_path:           Optional[Path]              = None,
    ):
        self.image_dir                                  = Path(image_dir)
        self.aoi_dir                                    = Path(aoi_dir)
        self.session_path                               = session_path or (self.image_dir / "aoi_session.json")
        self.csv_path                                   = self.image_dir / "aoi_boundaries.csv"

        # Scan HDR files in image_dir
        self.hdr_files:             List[dict]          = self._scan_hdr_files()
        self.current_hdr_idx:       int                 = 0

        # Image variant toggle state (HDR + associated TIFFs for current HDR)
        self.image_variants:        List[Path]          = []
        self.current_variant_idx:   int                 = 0
        self._rebuild_image_variants()

        # Room storage
        self.rooms:                 List[dict]          = []
        self.current_polygon_vertices                   = []
        self.selected_room_idx:     Optional[int]       = None

        # Zoom / pan state
        self.original_xlim                              = None
        self.original_ylim                              = None
        self._pan_active:           bool                = False
        self._pan_start:            Optional[tuple]     = None  # (x_event, y_event) in pixels

        # Image dimensions (set on first render)
        self._image_width:          int                 = 1
        self._image_height:         int                 = 1
        self._reference_view_w:     float               = 1.0  # view width at default zoom

        # Snap to existing polygon vertices (always on, no UI control)
        self._snap_distance_px:     float               = 10.0
        self.current_vertices:      np.ndarray          = np.array([])
        self.ortho_mode:            bool                = True
        self._pending_snap:         Optional[tuple]     = None

        # Draw mode (polygon drawing; toggled with 'd' when not in edit mode)
        self.draw_mode:             bool                = False

        # Vertex editing mode
        self.edit_mode:             bool                = False
        self.edit_room_idx:         Optional[int]       = None
        self.edit_vertex_idx:       Optional[int]       = None
        self.hover_room_idx:        Optional[int]       = None
        self.hover_vertex_idx:      Optional[int]       = None
        self.hover_edge_room_idx:   Optional[int]       = None
        self.hover_edge_idx:        Optional[int]       = None
        self.hover_edge_point:      Optional[tuple]     = None
        self.edit_edge_room_idx:    Optional[int]       = None
        self.edit_edge_idx:         Optional[int]       = None
        self.edit_edge_start:       Optional[tuple]     = None
        self._edit_drag_origin:     Optional[tuple]     = None  # vertex pos at drag start (for ortho)

        # Parent apartment selection
        self.selected_parent:       Optional[str]       = None
        self.parent_options:        List[str]           = []

        # Room list scroll state
        self.room_list_scroll_offset: int               = 0
        self._room_list_hit_boxes:  List[Tuple]         = []

        # Undo stack for edit-mode vertex operations
        # Each entry: (room_idx, vertices_snapshot)
        self._edit_undo_stack:      List[Tuple]         = []
        self._edit_undo_max:        int                 = 50

        # General-purpose undo stack for draw-mode operations
        # Each entry: (tag, ...data) where tag is one of:
        #   ('delete',    insertion_index, room_dict)
        #   ('type',      [(idx, old_type), ...])
        #   ('rename',    room_idx, old_name)
        #   ('create',    room_idx)
        self._draw_undo_stack:      List[Tuple]         = []
        self._draw_undo_max:        int                 = 50

        # Image cache: path → numpy array (avoids reloading from disk)
        self._image_cache:          dict                = {}

        # Cached matplotlib artists for incremental rendering
        self._room_patch_cache                          = {}
        self._room_label_cache                          = {}
        self._df_text_cache:        list                = []
        self._edit_vertex_scatter                       = None
        self._last_hover_check                          = 0.0
        self._last_drag_draw:       float               = 0.0
        # Pre-built hover arrays (rebuilt on full render)
        self._hover_all_verts:      Optional[np.ndarray]= None
        self._hover_vert_room_idx:  Optional[np.ndarray]= None
        self._hover_vert_local_idx: Optional[np.ndarray]= None
        self._hover_edge_starts:    Optional[np.ndarray]= None
        self._hover_edge_ends:      Optional[np.ndarray]= None
        self._hover_edge_room_idx_arr: Optional[np.ndarray]= None
        self._hover_edge_local_idx: Optional[np.ndarray]= None
        # Blitting state for drag operations
        self._blit_background                           = None
        self._blit_active:          bool                = False
        self._image_handle                              = None
        self.ax_legend                                  = None

        # Room type tagging (BED / LIVING — LIVING requires sub-rooms)
        self.room_type:             Optional[str]       = None
        self.multi_selected_room_idxs: set              = set()   # Ctrl+click multi-select

        # Daylight factor analysis — thresholds are fixed per room type
        self.DF_THRESHOLDS          = {'BED': 0.5, 'LIVING': 1.0}
        self._df_image:             Optional[np.ndarray]= None   # (H, W) DF% for current HDR
        self._df_image_cache:       dict                = {}     # hdr_path_str → np.ndarray
        self._room_df_results:      dict                = {}     # room_idx → list of result strings
        self._hdr2wpd:              Optional[Hdr2Wpd]   = None
        # Persistent DF cache: keyed by (hdr_file, vertices_tuple, thresholds_tuple)
        # Stored on each room dict as 'df_cache' = {'thresholds': [...], 'results': [...], 'vertices_hash': str}
        # This avoids recomputation when rooms haven't changed.

        # Room divider mode (sub-state of edit mode)
        self.divider_mode:          bool                = False
        self._divider_room_idx:     Optional[int]       = None   # room being divided
        self._divider_points:       list                = []     # placed (x,y) tuples
        self._divider_markers:      list                = []     # scatter artists per point
        self._divider_segments:     list                = []     # solid Line2D per segment
        self._divider_preview_line                      = None   # dashed Line2D to cursor
        self._divider_snap_pt:      Optional[tuple]     = None   # snapped vertex (x, y) or None
        self._divider_snap_marker                       = None   # highlight ring artist

    # === LAYOUT HELPERS ========================================================

    def _axes(self, x, y, w, h):
        """Create figure axes at (x, y) measured from top-left corner.

        x increases right, y increases down — unlike matplotlib's native
        bottom-left origin. Converts to fig.add_axes([left, bottom, w, h]).
        """
        return self.fig.add_axes((x, 1.0 - y - h, w, h))

    def _make_button(self, x, y, w, h, label, callback,
                     fontsize=7, color=None, hovercolor=None):
        """Create a styled Button with rounded corners, shadow, and hover effect."""
        ax = self._axes(x, y, w, h)
        base_color  = color or self._btn_color
        hover_color = hovercolor or self._btn_hover

        # Hide default axes chrome
        ax.set_facecolor('none')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.set_navigate(False)
        for spine in ax.spines.values():
            spine.set_visible(False)
        ax.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)

        # Drop shadow
        shadow = FancyBboxPatch(
            (0.03, 0.03), 0.94, 0.86,
            boxstyle='round,pad=0.04', facecolor='#00000015', edgecolor='none',
            transform=ax.transAxes, clip_on=False, zorder=1)
        ax.add_patch(shadow)

        # Main button body
        body = FancyBboxPatch(
            (0.02, 0.08), 0.96, 0.86,
            boxstyle='round,pad=0.04', facecolor=base_color,
            edgecolor='#B0B0A8', linewidth=0.8,
            transform=ax.transAxes, clip_on=False, zorder=2)
        ax.add_patch(body)

        # Highlight strip (subtle top gradient effect)
        highlight = FancyBboxPatch(
            (0.05, 0.55), 0.90, 0.34,
            boxstyle='round,pad=0.03', facecolor='#FFFFFF30', edgecolor='none',
            transform=ax.transAxes, clip_on=False, zorder=3)
        ax.add_patch(highlight)

        # Label
        txt = ax.text(0.5, 0.48, label, fontsize=fontsize,
                      fontweight='medium', color='#333333',
                      ha='center', va='center',
                      transform=ax.transAxes, zorder=4)

        # Build a button-like object to stay compatible with existing code
        btn = Button(ax, '', color='none', hovercolor='none')
        btn.label = txt
        btn.color = base_color
        btn.hovercolor = hover_color
        btn._body = body
        btn._shadow = shadow
        btn._highlight = highlight

        # Hover enter/leave callbacks
        def _on_enter(event):
            if event.inaxes == ax:
                body.set_facecolor(hover_color)
                body.set_edgecolor('#909088')
                shadow.set_facecolor('#00000025')
                ax.figure.canvas.draw_idle()

        def _on_leave(event):
            body.set_facecolor(btn.color)
            body.set_edgecolor('#B0B0A8')
            shadow.set_facecolor('#00000018')
            ax.figure.canvas.draw_idle()

        ax.figure.canvas.mpl_connect('axes_enter_event', _on_enter)
        ax.figure.canvas.mpl_connect('axes_leave_event', _on_leave)

        btn.on_clicked(callback)
        return btn

    def _make_label(self, x, y, w, h, text, fontsize=9, **kwargs):
        """Create a text label on an invisible axes."""
        ax = self._axes(x, y, w, h)
        ax.axis('off')
        txt = ax.text(0, 0.5, text, fontsize=fontsize, fontweight='bold',
                      color='#404040', **kwargs)
        return ax, txt

    def _is_room_on_current_hdr(self, room: dict) -> bool:
        """Check whether a room belongs to the currently displayed HDR file."""
        return room.get('hdr_file') == self.current_hdr_name

    def _reset_hover_state(self):
        """Clear all hover and drag tracking variables."""
        self.hover_room_idx      = None
        self.hover_vertex_idx    = None
        self.hover_edge_room_idx = None
        self.hover_edge_idx      = None
        self.hover_edge_point    = None
        self.edit_room_idx       = None
        self.edit_vertex_idx     = None
        self.edit_edge_room_idx  = None
        self.edit_edge_idx       = None
        self.edit_edge_start     = None
        self._edit_drag_origin   = None

    def _start_blit_drag(self, room_idx: int):
        """Capture background for blitting before starting a drag operation."""
        canvas = self.fig.canvas
        if not getattr(canvas, 'supports_blit', False):
            return
        patch = self._room_patch_cache.get(room_idx)
        if patch is None:
            return
        # Hide the patch, draw the static background, then restore
        patch.set_visible(False)
        label = self._room_label_cache.get(room_idx)
        if label is not None:
            label.set_visible(False)
        canvas.draw()
        self._blit_background = canvas.copy_from_bbox(self.ax.bbox)
        patch.set_visible(True)
        if label is not None:
            label.set_visible(True)
        self._blit_active = True

    def _end_blit_drag(self):
        """Clear blitting state after a drag ends."""
        self._blit_active = False
        self._blit_background = None

    def _update_dragged_patch(self, room_idx: int):
        """Update a cached patch/label after vertex or edge drag. Returns True if handled.
        Uses blitting when available for fast partial redraws.
        Label position is deferred to drag-end (force_full render)."""
        patch = self._room_patch_cache.get(room_idx)
        if patch is not None:
            patch.set_xy(self.rooms[room_idx]['vertices'])
            # Centroid/label update deferred to button release (_render_section force_full)
            now = time.monotonic()
            if now - self._last_drag_draw >= 0.033:  # ~30 fps throttle
                self._last_drag_draw = now
                canvas = self.fig.canvas
                if self._blit_active and self._blit_background is not None:
                    canvas.restore_region(self._blit_background)
                    self.ax.draw_artist(patch)
                    canvas.blit(self.ax.bbox)
                else:
                    canvas.draw_idle()
            return True
        self._render_section(force_full=True)
        return True

    # === COORDINATE MAPPING ====================================================

    def _load_from_aoi_files(self, aoi_dir: Path):
        """Load room boundaries from .aoi files (pixel coordinates already included).

        AOI format:
            AOI Points File: <name> <level>
            ASSOCIATED VIEW FILE: plan_ffl_<z_mm>.vp
            FFL z height(m): <z>
            CENTRAL x,y: <cx> <cy>
            NO. PERIMETER POINTS <n>: x,y pixel_x pixel_y positions
            <world_x> <world_y> <pixel_x> <pixel_y>
            ...
        """
        aoi_files = sorted(aoi_dir.glob('*.aoi'))
        for aoi_path in aoi_files:
            with open(aoi_path, 'r') as f:
                lines = [l.strip() for l in f.readlines()]
            if len(lines) < 6:
                continue

            # Header: name from line 0
            name_match = re.match(r'AOI Points File:\s*(.+)', lines[0])
            name       = name_match.group(1).strip() if name_match else aoi_path.stem

            # HDR file from view file reference (plan_ffl_28700.vp → model_plan_ffl_28700)
            vp_match = re.search(r'plan_ffl_(\d+)', lines[1])
            hdr_file = self.current_hdr_name
            if vp_match:
                ffl_val = vp_match.group(1)
                for entry in self.hdr_files:
                    if ffl_val in entry['name']:
                        hdr_file = entry['name']
                        break

            # Vertex lines: world_x world_y pixel_x pixel_y
            vertices = []
            for line in lines[5:]:
                parts = line.split()
                if len(parts) >= 4:
                    px, py = float(parts[2]), float(parts[3])
                    vertices.append([px, py])

            if len(vertices) >= 3:
                self.rooms.append({
                    'name':     name,
                    'parent':   None,
                    'vertices': vertices,
                    'hdr_file': hdr_file,
                })

        print(f"Loaded {len(self.rooms)} rooms from {len(aoi_files)} .aoi files in {aoi_dir}")

    # === IMAGE SCANNING & LOADING ==============================================

    def _scan_hdr_files(self) -> List[dict]:
        """Scan image_dir for HDR files and associated TIFFs.

        Returns:
            Sorted list of dicts with keys: hdr_path, tiff_paths, name (stem).
        """
        if not self.image_dir.exists():
            print(f"Warning: image_dir does not exist: {self.image_dir}")
            return []

        hdr_paths = sorted(self.image_dir.glob("*.hdr"))
        result = []
        for hdr_path in hdr_paths:
            stem = hdr_path.stem
            # Associated TIFFs: any .tiff in same dir whose stem starts with stem + '_'
            # Exclude previously-exported aoi_overlay files to avoid re-processing them
            tiff_paths = sorted(
                p for p in self.image_dir.glob("*.tiff")
                if p.stem.startswith(stem + "_") and not p.stem.endswith("_aoi_overlay")
            )
            result.append({
                'hdr_path': hdr_path,
                'tiff_paths': tiff_paths,
                'name': stem,
            })

        # Build legend map: key → Path for files matching '*_legend.tiff'
        # e.g. 'df_cntr' → Path('df_cntr_legend.tiff')
        self.legend_map: dict = {}
        for legend_path in sorted(self.image_dir.glob("*_legend.tiff")):
            key = legend_path.stem[: -len("_legend")]  # strip trailing '_legend'
            self.legend_map[key] = legend_path
        if self.legend_map:
            print(f"Found legend(s): {list(self.legend_map.keys())}")

        print(f"Found {len(result)} HDR file(s) in {self.image_dir}")
        return result

    def _rebuild_image_variants(self):
        """Rebuild the image_variants list for the current HDR index.

        Preserves the active layer type across HDR navigation by matching the
        suffix of the current variant (e.g. '_df_false') against the new HDR's
        variants. Falls back to index 0 (the HDR itself) if no match is found.
        """
        if not self.hdr_files:
            self.image_variants = []
            self.current_variant_idx = 0
            return

        # Remember which layer suffix is currently active before rebuilding
        active_suffix = None
        if self.image_variants and 0 < self.current_variant_idx < len(self.image_variants):
            old_path = self.image_variants[self.current_variant_idx]
            if old_path.suffix.lower() != '.hdr':
                # Extract the part after the old HDR stem, e.g. '_df_false'
                old_hdr_stem = self.hdr_files[self.current_hdr_idx]['name']
                active_suffix = old_path.stem[len(old_hdr_stem):]

        entry = self.hdr_files[self.current_hdr_idx]
        self.image_variants = [entry['hdr_path']] + list(entry['tiff_paths'])

        # Try to restore the same layer type in the new HDR's variants
        if active_suffix:
            new_hdr_stem = entry['name']
            for i, path in enumerate(self.image_variants):
                if path.suffix.lower() != '.hdr' and path.stem[len(new_hdr_stem):] == active_suffix:
                    self.current_variant_idx = i
                    return

        self.current_variant_idx = 0

    @property
    def current_hdr_name(self) -> str:
        """Stem of the currently active HDR file."""
        if not self.hdr_files:
            return ""
        return self.hdr_files[self.current_hdr_idx]['name']

    @property
    def current_variant_path(self) -> Optional[Path]:
        """Path of the currently displayed image variant."""
        if not self.image_variants:
            return None
        idx = self.current_variant_idx % len(self.image_variants)
        return self.image_variants[idx]

    def _load_image(self, path: Path) -> Optional[np.ndarray]:
        """Load an image file as a normalised float32 numpy array.

        Results are cached in memory so repeated renders don't hit disk.

        Args:
            path: Path to .hdr or .tiff file.

        Returns:
            Array of shape (H, W, 3) with values in [0, 1], or None on failure.
        """
        key = str(path)
        if key in self._image_cache:
            return self._image_cache[key]

        try:
            if path.suffix.lower() == '.hdr':
                import imageio.v2 as imageio
                img = imageio.imread(str(path)).astype(np.float32)
                if img.ndim == 2:
                    img = np.stack([img, img, img], axis=-1)
                p99 = np.percentile(img, 99)
                if p99 > 0:
                    img = img / p99
                img = np.clip(img ** (1.0 / 2.2), 0.0, 1.0)
            else:
                from PIL import Image
                pil_img = Image.open(path).convert('RGB')
                img = np.array(pil_img, dtype=np.float32) / 255.0

            self._image_cache[key] = img
            return img
        except Exception as exc:
            print(f"Warning: could not load image {path}: {exc}")
            return None

    def _get_legend_for_variant(self, path: Optional[Path]) -> Optional[Path]:
        """Return the legend TIFF matching the given image variant, or None.

        Matching rule: for a TIFF whose stem is '<hdr_stem>_<suffix>', find
        the first legend key that is a substring of '<suffix>'.
        HDR variants never have an associated legend.
        """
        if path is None or path.suffix.lower() == '.hdr':
            return None
        hdr_stem = self.current_hdr_name
        suffix = path.stem[len(hdr_stem) + 1:] if path.stem.startswith(hdr_stem + "_") else path.stem
        for key, legend_path in self.legend_map.items():
            if key in suffix:
                return legend_path
        return None

    # === LAUNCH ================================================================

    _WINDOW_TITLE = "Archilume - HDR AOI Editor"

    def launch(self):
        """Open the interactive editor window."""
        if not self.hdr_files:
            raise FileNotFoundError(f"No .hdr files found in {self.image_dir}")

        # Close any previous editor window that was left open
        for fig_num in plt.get_fignums():
            fig = plt.figure(fig_num)
            if getattr(fig, '_archilume_editor', False):
                plt.close(fig)

        # Initialise daylight factor analysis (Phase 3 from daylight_workflow_iesve)
        try:
            coordinate_map_path = utils.create_pixel_to_world_coord_map(self.image_dir)
            self._hdr2wpd = Hdr2Wpd(pixel_to_world_map=coordinate_map_path)
            print(f"DF analysis ready (area_per_pixel={self._hdr2wpd.area_per_pixel} m2)")
        except Exception as exc:
            print(f"Warning: DF analysis unavailable ({exc}). Results will not be shown.")
            self._hdr2wpd = None

        # Setup matplotlib figure — wide to match ~2.6:1 floor plan aspect ratio
        plt.rcParams['savefig.directory'] = str(self.image_dir)
        plt.rcParams['keymap.save']       = []  # disable 's' / 'ctrl+s' save-figure hotkey
        plt.rcParams['keymap.fullscreen'] = []  # disable 'f' fullscreen hotkey
        self.fig = plt.figure(figsize=(20, 8), facecolor='#F5F5F0')
        self.fig._archilume_editor = True
        self.fig.canvas.manager.set_window_title(self._WINDOW_TITLE)
        self.fig.subplots_adjust(left=0.02, right=0.98, top=0.95, bottom=0.02)

        # Maximise window
        try:
            import sys
            manager = plt.get_current_fig_manager()
            if sys.platform == "win32":
                manager.window.state('zoomed')
            else:
                manager.window.attributes('-zoomed', True)
            self.fig.canvas.mpl_connect('resize_event', self._on_resize)
            self.fig.canvas.get_tk_widget().after(100, self._force_resize_update)
            # Intercept the OS window-close (X button) on TkAgg so _on_close runs
            # before the window is destroyed (close_event alone does not fire for X).
            manager.window.protocol(
                'WM_DELETE_WINDOW',
                lambda: (self._on_close(None), plt.close(self.fig)))
        except AttributeError:
            try:
                manager = plt.get_current_fig_manager()
                manager.window.showMaximized()
                self.fig.canvas.mpl_connect('resize_event', self._on_resize)
            except AttributeError:
                pass

        # Main plot area — maximised to fill available space
        self.ax = self._axes(0.02, 0.21, 0.96, 0.69)
        self.ax.set_aspect('equal', adjustable='box')
        self.ax.set_facecolor('#FAFAF8')

        # DF% legend axes: top-right, just above the main image (rotated 90°)
        # DF legend — positioned in _setup_bottom_toolbar after buttons are laid out
        self.ax_legend = None

        # Setup side panel
        self._setup_side_panel()

        # Apply initial bevel styling to toggle buttons
        self._style_toggle_button(self.btn_edit_mode, self.edit_mode)
        self._style_toggle_button(self.btn_ortho, self.ortho_mode)
        self._update_room_type_buttons()

        # Initial render
        self._render_section()

        # Store original limits
        self.original_xlim = self.ax.get_xlim()
        self.original_ylim = self.ax.get_ylim()

        # Polygon selector for drawing — created inactive; 'd' activates draw mode
        self._create_polygon_selector()

        # Event handlers
        self.fig.canvas.mpl_connect('button_press_event', self._on_click_with_snap)
        self.fig.canvas.mpl_connect('button_press_event', self._on_list_click)
        self.fig.canvas.mpl_connect('button_release_event', self._on_button_release)
        self.fig.canvas.mpl_connect('motion_notify_event', self._on_mouse_motion)
        self.fig.canvas.mpl_connect('key_press_event', self._on_key_press)
        self.fig.canvas.mpl_connect('scroll_event', self._on_scroll)
        self.fig.canvas.mpl_connect('close_event', self._on_close)

        # Load existing session
        self._load_session()
        self._update_room_list()
        self._update_hdr_list()

        print("\n=== HDR Boundary Editor ===")
        print(f"Loaded {len(self.hdr_files)} HDR file(s) from {self.image_dir}")
        print("Use Up/Down to navigate HDR files, 't' to toggle image variant.")
        print("Scroll: zoom | Right-click: select room | s: save | d: delete | q: quit")
        print("===========================\n")
        plt.show()

    # === UI SETUP ==============================================================

    def _setup_side_panel(self):
        """Create the side panel with inputs, buttons, and room list."""
        self._btn_color    = '#E8E8E0'
        self._btn_hover    = '#D8D8D0'
        self._btn_on_color = '#8A8A84'
        self._btn_on_hover = '#7A7A74'

        self._setup_instructions_panel()
        status_y = self._setup_input_panels()
        action_bot_y = self._setup_action_buttons(status_y)
        self._setup_room_list_panel(action_bot_y)
        self._setup_bottom_toolbar()
        self._setup_colour_key_legend()

    # Layout constants shared across sub-methods
    _PL     = 0.02    # panel left
    _PW     = 0.28    # panel width
    _LBL_H  = 0.016
    _INP_H  = 0.032
    _GAP    = 0.008
    _SUB_H  = 0.014
    _ROW_Y  = 0.02    # top row y
    _INSTR_W = 0.12
    _COL2_X = _PL + _INSTR_W + 0.01
    _PRNT_X   = _COL2_X + 0.18
    _FIELD_W  = 0.150 * 0.75            # single field width (Parent / Name)
    _PRNT_W   = _FIELD_W * 2 + _GAP     # full row width (spans both fields)

    def _setup_instructions_panel(self):
        """Create the keyboard shortcut reference panel (top-left)."""
        ax = self._axes(self._PL, 0.02, self._INSTR_W, 0.15)
        ax.axis('off')
        ax.patch.set_visible(False)
        ax.text(0, 0.95, "HDR BOUNDARY EDITOR", fontsize=9, fontweight='bold',
                color='#404040', transform=ax.transAxes)
        controls = [
            ("\u2191/\u2193", "Navigate HDR files"),   ("t",           "Toggle image (HDR / TIFFs)"),
            ("Left-click",    "Place vertex / drag"),   ("Shift+click", "Drag edge (edit mode)"),
            ("Right-click",   "Select existing room"),  ("Scroll",      "Zoom centred on cursor"),
            ("s",             "Save room / confirm edit"),
            ("e",             "Toggle Edit Mode"),      ("d",           "Room divider (edit mode)"),
            ("ctrl+z",        "Undo (edit/type/name/create/del)"),
            ("o",             "Toggle orthogonal lines"), ("f",         "Fit zoom to selected room"),
            ("r",             "Reset zoom"),
            ("Ctrl+click",    "Multi-select rooms"),     ("ctrl+a",    "Select all rooms"),
            ("q",             "Quit"),
        ]
        for i, (key, desc) in enumerate(controls):
            y = 0.87 - i * 0.08
            ax.text(0.00, y, key,  fontsize=7.5, color='#404040', fontweight='bold', transform=ax.transAxes)
            ax.text(0.38, y, desc, fontsize=7.5, color='#505050', transform=ax.transAxes)

    def _setup_input_panels(self) -> float:
        """Create HDR nav, parent, name, room-type inputs. Returns status_y."""
        col2_x, row_y = self._COL2_X, self._ROW_Y
        lbl_h, inp_h, gap = self._LBL_H, self._INP_H, self._GAP
        prnt_x, prnt_w = self._PRNT_X, self._PRNT_W
        arrow_w = 0.025
        arrow_h = inp_h * 2 + gap

        # HDR FILES label + nav arrows + list
        self._make_label(col2_x, row_y, arrow_w * 2 + 0.004, lbl_h, "HDR FILES:")
        arrows_y = row_y + lbl_h + gap
        self.btn_next_hdr = self._make_button(col2_x, arrows_y, arrow_w, arrow_h,
                                              '\u25b2', self._on_next_hdr_click)
        self.btn_prev_hdr = self._make_button(col2_x + arrow_w + 0.004, arrows_y, arrow_w, arrow_h,
                                              '\u25bc', self._on_prev_hdr_click)
        hdr_list_x = col2_x + arrow_w * 2 + 0.008
        self.ax_hdr_list = self._axes(hdr_list_x, row_y, 0.060, lbl_h + gap + arrow_h)
        self.ax_hdr_list.axis('off')

        # Parent Apartment + Apartment Name (side by side)
        field_w = self._FIELD_W
        name_x  = prnt_x + field_w + gap

        self._make_label(prnt_x, row_y, field_w, lbl_h, "Parent Apartment:")
        self.btn_parent = self._make_button(prnt_x, row_y + lbl_h + gap, field_w, inp_h,
                                            '(None)', self._on_parent_cycle, fontsize=8)

        _, self.name_label_text = self._make_label(name_x, row_y, field_w, lbl_h, "Apartment Name:")
        ax_name = self._axes(name_x, row_y + lbl_h + gap, field_w, inp_h)
        self.name_textbox = TextBox(ax_name, '', initial='')
        self.name_textbox.on_text_change(self._on_name_changed)

        # Room Type (BED / LIVING / CIRC)
        rtype_y = row_y + lbl_h + gap + inp_h + gap
        self._make_label(prnt_x, rtype_y, prnt_w, lbl_h, "Room Type:")
        rtype_btn_y = rtype_y + lbl_h + gap
        n_type_btns = 3
        rtype_btn_w = (prnt_w - gap * (n_type_btns - 1)) / n_type_btns
        self.btn_room_type_bed = self._make_button(
            prnt_x, rtype_btn_y, rtype_btn_w, inp_h, 'BED',
            lambda e: self._on_room_type_toggle('BED'))
        self.btn_room_type_living = self._make_button(
            prnt_x + (rtype_btn_w + gap), rtype_btn_y, rtype_btn_w, inp_h, 'LIVING',
            lambda e: self._on_room_type_toggle('LIVING', requires_children=True))
        self.btn_room_type_circ = self._make_button(
            prnt_x + 2 * (rtype_btn_w + gap), rtype_btn_y, rtype_btn_w, inp_h, 'CIRC',
            lambda e: self._on_room_type_toggle('CIRCULATION'))

        # Status / preview
        status_y = rtype_y + lbl_h + gap + inp_h + gap
        ax_preview = self._axes(prnt_x, status_y, prnt_w, self._SUB_H)
        ax_preview.axis('off')
        self.name_preview_text = ax_preview.text(0, 0.5, "", fontsize=8, color='#666666', style='italic')
        ax_status = self._axes(prnt_x, status_y, prnt_w, self._SUB_H)
        ax_status.axis('off')
        self.status_text = ax_status.text(
            0, 0.5, "Status: Ready to draw", fontsize=8, color='blue', style='italic')
        return status_y

    def _setup_action_buttons(self, status_y: float) -> float:
        """Create Save / Clear / Delete button row below status. Returns bottom y."""
        gap    = self._GAP
        prnt_x = self._PRNT_X
        prnt_w = self._PRNT_W
        btn_y  = status_y + self._SUB_H + gap
        btn_h  = self._INP_H
        n_btns = 3
        btn_w  = (prnt_w - gap * (n_btns - 1)) / n_btns

        for i, (attr, label, cb) in enumerate([
            ('btn_save',   'Save',   self._on_save_click),
            ('btn_clear',  'Clear',  self._on_clear_click),
            ('btn_delete', 'Delete', self._on_delete_click),
        ]):
            setattr(self, attr, self._make_button(
                prnt_x + i * (btn_w + gap), btn_y, btn_w, btn_h, label, cb))
        return btn_y + btn_h

    def _setup_room_list_panel(self, action_bot_y: float):
        """Create the scrollable saved-rooms list (right of input column)."""
        gap     = self._GAP
        list_x  = self._PRNT_X + self._PRNT_W + gap
        list_w  = 0.270
        list_y  = self._ROW_Y
        label_h = 0.020
        self._make_label(list_x, list_y, list_w, label_h, "SAVED ROOMS:")
        list_top = list_y + label_h + gap * 0.5
        list_h   = action_bot_y - list_top
        self.ax_list = self._axes(list_x, list_top, list_w, list_h)
        self.ax_list.set_facecolor('#FAFAF8')
        self.ax_list.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
        for spine in self.ax_list.spines.values():
            spine.set_edgecolor('#CCCCCC')
            spine.set_linewidth(0.5)

    def _setup_bottom_toolbar(self):
        """Create the two-row bottom toolbar with DF legend filling the right."""
        gap   = self._GAP
        tr_x  = 0.02                     # push buttons to far left
        tr_h  = 0.025                     # single row height
        row1_y = 0.91                     # top row
        row2_y = row1_y + tr_h + gap * 2    # bottom row (extra spacing for button shadows)

        # Button width: 4 buttons across on top row
        n_cols   = 4
        btn_w    = 0.130
        btn_step = btn_w + gap

        # Row 1: Toggle, Edit, Draw, Ortho
        self.btn_image_toggle = self._make_button(
            tr_x, row1_y, btn_w, tr_h,
            'Toggle Image Layers: HDR (Press T)', self._on_image_toggle_click)
        self.btn_edit_mode = self._make_button(
            tr_x + btn_step, row1_y, btn_w, tr_h,
            'Edit Mode: OFF (Press E)', self._on_edit_mode_toggle)
        self.btn_draw_mode = self._make_button(
            tr_x + 2 * btn_step, row1_y, btn_w, tr_h,
            'Draw Mode: OFF (Press D)', self._on_draw_mode_toggle)

        ortho_label = 'Ortho Lines: ON (Press O)' if self.ortho_mode else 'Ortho Lines: OFF (Press O)'
        ortho_color = self._btn_on_color if self.ortho_mode else self._btn_color
        ortho_hover = self._btn_on_hover if self.ortho_mode else self._btn_hover
        self.btn_ortho = self._make_button(
            tr_x + 3 * btn_step, row1_y, btn_w, tr_h,
            ortho_label, self._on_ortho_toggle, color=ortho_color, hovercolor=ortho_hover)

        # Row 2: Export & Archive, Extract Archive, Reset Zoom
        self.btn_export = self._make_button(
            tr_x, row2_y, btn_w, tr_h,
            'Export & Archive', self._on_export_report, color='#C8E6C9', hovercolor='#A5D6A7')
        self.btn_extract = self._make_button(
            tr_x + btn_step, row2_y, btn_w, tr_h,
            'Extract Archive', self._on_extract_click)
        self.btn_reset_zoom = self._make_button(
            tr_x + 2 * btn_step, row2_y, btn_w, tr_h,
            'Reset Zoom', self._on_reset_zoom_click)

        # DF% legend — starts after all top-row buttons, no overlap
        legend_x = tr_x + n_cols * btn_step
        legend_w = 0.98 - legend_x
        legend_h = 0.99 - row1_y
        self.ax_legend = self._axes(legend_x, row1_y, legend_w, legend_h)
        self.ax_legend.axis('off')
        self.ax_legend.set_visible(False)

        # Progress bar (hidden until export)
        prog_y = row2_y + tr_h + gap * 0.5
        self.ax_progress = self._axes(tr_x, prog_y, 0.96, 0.010)
        self.ax_progress.set_xlim(0, 1)
        self.ax_progress.set_ylim(0, 1)
        self.ax_progress.axis('off')
        self.ax_progress.set_visible(False)
        self._progress_bar_patch = None
        self._progress_text = None

    def _setup_colour_key_legend(self):
        """Create the colour-key legend (far right, vertical stack)."""
        ax_legend = self._axes(0.88, 0.02, 0.11, 0.14)
        ax_legend.axis('off')
        ax_legend.set_facecolor('#F0F0EC')
        ax_legend.text(0.05, 0.95, "LEGEND", fontsize=7, fontweight='bold', color='#404040',
                       va='top', transform=ax_legend.transAxes)
        items = [
            ('red',     0.6, 'Room boundary'),
            ('yellow',  0.5, 'Selected'),
            ('cyan',    0.5, 'Being edited'),
            ('magenta', 0.5, 'Hover (edit mode)'),
        ]
        n = len(items)
        for i, (color, alpha, label) in enumerate(items):
            y0 = 0.78 - i * (0.75 / max(n - 1, 1))
            rect = FancyBboxPatch((0.05, y0 - 0.04), 0.08, 0.08,
                                  boxstyle='round,pad=0.01',
                                  facecolor=color, edgecolor=color, alpha=alpha,
                                  transform=ax_legend.transAxes, clip_on=False)
            ax_legend.add_patch(rect)
            ax_legend.text(0.18, y0, label, fontsize=6, color='#404040',
                           va='center', transform=ax_legend.transAxes)

    # === HDR NAVIGATION ========================================================

    def _on_prev_hdr_click(self, event):
        """Navigate to the previous HDR file (lower index)."""
        if not self.hdr_files:
            self._update_status("No HDR files loaded", 'red')
            return
        if self.current_hdr_idx > 0:
            self._jump_to_hdr(self.current_hdr_idx - 1)
        else:
            self._update_status("Already at first HDR file", 'orange')

    def _on_next_hdr_click(self, event):
        """Navigate to the next HDR file (higher index)."""
        if not self.hdr_files:
            self._update_status("No HDR files loaded", 'red')
            return
        if self.current_hdr_idx < len(self.hdr_files) - 1:
            self._jump_to_hdr(self.current_hdr_idx + 1)
        else:
            self._update_status("Already at last HDR file", 'orange')

    def _jump_to_hdr(self, hdr_idx: int):
        """Switch to the given HDR file index."""
        if self.divider_mode:
            self._exit_divider_mode(cancelled=True)
        if not (0 <= hdr_idx < len(self.hdr_files)):
            return
        self.current_hdr_idx = hdr_idx
        self._rebuild_image_variants()
        self._update_image_toggle_label()
        self.selected_parent = None
        self.btn_parent.label.set_text('(None)')
        self.name_label_text.set_text("Apartment Name:")
        self._update_status(f"HDR: {self.current_hdr_name}", 'green')
        self._update_hdr_list()
        self._update_room_list()
        self._render_section(reset_view=True, force_full=True)
        self._create_polygon_selector()

    def _update_hdr_list(self):
        """Refresh the HDR file list in the side panel."""
        self.ax_hdr_list.clear()
        self.ax_hdr_list.axis('off')

        if not self.hdr_files:
            self.ax_hdr_list.text(0, 0.95, "(no HDR files)", fontsize=8, style='italic', color='gray')
        else:
            max_display = 8
            total = len(self.hdr_files)
            for display_idx in range(min(max_display, total)):
                i    = total - 1 - display_idx  # descending order (latest first)
                name = self.hdr_files[i]['name']
                y_pos = 0.95 - (display_idx * 0.12)
                is_current = (self.current_hdr_idx == i)
                room_count = sum(1 for r in self.rooms if r.get('hdr_file') == name)
                indicator  = "*" if is_current else "o"
                text       = f"{indicator} {name}"
                if room_count > 0:
                    text += f" ({room_count})"
                self.ax_hdr_list.text(
                    0, y_pos, text, fontsize=7,
                    fontweight='bold' if is_current else 'normal',
                    color='green' if is_current else 'darkgray',
                )
            if total > max_display:
                self.ax_hdr_list.text(0, 0.02, f"... and {total - max_display} more",
                                      fontsize=7, style='italic')

        self.fig.canvas.draw_idle()

    # === IMAGE TOGGLE ==========================================================

    def _on_image_toggle_click(self, event):
        """Cycle through image variants (HDR then associated TIFFs)."""
        if not self.image_variants:
            return
        self.current_variant_idx = (self.current_variant_idx + 1) % len(self.image_variants)
        self._update_image_toggle_label()
        self._render_section(force_full=True)

    def _update_image_toggle_label(self):
        """Update the toggle button label to show the active variant."""
        if not self.image_variants:
            self.btn_image_toggle.label.set_text('Toggle Image Layers: (none)')
            return
        idx  = self.current_variant_idx % len(self.image_variants)
        path = self.image_variants[idx]
        # Show a short label: HDR stem or TIFF suffix portion
        if path.suffix.lower() == '.hdr':
            label = 'Toggle Image Layers: HDR (Press T)'
        else:
            # e.g. model_plan_ffl_25300_df_false → show "df_false"
            hdr_stem = self.hdr_files[self.current_hdr_idx]['name']
            suffix   = path.stem[len(hdr_stem):]  # e.g. "_df_false"
            suffix   = suffix.lstrip('_')
            label    = f'Toggle Image Layers: {suffix} (Press T)'
        self.btn_image_toggle.label.set_text(label)
        self.fig.canvas.draw_idle()

    # === ROOM DATA MANAGEMENT ==================================================

    def _get_apartments_for_hdr(self, hdr_name: str) -> List[str]:
        """Return names of apartment-level rooms (no parent) for the given HDR file."""
        return [
            room['name'] for room in self.rooms
            if room.get('parent') is None and room.get('hdr_file') == hdr_name
        ]

    def _get_children(self, parent_name: str) -> List[dict]:
        """Return all rooms that have the given parent."""
        return [r for r in self.rooms if r.get('parent') == parent_name]

    def _get_parent_room(self, parent_name: str) -> Optional[dict]:
        """Return the room dict for the given parent name, or None."""
        for r in self.rooms:
            if r.get('name') == parent_name and r.get('parent') is None:
                return r
        return None

    # --- Name helpers ---

    def _make_unique_name(self, base_name: str, exclude_idx: Optional[int] = None) -> str:
        """Ensure room name is unique by appending numeric suffix if needed."""
        existing = {r['name'] for i, r in enumerate(self.rooms) if i != exclude_idx}
        if base_name not in existing:
            return base_name

        match = re.match(r'^(.*?)(\d+)$', base_name)
        root  = match.group(1) if match else base_name
        counter = 1
        while f"{root}{counter}" in existing:
            counter += 1
        return f"{root}{counter}"

    def _enforce_unique_names(self) -> int:
        """Ensure all room names are unique; returns count of rooms renamed."""
        seen       = set()
        renamed    = 0
        for room in self.rooms:
            name = room['name']
            if name not in seen:
                seen.add(name)
                continue
            match = re.match(r'^(.*?)(\d+)$', name)
            root  = match.group(1) if match else name
            counter = 1
            while f"{root}{counter}" in seen:
                counter += 1
            new_name = f"{root}{counter}"
            print(f"Renamed duplicate '{name}' -> '{new_name}'")
            room['name'] = new_name
            seen.add(new_name)
            renamed += 1
        return renamed

    def _check_boundary_containment(self, child_verts, parent_verts) -> bool:
        """Return True if all child vertices lie inside the parent polygon."""
        if not parent_verts or len(parent_verts) < 3:
            return True
        if not child_verts or len(child_verts) < 3:
            return True
        arr = np.array(parent_verts)
        if not np.allclose(arr[0], arr[-1]):
            arr = np.vstack([arr, arr[0]])
        path = MplPath(arr)
        return all(path.contains_point(v) for v in child_verts)

    def _update_parent_options(self):
        """Update the list of available parent apartments for the current HDR file."""
        self.parent_options = self._get_apartments_for_hdr(self.current_hdr_name)

    def _on_parent_cycle(self, event):
        """Cycle through parent apartment options."""
        self._update_parent_options()
        if not self.parent_options:
            self.selected_parent = None
            self.btn_parent.label.set_text('(None - New Apartment)')
            self.name_label_text.set_text("Apartment Name (Space ID):")
        else:
            options = [None] + self.parent_options
            try:
                current_idx = options.index(self.selected_parent)
                next_idx    = (current_idx + 1) % len(options)
            except ValueError:
                next_idx = 0
            self.selected_parent = options[next_idx]
            if self.selected_parent is None:
                self.btn_parent.label.set_text('(None - New Apartment)')
                self.name_label_text.set_text("Apartment Name (Space ID):")
            else:
                self.btn_parent.label.set_text(self.selected_parent)
                self.name_label_text.set_text("Room Name:")
        self._update_name_preview()
        self.fig.canvas.draw_idle()

    def _on_name_changed(self, _text):
        """Update name preview when name textbox changes."""
        self._update_name_preview()

    def _update_name_preview(self):
        """Update the name preview text, hiding status while preview is shown."""
        name = self.name_textbox.text.strip().upper()
        if not name:
            self.name_preview_text.set_text("")
            self.status_text.set_visible(True)
        elif self.selected_parent:
            self.name_preview_text.set_text(f"Will save as: {self.selected_parent}_{name}")
            self.status_text.set_visible(False)
        else:
            self.name_preview_text.set_text(f"Will save as: {name}")
            self.status_text.set_visible(False)
        self.fig.canvas.draw_idle()

    # === RENDERING =============================================================

    def _render_section(self, reset_view: bool = False, force_full: bool = False):
        """Render the current image and room overlays."""
        xlim = self.ax.get_xlim()
        ylim = self.ax.get_ylim()

        need_full = force_full or reset_view
        if need_full:
            self._do_full_render(xlim, ylim, reset_view)
        else:
            self._update_room_visuals()
            self.fig.canvas.draw_idle()

    def _do_full_render(self, xlim, ylim, reset_view: bool):
        """Complete redraw of image and room polygons."""
        # Disconnect old selector BEFORE ax.clear() destroys its artists,
        # so it can't respond to events in a broken state.
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.disconnect_events()
            self.selector = None
        self.ax.clear()
        self._room_patch_cache.clear()
        self._room_label_cache.clear()
        self._df_text_cache.clear()
        self._edit_vertex_scatter = None
        self._image_handle = None

        # Load and display current image
        path = self.current_variant_path
        if path is not None:
            img = self._load_image(path)
        else:
            img = None

        if img is not None:
            H, W = img.shape[:2]
            self._image_width  = W
            self._image_height = H
            self._reference_view_w = float(W)
            self._image_handle = self.ax.imshow(
                img, origin='upper',
                extent=[0, W, H, 0],
                aspect='equal', zorder=0,
            )
            self.ax.set_xlim(0, W)
            self.ax.set_ylim(H, 0)
        else:
            # Blank canvas
            self.ax.set_xlim(0, 1000)
            self.ax.set_ylim(1000, 0)

        # Rebuild snap vertex + edge pools from current-HDR room vertices
        all_verts, edge_starts, edge_ends = [], [], []
        for room in self.rooms:
            if self._is_room_on_current_hdr(room):
                verts = room['vertices']
                all_verts.extend(verts)
                n = len(verts)
                for j in range(n):
                    edge_starts.append(verts[j])
                    edge_ends.append(verts[(j + 1) % n])
        self.current_vertices = np.array(all_verts) if all_verts else np.array([])
        self._snap_edge_starts = np.array(edge_starts, dtype=float) if edge_starts else None
        self._snap_edge_ends = np.array(edge_ends, dtype=float) if edge_ends else None

        # Load DF image for current HDR (cached)
        self._load_current_df_image()
        self._compute_all_room_df_results()

        # Draw room polygons
        self._draw_all_room_polygons()

        # Pre-build vectorized hover arrays for fast hit-testing
        self._rebuild_hover_arrays()

        self.ax.set_aspect('equal', adjustable='box')
        self.ax.axis('off')

        # Restore or reset zoom
        if reset_view or not hasattr(self, 'original_xlim') or xlim == (0.0, 1.0):
            self.ax.autoscale()
        elif img is not None and xlim != (0.0, 1.0):
            self.ax.set_xlim(xlim)
            self.ax.set_ylim(ylim)

        # Reapply zoom-aware font sizes now that limits are restored
        self._apply_zoom_fontsizes()

        self._render_df_legend()

        # Title
        hdr_name = self.current_hdr_name or "(no HDR)"
        variant  = self.current_variant_path
        variant_label = variant.stem if variant else ""
        # Title removed to maximise image area

        # Always recreate the polygon selector after ax.clear() destroyed the old one
        if not self.edit_mode:
            self._create_polygon_selector()

        self.fig.canvas.draw_idle()

    def _render_df_legend(self):
        """Show the DF% legend strip in the bottom toolbar."""
        if self.ax_legend is None:
            return
        self.ax_legend.clear()
        legend_path = self._get_legend_for_variant(self.current_variant_path)
        if legend_path is None:
            self.ax_legend.set_visible(False)
            return
        legend_img = self._load_image(legend_path)
        if legend_img is None:
            self.ax_legend.set_visible(False)
            return
        legend_img = np.rot90(legend_img, k=1)
        lH, lW = legend_img.shape[:2]
        # Scale up 30% while preserving aspect ratio
        scale = 1.3
        sW, sH = lW * scale, lH * scale
        self.ax_legend.imshow(legend_img, origin='upper', extent=[0, sW, sH, 0], aspect='equal')
        self.ax_legend.set_xlim(0, sW)
        self.ax_legend.set_ylim(sH, 0)
        self.ax_legend.axis('off')
        self.ax_legend.set_visible(True)

    def _load_current_df_image(self):
        """Load the DF% image for the current HDR file (cached)."""
        if self._hdr2wpd is None or not self.hdr_files:
            self._df_image = None
            return
        hdr_path = self.hdr_files[self.current_hdr_idx]['hdr_path']
        key = str(hdr_path)
        if key in self._df_image_cache:
            self._df_image = self._df_image_cache[key]
            return
        self._df_image = Hdr2Wpd.load_df_image(hdr_path)
        if self._df_image is not None:
            self._df_image_cache[key] = self._df_image

    @staticmethod
    def _vertices_hash(vertices: list) -> str:
        """Return a compact hash string for a list of vertex coordinates."""
        return str([(round(v[0], 2), round(v[1], 2)) for v in vertices])

    def _effective_room_type(self, room: dict) -> Optional[str]:
        """Return the effective room type for a room (sub-rooms inherit parent's type)."""
        rtype = room.get('room_type')
        if rtype:
            return rtype
        parent_name = room.get('parent')
        if parent_name:
            for r in self.rooms:
                if r.get('name') == parent_name:
                    return r.get('room_type')
        return None

    def _threshold_for_type(self, room_type: Optional[str]) -> Optional[float]:
        """Return the DF threshold for a given room type, or None if untagged."""
        return self.DF_THRESHOLDS.get(room_type)

    def _compute_all_room_df_results(self):
        """Compute DF threshold results for every room on the current HDR floor.

        Each room type has a single fixed threshold (BED=0.5%, LIVING=1.0%).
        Only rooms with a type (or inherited type) get results computed.
        """
        self._room_df_results.clear()
        if self._df_image is None or self._hdr2wpd is None:
            return
        any_new = False
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room['vertices']
            if len(verts) < 3:
                continue
            eff_type = self._effective_room_type(room)
            threshold = self._threshold_for_type(eff_type)
            if threshold is None:
                continue
            had_cache = bool(room.get('df_cache'))
            lines = self._compute_room_df(room, verts, (threshold,))
            self._room_df_results[i] = lines
            if not had_cache and room.get('df_cache'):
                any_new = True
        # Persist newly computed DF caches to the session JSON
        if any_new:
            self._save_session()

    def _compute_room_df(self, room: dict, verts: list, thresholds_key: tuple) -> list:
        """Compute DF results for a single room, using its cached df_cache if valid.

        For parent rooms with children, sub-room areas are subtracted so the
        result represents only the remaining (non-sub-room) area.
        """
        # Build hash including child vertices for parent rooms
        children = self._get_children(room.get('name', ''))
        child_verts_list = [c['vertices'] for c in children if len(c.get('vertices', [])) >= 3]

        verts_hash = self._vertices_hash(verts)
        if child_verts_list:
            child_hashes = tuple(self._vertices_hash(cv) for cv in child_verts_list)
            verts_hash = hash((verts_hash, child_hashes))

        _DF_DISPLAY_FMT = 2  # bump to invalidate cached display_lines
        cache = room.get('df_cache')
        if (cache
                and cache.get('vertices_hash') == verts_hash
                and tuple(cache.get('thresholds', [])) == thresholds_key):
            if cache.get('display_fmt') == _DF_DISPLAY_FMT:
                return cache['display_lines']
            # Geometry unchanged — regenerate display lines from cached raw_result
            if cache.get('raw_result'):
                lines = self._format_df_lines(cache['raw_result'])
                cache['display_lines'] = lines
                cache['display_fmt'] = _DF_DISPLAY_FMT
                return lines

        # Cache miss — recompute
        if child_verts_list:
            result = self._hdr2wpd.compute_df_for_polygon_excluding(
                self._df_image, verts, child_verts_list, thresholds_key)
        else:
            result = self._hdr2wpd.compute_df_for_polygon(
                self._df_image, verts, thresholds_key)
        lines = self._format_df_lines(result)
        room['df_cache'] = {
            'vertices_hash':  verts_hash,
            'thresholds':     list(thresholds_key),
            'display_fmt':    _DF_DISPLAY_FMT,
            'display_lines':  lines,
            'raw_result':     result,
        }
        return lines

    @staticmethod
    def _format_df_lines(result: dict) -> list:
        """Format DF results as display lines: area (%) then threshold."""
        total_area = result['total_area_m2']
        lines = []
        for tr in result['thresholds']:
            pct = (tr['area_m2'] / total_area * 100) if total_area > 0 else 0.0
            lines.append(f"{tr['area_m2']:.1f} m\u00b2 ({pct:.0f}%)")
            lines.append(f"@ {tr['threshold']:g}% DF")
        return lines

    def _invalidate_room_df_cache(self, room_idx: int):
        """Clear the DF cache for a specific room, forcing recomputation.

        Also invalidates the parent room's cache if the edited room is a
        sub-room, since LIVING-type parents depend on child geometry.
        """
        if 0 <= room_idx < len(self.rooms):
            room = self.rooms[room_idx]
            room.pop('df_cache', None)
            # Invalidate parent if this is a sub-room
            parent_name = room.get('parent')
            if parent_name:
                for i, r in enumerate(self.rooms):
                    if r.get('name') == parent_name:
                        r.pop('df_cache', None)
                        break

    def _draw_all_room_polygons(self):
        """Draw all room polygons for the current view."""
        for i, room in enumerate(self.rooms):
            if self._is_room_on_current_hdr(room):
                self._draw_room_polygon(room, i, is_current_floor=True)

    def _draw_room_polygon(self, room: dict, idx: int, is_current_floor: bool):
        """Draw a single room polygon with its label."""
        verts = room['vertices']
        if len(verts) < 3:
            return

        is_selected = (idx == self.selected_room_idx or idx in self.multi_selected_room_idxs)
        is_hover    = (idx == self.hover_room_idx)
        is_editing  = (idx == self.edit_room_idx and self.edit_mode)
        is_subroom  = room.get('parent') is not None
        is_div      = '_DIV' in room.get('name', '')

        if is_editing:
            edge_color, lw, linestyle, alpha = 'cyan',    3, '-',  1.0
        elif is_selected:
            edge_color, lw, linestyle, alpha = 'yellow',  4, '-',  1.0
        elif is_hover and self.edit_mode:
            edge_color, lw, linestyle, alpha = 'magenta', 2, '-',  1.0
        elif is_div and is_current_floor:
            edge_color, lw, linestyle, alpha = 'red',     self._zoom_linewidth(), '--', 0.6
        elif is_current_floor:
            edge_color, lw, linestyle, alpha = 'red',     self._zoom_linewidth(), '-',  1.0
        else:
            edge_color, lw, linestyle, alpha = 'gray',    self._zoom_linewidth(base=1.0), '-', 1.0

        poly = Polygon(verts, closed=True,
                       edgecolor=edge_color, facecolor='none', alpha=alpha, linewidth=lw,
                       linestyle=linestyle, clip_on=True)
        poly._patch_type = 'current' if is_current_floor else 'other'
        self.ax.add_patch(poly)
        self._room_patch_cache[idx] = poly

        # Edit mode: draw vertex handles
        if self.edit_mode and is_current_floor:
            verts_array = np.array(verts)
            xs, ys, colors, sizes = [], [], [], []
            for v_idx, (vx, vy) in enumerate(verts_array):
                is_hovered  = (idx == self.hover_room_idx  and v_idx == self.hover_vertex_idx)
                is_dragging = (idx == self.edit_room_idx   and v_idx == self.edit_vertex_idx)
                if is_dragging:
                    c, s = 'yellow', 9
                elif is_hovered:
                    c, s = 'red', 8
                else:
                    c, s = 'cyan', 5
                xs.append(vx); ys.append(vy); colors.append(c); sizes.append(s ** 2)
            if xs:
                self.ax.scatter(xs, ys, c=colors, s=sizes,
                                edgecolors='black', linewidths=1.0, zorder=100, picker=5)

            # Highlight hovered edge
            if idx == self.hover_edge_room_idx and self.hover_edge_idx is not None:
                n = len(verts)
                j = self.hover_edge_idx
                ex = [verts[j][0], verts[(j + 1) % n][0]]
                ey = [verts[j][1], verts[(j + 1) % n][1]]
                self.ax.plot(ex, ey, '-', color='lime', linewidth=3, zorder=99, alpha=0.8)
                if self.hover_edge_point is not None:
                    self.ax.plot([self.hover_edge_point[0]], [self.hover_edge_point[1]], 'o',
                                 markersize=8, color='lime', markeredgecolor='darkgreen',
                                 markeredgewidth=1.5, zorder=101, alpha=0.9)

        # Label — use true centroid (centre of mass), guaranteed inside polygon
        centroid = self._polygon_label_point(verts)
        label    = room.get('name', '')
        if not is_current_floor:
            hf = room.get('hdr_file', '')
            label += f"\n({hf})"

        fs_name = self._zoom_fontsize()
        label_text = self.ax.text(
            centroid[0], centroid[1], label,
            color='red', fontsize=fs_name,
            ha='center', va='center', clip_on=True,
            path_effects=[patheffects.withStroke(linewidth=fs_name * 0.06, foreground='black')],
        )
        label_text.set_clip_path(self.ax.patch)
        self._room_label_cache[idx] = label_text

        # DF results as smaller subtext below the room name (only if room type is set)
        df_lines = self._room_df_results.get(idx, [])
        if df_lines and is_current_floor:
            line_step = self._df_line_step()
            for line_i, line in enumerate(df_lines):
                dy = 1.2 * line_step + line_i * line_step * 0.7
                fs_df = self._zoom_fontsize(base=6.5)
                df_text = self.ax.text(
                    centroid[0], centroid[1] + dy, line,
                    color='red', fontsize=fs_df,
                    ha='center', va='center', clip_on=True, alpha=0.85,
                    path_effects=[patheffects.withStroke(linewidth=fs_df * 0.06, foreground='black')],
                )
                df_text.set_clip_path(self.ax.patch)
                # Store centroid + line index for zoom repositioning
                df_text._df_centroid = centroid
                df_text._df_line_i  = line_i
                self._df_text_cache.append(df_text)

    def _update_room_visuals(self):
        """Update room colours without a full redraw (for hover/selection changes)."""
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            patch = self._room_patch_cache.get(i)
            if patch is None:
                continue
            is_selected = (i == self.selected_room_idx or i in self.multi_selected_room_idxs)
            is_hover    = (i == self.hover_room_idx)
            is_editing  = (i == self.edit_room_idx and self.edit_mode)
            is_subroom  = room.get('parent') is not None

            if is_editing:
                patch.set_edgecolor('cyan');    patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(3)
                patch._patch_type = 'special'
            elif is_selected:
                patch.set_edgecolor('yellow');  patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(4)
                patch._patch_type = 'special'
            elif is_hover and self.edit_mode:
                patch.set_edgecolor('magenta'); patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(2)
                patch._patch_type = 'special'
            else:
                patch.set_edgecolor('red');     patch.set_facecolor('none'); patch.set_alpha(1.0); patch.set_linewidth(self._zoom_linewidth())
                patch._patch_type = 'current'

        self.fig.canvas.draw_idle()

    # === UNDO & SNAPPING =======================================================

    def _push_undo(self, room_idx: int):
        """Push a snapshot of a room's vertices onto the undo stack."""
        verts_copy = [list(v) for v in self.rooms[room_idx]['vertices']]
        self._edit_undo_stack.append((room_idx, verts_copy))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)

    def _undo_edit(self):
        """Restore the last undone vertex state (Ctrl+Z in edit mode)."""
        if not self.edit_mode:
            return
        if not self._edit_undo_stack:
            self._update_status("Nothing to undo", 'orange')
            return
        entry = self._edit_undo_stack.pop()

        # Divider undo: full rooms-list snapshot
        if isinstance(entry, tuple) and len(entry) == 2 and entry[0] == 'divider':
            self.rooms = entry[1]
            self._reset_hover_state()
            self._update_status("Undid room division", 'blue')
            self._save_session()
            self._update_room_list()
            self._update_hdr_list()
            self._render_section(force_full=True)
            return

        room_idx, verts_snapshot = entry
        self.rooms[room_idx]['vertices'] = verts_snapshot
        self._invalidate_room_df_cache(room_idx)
        self._reset_hover_state()
        room_name = self.rooms[room_idx].get('name', 'unnamed')
        remaining = len(self._edit_undo_stack)
        self._update_status(
            f"Undid change to '{room_name}' ({remaining} undo{'s' if remaining != 1 else ''} left)", 'blue')
        self._save_session()
        self._render_section(force_full=True)

    def _push_draw_undo(self, entry: tuple):
        """Push an entry onto the draw-mode undo stack."""
        self._draw_undo_stack.append(entry)
        if len(self._draw_undo_stack) > self._draw_undo_max:
            self._draw_undo_stack.pop(0)

    def _undo_draw(self):
        """Undo the last draw-mode operation (Ctrl+Z outside edit mode)."""
        if not self._draw_undo_stack:
            self._update_status("Nothing to undo", 'orange')
            return
        entry = self._draw_undo_stack.pop()
        tag   = entry[0]

        if tag == 'delete':
            _, idx, room = entry
            idx = min(idx, len(self.rooms))
            self.rooms.insert(idx, room)
            name = room.get('name', 'unnamed')
            self._update_status(f"Restored '{name}'", 'blue')
            print(f"Restored deleted room '{name}'")

        elif tag == 'type':
            _, old_types = entry
            for idx, old_type in old_types:
                if idx < len(self.rooms):
                    self.rooms[idx]['room_type'] = old_type
                    self._invalidate_room_df_cache(idx)
            count = len(old_types)
            self._update_status(f"Undid type change ({count} room{'s' if count != 1 else ''})", 'blue')

        elif tag == 'rename':
            _, idx, old_name = entry
            if idx < len(self.rooms):
                self.rooms[idx]['name'] = old_name
                self._update_status(f"Undid rename → '{old_name}'", 'blue')

        elif tag == 'create':
            _, idx, parent_type_changed = entry
            if idx < len(self.rooms):
                name = self.rooms[idx].get('name', 'unnamed')
                self.rooms.pop(idx)
                self.selected_room_idx = None
                # Revert auto-assigned parent type if applicable
                if parent_type_changed is not None:
                    p_idx, p_old_type = parent_type_changed
                    if p_idx < len(self.rooms):
                        self.rooms[p_idx]['room_type'] = p_old_type
                        self._invalidate_room_df_cache(p_idx)
                self._update_status(f"Undid creation of '{name}'", 'blue')

        else:
            self._update_status("Unknown undo entry", 'red')
            return

        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)

    def _snap_to_vertex(self, x: float, y: float) -> tuple:
        """Snap to nearest existing polygon vertex within snap_distance_px pixels."""
        if len(self.current_vertices) == 0:
            return x, y
        dists = np.hypot(self.current_vertices[:, 0] - x, self.current_vertices[:, 1] - y)
        min_idx  = int(np.argmin(dists))
        min_dist = dists[min_idx]
        if min_dist <= self._snap_distance_px:
            return float(self.current_vertices[min_idx, 0]), float(self.current_vertices[min_idx, 1])
        return x, y

    def _snap_to_edge(self, x: float, y: float) -> tuple:
        """Snap to nearest existing polygon edge within snap_distance_px pixels."""
        if self._snap_edge_starts is None or self._snap_edge_ends is None:
            return x, y
        A = self._snap_edge_starts
        B = self._snap_edge_ends
        AB = B - A
        AP = np.array([[x, y]]) - A
        seg_len_sq = np.sum(AB * AB, axis=1)
        safe_len = np.where(seg_len_sq == 0, 1.0, seg_len_sq)
        t = np.clip(np.sum(AP * AB, axis=1) / safe_len, 0.0, 1.0)
        t = np.where(seg_len_sq == 0, 0.0, t)
        proj = A + t[:, None] * AB
        dists = np.hypot(x - proj[:, 0], y - proj[:, 1])
        min_idx = int(np.argmin(dists))
        if dists[min_idx] <= self._snap_distance_px:
            return float(proj[min_idx, 0]), float(proj[min_idx, 1])
        return x, y

    def _point_to_segment_dist(self, px, py, ax, ay, bx, by):
        """Return (distance, proj_x, proj_y) from point P to segment A→B."""
        dx, dy    = bx - ax, by - ay
        seg_len_sq = dx*dx + dy*dy
        if seg_len_sq == 0:
            return np.hypot(px - ax, py - ay), ax, ay
        t = max(0.0, min(1.0, ((px - ax)*dx + (py - ay)*dy) / seg_len_sq))
        proj_x, proj_y = ax + t*dx, ay + t*dy
        return np.hypot(px - proj_x, py - proj_y), proj_x, proj_y

    @staticmethod
    def _polygon_label_point(verts) -> np.ndarray:
        """Return a point guaranteed to be inside the polygon for label placement.

        Uses the true geometric centroid (centre of mass via the shoelace
        formula) rather than the simple average of vertices.  If the centroid
        falls outside a concave polygon, falls back to scanning interior
        candidate points to find one that is inside and as far from the edges
        as possible.
        """
        pts = np.array(verts, dtype=float)
        n = len(pts)
        if n < 3:
            return pts.mean(axis=0)

        # --- Signed area via shoelace formula ---
        x, y = pts[:, 0], pts[:, 1]
        x1, y1 = np.roll(x, -1), np.roll(y, -1)
        cross = x * y1 - x1 * y
        signed_area = cross.sum() / 2.0

        if abs(signed_area) < 1e-10:
            # Degenerate polygon — fall back to vertex average
            return pts.mean(axis=0)

        # --- True centroid (centre of mass) ---
        cx = ((x + x1) * cross).sum() / (6.0 * signed_area)
        cy = ((y + y1) * cross).sum() / (6.0 * signed_area)
        centroid = np.array([cx, cy])

        # Check if centroid is inside the polygon
        path = MplPath(pts)
        if path.contains_point(centroid):
            return centroid

        # --- Fallback: find best interior point for concave polygons ---
        # Sample candidate points on a grid inside the bounding box and pick
        # the one with the greatest distance to the nearest edge (visual
        # "pole of inaccessibility" approximation).
        xmin, ymin = pts.min(axis=0)
        xmax, ymax = pts.max(axis=0)
        w, h = xmax - xmin, ymax - ymin
        # Use ~20 steps along the longer dimension
        steps = 20
        xs = np.linspace(xmin, xmax, max(steps, 4))
        ys = np.linspace(ymin, ymax, max(steps, 4))
        grid_x, grid_y = np.meshgrid(xs, ys)
        candidates = np.column_stack([grid_x.ravel(), grid_y.ravel()])

        inside_mask = path.contains_points(candidates)
        inside_pts = candidates[inside_mask]

        if len(inside_pts) == 0:
            # Extremely unlikely — return vertex average as last resort
            return pts.mean(axis=0)

        # For each inside point, compute min distance to any polygon edge
        best_point = inside_pts[0]
        best_dist  = -1.0
        for cp in inside_pts:
            min_d = np.inf
            for i in range(n):
                j = (i + 1) % n
                a, b = pts[i], pts[j]
                edge = b - a
                edge_len_sq = edge @ edge
                if edge_len_sq < 1e-12:
                    d = np.linalg.norm(cp - a)
                else:
                    t = np.clip((cp - a) @ edge / edge_len_sq, 0, 1)
                    proj = a + t * edge
                    d = np.linalg.norm(cp - proj)
                if d < min_d:
                    min_d = d
            if min_d > best_dist:
                best_dist  = min_d
                best_point = cp

        return best_point

    @staticmethod
    def _snap_to_pixel(x: float, y: float) -> tuple:
        """Snap coordinates to the nearest pixel centre (int + 0.5)."""
        return int(x) + 0.5, int(y) + 0.5

    # === EVENT HANDLERS ========================================================

    def _on_click_with_snap(self, event):
        """Handle mouse clicks; snap left-clicks to existing polygon vertices."""
        if event.inaxes != self.ax:
            return

        # Middle-click: start pan
        if event.button == 2:
            self._pan_active = True
            self._pan_start = (event.x, event.y)
            self._pan_xlim = self.ax.get_xlim()
            self._pan_ylim = self.ax.get_ylim()
            return

        # Right-click in divider mode: undo last placed point
        if event.button == 3 and self.divider_mode:
            self._divider_undo_last_point()
            return

        # Right-click in edit mode: delete hovered vertex
        if event.button == 3 and self.edit_mode:
            if self.hover_vertex_idx is not None and self.hover_room_idx is not None:
                room = self.rooms[self.hover_room_idx]
                if len(room['vertices']) > 3:
                    self._push_undo(self.hover_room_idx)
                    room['vertices'].pop(self.hover_vertex_idx)
                    rname = room.get('name', 'unnamed')
                    self.hover_vertex_idx = None
                    self.hover_room_idx   = None
                    self._update_status(f"Removed vertex from '{rname}' - press 's' to save", 'green')
                    self._save_session()
                    self._render_section(force_full=True)
                else:
                    self._update_status("Cannot remove - polygon must have at least 3 vertices", 'red')
            return

        # Left-click in divider mode: place division line endpoints
        if event.button == 1 and self.divider_mode and event.xdata is not None and event.ydata is not None:
            self._on_divider_click(event.xdata, event.ydata)
            return

        # Left-click in default mode (not draw, not edit): select room
        if event.button == 1 and not self.draw_mode and not self.edit_mode:
            self._select_room_at(event.xdata, event.ydata)
            return

        # Left-click in edit mode: drag vertex or insert vertex on edge
        if event.button == 1 and self.edit_mode and event.xdata is not None and event.ydata is not None:
            if self.hover_vertex_idx is not None and self.hover_room_idx is not None:
                # Begin vertex drag
                self._push_undo(self.hover_room_idx)
                self.edit_room_idx   = self.hover_room_idx
                self.edit_vertex_idx = self.hover_vertex_idx
                v = self.rooms[self.hover_room_idx]['vertices'][self.hover_vertex_idx]
                self._edit_drag_origin = (float(v[0]), float(v[1]))
                room_name = self.rooms[self.edit_room_idx].get('name', 'unnamed')
                self._update_status(f"Dragging vertex in '{room_name}'", 'cyan')
                self._render_section(force_full=True)
                self._start_blit_drag(self.edit_room_idx)
                return

            if self.hover_edge_room_idx is not None and self.hover_edge_idx is not None and self.hover_edge_point is not None:
                if event.key == 'shift':
                    # Shift+click: begin edge drag (move both endpoints together)
                    self._push_undo(self.hover_edge_room_idx)
                    self.edit_edge_room_idx = self.hover_edge_room_idx
                    self.edit_edge_idx      = self.hover_edge_idx
                    self.edit_edge_start    = (event.xdata, event.ydata)
                    room_name = self.rooms[self.edit_edge_room_idx].get('name', 'unnamed')
                    self._update_status(f"Dragging edge in '{room_name}'", 'cyan')
                    self._start_blit_drag(self.edit_edge_room_idx)
                    return
                # Click: insert two vertices on edge (straddling the click point)
                self._push_undo(self.hover_edge_room_idx)
                room  = self.rooms[self.hover_edge_room_idx]
                j     = self.hover_edge_idx
                verts = room['vertices']
                ax_, ay_ = verts[j]
                bx_, by_ = verts[(j + 1) % len(verts)]
                cx, cy = self.hover_edge_point

                # Place two points offset along the edge from the click point
                edge_dx, edge_dy = bx_ - ax_, by_ - ay_
                edge_len = (edge_dx ** 2 + edge_dy ** 2) ** 0.5
                offset = min(15.0, edge_len * 0.15)  # 15 px or 15% of edge
                if edge_len > 0:
                    ux, uy = edge_dx / edge_len, edge_dy / edge_len
                else:
                    ux, uy = 1.0, 0.0
                p1 = [float(cx - ux * offset), float(cy - uy * offset)]
                p2 = [float(cx + ux * offset), float(cy + uy * offset)]

                room['vertices'].insert(j + 1, p1)
                room['vertices'].insert(j + 2, p2)
                # Start dragging the second vertex (closer to b) for immediate repositioning
                self.edit_room_idx   = self.hover_edge_room_idx
                self.edit_vertex_idx = j + 2
                self._edit_drag_origin = (float(p2[0]), float(p2[1]))
                self.hover_edge_room_idx = None
                self.hover_edge_idx      = None
                self.hover_edge_point    = None
                self._update_status("Inserted 2 vertices — drag to reposition, right-click to remove", 'cyan')
                self._render_section(force_full=True)
                return
            return  # click on empty space in edit mode

        # Left-click in draw mode: store snapped position for correction on release
        if event.button == 1 and not self.edit_mode and event.xdata is not None and event.ydata is not None:
            x, y = event.xdata, event.ydata

            # Orthogonal constraint: lock to horizontal or vertical from last vertex
            if self.ortho_mode and hasattr(self, 'selector') and self.selector is not None and self.selector.verts:
                last_x, last_y = self.selector.verts[-1]
                dx, dy = abs(x - last_x), abs(y - last_y)
                if dx >= dy:
                    y = last_y   # horizontal line
                else:
                    x = last_x   # vertical line

            x, y = self._snap_to_pixel(x, y)
            snapped_x, snapped_y = self._snap_to_vertex(x, y)
            # If vertex snap didn't fire, try edge snap
            if snapped_x == x and snapped_y == y:
                snapped_x, snapped_y = self._snap_to_edge(x, y)
            self._pending_snap = (snapped_x, snapped_y)

            # Auto-detect parent room on first point of a new polygon
            if (not self.selected_parent
                    and hasattr(self, 'selector') and self.selector is not None
                    and not self.selector.verts):
                for room in self.rooms:
                    if (room.get('parent') is None
                            and self._is_room_on_current_hdr(room)
                            and len(room['vertices']) >= 3):
                        path = MplPath(np.array(room['vertices']))
                        if path.contains_point((snapped_x, snapped_y)):
                            self.selected_parent = room['name']
                            self.btn_parent.label.set_text(room['name'])
                            self._update_status(
                                f"Auto-detected parent: {room['name']}", 'blue')
                            break

    def _on_button_release(self, event):
        """Handle mouse button release (end of drag, pan, or snap correction)."""
        # End pan (check before inaxes guard — mouse may have left the axes)
        if event.button == 2 and self._pan_active:
            self._pan_active = False
            self._pan_start = None
            return

        if event.inaxes != self.ax:
            return

        # Correct the vertex the PolygonSelector just placed with our snapped position
        if (event.button == 1 and not self.edit_mode
                and hasattr(self, '_pending_snap') and self._pending_snap is not None):
            sx, sy = self._pending_snap
            self._pending_snap = None
            # _xys[-1] is the cursor tracking point; the just-added vertex is [-2]
            if self.selector is not None and len(self.selector._xys) >= 2 and not self.selector._selection_completed:
                self.selector._xys[-2] = (sx, sy)
                self.selector._draw_polygon()

        if self.edit_vertex_idx is not None and self.edit_room_idx is not None:
            self._end_blit_drag()
            room = self.rooms[self.edit_room_idx]
            current_pos = room['vertices'][self.edit_vertex_idx]
            # Snap to pixel centre, then to existing vertex if close
            px, py = self._snap_to_pixel(current_pos[0], current_pos[1])
            sx, sy = self._snap_to_vertex(px, py)
            room['vertices'][self.edit_vertex_idx] = [float(sx), float(sy)]
            self.edit_vertex_idx = None
            self._edit_drag_origin = None
            self._save_session()
            room_name = room.get('name', 'unnamed')
            self._update_status(f"Moved vertex in '{room_name}'", 'green')
            self._render_section(force_full=True)
            self._create_polygon_selector()

        if self.edit_edge_room_idx is not None and self.edit_edge_idx is not None:
            self._end_blit_drag()
            room = self.rooms[self.edit_edge_room_idx]
            j  = self.edit_edge_idx
            j2 = (j + 1) % len(room['vertices'])
            # Snap both endpoints to pixel centre
            for vi in (j, j2):
                px, py = self._snap_to_pixel(room['vertices'][vi][0], room['vertices'][vi][1])
                room['vertices'][vi] = [float(px), float(py)]
            self.edit_edge_room_idx = None
            self.edit_edge_idx      = None
            self.edit_edge_start    = None
            self._save_session()
            room_name = room.get('name', 'unnamed')
            self._update_status(f"Moved edge in '{room_name}'", 'green')
            self._render_section(force_full=True)
            self._create_polygon_selector()

    def _on_mouse_motion(self, event):
        """Handle mouse movement — dispatches to drag, pan, or hover handlers."""
        # Pan: use raw pixel coords so it works even when inaxes changes during drag
        if self._pan_active and self._pan_start is not None:
            dx_px = event.x - self._pan_start[0]
            dy_px = event.y - self._pan_start[1]
            # Convert pixel delta to data-coordinate delta
            fig_w, fig_h = self.fig.get_size_inches() * self.fig.dpi
            bbox = self.ax.get_position()
            ax_w_px = bbox.width * fig_w
            ax_h_px = bbox.height * fig_h
            xlim = self._pan_xlim
            ylim = self._pan_ylim
            dx_data = (xlim[1] - xlim[0]) * dx_px / ax_w_px
            dy_data = (ylim[1] - ylim[0]) * dy_px / ax_h_px
            self.ax.set_xlim(xlim[0] - dx_data, xlim[1] - dx_data)
            self.ax.set_ylim(ylim[0] - dy_data, ylim[1] - dy_data)
            self.fig.canvas.draw_idle()
            return

        if event.inaxes != self.ax:
            return
        x, y = event.xdata, event.ydata
        if x is None or y is None:
            return

        # Active drag operations (vertex or edge)
        if self.edit_vertex_idx is not None and self.edit_room_idx is not None:
            # Ortho constraint: lock to H or V from drag origin
            if self.ortho_mode and self._edit_drag_origin is not None:
                ox, oy = self._edit_drag_origin
                dx, dy = abs(x - ox), abs(y - oy)
                if dx >= dy:
                    y = oy   # horizontal movement
                else:
                    x = ox   # vertical movement
            self.rooms[self.edit_room_idx]['vertices'][self.edit_vertex_idx] = [float(x), float(y)]
            self._update_dragged_patch(self.edit_room_idx)
            return
        if self.edit_edge_room_idx is not None and self.edit_edge_idx is not None:
            self._handle_edge_drag(x, y)
            return

        # Divider mode: update preview line and vertex snap highlight
        if self.divider_mode:
            self._update_divider_preview(x, y)
            return

        if not self.edit_mode:
            return

        # Throttle hover detection (~15 fps)
        now = time.monotonic()
        if now - self._last_hover_check < 0.067:
            return
        self._last_hover_check = now
        self._handle_hover_detection(x, y)

    def _handle_edge_drag(self, x: float, y: float):
        """Move both endpoints of the hovered edge perpendicular to it.

        The drag is constrained to the direction orthogonal to the edge,
        so horizontal edges move only vertically and vice-versa.
        """
        dx = x - self.edit_edge_start[0]
        dy = y - self.edit_edge_start[1]
        room = self.rooms[self.edit_edge_room_idx]
        j  = self.edit_edge_idx
        j2 = (j + 1) % len(room['vertices'])

        # Edge direction vector
        ex = room['vertices'][j2][0] - room['vertices'][j][0]
        ey = room['vertices'][j2][1] - room['vertices'][j][1]
        edge_len_sq = ex * ex + ey * ey

        if edge_len_sq > 0:
            # Normal (perpendicular) to the edge: (-ey, ex)
            nx, ny = -ey, ex
            # Project the drag delta onto the normal direction
            proj = (dx * nx + dy * ny) / edge_len_sq
            dx = proj * nx
            dy = proj * ny

        room['vertices'][j][0]  += dx;  room['vertices'][j][1]  += dy
        room['vertices'][j2][0] += dx;  room['vertices'][j2][1] += dy
        self.edit_edge_start = (self.edit_edge_start[0] + dx, self.edit_edge_start[1] + dy)
        self._update_dragged_patch(self.edit_edge_room_idx)

    def _rebuild_hover_arrays(self):
        """Pre-build combined numpy arrays for vectorized hover hit-testing."""
        all_verts, vert_room, vert_local = [], [], []
        edge_starts, edge_ends, edge_room, edge_local = [], [], [], []
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = room['vertices']
            n = len(verts)
            for vi, v in enumerate(verts):
                all_verts.append(v)
                vert_room.append(i)
                vert_local.append(vi)
            for j in range(n):
                edge_starts.append(verts[j])
                edge_ends.append(verts[(j + 1) % n])
                edge_room.append(i)
                edge_local.append(j)
        self._hover_all_verts       = np.array(all_verts, dtype=float) if all_verts else None
        self._hover_vert_room_idx   = np.array(vert_room, dtype=int) if vert_room else None
        self._hover_vert_local_idx  = np.array(vert_local, dtype=int) if vert_local else None
        self._hover_edge_starts     = np.array(edge_starts, dtype=float) if edge_starts else None
        self._hover_edge_ends       = np.array(edge_ends, dtype=float) if edge_ends else None
        self._hover_edge_room_idx_arr = np.array(edge_room, dtype=int) if edge_room else None
        self._hover_edge_local_idx  = np.array(edge_local, dtype=int) if edge_local else None

    def _handle_hover_detection(self, x: float, y: float):
        """Detect vertex or edge hover using pre-built vectorized arrays."""
        hover_threshold = max(5.0, max(self._image_width, self._image_height) * 0.01)

        # --- Vertex hover (single vectorized distance computation) ---
        closest_dist, closest_room, closest_vidx = float('inf'), None, None
        if self._hover_all_verts is not None and len(self._hover_all_verts) > 0:
            dists = np.hypot(self._hover_all_verts[:, 0] - x, self._hover_all_verts[:, 1] - y)
            min_idx = int(np.argmin(dists))
            closest_dist = dists[min_idx]
            closest_room = int(self._hover_vert_room_idx[min_idx])
            closest_vidx = int(self._hover_vert_local_idx[min_idx])

        if closest_dist < hover_threshold:
            if self.hover_room_idx != closest_room or self.hover_vertex_idx != closest_vidx:
                self.hover_room_idx   = closest_room
                self.hover_vertex_idx = closest_vidx
                self.hover_edge_room_idx = self.hover_edge_idx = self.hover_edge_point = None
                name = self.rooms[closest_room].get('name', 'unnamed')
                self._update_status(f"Vertex in '{name}' - drag or right-click to remove", 'cyan')
                self._render_section()
            return

        # Clear stale vertex hover
        vertex_was_hovered = self.hover_vertex_idx is not None or self.hover_room_idx is not None
        if vertex_was_hovered:
            self.hover_vertex_idx = self.hover_room_idx = None

        # --- Edge hover (vectorized point-to-segment distance) ---
        edge_threshold = hover_threshold * 1.3
        best_dist, best_room, best_edge, best_pt = float('inf'), None, None, None
        if self._hover_edge_starts is not None and len(self._hover_edge_starts) > 0:
            A = self._hover_edge_starts
            B = self._hover_edge_ends
            AB = B - A
            P = np.array([[x, y]])
            AP = P - A
            seg_len_sq = np.sum(AB * AB, axis=1)
            # Avoid division by zero for degenerate edges
            safe_len = np.where(seg_len_sq == 0, 1.0, seg_len_sq)
            t = np.clip(np.sum(AP * AB, axis=1) / safe_len, 0.0, 1.0)
            # For zero-length edges, t=0 projects to A
            t = np.where(seg_len_sq == 0, 0.0, t)
            proj = A + t[:, None] * AB
            edge_dists = np.hypot(x - proj[:, 0], y - proj[:, 1])
            min_edge_idx = int(np.argmin(edge_dists))
            best_dist = edge_dists[min_edge_idx]
            best_room = int(self._hover_edge_room_idx_arr[min_edge_idx])
            best_edge = int(self._hover_edge_local_idx[min_edge_idx])
            best_pt   = (float(proj[min_edge_idx, 0]), float(proj[min_edge_idx, 1]))

        if best_dist < edge_threshold:
            changed = (self.hover_edge_room_idx != best_room or self.hover_edge_idx != best_edge)
            self.hover_edge_room_idx = best_room
            self.hover_edge_idx      = best_edge
            self.hover_edge_point    = best_pt
            if changed:
                self._update_status("Click to insert vertex, Shift+click to drag edge", 'blue')
                self._update_cursor()
                self._render_section()
        else:
            need_redraw = self.hover_edge_room_idx is not None or vertex_was_hovered
            self.hover_edge_room_idx = self.hover_edge_idx = self.hover_edge_point = None
            if need_redraw:
                self._update_status("Edit Mode: Hover over any vertex to drag", 'blue')
                self._update_cursor()
                self._render_section()

    def _on_key_press(self, event):
        """Handle keyboard shortcuts."""
        if event.key in ('backspace', 'delete', 'escape'):
            if event.key == 'escape':
                if self.divider_mode:
                    self._exit_divider_mode(cancelled=True)
                elif self.draw_mode:
                    self._exit_draw_mode()
                elif self.edit_mode:
                    # Exit edit mode without saving
                    self.edit_mode = False
                    self.btn_edit_mode.label.set_text('Edit Mode: OFF (Press E)')
                    self._style_toggle_button(self.btn_edit_mode, False)
                    self._reset_hover_state()
                    self._edit_undo_stack.clear()
                    self._create_polygon_selector()
                    self._update_cursor()
                    self._update_status("Edit Mode cancelled (no save)", 'orange')
                    self._render_section(force_full=True)
                else:
                    self._deselect_room()
            return
        if event.key == 's':
            if self.divider_mode:
                if len(self._divider_points) >= 2:
                    self._finalize_division()
                else:
                    self._update_status(
                        "Need at least 2 points to finalize divider — keep clicking or Esc to cancel",
                        'orange')
            else:
                self._on_save_click(None)
        elif event.key == 'S':
            self._save_session()
        elif event.key == 'r':
            self._on_reset_zoom_click(None)
        elif event.key == 'q':
            plt.close(self.fig)
        elif event.key == 'up':
            self._on_next_hdr_click(None)
        elif event.key == 'down':
            self._on_prev_hdr_click(None)
        elif event.key == 't':
            self._on_image_toggle_click(None)
        elif event.key == 'e':
            self._on_edit_mode_toggle(None)
        elif event.key == 'o':
            self._on_ortho_toggle(None)
        elif event.key == 'd':
            if self.edit_mode:
                if self.divider_mode:
                    self._exit_divider_mode(cancelled=True)
                else:
                    self._enter_divider_mode()
            else:
                if self.draw_mode:
                    self._exit_draw_mode()
                else:
                    self._enter_draw_mode()
        elif event.key == 'f':
            self._fit_to_selected_room()
        elif event.key == 'ctrl+z':
            if self.edit_mode:
                self._undo_edit()
            else:
                self._undo_draw()
        elif event.key == 'ctrl+a':
            self._select_all_rooms()

    def _on_scroll(self, event):
        """Handle scroll wheel for zooming or room list scrolling."""
        if event.inaxes == self.ax_list:
            if event.button == 'down':
                self.room_list_scroll_offset += 1
            else:
                self.room_list_scroll_offset = max(0, self.room_list_scroll_offset - 1)
            now = time.monotonic()
            if now - getattr(self, '_last_list_scroll_draw', 0.0) >= 0.05:
                self._last_list_scroll_draw = now
                self._update_room_list()
            return

        if event.inaxes != self.ax:
            return
        scale  = 1.2 if event.button == 'down' else 1 / 1.2
        xlim   = self.ax.get_xlim()
        ylim   = self.ax.get_ylim()
        xdata, ydata = event.xdata, event.ydata
        new_w  = (xlim[1] - xlim[0]) * scale
        new_h  = (ylim[1] - ylim[0]) * scale
        relx   = (xdata - xlim[0]) / (xlim[1] - xlim[0])
        rely   = (ydata - ylim[0]) / (ylim[1] - ylim[0])
        self.ax.set_xlim([xdata - new_w * relx,   xdata + new_w * (1 - relx)])
        self.ax.set_ylim([ydata - new_h * rely,   ydata + new_h * (1 - rely)])
        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()

        now = time.monotonic()
        if now - getattr(self, '_last_scroll_draw', 0.0) >= 0.033:
            self._last_scroll_draw = now
            self.fig.canvas.draw_idle()
        else:
            if hasattr(self, '_scroll_draw_timer') and self._scroll_draw_timer:
                self._scroll_draw_timer.stop()
            self._scroll_draw_timer = self.fig.canvas.new_timer(interval=60)
            self._scroll_draw_timer.single_shot = True
            self._scroll_draw_timer.add_callback(lambda: self.fig.canvas.draw_idle())
            self._scroll_draw_timer.start()

    def _on_resize(self, event):
        """Handle window resize events."""
        self._blit_background = None
        if not getattr(self, '_launch_complete', False):
            return
        if event.width > 0 and event.height > 0:
            now = time.monotonic()
            if now - getattr(self, '_last_resize_draw', 0.0) >= 0.1:
                self._last_resize_draw = now
                self.fig.canvas.draw_idle()

    def _force_resize_update(self):
        """Force update after window maximisation."""
        try:
            self._launch_complete = True
            canvas    = self.fig.canvas
            tk_widget = canvas.get_tk_widget()
            tk_widget.update_idletasks()
            width  = tk_widget.winfo_width()
            height = tk_widget.winfo_height()
            if width > 1 and height > 1:
                canvas.resize(width, height)
            canvas.draw_idle()
        except Exception:
            pass

    # === ROOM SELECTION ========================================================

    def _select_room_at(self, x, y):
        """Select the room polygon at the given point."""
        if x is None or y is None:
            self._deselect_room()
            return
        for i, room in enumerate(self.rooms):
            if not self._is_room_on_current_hdr(room):
                continue
            verts = np.array(room['vertices'])
            if MplPath(verts).contains_point((x, y)):
                self._select_room(i)
                return
        self._deselect_room()

    def _select_room(self, idx: int):
        """Select a room by index and populate the name textbox and parent."""
        self._deselect_room()
        self.selected_room_idx = idx
        room = self.rooms[idx]
        self.name_textbox.set_val(room.get('name', ''))
        self.room_type = room.get('room_type')
        self._update_room_type_buttons()

        # Auto-populate parent apartment selector
        parent = room.get('parent')
        if parent:
            # Sub-room: show its parent
            self.selected_parent = parent
            self.btn_parent.label.set_text(parent)
            self.name_label_text.set_text("Room Name:")
        else:
            # This IS a parent apartment — set it as the active parent
            self.selected_parent = room.get('name', '')
            self.btn_parent.label.set_text(self.selected_parent)
            self.name_label_text.set_text("Room Name:")

        self._update_status(f"Selected: {room.get('name', 'unnamed')}", 'orange')
        self._update_room_list()
        self._render_section()

    def _deselect_room(self):
        """Deselect any selected room and clear multi-selection."""
        changed = self.selected_room_idx is not None or self.multi_selected_room_idxs
        self.multi_selected_room_idxs.clear()
        if self.selected_room_idx is not None:
            self.selected_room_idx = None
            self.name_textbox.set_val('')
            self.room_type = None
            self._update_room_type_buttons()
        if changed:
            self._update_status("Ready to draw", 'blue')
            self._update_room_list()
            self.fig.canvas.draw_idle()

    def _select_all_rooms(self):
        """Select all rooms on the current HDR (Ctrl+A)."""
        self.multi_selected_room_idxs = {
            i for i, r in enumerate(self.rooms) if self._is_room_on_current_hdr(r)
        }
        n = len(self.multi_selected_room_idxs)
        self._update_status(f"Multi-select: {n} room(s) on current HDR", 'blue')
        self._update_room_list()
        self.fig.canvas.draw_idle()

    # === POLYGON SELECTOR ======================================================

    def _create_polygon_selector(self):
        """Create or recreate the polygon selector."""
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.disconnect_events()
        self.selector = PolygonSelector(
            self.ax,
            self._on_polygon_select,
            useblit=True,
            props=dict(color='cyan', linestyle='-', linewidth=2, alpha=0.5),
            handle_props=dict(markersize=8, markerfacecolor='lime',
                              markeredgecolor='darkgreen', markeredgewidth=1.5),
        )
        # Monkey-patch _onmove to apply ortho constraint to the rubber-band
        # line in real time, so the preview matches the final placement.
        _orig_onmove = self.selector._onmove
        editor = self

        def _ortho_onmove(event):
            _orig_onmove(event)
            sel = editor.selector
            if (editor.ortho_mode
                    and not sel._selection_completed
                    and len(sel._xys) >= 2):
                # _xys[-1] is the cursor tracking point; previous vertex is [-2]
                last_x, last_y = sel._xys[-2]
                cx, cy = sel._xys[-1]
                dx, dy = abs(cx - last_x), abs(cy - last_y)
                if dx >= dy:
                    cy = last_y   # horizontal
                else:
                    cx = last_x   # vertical
                sel._xys[-1] = (cx, cy)
                sel._draw_polygon()

        self.selector._onmove = _ortho_onmove
        # Honour draw_mode: selector is inactive by default until 'd' is pressed
        self.selector.set_active(getattr(self, 'draw_mode', False))

    def _on_polygon_select(self, vertices):
        """Callback when a polygon drawing is completed."""
        if len(vertices) < 3:
            return
        self.current_polygon_vertices = list(vertices)
        if self.selected_room_idx is not None:
            self._deselect_room()
        area = 0.5 * abs(
            np.dot([v[0] for v in vertices], np.roll([v[1] for v in vertices], 1))
            - np.dot([v[1] for v in vertices], np.roll([v[0] for v in vertices], 1))
        )
        self._update_status(f"Polygon ready: {len(vertices)} pts, {area:.0f} px2", 'green')

    # === ROOM OPERATIONS =======================================================

    def _on_save_click(self, event):
        """Save room, update selected room, or save edited boundary."""
        if self.divider_mode:
            self._update_status("Complete or cancel divider mode first (Esc to cancel)", 'orange')
            return
        if self.edit_mode and self.edit_room_idx is not None:
            self._save_edited_room()
        elif self.selected_room_idx is not None:
            self._update_selected_room()
        else:
            self._save_current_room()

    def _save_edited_room(self):
        """Save changes to an edited room boundary (also applies name edit)."""
        if self.edit_room_idx is None:
            return
        room = self.rooms[self.edit_room_idx]

        # Apply name from textbox if provided
        typed_name = self.name_textbox.text.strip().upper()
        if typed_name and typed_name != room.get('name', ''):
            old_name = room['name']
            new_name = self._make_unique_name(typed_name, exclude_idx=self.edit_room_idx)
            self._push_draw_undo(('rename', self.edit_room_idx, old_name))
            room['name'] = new_name

        name = room.get('name', 'unnamed')
        self._save_session()
        self.edit_room_idx    = None
        self.hover_vertex_idx = None
        self._update_status(f"Saved changes for '{name}'", 'green')
        self._render_section(force_full=True)
        self._create_polygon_selector()
        self._update_room_list()
        self._update_hdr_list()
        print(f"Saved edited room '{name}'")

    def _save_current_room(self):
        """Save the currently drawn polygon as a new room."""
        # Auto-close: if polygon isn't completed but has 3+ vertices, grab them
        if (len(self.current_polygon_vertices) < 3
                and hasattr(self, 'selector') and self.selector is not None
                and not self.selector._selection_completed
                and len(self.selector.verts) >= 3):
            self.current_polygon_vertices = list(self.selector.verts)

        if len(self.current_polygon_vertices) < 3:
            self._update_status("No polygon to save - draw one first", 'red')
            return

        name = self.name_textbox.text.strip().upper()
        if not name:
            name = f"ROOM_{len(self.rooms) + 1:03d}"

        if self.selected_parent:
            full_name = f"{self.selected_parent}_{name}"
        else:
            full_name = name

        full_name = self._make_unique_name(full_name)
        vertices  = [[float(x), float(y)] for x, y in self.current_polygon_vertices]

        # Boundary containment check
        warning_msg      = ""
        is_outside_parent = False
        if self.selected_parent:
            parent_room = self._get_parent_room(self.selected_parent)
            if parent_room:
                is_outside_parent = not self._check_boundary_containment(vertices, parent_room['vertices'])
                if is_outside_parent:
                    warning_msg = " (WARNING: extends outside parent boundary!)"
                    print(f"WARNING: Room '{full_name}' extends outside parent '{self.selected_parent}'")
            else:
                print(f"Warning: Parent room '{self.selected_parent}' not found")

        # Auto-assign types: sub-rooms default to BED, parent defaults to LIVING
        room_type = self.room_type
        parent_type_changed = None          # (parent_idx, old_type) for undo
        if self.selected_parent:
            if not room_type:
                room_type = 'BED'
            parent_room = self._get_parent_room(self.selected_parent)
            if parent_room and not parent_room.get('room_type'):
                parent_idx = self.rooms.index(parent_room)
                parent_type_changed = (parent_idx, parent_room.get('room_type'))
                parent_room['room_type'] = 'LIVING'

        room = {
            'name':      full_name,
            'parent':    self.selected_parent,
            'vertices':  vertices,
            'hdr_file':  self.current_hdr_name,
            'room_type': room_type,
        }
        self.rooms.append(room)
        new_idx = len(self.rooms) - 1
        self._push_draw_undo(('create', new_idx, parent_type_changed))

        status_color = 'orange' if is_outside_parent else 'green'
        self._update_status(f"Saved '{full_name}'{warning_msg}", status_color)
        print(f"Saved room '{full_name}' on HDR '{self.current_hdr_name}'")

        # Reset drawing state
        self.current_polygon_vertices = []
        self.name_textbox.set_val('')
        self.room_type = None
        self._update_room_type_buttons()
        self._update_parent_options()
        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)
        self._create_polygon_selector()

    def _update_selected_room(self):
        """Update the name of the selected room."""
        if self.selected_room_idx is None:
            return
        idx      = self.selected_room_idx
        old_name = self.rooms[idx]['name']
        new_name = self.name_textbox.text.strip().upper() or old_name
        new_name = self._make_unique_name(new_name, exclude_idx=idx)
        if new_name == old_name:
            return
        self._push_draw_undo(('rename', idx, old_name))
        self.rooms[idx]['name'] = new_name
        self._update_status(f"Renamed '{old_name}' → '{new_name}'", 'green')
        self._update_room_list()
        self._render_section(force_full=True)
        self._save_session()

    def _on_clear_click(self, event):
        """Clear the current polygon drawing."""
        self._deselect_room()
        self.current_polygon_vertices = []
        if self.selector is not None:
            self.selector.clear()
        self._update_status("Cleared - ready to draw", 'blue')

    def _on_delete_click(self, event):
        """Delete the selected room (Ctrl+Z to undo)."""
        if self.selected_room_idx is None:
            self._update_status("No room selected to delete", 'red')
            return
        idx  = self.selected_room_idx
        room = self.rooms.pop(idx)
        name = room.get('name', 'unnamed')
        self._push_draw_undo(('delete', idx, room))
        self.selected_room_idx = None
        self.current_polygon_vertices = []
        self._update_status(f"Deleted '{name}' (Ctrl+Z to undo)", 'green')
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)
        self._create_polygon_selector()
        print(f"Deleted room '{name}'")
        self._save_session()

    # === ZOOM & DISPLAY ========================================================

    def _view_ratio(self) -> float:
        """Return view_w / image_width. 1.0 = fully zoomed out, <1 = zoomed in."""
        if self._image_width <= 1:
            return 1.0
        view_w = abs(self.ax.get_xlim()[1] - self.ax.get_xlim()[0])
        return max(0.01, view_w / self._image_width)

    def _zoom_fontsize(self, base: float = 8.0) -> float:
        """Return font size scaled relative to the default (full-image) zoom.

        At default zoom returns base. Scales down when zoomed out further,
        scales up when zoomed in. Clamps to [4, 20].
        """
        if self._reference_view_w <= 1:
            return base
        view_w = abs(self.ax.get_xlim()[1] - self.ax.get_xlim()[0])
        scale  = self._reference_view_w / max(view_w, 1.0)
        return max(4.0, min(20.0, base * scale))

    def _df_line_step(self) -> float:
        """Return zoom-aware vertical spacing (in data coords) between DF result lines.

        The spacing must stay proportional to the rendered font height.
        Because _zoom_fontsize clamps at 4 pt, a pure view_w fraction
        would shrink below the text height when zoomed in. Instead we
        derive spacing from the actual DF font size converted back to
        data coordinates, guaranteeing no overlap at any zoom level.
        """
        view_w = abs(self.ax.get_xlim()[1] - self.ax.get_xlim()[0])
        fs = self._zoom_fontsize(base=6.5)          # actual DF font size (pts)
        ref_fs = 6.5                                  # font size at default zoom
        ref_step = self._reference_view_w * 0.009     # spacing at default zoom
        if ref_fs <= 0 or self._reference_view_w <= 1:
            return view_w * 0.009
        # Scale the default-zoom spacing by (current font / default font)
        # and by (current view / default view) to stay in data coords
        return ref_step * (fs / ref_fs) * (view_w / self._reference_view_w)

    def _apply_zoom_fontsizes(self):
        """Reapply zoom-dependent font sizes and stroke widths to all cached text."""
        fs = self._zoom_fontsize()
        stroke_name = [patheffects.withStroke(linewidth=fs * 0.06, foreground='black')]
        for label in self._room_label_cache.values():
            if label is not None:
                label.set_fontsize(fs)
                label.set_path_effects(stroke_name)
        fs_df = self._zoom_fontsize(base=6.5)
        stroke_df = [patheffects.withStroke(linewidth=fs_df * 0.06, foreground='black')]
        line_step = self._df_line_step()
        for dt in self._df_text_cache:
            dt.set_fontsize(fs_df)
            dt.set_path_effects(stroke_df)
            # Reposition based on current zoom-aware line step
            centroid = getattr(dt, '_df_centroid', None)
            line_i   = getattr(dt, '_df_line_i', None)
            if centroid is not None and line_i is not None:
                dt.set_position((centroid[0], centroid[1] + 1.2 * line_step + line_i * line_step * 0.7))

    def _zoom_linewidth(self, base: float = 1.5) -> float:
        """Return linewidth that stays visually constant regardless of zoom.

        Matplotlib linewidths are in *points* (screen units) but the polygon
        edges are in *data* coordinates, so when you zoom in the data-space
        features grow on-screen while the point-width stays fixed — making
        lines appear thinner.  To compensate we scale inversely with the
        view ratio: zooming in (smaller view_ratio) → larger linewidth.
        Clamped to [base * 0.5, base * 4] to stay reasonable at extremes.
        """
        ratio = self._view_ratio()                       # 1.0 = full image, <1 = zoomed in
        lw = base / max(ratio, 0.01)                     # inverse: zoom in → thicker
        return max(base * 0.5, min(base * 4.0, lw))

    def _apply_zoom_linewidths(self):
        """Reapply zoom-dependent linewidths to all cached room patches."""
        lw_current = self._zoom_linewidth()
        lw_other = self._zoom_linewidth(base=1.0)
        for patch in self._room_patch_cache.values():
            ptype = getattr(patch, '_patch_type', None)
            if ptype == 'current':
                patch.set_linewidth(lw_current)
            elif ptype == 'other':
                patch.set_linewidth(lw_other)


    def _on_reset_zoom_click(self, event):
        """Reset zoom to the full image extent."""
        self.ax.set_xlim(0, self._image_width)
        self.ax.set_ylim(self._image_height, 0)
        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()
        self.fig.canvas.draw_idle()

    def _fit_to_selected_room(self):
        """Zoom to fit the selected room boundary on screen (press f)."""
        if self.selected_room_idx is None:
            self._update_status("Select a room first, then press 'f' to fit", 'red')
            return
        room = self.rooms[self.selected_room_idx]
        verts = room['vertices']
        if len(verts) < 3:
            return

        xs = [v[0] for v in verts]
        ys = [v[1] for v in verts]
        min_x, max_x = min(xs), max(xs)
        min_y, max_y = min(ys), max(ys)

        # Add 10% padding around the bounding box
        pad_x = max((max_x - min_x) * 0.10, 20)
        pad_y = max((max_y - min_y) * 0.10, 20)
        self.ax.set_xlim(min_x - pad_x, max_x + pad_x)
        self.ax.set_ylim(max_y + pad_y, min_y - pad_y)  # y-axis is inverted (origin='upper')

        self._apply_zoom_linewidths()
        self._apply_zoom_fontsizes()
        room_name = room.get('name', 'unnamed')
        self._update_status(f"Fit to '{room_name}' (r to reset zoom)", 'blue')
        self.fig.canvas.draw_idle()

    # === TOGGLE BUTTONS & ROOM TYPE ============================================

    def _update_cursor(self):
        """Set the canvas cursor based on the active editing mode.

        Matplotlib's TkAgg backend overrides the cursor on every mouse-move
        via its own set_cursor(). We patch that method on the canvas instance
        so our custom cursor is not stomped, then restore normal behaviour
        when no custom cursor is needed.
        """
        try:
            canvas = self.fig.canvas
            tkcanvas = canvas._tkcanvas
        except Exception:
            return

        if self.divider_mode:
            tkcanvas.configure(cursor='crosshair')
            canvas.set_cursor = lambda cursor: None          # suppress mpl overrides
        elif self.edit_mode and self.hover_edge_room_idx is not None:
            tkcanvas.configure(cursor='plus')
            canvas.set_cursor = lambda cursor: None          # suppress mpl overrides
        else:
            # Restore matplotlib's own set_cursor before resetting
            import types
            from matplotlib.backends._backend_tk import FigureCanvasTk
            canvas.set_cursor = types.MethodType(FigureCanvasTk.set_cursor, canvas)
            tkcanvas.configure(cursor='arrow')

    def _style_toggle_button(self, btn, is_on: bool):
        """Apply pressed/depressed styling to a toggle button.

        ON  = sunken look: darker background, bold white text.
        OFF = raised look: lighter background, normal dark text.
        """
        body = getattr(btn, '_body', None)
        highlight = getattr(btn, '_highlight', None)
        if is_on:
            btn.color      = self._btn_on_color
            btn.hovercolor = self._btn_on_hover
            if body:
                body.set_facecolor(self._btn_on_color)
                body.set_edgecolor('#606058')
            if highlight:
                highlight.set_facecolor('#FFFFFF15')
            btn.label.set_fontweight('bold')
            btn.label.set_color('#FFFFFF')
        else:
            btn.color      = self._btn_color
            btn.hovercolor = self._btn_hover
            if body:
                body.set_facecolor(self._btn_color)
                body.set_edgecolor('#B0B0A8')
            if highlight:
                highlight.set_facecolor('#FFFFFF30')
            btn.label.set_fontweight('medium')
            btn.label.set_color('#333333')

    def _on_edit_mode_toggle(self, event):
        """Toggle edit mode for modifying existing room boundaries."""
        if self.divider_mode:
            self._exit_divider_mode(cancelled=True)
        if self.draw_mode:
            self._exit_draw_mode()

        self.edit_mode = not self.edit_mode
        self.btn_edit_mode.label.set_text(
            'Edit Mode: ON (Press E)' if self.edit_mode else 'Edit Mode: OFF (Press E)')
        self._style_toggle_button(self.btn_edit_mode, self.edit_mode)

        if self.edit_mode:
            self._edit_undo_stack.clear()
            if hasattr(self, 'selector') and self.selector is not None:
                self.selector.set_active(False)
            self._update_status("Edit Mode: Hover over any vertex to drag (all rooms editable)", 'cyan')
        else:
            self._reset_hover_state()
            self._edit_undo_stack.clear()
            self._save_session()
            self._create_polygon_selector()   # set_active(False) applied inside via draw_mode=False
            self._update_status("Edit Mode OFF — press 'd' to draw, left-click to select", 'blue')
            self._update_cursor()

        self._render_section(force_full=True)

    def _on_draw_mode_toggle(self, event):
        """Toggle draw mode via the UI button."""
        if self.edit_mode:
            return   # draw mode not available in edit mode
        if self.draw_mode:
            self._exit_draw_mode()
        else:
            self._enter_draw_mode()

    def _on_ortho_toggle(self, event):
        """Toggle orthogonal lines mode."""
        self.ortho_mode = not self.ortho_mode
        self.btn_ortho.label.set_text(
            'Ortho Lines: ON (Press O)' if self.ortho_mode else 'Ortho Lines: OFF (Press O)')
        self._style_toggle_button(self.btn_ortho, self.ortho_mode)
        state = "ON" if self.ortho_mode else "OFF"
        self._update_status(f"Orthogonal mode: {state}", 'blue')
        self.fig.canvas.draw_idle()

    def _on_room_type_toggle(self, room_type: str, **_kw):
        """Toggle a room type on/off.  Supports bulk tagging when multi-selected."""
        # In multi-select mode: always set (no toggle-off), apply to all
        if self.multi_selected_room_idxs:
            self.room_type = room_type
            self._apply_room_type_change()
            return

        # Single-select: toggle behaviour
        if self.room_type == room_type:
            self.room_type = None
            self._apply_room_type_change()
            return
        self.room_type = room_type
        self._apply_room_type_change()

    def _apply_room_type_change(self):
        """Update UI and persist room type change.

        When ``multi_selected_room_idxs`` is non-empty the change is applied
        to every room in the set.
        """
        self._update_room_type_buttons()

        # Collect target indices (multi-select takes priority)
        if self.multi_selected_room_idxs:
            target_idxs = list(self.multi_selected_room_idxs)
        elif self.selected_room_idx is not None:
            target_idxs = [self.selected_room_idx]
        else:
            target_idxs = []

        if not target_idxs:
            self.fig.canvas.draw_idle()
            return

        # Snapshot old types for undo
        old_types = [(idx, self.rooms[idx].get('room_type')) for idx in target_idxs]
        self._push_draw_undo(('type', old_types))

        for idx in target_idxs:
            self.rooms[idx]['room_type'] = self.room_type
            self._invalidate_room_df_cache(idx)

        self._save_session()
        self._update_room_list()
        self._render_section(force_full=True)

        # Status feedback
        tag = self.room_type or 'None'
        applied = len(target_idxs)
        if applied > 1:
            self._update_status(f"Set {applied} rooms to {tag}", 'green')
        elif applied == 1:
            name = self.rooms[target_idxs[0]].get('name', '')
            self._update_status(f"{name} → {tag}", 'green')

        self.fig.canvas.draw_idle()

    def _update_room_type_buttons(self):
        """Style BED/LIVING/CIRC buttons to reflect current room_type."""
        self._style_toggle_button(self.btn_room_type_bed, self.room_type == 'BED')
        self._style_toggle_button(self.btn_room_type_living, self.room_type == 'LIVING')
        self._style_toggle_button(self.btn_room_type_circ, self.room_type == 'CIRCULATION')

    # === ROOM DIVIDER ==========================================================

    @staticmethod
    def _points_close(a, b, tol=1.0) -> bool:
        """Return True if two points are within *tol* pixels of each other."""
        return abs(a[0] - b[0]) < tol and abs(a[1] - b[1]) < tol

    @staticmethod
    def _point_on_segment(pt, a, b, tol=2.0) -> bool:
        """Return True if *pt* lies on segment a→b within pixel tolerance."""
        px, py = float(pt[0]), float(pt[1])
        ax_, ay_ = float(a[0]), float(a[1])
        bx_, by_ = float(b[0]), float(b[1])
        # Collinearity via cross product
        cross = abs((bx_ - ax_) * (py - ay_) - (by_ - ay_) * (px - ax_))
        seg_len = ((bx_ - ax_) ** 2 + (by_ - ay_) ** 2) ** 0.5
        if seg_len > 0 and cross / seg_len > tol:
            return False
        # Bounding box check
        min_x, max_x = min(ax_, bx_) - tol, max(ax_, bx_) + tol
        min_y, max_y = min(ay_, by_) - tol, max(ay_, by_) + tol
        return min_x <= px <= max_x and min_y <= py <= max_y

    @staticmethod
    def _line_polygon_intersections(line_start, line_end, polygon_verts):
        """Find all points where an infinite line (through *line_start* and
        *line_end*) intersects the edges of *polygon_verts*.

        Returns a list of (x, y) tuples sorted by distance from *line_start*.
        Near-duplicate points (within 1 px) are filtered out.
        """
        intersections = []
        sx, sy = float(line_start[0]), float(line_start[1])
        ex, ey = float(line_end[0]),   float(line_end[1])
        dx, dy = ex - sx, ey - sy

        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]),           float(polygon_verts[i][1])
            bx_, by_ = float(polygon_verts[(i + 1) % n][0]), float(polygon_verts[(i + 1) % n][1])
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = dx * edge_dy - dy * edge_dx
            if abs(denom) < 1e-10:
                continue  # parallel / collinear
            u = ((ax_ - sx) * dy - (ay_ - sy) * dx) / denom
            if -1e-10 <= u <= 1.0 + 1e-10:
                ix = ax_ + u * edge_dx
                iy = ay_ + u * edge_dy
                intersections.append((float(ix), float(iy)))

        # Deduplicate (adjacent edges sharing a vertex produce two hits)
        if len(intersections) > 1:
            unique = [intersections[0]]
            for pt in intersections[1:]:
                if not any(abs(pt[0] - up[0]) < 1.0 and abs(pt[1] - up[1]) < 1.0 for up in unique):
                    unique.append(pt)
            intersections = unique

        intersections.sort(key=lambda p: (p[0] - sx) ** 2 + (p[1] - sy) ** 2)
        return intersections

    def _split_polygon_by_line(self, verts, p1, p2):
        """Split a closed polygon into two sub-polygons along segment p1→p2.

        *p1* and *p2* must lie on the polygon boundary.
        Returns ``(poly_a, poly_b)`` as lists of ``[x, y]``, or
        ``(None, None)`` on failure.
        """
        n = len(verts)
        augmented = []
        p1_idx = None
        p2_idx = None

        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1

            # Check if p1 lies on edge a→b
            if p1_idx is None and self._point_on_segment(p1, a, b):
                if self._points_close(p1, a):
                    p1_idx = cur
                elif self._points_close(p1, b):
                    pass  # will be picked up when b is appended
                else:
                    augmented.append([float(p1[0]), float(p1[1])])
                    p1_idx = len(augmented) - 1

            # Check if p2 lies on edge a→b
            if p2_idx is None and self._point_on_segment(p2, a, b):
                if self._points_close(p2, a):
                    p2_idx = cur
                elif self._points_close(p2, b):
                    pass  # will be picked up when b is appended
                else:
                    augmented.append([float(p2[0]), float(p2[1])])
                    p2_idx = len(augmented) - 1

        # Handle the deferred "close to b" cases
        if p1_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p1, v):
                    p1_idx = i
                    break
        if p2_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p2, v):
                    p2_idx = i
                    break

        if p1_idx is None or p2_idx is None:
            return None, None

        # Walk forward from p1→p2 to collect poly_a
        m = len(augmented)
        poly_a = []
        idx = p1_idx
        while True:
            poly_a.append(augmented[idx])
            if idx == p2_idx and len(poly_a) > 1:
                break
            idx = (idx + 1) % m

        # Walk forward from p2→p1 to collect poly_b
        poly_b = []
        idx = p2_idx
        while True:
            poly_b.append(augmented[idx])
            if idx == p1_idx and len(poly_b) > 1:
                break
            idx = (idx + 1) % m

        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    # --- draw mode lifecycle ---------------------------------------------------

    def _enter_draw_mode(self):
        """Activate polygon drawing mode."""
        self.draw_mode = True
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.set_active(True)
        if hasattr(self, 'btn_draw_mode'):
            self.btn_draw_mode.label.set_text('Draw Mode: ON (Press D)')
            self._style_toggle_button(self.btn_draw_mode, True)
        self._update_status("Draw Mode: left-click to place vertices, s=save, Esc=exit", 'cyan')
        self.fig.canvas.draw_idle()

    def _exit_draw_mode(self):
        """Deactivate polygon drawing mode and clear any in-progress polygon."""
        self.draw_mode = False
        if hasattr(self, 'selector') and self.selector is not None:
            self.selector.set_active(False)
            self.selector.clear()
        self.current_polygon_vertices = []
        self._pending_snap = None
        if hasattr(self, 'btn_draw_mode'):
            self.btn_draw_mode.label.set_text('Draw Mode: OFF (Press D)')
            self._style_toggle_button(self.btn_draw_mode, False)
        self._update_status("Draw Mode OFF — press 'd' to draw, left-click to select", 'blue')
        self.fig.canvas.draw_idle()

    # --- divider mode lifecycle ------------------------------------------------

    def _enter_divider_mode(self):
        """Enter divider mode: draw an ortho polyline to split the selected room."""
        if not self.edit_mode:
            self._update_status("Must be in edit mode to use divider", 'red')
            return
        if self.selected_room_idx is None:
            self._update_status("Select a room first (right-click), then press 'd'", 'red')
            return

        room = self.rooms[self.selected_room_idx]
        if len(room['vertices']) < 3:
            self._update_status("Selected room has too few vertices to divide", 'red')
            return

        self.divider_mode          = True
        self._divider_room_idx     = self.selected_room_idx
        self._divider_points       = []
        self._divider_markers      = []
        self._divider_segments     = []
        self._divider_preview_line = None

        self._reset_hover_state()
        self._update_cursor()

        room_name = room.get('name', 'unnamed')
        self._update_status(
            f"DIVIDER: Click points to build ortho path on '{room_name}' "
            f"(s=finish, right-click=undo, Esc=cancel)", 'magenta')
        self.fig.canvas.draw_idle()

    def _exit_divider_mode(self, cancelled: bool = False):
        """Exit divider mode, cleaning up all preview artists."""
        self.divider_mode      = False
        self._divider_room_idx = None
        self._divider_points   = []
        self._update_cursor()

        for artist in self._divider_markers:
            if artist is not None:
                try: artist.remove()
                except ValueError: pass
        self._divider_markers = []

        for artist in self._divider_segments:
            if artist is not None:
                try: artist.remove()
                except ValueError: pass
        self._divider_segments = []

        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        if self._divider_snap_marker is not None:
            try: self._divider_snap_marker.remove()
            except ValueError: pass
            self._divider_snap_marker = None
        self._divider_snap_pt = None

        if cancelled:
            self._update_status("Divider mode cancelled", 'blue')
        self._render_section(force_full=True)

    # --- divider click / preview / undo ----------------------------------------

    def _on_divider_click(self, x: float, y: float):
        """Handle a left-click during divider mode — accumulate polyline points."""
        if self._divider_room_idx is None:
            return

        # Use highlighted snap vertex if one is active, else normal snap pipeline
        if self._divider_snap_pt is not None:
            x, y = self._divider_snap_pt
        else:
            x, y = self._snap_to_pixel(x, y)
            x, y = self._snap_to_vertex(x, y)

        # Ortho constraint from last placed point
        if self._divider_points:
            lx, ly = self._divider_points[-1]
            dx, dy = abs(x - lx), abs(y - ly)
            if dx >= dy:
                y = ly   # horizontal
            else:
                x = lx   # vertical

        # Reject zero-length segment
        if self._divider_points:
            lx, ly = self._divider_points[-1]
            if abs(x - lx) < 0.5 and abs(y - ly) < 0.5:
                return

        self._divider_points.append((x, y))

        # Draw marker dot
        marker = self.ax.plot(
            x, y, 'o', color='magenta', markersize=8,
            markeredgecolor='white', markeredgewidth=1.5, zorder=200)[0]
        self._divider_markers.append(marker)

        # Draw solid segment from previous point
        if len(self._divider_points) >= 2:
            px, py = self._divider_points[-2]
            seg, = self.ax.plot(
                [px, x], [py, y],
                '-', color='magenta', linewidth=2, alpha=0.9, zorder=195)
            self._divider_segments.append(seg)

        # Clear old preview line (becomes solid segment)
        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        n = len(self._divider_points)
        self._update_status(
            f"DIVIDER: {n} pt{'s' if n != 1 else ''} — "
            f"click=add, s=finish, right-click=undo last, Esc=cancel", 'magenta')
        self.fig.canvas.draw_idle()

    def _divider_undo_last_point(self):
        """Remove the last placed divider point (right-click during divider mode)."""
        if not self._divider_points:
            self._update_status("No points to undo — Esc to cancel divider", 'orange')
            return

        self._divider_points.pop()

        if self._divider_markers:
            artist = self._divider_markers.pop()
            if artist is not None:
                try: artist.remove()
                except ValueError: pass

        if self._divider_segments:
            artist = self._divider_segments.pop()
            if artist is not None:
                try: artist.remove()
                except ValueError: pass

        if self._divider_preview_line is not None:
            try: self._divider_preview_line.remove()
            except ValueError: pass
            self._divider_preview_line = None

        n = len(self._divider_points)
        if n == 0:
            self._update_status(
                "DIVIDER: All points removed — click to start again, Esc to cancel", 'magenta')
        else:
            self._update_status(
                f"DIVIDER: Undid last point, {n} remaining", 'magenta')
        self.fig.canvas.draw_idle()

    def _update_divider_preview(self, x: float, y: float):
        """Draw / update dashed preview from last placed point to cursor.

        Also detects nearby vertices from any room and shows a highlight ring
        so the user can snap the next divider point to an existing vertex.
        """
        # --- Vertex snap detection (active regardless of whether first point is placed) ---
        snap_x, snap_y = self._snap_to_vertex(x, y)
        snapped = (snap_x != x or snap_y != y)
        self._divider_snap_pt = (snap_x, snap_y) if snapped else None

        if snapped:
            if self._divider_snap_marker is None:
                self._divider_snap_marker, = self.ax.plot(
                    snap_x, snap_y, 'o',
                    color='none', markersize=14,
                    markeredgecolor='yellow', markeredgewidth=2.5,
                    zorder=300)
            else:
                self._divider_snap_marker.set_data([snap_x], [snap_y])
                self._divider_snap_marker.set_visible(True)
        else:
            if self._divider_snap_marker is not None:
                self._divider_snap_marker.set_visible(False)

        # --- Preview line (only once at least one point is placed) ---
        if not self._divider_points:
            now = time.monotonic()
            if now - self._last_drag_draw >= 0.033:
                self._last_drag_draw = now
                self.fig.canvas.draw_idle()
            return

        lx, ly = self._divider_points[-1]

        # Ortho constraint applied to the (possibly snapped) target
        tx, ty = (snap_x, snap_y) if snapped else (x, y)
        dx, dy = abs(tx - lx), abs(ty - ly)
        if dx >= dy:
            ty = ly
        else:
            tx = lx

        if self._divider_preview_line is not None:
            self._divider_preview_line.set_data([lx, tx], [ly, ty])
        else:
            line, = self.ax.plot(
                [lx, tx], [ly, ty],
                '--', color='magenta', linewidth=2, alpha=0.8, zorder=150)
            self._divider_preview_line = line

        now = time.monotonic()
        if now - self._last_drag_draw >= 0.033:
            self._last_drag_draw = now
            self.fig.canvas.draw_idle()

    # --- ray intersection helper -----------------------------------------------

    @staticmethod
    def _ray_polygon_intersection(origin, direction, polygon_verts):
        """Find the first intersection of a ray with the polygon boundary.

        The ray starts at *origin* and extends in *direction*.
        Returns ``(x, y)`` of the nearest hit, or ``None``.
        """
        ox, oy   = float(origin[0]),    float(origin[1])
        rdx, rdy = float(direction[0]), float(direction[1])
        if abs(rdx) < 1e-12 and abs(rdy) < 1e-12:
            return None

        best_t  = float('inf')
        best_pt = None

        n = len(polygon_verts)
        for i in range(n):
            ax_, ay_ = float(polygon_verts[i][0]),           float(polygon_verts[i][1])
            bx_, by_ = float(polygon_verts[(i + 1) % n][0]), float(polygon_verts[(i + 1) % n][1])
            edge_dx, edge_dy = bx_ - ax_, by_ - ay_
            denom = rdx * edge_dy - rdy * edge_dx
            if abs(denom) < 1e-10:
                continue
            t = ((ax_ - ox) * edge_dy - (ay_ - oy) * edge_dx) / denom
            u = ((ax_ - ox) * rdy   - (ay_ - oy) * rdx)       / denom
            if t > 1e-6 and -1e-10 <= u <= 1.0 + 1e-10:
                if t < best_t:
                    best_t  = t
                    best_pt = (float(ax_ + u * edge_dx), float(ay_ + u * edge_dy))

        return best_pt

    # --- finalize & split along polyline ---------------------------------------

    def _finalize_division(self):
        """Finalize the multi-segment divider: find boundary intersections
        for the first and last segments, then split the room polygon."""
        if self._divider_room_idx is None:
            return
        if len(self._divider_points) < 2:
            self._update_status("Need at least 2 points to finalize — keep clicking", 'orange')
            return

        room  = self.rooms[self._divider_room_idx]
        verts = room['vertices']
        pts   = list(self._divider_points)

        # --- boundary intersection for FIRST segment (extend pts[0] outward) ---
        boundary_start = self._find_boundary_hit(
            pts[0], pts[1], verts, outward=True)

        # --- boundary intersection for LAST segment (extend pts[-1] outward) ---
        boundary_end = self._find_boundary_hit(
            pts[-1], pts[-2], verts, outward=True)

        if boundary_start is None:
            self._update_status("Could not find boundary intersection for first segment", 'red')
            return
        if boundary_end is None:
            self._update_status("Could not find boundary intersection for last segment", 'red')
            return

        # Build full polyline: boundary_start → user pts → boundary_end
        polyline = list(pts)
        if self._points_close(boundary_start, polyline[0], tol=2.0):
            polyline[0] = boundary_start
        else:
            polyline.insert(0, boundary_start)

        if self._points_close(boundary_end, polyline[-1], tol=2.0):
            polyline[-1] = boundary_end
        else:
            polyline.append(boundary_end)

        polyline = [(float(p[0]), float(p[1])) for p in polyline]
        if len(polyline) < 2:
            self._update_status("Divider polyline too short", 'red')
            return

        poly_a, poly_b = self._split_polygon_by_polyline(verts, polyline)
        if poly_a is None or poly_b is None:
            self._update_status("Division failed — could not split polygon along polyline", 'red')
            return

        self._apply_division_with_polys(self._divider_room_idx, poly_a, poly_b)

    def _find_boundary_hit(self, tip, anchor, polygon_verts, outward=True):
        """Find where the segment anchor→tip (extended) crosses the polygon boundary.

        First checks if *tip* is already on the boundary.  If not, casts a ray
        from *tip* outward (direction tip−anchor) and returns the first hit.
        Falls back to the reverse direction if the outward ray misses.
        """
        # Check if tip is already on boundary
        n = len(polygon_verts)
        for i in range(n):
            a = polygon_verts[i]
            b = polygon_verts[(i + 1) % n]
            if self._point_on_segment(tip, a, b, tol=3.0):
                return (float(tip[0]), float(tip[1]))

        # Ray outward from tip (direction = tip − anchor)
        direction = (float(tip[0]) - float(anchor[0]),
                     float(tip[1]) - float(anchor[1]))
        hit = self._ray_polygon_intersection(tip, direction, polygon_verts)
        if hit is not None:
            return hit

        # Fallback: try inward direction (point may be outside the room)
        inward = (-direction[0], -direction[1])
        hit = self._ray_polygon_intersection(tip, inward, polygon_verts)
        return hit

    def _split_polygon_by_polyline(self, verts, polyline):
        """Split a closed polygon along a multi-segment polyline.

        *polyline[0]* and *polyline[-1]* must lie on the polygon boundary.
        Interior points may be inside the polygon.

        Returns ``(poly_a, poly_b)`` or ``(None, None)`` on failure.
        """
        p_entry  = polyline[0]
        p_exit   = polyline[-1]
        interior = polyline[1:-1]

        n = len(verts)
        augmented = []
        entry_idx = None
        exit_idx  = None

        for i in range(n):
            a = verts[i]
            b = verts[(i + 1) % n]
            augmented.append(list(a))
            cur = len(augmented) - 1

            if entry_idx is None and self._point_on_segment(p_entry, a, b):
                if self._points_close(p_entry, a):
                    entry_idx = cur
                elif self._points_close(p_entry, b):
                    pass
                else:
                    augmented.append([float(p_entry[0]), float(p_entry[1])])
                    entry_idx = len(augmented) - 1

            if exit_idx is None and self._point_on_segment(p_exit, a, b):
                if self._points_close(p_exit, a):
                    exit_idx = cur
                elif self._points_close(p_exit, b):
                    pass
                else:
                    augmented.append([float(p_exit[0]), float(p_exit[1])])
                    exit_idx = len(augmented) - 1

        # Deferred "close to b" resolution
        if entry_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_entry, v):
                    entry_idx = i
                    break
        if exit_idx is None:
            for i, v in enumerate(augmented):
                if self._points_close(p_exit, v):
                    exit_idx = i
                    break

        if entry_idx is None or exit_idx is None:
            return None, None

        m = len(augmented)

        # Walk boundary entry → exit
        bnd_a = []
        idx = entry_idx
        while True:
            bnd_a.append(augmented[idx])
            if idx == exit_idx and len(bnd_a) > 1:
                break
            idx = (idx + 1) % m

        # Walk boundary exit → entry
        bnd_b = []
        idx = exit_idx
        while True:
            bnd_b.append(augmented[idx])
            if idx == entry_idx and len(bnd_b) > 1:
                break
            idx = (idx + 1) % m

        interior_fwd = [[float(p[0]), float(p[1])] for p in interior]
        interior_rev = list(reversed(interior_fwd))

        # poly_a: boundary entry→exit, then interior reversed (exit→entry)
        poly_a = bnd_a + interior_rev
        # poly_b: boundary exit→entry, then interior forward (entry→exit)
        poly_b = bnd_b + interior_fwd

        if len(poly_a) < 3 or len(poly_b) < 3:
            return None, None
        return poly_a, poly_b

    # --- apply division --------------------------------------------------------

    def _apply_division_with_polys(self, room_idx: int, poly_a: list, poly_b: list):
        """Apply a division using pre-computed sub-polygons.

        Keeps the larger polygon as the original room (resized) and creates
        only the smaller polygon as a new DIV sub-room.
        Exits divider mode **and** edit mode, then saves the session.
        """
        room            = self.rooms[room_idx]
        original_name   = room.get('name', 'unnamed')
        original_parent = room.get('parent')
        hdr_file        = room.get('hdr_file', self.current_hdr_name)

        if original_parent is None:
            division_parent = original_name
        else:
            division_parent = original_parent

        # Snapshot entire rooms list for undo
        rooms_snapshot = [dict(r, vertices=[list(v) for v in r['vertices']]) for r in self.rooms]
        self._edit_undo_stack.append(('divider', rooms_snapshot))
        if len(self._edit_undo_stack) > self._edit_undo_max:
            self._edit_undo_stack.pop(0)

        # Determine which polygon is smaller (the cut-off piece)
        def _shoelace(poly):
            n = len(poly)
            if n < 3:
                return 0.0
            s = 0.0
            for i in range(n):
                x0, y0 = poly[i]
                x1, y1 = poly[(i + 1) % n]
                s += float(x0) * float(y1) - float(x1) * float(y0)
            return abs(s) / 2.0
        area_a = _shoelace(poly_a)
        area_b = _shoelace(poly_b)

        if area_a <= area_b:
            small_poly, large_poly = poly_a, poly_b
        else:
            small_poly, large_poly = poly_b, poly_a

        # Create only the smaller piece as a new DIV sub-room
        base_name = f"{division_parent}_DIV"
        div_name = self._make_unique_name(f"{base_name}1")

        div_room = {
            'name':      div_name,
            'parent':    division_parent,
            'vertices':  [[float(x), float(y)] for x, y in small_poly],
            'hdr_file':  hdr_file,
            'room_type': None,
        }

        # Resize the original room to the larger polygon
        room['vertices'] = [[float(x), float(y)] for x, y in large_poly]

        # Insert the new DIV sub-room after the original
        self.rooms.insert(room_idx + 1, div_room)

        self.selected_room_idx = None
        self._exit_divider_mode(cancelled=False)

        # Exit edit mode and save (per requirement)
        self.edit_mode = False
        self.btn_edit_mode.label.set_text('Edit Mode: OFF (Press E)')
        self._style_toggle_button(self.btn_edit_mode, False)
        self._reset_hover_state()
        self._create_polygon_selector()

        self._save_session()
        self._update_room_list()
        self._update_hdr_list()
        self._render_section(force_full=True)

        self._update_status(
            f"Divided '{original_name}' → '{div_name}' (smaller piece)", 'green')
        print(f"Room divider: '{original_name}' -> kept larger side, created '{div_name}'")

    # === STATUS & ROOM LIST ====================================================

    def _update_status(self, message: str, color: str = 'blue'):
        """Update the status text in the side panel."""
        if hasattr(self, 'status_text') and self.status_text is not None:
            self.status_text.set_text(f"Status: {message}")
            self.status_text.set_color(color)
            self.status_text.set_visible(True)
            # Clear the name preview so they don't overlap
            if hasattr(self, 'name_preview_text') and self.name_preview_text is not None:
                self.name_preview_text.set_text("")
            self.fig.canvas.draw_idle()

    def _update_room_list(self):
        """Refresh the saved rooms list as a scrollable, click-to-select panel."""
        self.ax_list.clear()
        self.ax_list.set_facecolor('#FAFAF8')
        self.ax_list.set_xlim(0, 1)
        self.ax_list.tick_params(left=False, bottom=False, labelleft=False, labelbottom=False)
        for spine in self.ax_list.spines.values():
            spine.set_edgecolor('#CCCCCC'); spine.set_linewidth(0.5)
        self._room_list_hit_boxes = []

        hdr_rooms = [(i, r) for i, r in enumerate(self.rooms) if self._is_room_on_current_hdr(r)]
        if not hdr_rooms:
            self.ax_list.set_ylim(0, 1)
            self.ax_list.text(0.05, 0.5, "(no rooms on this HDR)", fontsize=7,
                              style='italic', color='gray', va='center')
            self.fig.canvas.draw_idle()
            return

        flat_items, children_by_parent = self._build_room_tree(hdr_rooms)
        self._render_room_list_rows(flat_items, children_by_parent)
        self.fig.canvas.draw_idle()

    def _build_room_tree(self, hdr_rooms):
        """Build a flat parent→children tree for the room list. Returns (items, children_map).
        Each item is (room_idx, indent, is_last_child)."""
        apartments         = [(i, r) for i, r in hdr_rooms if r.get('parent') is None]
        children_by_parent = {}
        for i, room in hdr_rooms:
            parent = room.get('parent')
            if parent is not None:
                children_by_parent.setdefault(parent, []).append((i, room))
        flat_items = []
        for apt_idx, apt in apartments:
            flat_items.append((apt_idx, 0, False))
            kids = children_by_parent.get(apt.get('name', ''), [])
            for ki, (child_idx, _) in enumerate(kids):
                is_last = (ki == len(kids) - 1)
                flat_items.append((child_idx, 1, is_last))
        return flat_items, children_by_parent

    def _render_room_list_rows(self, flat_items, children_by_parent):
        """Render visible rows in two columns with scrollbar into ax_list."""
        # Row heights: parents get 1.0 unit, children get 0.5 unit
        pad_top, pad_bot = 0.01, 0.03
        usable = 1.0 - pad_top - pad_bot
        total = len(flat_items)

        # Compute total weight for scroll capacity (parents=1, children=0.5)
        weights = [0.5 if indent > 0 else 1.0 for (_, indent, _) in flat_items]

        # Determine how many items fit in ONE column
        unit_h = usable / 12.0
        cumulative = 0.0
        col_capacity = 0
        for w in weights:
            cumulative += w
            if cumulative * unit_h > usable:
                break
            col_capacity += 1
        col_capacity = max(1, col_capacity)

        # Two columns = 2x capacity
        visible_count = min(total, col_capacity * 2)

        # Scroll offset
        max_offset = max(0, total - visible_count)
        self.room_list_scroll_offset = max(0, min(self.room_list_scroll_offset, max_offset))
        self.ax_list.set_ylim(0, 1)

        # Column layout constants
        col_w     = 0.48          # each column width in axes fraction
        col_x     = [0.01, 0.51]  # left edges for col 0 and col 1
        divider_x = 0.50          # vertical divider between columns

        # Split visible items into two columns
        visible_slice = flat_items[self.room_list_scroll_offset:self.room_list_scroll_offset + visible_count]
        col0_items = visible_slice[:col_capacity]
        col1_items = visible_slice[col_capacity:]

        # Draw vertical divider between columns (if there are items in col 1)
        if col1_items:
            self.ax_list.plot(
                [divider_x, divider_x], [pad_bot, 1.0 - pad_top],
                color='#DDDDDD', linewidth=0.5,
                transform=self.ax_list.transAxes, clip_on=True)

        # Render each column
        for col_idx, col_items in enumerate([col0_items, col1_items]):
            cx = col_x[col_idx]
            y_cursor = 1.0 - pad_top

            for row_i, (room_idx, indent, is_last_child) in enumerate(col_items):
                room   = self.rooms[room_idx]
                name   = room.get('name', 'unnamed')
                is_sel   = (room_idx == self.selected_room_idx)
                is_multi = (room_idx in self.multi_selected_room_idxs)
                row_h   = unit_h * (0.5 if indent > 0 else 1.0)
                row_top = y_cursor
                row_bot = y_cursor - row_h
                row_mid = (row_top + row_bot) / 2
                y_cursor = row_bot

                box_pad = 0.005 if indent > 0 else 0.01
                box_inset = 0.001 if indent > 0 else 0.002
                if is_sel:
                    self.ax_list.add_patch(FancyBboxPatch(
                        (cx, row_bot + box_inset), col_w, row_h - box_inset * 2,
                        boxstyle=f'round,pad={box_pad}', facecolor='#FFE082', edgecolor='orange',
                        linewidth=1.0, transform=self.ax_list.transAxes, clip_on=True))
                elif is_multi:
                    self.ax_list.add_patch(FancyBboxPatch(
                        (cx, row_bot + box_inset), col_w, row_h - box_inset * 2,
                        boxstyle=f'round,pad={box_pad}', facecolor='#BBDEFB', edgecolor='#42A5F5',
                        linewidth=0.8, transform=self.ax_list.transAxes, clip_on=True))

                rtype    = room.get('room_type', '')
                type_tag = f" : {rtype}" if rtype else ""
                if indent > 0:
                    parent_name = room.get('parent', '')
                    short = name[len(parent_name) + 1:] if name.startswith(f"{parent_name}_") else name
                    text   = f"{short}{type_tag}"
                    color, fs, fw = ('#E65100' if is_sel else '#0D47A1'), 6.5, 'normal'

                    # Draw vertical connector line from this row up to the previous row
                    line_x = cx + 0.02 + 0.02
                    line_bot = row_mid
                    line_top = row_top + row_h * 0.5
                    self.ax_list.plot(
                        [line_x, line_x], [line_bot, line_top],
                        color='#90A4AE', linewidth=0.7, solid_capstyle='round',
                        transform=self.ax_list.transAxes, clip_on=True)
                    # Horizontal tick from vertical line to the label
                    tick_end = line_x + 0.02
                    self.ax_list.plot(
                        [line_x, tick_end], [row_mid, row_mid],
                        color='#90A4AE', linewidth=0.7, solid_capstyle='round',
                        transform=self.ax_list.transAxes, clip_on=True)

                    text_x = tick_end + 0.01
                else:
                    n_kids = len(children_by_parent.get(name, []))
                    suffix = f" ({n_kids})" if n_kids else ""
                    text   = f"{name}{suffix}{type_tag}"
                    color, fs, fw = ('#E65100' if is_sel else '#1B5E20'), 7, 'bold'
                    text_x = cx + 0.02

                self.ax_list.text(text_x, row_mid, text, fontsize=fs,
                                  fontweight=fw, color=color, va='center',
                                  transform=self.ax_list.transAxes, clip_on=True)
                self._room_list_hit_boxes.append((row_bot, row_top, cx, cx + col_w, room_idx))

        # Scrollbar
        if total > visible_count:
            pct = self.room_list_scroll_offset / max(1, max_offset)
            ind_h = visible_count / total
            ind_y = (1.0 - ind_h) * (1.0 - pct)
            self.ax_list.add_patch(FancyBboxPatch(
                (0.965, ind_y), 0.025, ind_h, boxstyle='round,pad=0.005',
                facecolor='#AAAAAA', edgecolor='none',
                transform=self.ax_list.transAxes, clip_on=True))
            self.ax_list.text(
                0.5, 0.01,
                f"\u2191\u2193 scroll  ({self.room_list_scroll_offset + 1}"
                f"-{min(self.room_list_scroll_offset + visible_count, total)} of {total})",
                fontsize=6, color='#888888', ha='center', va='bottom',
                transform=self.ax_list.transAxes)

    def _on_list_click(self, event):
        """Handle clicks on the saved rooms list to select a room.

        Ctrl+click toggles rooms in/out of the multi-selection set,
        allowing bulk room-type tagging.  Plain click reverts to
        single-select behaviour.
        """
        if event.inaxes != self.ax_list:
            return
        if event.xdata is None or event.ydata is None:
            return
        x = event.xdata
        y = event.ydata
        ctrl = event.key == 'control'
        for hit in self._room_list_hit_boxes:
            if len(hit) == 5:
                y_min, y_max, x_min, x_max, room_idx = hit
            else:
                y_min, y_max, room_idx = hit
                x_min, x_max = 0, 1
            if y_min <= y <= y_max and x_min <= x <= x_max:
                if ctrl:
                    # Ctrl+click: toggle room in multi-selection
                    if room_idx in self.multi_selected_room_idxs:
                        self.multi_selected_room_idxs.discard(room_idx)
                    else:
                        self.multi_selected_room_idxs.add(room_idx)
                    # Also add current primary selection to multi-set
                    if self.selected_room_idx is not None:
                        self.multi_selected_room_idxs.add(self.selected_room_idx)
                    n = len(self.multi_selected_room_idxs)
                    self._update_status(f"Multi-select: {n} room(s)", 'blue')
                else:
                    # Plain click: clear multi-selection, single-select
                    self.multi_selected_room_idxs.clear()
                    if self.selected_room_idx == room_idx:
                        self._deselect_room()
                    else:
                        self._select_room(room_idx)
                self._update_room_list()
                return

    # === SESSION PERSISTENCE ===================================================

    def _save_session(self):
        """Save all room boundaries and DF cache to JSON."""
        self.session_path.parent.mkdir(parents=True, exist_ok=True)
        data = {
            'image_dir':      str(self.image_dir),
            'df_thresholds':  self.DF_THRESHOLDS,
            'rooms':          self.rooms,
        }
        with open(self.session_path, 'w') as f:
            json.dump(data, f, indent=2)
        print(f"Session saved to {self.session_path}")
        # Remove stale CSV so the JSON is the single source of truth
        if self.csv_path.exists():
            self.csv_path.unlink()
            print(f"Removed stale CSV: {self.csv_path}")

    def _load_session(self):
        """Load room boundaries and cached DF results from JSON session or AOI files."""
        if self.session_path.exists():
            with open(self.session_path, 'r') as f:
                data = json.load(f)
            self.rooms = data.get('rooms', [])
            # df_thresholds from old sessions are ignored — now fixed per room type
            source = "session"
            cached = sum(1 for r in self.rooms if r.get('df_cache'))
            if cached:
                print(f"Restored DF cache for {cached}/{len(self.rooms)} rooms")
        elif self.aoi_dir.exists() and list(self.aoi_dir.glob('*.aoi')):
            self._load_from_aoi_files(self.aoi_dir)
            source = "AOI files"
        else:
            return

        renamed_count = self._enforce_unique_names()
        if renamed_count > 0:
            print(f"Renamed {renamed_count} rooms to ensure uniqueness")
            self._save_session()

        self._update_status(f"Loaded {len(self.rooms)} rooms from {source}", 'green')
        if hasattr(self, 'ax'):
            self._render_section(force_full=True)
        print(f"Loaded {len(self.rooms)} rooms from {source}")


    # === ARCHIVE ===============================================================

    def _on_close(self, event):
        """Handle editor window close."""
        pass

    # === EXPORT ================================================================

    def _show_progress(self, fraction: float, label: str = ""):
        """Update the export progress bar (0.0–1.0) and optional label."""
        if not hasattr(self, 'ax_progress'):
            return
        self.ax_progress.clear()
        self.ax_progress.set_xlim(0, 1)
        self.ax_progress.set_ylim(0, 1)
        self.ax_progress.axis('off')

        # Background track
        self.ax_progress.add_patch(FancyBboxPatch(
            (0, 0.1), 1.0, 0.8, boxstyle='round,pad=0.02',
            facecolor='#E0E0E0', edgecolor='#CCCCCC', linewidth=0.5,
        ))
        # Filled portion
        if fraction > 0:
            self.ax_progress.add_patch(FancyBboxPatch(
                (0, 0.1), max(0.01, min(fraction, 1.0)), 0.8,
                boxstyle='round,pad=0.02',
                facecolor='#4CAF50', edgecolor='none',
            ))
        # Label
        display = label or f"{fraction * 100:.0f}%"
        self.ax_progress.text(
            0.5, 0.5, display, fontsize=7, color='#333333',
            ha='center', va='center', fontweight='bold',
        )
        self.ax_progress.set_visible(True)

    def _hide_progress(self):
        """Hide the export progress bar."""
        if hasattr(self, 'ax_progress'):
            self.ax_progress.set_visible(False)

    def _on_export_report(self, event):
        """Export per-pixel illuminance and DF data for every room to Excel.

        The heavy work (HDR loading, rasterisation, Excel writing) runs in a
        background thread so the UI stays responsive.  A matplotlib timer
        polls for completion and updates the progress bar.
        """
        if self._hdr2wpd is None:
            self._update_status("Export unavailable - DF analysis not initialised", 'red')
            return
        if not self.rooms:
            self._update_status("No rooms to export", 'red')
            return
        if getattr(self, '_export_thread', None) and self._export_thread.is_alive():
            self._update_status("Export already in progress...", 'orange')
            return

        n_rooms = len(self.rooms)
        self._update_status(f"Exporting {n_rooms} rooms...", 'orange')
        self.btn_export.label.set_text('Exporting...')
        self._show_progress(0.0, f"0 / {n_rooms} rooms")
        self.fig.canvas.draw_idle()

        progress = {'done': 0, 'total': n_rooms, 'phase': 'pixels', 'error': None}
        self._export_thread = threading.Thread(
            target=self._export_worker, args=(progress,), daemon=True)
        self._export_thread.start()
        self._start_export_progress_poll(progress, n_rooms)

    @staticmethod
    def _compute_summary_for_hdr(hdr_name, hdr_path, rooms_for_hdr):
        """Load one HDR's DF image and compute per-room summary statistics.

        Designed to run in a ThreadPoolExecutor — the heavy work is the
        subprocess call to `pvalue` (I/O-bound, releases the GIL) and numpy
        array operations (also release the GIL).

        Args:
            rooms_for_hdr: list of (parent, name, room_type, verts, threshold_or_None,
                           child_verts_list).
                           child_verts_list is a list of vertex lists for sub-rooms to
                           exclude from this room's pixel set (mirrors the UI behaviour).
                           Pass an empty list for rooms that have no children.

        Returns:
            (summary_rows, pixel_chunks) where summary_rows is a list of dicts
            (one per room) and pixel_chunks is a list of
            (room_name, lux_array, df_pct_array).
            Child polygon pixels are excluded from parent room results in both outputs,
            so no pixel appears in both a parent room and its sub-rooms.
        """
        from skimage.draw import polygon as skimage_polygon

        df_img = Hdr2Wpd.load_df_image(hdr_path)
        if df_img is None:
            return [], []

        h, w = df_img.shape
        summary_rows = []
        pixel_chunks = []   # list of (room_name, lux_array, df_pct_array)
        for parent, name, room_type, verts, threshold, child_verts_list in rooms_for_hdr:
            xs = [int(round(v[0])) for v in verts]
            ys = [int(round(v[1])) for v in verts]
            rr, cc = skimage_polygon(ys, xs, shape=(h, w))
            if len(rr) == 0:
                continue

            if child_verts_list:
                # Build a boolean mask for the parent polygon then punch out each child,
                # so child pixels are never counted in the parent's results.
                mask = np.zeros((h, w), dtype=bool)
                mask[rr, cc] = True
                for child_verts in child_verts_list:
                    cxs = [int(round(v[0])) for v in child_verts]
                    cys = [int(round(v[1])) for v in child_verts]
                    crr, ccc = skimage_polygon(cys, cxs, shape=(h, w))
                    mask[crr, ccc] = False
                rr, cc = np.where(mask)

            n = len(rr)
            if n == 0:
                continue

            df_vals  = df_img[rr, cc]   # already DF% (e.g. 0.98 = 0.98%)
            lux_vals = df_vals * 100.0  # DF% → lux  (df% = lux/10000*100 → lux = df%*100)

            if threshold is not None:
                passing     = int((df_vals >= threshold).sum())
                passing_pct = round(passing / n * 100, 1)
            else:
                passing     = None
                passing_pct = None

            summary_rows.append({
                'HDR File':            hdr_name,
                'Parent':              parent,
                'Room':                name,
                'Room Type':           room_type or '',
                'Total Pixels':        n,
                'DF Threshold (%)':    threshold,
                'Pixels >= Threshold': passing,
                '% Area >= Threshold': passing_pct,
            })
            pixel_chunks.append((name, np.round(lux_vals, 2), np.round(df_vals, 4)))

        return summary_rows, pixel_chunks

    def _export_worker(self, progress: dict):
        """Background thread: compute per-room DF summary stats and write outputs.

        Groups rooms by HDR file and uses ThreadPoolExecutor to parallelise
        DF image loading (subprocess I/O) and rasterisation (numpy, GIL-free).

        Outputs written to WPD_DIR:
          - aoi_report_daylight.xlsx  — one summary row per room
          - <room_name>_pixels.csv    — one file per room with per-pixel lux + DF%
        """
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import pandas as pd
        import csv as csv_mod

        try:
            # Build HDR name → path lookup
            hdr_lookup = {entry['name']: entry['hdr_path'] for entry in self.hdr_files}

            # Build a lookup of room name → child vertex lists for exclusion.
            # Only rooms with children (i.e. LIVING/parent rooms) will have entries.
            children_verts_by_name = {}  # room_name → [child_verts, ...]
            for room in self.rooms:
                p = room.get('parent', '')
                if p and len(room.get('vertices', [])) >= 3:
                    children_verts_by_name.setdefault(p, []).append(room['vertices'])

            # Group rooms by HDR file; carry per-room threshold and child verts for exclusion
            hdr_room_groups = {}  # hdr_name → list of (parent, name, room_type, verts, threshold, child_verts_list)
            for room in self.rooms:
                verts    = room.get('vertices', [])
                hdr_name = room.get('hdr_file', '')
                if len(verts) < 3 or hdr_name not in hdr_lookup:
                    continue
                parent    = room.get('parent', '')
                name      = room.get('name', 'unnamed')
                room_type = self._effective_room_type(room)
                threshold = self._threshold_for_type(room_type)   # None if untyped
                # Child verts to subtract: only applicable to parent rooms
                child_verts_list = children_verts_by_name.get(name, [])
                hdr_room_groups.setdefault(hdr_name, []).append(
                    (parent, name, room_type or '', verts, threshold, child_verts_list))

            # Phase 1: Parallel extraction (pvalue + numpy are GIL-free)
            all_summary_rows = []
            all_pixel_chunks = []   # list of (room_name, lux_array, df_pct_array)
            max_workers = min(len(hdr_room_groups), 4) if hdr_room_groups else 1
            completed_rooms = 0
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = {
                    pool.submit(
                        self._compute_summary_for_hdr,
                        hdr_name, hdr_lookup[hdr_name], rooms_list
                    ): rooms_list
                    for hdr_name, rooms_list in hdr_room_groups.items()
                }
                for future in as_completed(futures):
                    summary_rows, pixel_chunks = future.result()
                    all_summary_rows.extend(summary_rows)
                    all_pixel_chunks.extend(pixel_chunks)
                    completed_rooms += len(futures[future])
                    progress['done'] = completed_rooms

            # Phase 2: Write outputs
            progress['phase'] = 'writing'
            output_dir = config.WPD_DIR
            output_dir.mkdir(parents=True, exist_ok=True)

            # Excel summary (one row per room)
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.table import Table, TableStyleInfo
            df = pd.DataFrame(all_summary_rows) if all_summary_rows else pd.DataFrame()
            output_path = output_dir / 'aoi_report_daylight.xlsx'
            df.to_excel(output_path, sheet_name='Room Summary', index=False)
            wb = load_workbook(output_path)
            ws = wb.active
            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4
            # Convert to a structured Excel table with filters
            if ws.max_row > 1:
                last_col = get_column_letter(ws.max_column)
                table_ref = f"A1:{last_col}{ws.max_row}"
                table = Table(displayName="RoomSummary", ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(table)
            wb.save(output_path)
            print(f"Report saved to {output_path}")

            # Per-room CSVs (one file per AOI, all pixel values)
            csv_subdir = output_dir / 'aoi_pixel_data'
            csv_subdir.mkdir(parents=True, exist_ok=True)
            for room_name, lux_vals, df_pct_vals in all_pixel_chunks:
                safe_name = room_name.replace('/', '_').replace('\\', '_')
                csv_path = csv_subdir / f"{safe_name}_pixels.csv"
                with open(csv_path, 'w', newline='') as f:
                    writer = csv_mod.writer(f)
                    writer.writerow(['Room', 'Illuminance (Lux)', 'Daylight Factor (%)'])
                    for lux, df_pct in zip(lux_vals, df_pct_vals):
                        writer.writerow([room_name, lux, df_pct])
            print(f"Per-room CSVs saved to {csv_subdir}")

            progress['phase'] = 'overlays'
            self._export_overlay_images(progress)
            progress['phase'] = 'done'
        except Exception as exc:
            progress['error'] = exc

    def _start_export_progress_poll(self, progress: dict, n_rooms: int):
        """Poll the background export thread and update the progress bar."""
        def _check():
            if self._export_thread.is_alive():
                if progress['phase'] == 'overlays':
                    self._show_progress(0.97, "Rendering overlay images...")
                elif progress['phase'] == 'writing':
                    self._show_progress(0.95, "Writing Excel file...")
                else:
                    frac = progress['done'] / max(progress['total'], 1) * 0.9
                    self._show_progress(frac, f"{progress['done']} / {progress['total']} rooms")
                self.fig.canvas.draw_idle()
                return
            self._export_poll_timer.stop()
            if progress['error']:
                self._update_status(f"Export failed: {progress['error']}", 'red')
                self._show_progress(1.0, "Export failed")
            else:
                self._show_progress(0.98, "Archiving outputs...")
                self.fig.canvas.draw_idle()
                archive_path = self._run_archive()
                if archive_path:
                    self._update_status(f"Export & archive complete → {archive_path.name}", 'green')
                else:
                    self._update_status("Export complete (archive failed)", 'orange')
                self._show_progress(1.0, "Complete")
            self.btn_export.label.set_text('Export & Archive')
            self.fig.canvas.draw_idle()
            hide = self.fig.canvas.new_timer(interval=4000)
            hide.single_shot = True
            hide.add_callback(lambda: (self._hide_progress(), self.fig.canvas.draw_idle()))
            hide.start()

        self._export_poll_timer = self.fig.canvas.new_timer(interval=300)
        self._export_poll_timer.add_callback(_check)
        self._export_poll_timer.start()

    # === ARCHIVE / EXTRACT =====================================================

    def _run_archive(self) -> Optional[Path]:
        """Zip the outputs directory into the archive folder. Returns the zip path or None."""
        from archilume.config import OUTPUTS_DIR, ARCHIVE_DIR
        try:
            ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
            timestamp = time.strftime('%Y%m%d_%H%M%S')
            zip_name  = f"archilume_export_{timestamp}"
            zip_path  = ARCHIVE_DIR / f"{zip_name}.zip"
            shutil.make_archive(str(ARCHIVE_DIR / zip_name), 'zip', str(OUTPUTS_DIR))
            print(f"Archive created: {zip_path}")
            return zip_path
        except Exception as e:
            print(f"Archive failed: {e}")
            return None

    def _on_extract_click(self, event):
        """Open a file picker to select a zip archive, extract it to outputs, and reload."""
        from archilume.config import OUTPUTS_DIR, ARCHIVE_DIR
        from tkinter import Tk, filedialog, messagebox
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        initial_dir = str(ARCHIVE_DIR) if ARCHIVE_DIR.exists() else str(Path.cwd())
        zip_path = filedialog.askopenfilename(
            title="Select archive to extract",
            initialdir=initial_dir,
            filetypes=[("Zip files", "*.zip"), ("All files", "*.*")],
        )
        if not zip_path:
            root.destroy()
            return
        confirmed = messagebox.askyesno(
            title="Overwrite outputs?",
            message=(
                f"Extract '{Path(zip_path).name}' and overwrite all files in:\n\n"
                f"{OUTPUTS_DIR}\n\n"
                "This will replace all current output data. Continue?"
            ),
            parent=root,
        )
        root.destroy()
        if not confirmed:
            return
        zip_path = Path(zip_path)
        try:
            # Clear outputs directory (the zip contains an outputs/ prefix so we
            # extract into the parent; wipe first to avoid stale files)
            if OUTPUTS_DIR.exists():
                shutil.rmtree(OUTPUTS_DIR)

            # Extract into parent — zip root is outputs/ so it lands at the right place
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(str(OUTPUTS_DIR.parent))
            print(f"Extracted archive: {zip_path} → {OUTPUTS_DIR}")

            # Full editor reload
            self.rooms = []
            self.hdr_files        = self._scan_hdr_files()
            self.current_hdr_idx  = 0
            self._rebuild_image_variants()
            self._load_session()
            self._reset_hover_state()
            self._update_hdr_list()
            self._update_room_list()
            self._render_section(reset_view=True, force_full=True)
            self._create_polygon_selector()
            self._update_status(f"Extracted and reloaded: {zip_path.name}", 'green')
        except Exception as e:
            self._update_status(f"Extract failed: {e}", 'red')
            print(f"Extract failed: {e}")
        self.fig.canvas.draw_idle()

    @staticmethod
    def _render_single_overlay(tiff_path, rooms_data, output_dir):
        """Render room boundary overlays onto a single TIFF and save.

        Designed to run in a ThreadPoolExecutor (I/O + PIL drawing).

        Args:
            tiff_path: Path to source TIFF image.
            rooms_data: list of (name, vertices, df_display_lines) tuples.
            output_dir: directory to save the overlay TIFF.
        """
        from PIL import Image, ImageDraw, ImageFont

        if not tiff_path.exists():
            return
        img = Image.open(tiff_path).convert('RGB')
        draw = ImageDraw.Draw(img)

        font_size = max(12, int(img.height * 0.012))
        font_size_small = max(10, int(font_size * 0.8))
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", font_size)
            font_sm = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", font_size_small)
        except (OSError, IOError):
            font = ImageFont.load_default()
            font_sm = font

        red = (255, 0, 0)
        black = (0, 0, 0)
        outline_w = max(1, font_size // 12)

        def _outlined_text(x, y, text, fnt):
            for ox in range(-outline_w, outline_w + 1):
                for oy in range(-outline_w, outline_w + 1):
                    if ox or oy:
                        draw.text((x + ox, y + oy), text, fill=black, font=fnt)
            draw.text((x, y), text, fill=red, font=fnt)

        for name, verts, df_lines in rooms_data:
            pts = [(int(round(v[0])), int(round(v[1]))) for v in verts]
            pts.append(pts[0])
            draw.line(pts, fill=red, width=1)

            centroid = HdrAoiEditor._polygon_label_point(verts)
            cx, cy = int(round(centroid[0])), int(round(centroid[1]))

            bbox = draw.textbbox((0, 0), name, font=font)
            tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
            _outlined_text(cx - tw // 2, cy - th // 2, name, font)

            y_offset = cy + th
            for line in df_lines:
                bbox_s = draw.textbbox((0, 0), line, font=font_sm)
                tw_s = bbox_s[2] - bbox_s[0]
                th_s = bbox_s[3] - bbox_s[1]
                _outlined_text(cx - tw_s // 2, y_offset, line, font_sm)
                y_offset += th_s + 2

        out_path = output_dir / f"{tiff_path.stem}_aoi_overlay.tiff"
        if out_path.exists():
            out_path.unlink()
        img.save(out_path)
        print(f"Overlay saved: {out_path}")

    def _export_overlay_images(self, progress: dict):
        """Render room boundary overlays onto each TIFF using ThreadPoolExecutor."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        output_dir = self.image_dir

        # Collect all (tiff_path, rooms_data) jobs
        jobs = []
        for entry in self.hdr_files:
            hdr_name = entry['name']
            rooms_on_hdr = [
                (r.get('name', ''), r['vertices'],
                 r.get('df_cache', {}).get('display_lines', []))
                for r in self.rooms
                if r.get('hdr_file') == hdr_name and len(r.get('vertices', [])) >= 3
            ]
            if not rooms_on_hdr:
                continue
            for tiff_path in entry.get('tiff_paths', []):
                jobs.append((tiff_path, rooms_on_hdr, output_dir))

        max_workers = min(len(jobs), 4) if jobs else 1
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(self._render_single_overlay, *job) for job in jobs]
            for future in as_completed(futures):
                future.result()  # propagate exceptions

    def _export_rooms_as_aoi_files(self) -> int:
        """Write editor rooms as .aoi files compatible with Hdr2Wpd.

        Each .aoi file follows the standard format with headers that
        Hdr2Wpd._group_aoi_by_view() and _process_daylight_view_groups()
        expect. Pixel→world conversion uses the coordinate map header.

        Returns:
            Number of .aoi files written.
        """
        # Parse pixel→world mapping parameters from the coordinate map header
        coord_map_path = self.aoi_dir / "pixel_to_world_coordinate_map.txt"
        vp_x = vp_y = vh = vv = 0.0
        img_w = img_h = 1
        if coord_map_path.exists():
            with open(coord_map_path, 'r') as f:
                header_lines = [f.readline() for _ in range(4)]
            # Line 0: # VIEW: VIEW= -vtl v -vp X Y Z ... -vh VH -vv VV
            view_line = header_lines[0]
            vp_match = re.search(r'-vp\s+([\d.-]+)\s+([\d.-]+)\s+([\d.-]+)', view_line)
            if vp_match:
                vp_x, vp_y = float(vp_match.group(1)), float(vp_match.group(2))
            vh_match = re.search(r'-vh\s+([\d.-]+)', view_line)
            vv_match = re.search(r'-vv\s+([\d.-]+)', view_line)
            if vh_match:
                vh = float(vh_match.group(1))
            if vv_match:
                vv = float(vv_match.group(1))
            # Line 1: # Image dimensions in pixels: width=W, height=H
            dim_line = header_lines[1]
            img_w = int(dim_line.split('width=')[1].split(',')[0])
            img_h = int(dim_line.split('height=')[1])

        self.aoi_dir.mkdir(parents=True, exist_ok=True)
        count = 0

        for room in self.rooms:
            name     = room.get('name', 'unnamed')
            hdr_name = room.get('hdr_file', '')
            verts    = room['vertices']
            if len(verts) < 3:
                continue

            # Derive associated view file from HDR name (e.g. model_plan_ffl_25300 → plan_ffl_25300.vp)
            ffl_match = re.search(r'plan_ffl_(\d+)', hdr_name)
            associated_vp = f"plan_ffl_{ffl_match.group(1)}.vp" if ffl_match else f"{hdr_name}.vp"

            # Compute centroid in pixel space (true centre of mass)
            centroid_pt = self._polygon_label_point(verts)
            cx, cy = centroid_pt[0], centroid_pt[1]

            # Build vertex lines: world_x world_y pixel_x pixel_y
            vertex_lines = []
            for px, py in verts:
                if vh > 0 and vv > 0:
                    world_x = vp_x + (px - img_w / 2) * (vh / img_w)
                    world_y = vp_y + (img_h / 2 - py) * (vv / img_h)
                else:
                    world_x, world_y = px, py
                vertex_lines.append(f"{world_x:.4f} {world_y:.4f} {int(round(px))} {int(round(py))}")

            content = "\n".join([
                f"AOI Points File: {name}",
                f"ASSOCIATED VIEW FILE: {associated_vp}",
                f"FFL z height(m): 0.0",
                f"CENTRAL x,y: {cx:.4f} {cy:.4f}",
                f"NO. PERIMETER POINTS {len(verts)}: x,y pixel_x pixel_y positions",
            ] + vertex_lines)

            clean_name = re.sub(r'[^\w\s-]', '', name).strip()
            clean_name = re.sub(r'[-\s]+', '_', clean_name)
            filepath = self.aoi_dir / f"{clean_name}.aoi"
            with open(filepath, 'w') as f:
                f.write(content)
            count += 1

        print(f"Wrote {count} .aoi files to {self.aoi_dir}")
        return count

    def export_room_boundaries_csv(self, output_path: Optional[Path] = None):
        """Export room boundaries as CSV.

        Format: name, parent, hdr_file, X_px Y_px, X_px Y_px, ...
        """
        if not self.rooms:
            return

        output_path = Path(output_path) if output_path else self.csv_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        rows = []
        for room in self.rooms:
            coord_strings = [f"X_{x:.3f} Y_{y:.3f}" for x, y in room['vertices']]
            parent   = room.get('parent') or ''
            hdr_file = room.get('hdr_file', '')
            rows.append([room['name'], parent, hdr_file] + coord_strings)

        max_cols = max(len(r) for r in rows)
        with open(output_path, 'w', newline='') as f:
            writer = csv.writer(f)
            for row in rows:
                row += [''] * (max_cols - len(row))
                writer.writerow(row)

        self._update_status(f"Exported CSV ({len(self.rooms)} rooms)", 'green')
        print(f"Exported room boundaries CSV to {output_path}")
